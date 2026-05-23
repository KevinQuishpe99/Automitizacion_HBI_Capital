"""Escritura del Excel de control Merge tras Notify (update_merge_control_workbook_after_notify)."""

from __future__ import annotations

import asyncio
from datetime import date
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock

from openpyxl import Workbook, load_workbook

from app.application.sharepoint_resolution import encode_graph_drive_path
from app.application.use_cases.merge_control_workbook_notify import update_merge_control_workbook_after_notify
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_COLUMNS,
    SHEET_NAME,
    apply_merge_control_worksheet_protection,
    merge_control_workbook_relative_path,
)


def _row2_for_workbook(
    *,
    estado: str,
    is_active: bool | str,
    title: str = "Prev",
    hist: str = "old/hist.xlsx",
    pdf: str = "old/mail.pdf",
) -> list:
    m = {c: None for c in MERGE_CONTROL_COLUMNS}
    m["Title"] = title
    m["EstadoProceso"] = estado
    m["IsActive"] = is_active
    m["HistoricalFilePath"] = hist
    m["EmailPdfPath"] = pdf
    m["MergeOutputCount"] = 0
    m["MergeSkippedCount"] = 0
    m["LastErrorUserMessage"] = ""
    m["LastErrorNextAction"] = ""
    m["CreatedAtProceso"] = ""
    m["LastUpdatedAtProceso"] = ""
    return [m[c] for c in MERGE_CONTROL_COLUMNS]


def _control_workbook_bytes(*, estado: str, is_active: bool | str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(list(MERGE_CONTROL_COLUMNS))
    ws.append(_row2_for_workbook(estado=estado, is_active=is_active))
    apply_merge_control_worksheet_protection(ws)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class _RecordingGraph:
    def __init__(self, download_bytes: bytes) -> None:
        self._download = download_bytes
        self.get_endpoints: list[str] = []
        self.put_calls: list[tuple[str, bytes, str | None]] = []

    async def get_bytes(self, endpoint: str) -> bytes:
        self.get_endpoints.append(endpoint)
        return self._download

    async def put_bytes(self, endpoint: str, data: bytes, content_type: str | None = None) -> None:
        self.put_calls.append((endpoint, data, content_type))


def test_notify_updates_merge_control_workbook_when_completed():
    async def run():
        initial = _control_workbook_bytes(estado="VACIO", is_active=False)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "site1",
            "drive1",
            historical_file_path="NEW/hist.xlsx",
            email_pdf_path="NEW/mail.pdf",
            report_d=date(2026, 5, 12),
        )
        assert out.merge_control_updated is True
        assert out.merge_control_error_code is None
        assert len(g.put_calls) == 1

    asyncio.run(run())


def test_notify_writes_historical_file_path_and_email_pdf_path_to_control_row():
    async def run():
        initial = _control_workbook_bytes(estado="", is_active=False)
        g = _RecordingGraph(initial)
        hist = "HIST/cartera.xlsx"
        pdf = "05 EMAIL/out.pdf"
        await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path=hist,
            email_pdf_path=pdf,
            report_d=date(2026, 3, 1),
        )
        _, uploaded, _ = g.put_calls[0]
        wb = load_workbook(BytesIO(uploaded), data_only=True)
        try:
            ws = wb[SHEET_NAME]
            col_hist = MERGE_CONTROL_COLUMNS.index("HistoricalFilePath") + 1
            col_pdf = MERGE_CONTROL_COLUMNS.index("EmailPdfPath") + 1
            assert ws.cell(2, col_hist).value == hist
            assert ws.cell(2, col_pdf).value == pdf
        finally:
            wb.close()

    asyncio.run(run())


def test_notify_sets_merge_control_status_pending_asientos():
    async def run():
        initial = _control_workbook_bytes(estado="CONSOLIDADO", is_active=False)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 1, 2),
        )
        assert out.merge_control_status == "PENDIENTE_ASIENTOS"
        _, uploaded, _ = g.put_calls[0]
        wb = load_workbook(BytesIO(uploaded), data_only=True)
        try:
            ws = wb[SHEET_NAME]
            c_est = MERGE_CONTROL_COLUMNS.index("EstadoProceso") + 1
            assert ws.cell(2, c_est).value == "PENDIENTE_ASIENTOS"
        finally:
            wb.close()

    asyncio.run(run())


def test_notify_does_not_overwrite_active_pending_merge_control():
    async def run():
        initial = _control_workbook_bytes(estado="PENDIENTE_ASIENTOS", is_active=True)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="other/hist.xlsx",
            email_pdf_path="other/mail.pdf",
            report_d=date(2026, 6, 1),
        )
        assert out.merge_control_updated is False
        assert out.merge_control_error_code == "merge_control_active_process_exists"
        assert g.put_calls == []

    asyncio.run(run())


def test_notify_allows_overwrite_when_control_is_vacio():
    async def run():
        initial = _control_workbook_bytes(estado="VACIO", is_active=True)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 4, 4),
        )
        assert out.merge_control_updated is True
        assert len(g.put_calls) == 1

    asyncio.run(run())


def test_notify_allows_overwrite_when_control_is_consolidado_inactive():
    async def run():
        initial = _control_workbook_bytes(estado="CONSOLIDADO", is_active=False)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 4, 4),
        )
        assert out.merge_control_updated is True

    asyncio.run(run())


def test_notify_requires_email_pdf_path_before_registering_merge_control():
    async def run():
        g = MagicMock()
        g.get_bytes = AsyncMock(side_effect=AssertionError("get_bytes no debe llamarse sin PDF"))

        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="   ",
            report_d=date(2026, 2, 2),
        )
        assert out.merge_control_updated is False
        assert out.merge_control_error_code == "missing_email_pdf_path_for_merge_control"
        g.get_bytes.assert_not_called()

    asyncio.run(run())


def test_notify_preserves_or_reapplies_sheet_protection():
    async def run():
        initial = _control_workbook_bytes(estado="CANCELADO", is_active=False)
        g = _RecordingGraph(initial)
        await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 7, 7),
        )
        _, uploaded, _ = g.put_calls[0]
        wb = load_workbook(BytesIO(uploaded), data_only=False)
        try:
            ws = wb[SHEET_NAME]
            assert ws.protection.sheet is True
            for c in range(1, len(MERGE_CONTROL_COLUMNS) + 1):
                assert ws.cell(1, c).protection.locked is True
                assert ws.cell(2, c).protection.locked is True
        finally:
            wb.close()

    asyncio.run(run())


def test_notify_result_includes_merge_control_fields():
    """Contrato de resultado alineado con el router (campos merge_control_*)."""
    result = {
        "status": "ok",
        "message": "Ejecutado con éxito",
        "report_date": "12/05/2026",
        "historical_file_path": "HIST/x.xlsx",
        "historical_file_source": "explicit",
        "historico_excel_path": "HIST/x.xlsx",
        "rows_included": 1,
        "subject": "S",
        "attachments_count": 0,
        "email_pdf_path": "MAIL/x.pdf",
        "email_pdf_error": None,
        "graph_sendmail_http_status": 202,
        "mail_sender": "a@b.com",
        "mail_to": "c@d.com",
        "merge_control_updated": True,
        "merge_control_file_path": "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL/control_merge_pdfs.xlsx",
        "merge_control_status": "PENDIENTE_ASIENTOS",
        "merge_control_warning": None,
        "merge_control_error_code": None,
    }
    for k in (
        "merge_control_updated",
        "merge_control_file_path",
        "merge_control_status",
        "merge_control_warning",
        "merge_control_error_code",
    ):
        assert k in result


def test_notify_existing_response_fields_remain_backward_compatible():
    """Los clientes que lean solo campos históricos siguen encontrándolos."""
    result = {
        "status": "ok",
        "message": "Ejecutado con éxito",
        "report_date": "12/05/2026",
        "historical_file_path": "HIST/x.xlsx",
        "historical_file_source": "explicit",
        "historico_excel_path": "HIST/x.xlsx",
        "rows_included": 1,
        "subject": "S",
        "attachments_count": 0,
        "email_pdf_path": "MAIL/x.pdf",
        "email_pdf_error": None,
        "graph_sendmail_http_status": 202,
        "mail_sender": "a@b.com",
        "mail_to": "c@d.com",
        "merge_control_updated": True,
        "merge_control_file_path": "CONTROL/control_merge_pdfs.xlsx",
        "merge_control_status": "PENDIENTE_ASIENTOS",
        "merge_control_warning": None,
        "merge_control_error_code": None,
    }
    assert result["historical_file_path"]
    assert result["historico_excel_path"] == result["historical_file_path"]
    assert result["historical_file_source"] == "explicit"
    assert "subject" in result
    assert "attachments_count" in result


def test_merge_control_respects_graph_merge_control_workbook_path_env(monkeypatch):
    async def run():
        custom = "CUSTOM/path/control_merge_pdfs.xlsx"
        monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", custom)
        initial = _control_workbook_bytes(estado="VACIO", is_active=False)
        g = _RecordingGraph(initial)
        await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 8, 8),
        )
        enc = encode_graph_drive_path(custom)
        assert g.get_endpoints and enc in g.get_endpoints[0]
        put_ep, _, _ = g.put_calls[0]
        assert enc in put_ep

    asyncio.run(run())


def test_completed_example_merge_control_updated_true_document():
    """Ejemplo de documento completed con merge_control_updated true (referencia humana / contrato)."""
    doc = {
        "status": "completed",
        "result": {
            "status": "ok",
            "merge_control_updated": True,
            "merge_control_file_path": (
                "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL/control_merge_pdfs.xlsx"
            ),
            "merge_control_status": "PENDIENTE_ASIENTOS",
            "merge_control_warning": None,
            "merge_control_error_code": None,
            "historical_file_path": "HIST/validacion.xlsx",
            "email_pdf_path": "05 EMAIL/ABONOS BANCO BOGOTA 2026-05-12.pdf",
        },
    }
    assert doc["result"]["merge_control_updated"] is True
    assert doc["result"]["merge_control_status"] == "PENDIENTE_ASIENTOS"


def test_warning_example_active_merge_process_document():
    """Ejemplo de resultado con proceso activo (no se sobrescribe el control)."""
    doc = {
        "status": "completed",
        "result": {
            "status": "ok",
            "merge_control_updated": False,
            "merge_control_error_code": "merge_control_active_process_exists",
            "merge_control_warning": (
                "Ya existe un proceso pendiente de consolidación de PDFs en el archivo de control. "
                "Termine o cancele el proceso pendiente antes de registrar uno nuevo."
            ),
            "merge_control_status": "PENDIENTE_ASIENTOS",
        },
    }
    assert doc["result"]["merge_control_updated"] is False
    assert doc["result"]["merge_control_error_code"] == "merge_control_active_process_exists"


def test_env_default_path_when_graph_merge_control_workbook_path_unset(monkeypatch):
    monkeypatch.delenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", raising=False)
    p = merge_control_workbook_relative_path()
    assert p.endswith("control_merge_pdfs.xlsx")
    assert "00 CONTROL" in p


def test_notify_allows_overwrite_when_estado_error_merge_but_inactive():
    """ERROR_MERGE solo bloquea si IsActive es verdadero."""

    async def run():
        initial = _control_workbook_bytes(estado="ERROR_MERGE", is_active=False)
        g = _RecordingGraph(initial)
        out = await update_merge_control_workbook_after_notify(
            g,
            "s",
            "d",
            historical_file_path="h.xlsx",
            email_pdf_path="p.pdf",
            report_d=date(2026, 1, 1),
        )
        assert out.merge_control_updated is True

    asyncio.run(run())
