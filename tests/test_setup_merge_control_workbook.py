"""Tests del setup idempotente del Excel control_merge_pdfs.xlsx."""

import asyncio
import io

import httpx
import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.adapters.primary.http.deps import init_graph_client
from app.adapters.primary.http.routers.payment_validation import router
from app.application.use_cases import setup_merge_control_workbook as smcw
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_COLUMNS,
    MERGE_CONTROL_WORKBOOK_RELATIVE_PATH,
    SECURITY_WARNING,
    SHEET_NAME,
    TABLE_DISPLAY_NAME,
    _build_workbook_bytes,
    setup_merge_control_workbook,
)


def _http_error(status: int) -> httpx.HTTPStatusError:
    req = httpx.Request("GET", "https://graph.microsoft.com/test")
    resp = httpx.Response(status, request=req, text=f"HTTP {status} body")
    return httpx.HTTPStatusError("err", request=req, response=resp)


class MockGraphSetup:
    def __init__(self) -> None:
        self.files: dict[str, bytes] = {}
        self.put_count = 0
        self.post_json_count = 0
        self.web_url = "https://comwareec.sharepoint.com/sites/x/control_merge_pdfs.xlsx"
        self.get_item_meta: dict = {"webUrl": self.web_url}
        self.put_response: dict = {"webUrl": self.web_url}
        self.force_403_on_put = False
        self.force_403_on_list = False

    def _is_file_content(self, endpoint: str) -> bool:
        low = endpoint.lower()
        return ":/content" in endpoint and "control_merge_pdfs" in low

    def _is_file_item(self, endpoint: str) -> bool:
        low = endpoint.lower()
        return endpoint.endswith(":") and "control_merge_pdfs" in low and "/content" not in low

    async def get(self, endpoint: str, params=None):
        if self.force_403_on_list:
            raise _http_error(403)
        if self._is_file_item(endpoint):
            return dict(self.get_item_meta)
        return {"value": []}

    async def get_bytes(self, endpoint: str, params=None):
        if self.force_403_on_list:
            raise _http_error(403)
        if self._is_file_content(endpoint):
            if MERGE_CONTROL_WORKBOOK_RELATIVE_PATH not in self.files:
                raise _http_error(404)
            return self.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH]
        raise _http_error(404)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        if self.force_403_on_put:
            raise _http_error(403)
        self.put_count += 1
        self.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH] = content
        return dict(self.put_response)

    async def post_json(self, endpoint: str, body: dict):
        self.post_json_count += 1
        return {}, 201


@pytest.fixture
def mock_resolve(monkeypatch):
    async def _fake_resolve(client, site_search, drive_name, path):
        return {
            "site_id": "site-1",
            "drive_id": "drive-1",
            "path_encoded": "enc",
            "file_path": path,
        }

    monkeypatch.setattr(smcw, "resolve_sharepoint_path", _fake_resolve)


@pytest.fixture(autouse=True)
def env_site(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST_SITE")
    monkeypatch.setenv("GRAPH_SHAREPOINT_DRIVE_NAME", "")


def test_setup_merge_control_workbook_creates_file_when_missing(mock_resolve):
    g = MockGraphSetup()
    out = asyncio.run(setup_merge_control_workbook(g))
    assert out["created"] is True
    assert out["status"] == "success"
    assert g.put_count == 1
    assert MERGE_CONTROL_WORKBOOK_RELATIVE_PATH in g.files


def test_setup_merge_control_workbook_does_not_overwrite_existing_file(mock_resolve):
    g = MockGraphSetup()
    asyncio.run(setup_merge_control_workbook(g))
    first_puts = g.put_count
    out2 = asyncio.run(setup_merge_control_workbook(g))
    assert out2["created"] is False
    assert g.put_count == first_puts


def test_setup_merge_control_workbook_creates_expected_sheet_table_and_columns(mock_resolve):
    g = MockGraphSetup()
    asyncio.run(setup_merge_control_workbook(g))
    wb = openpyxl.load_workbook(io.BytesIO(g.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH]), data_only=True)
    try:
        assert SHEET_NAME in wb.sheetnames
        ws = wb[SHEET_NAME]
        assert list(ws.tables) == [TABLE_DISPLAY_NAME]
        for i, name in enumerate(MERGE_CONTROL_COLUMNS, start=1):
            assert ws.cell(1, i).value == name
    finally:
        wb.close()


def test_setup_merge_control_workbook_initializes_single_control_row(mock_resolve):
    g = MockGraphSetup()
    asyncio.run(setup_merge_control_workbook(g))
    wb = openpyxl.load_workbook(io.BytesIO(g.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH]), data_only=True)
    try:
        ws = wb[SHEET_NAME]
        assert ws.cell(2, 1).value in ("", None)
        assert ws.cell(2, 2).value == "VACIO"
        assert str(ws.cell(2, 3).value).lower() == "false"
        assert ws.cell(2, 6).value == 0
        assert ws.cell(2, 7).value == 0
    finally:
        wb.close()


def test_setup_merge_control_workbook_applies_generate_like_full_sheet_protection(mock_resolve):
    g = MockGraphSetup()
    asyncio.run(setup_merge_control_workbook(g))
    wb = openpyxl.load_workbook(io.BytesIO(g.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH]), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        assert ws.protection.sheet is True
        assert not ws.protection.password
        for r in (1, 2):
            for c in range(1, 12):
                assert ws.cell(r, c).protection.locked is True
    finally:
        wb.close()


def test_setup_merge_control_workbook_header_cells_locked():
    b = _build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        for c in range(1, 12):
            assert ws.cell(1, c).protection.locked is True
    finally:
        wb.close()


def test_setup_merge_control_workbook_control_row_cells_locked():
    b = _build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        for c in range(1, 12):
            assert ws.cell(2, c).protection.locked is True
    finally:
        wb.close()


def test_setup_merge_control_workbook_all_data_cells_locked_for_manual_editing():
    b = _build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        assert ws.protection.sheet is True
        for r in (1, 2):
            for c in range(1, 12):
                assert ws.cell(r, c).protection.locked is True
    finally:
        wb.close()


def test_setup_merge_control_workbook_protection_has_no_password():
    b = _build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(b), data_only=False)
    try:
        assert wb[SHEET_NAME].protection.password in (None, "")
    finally:
        wb.close()


def test_setup_merge_control_workbook_returns_weburl_when_graph_returns_weburl(mock_resolve):
    g = MockGraphSetup()
    g.put_response = {"webUrl": "https://custom.web/url.xlsx"}
    out = asyncio.run(setup_merge_control_workbook(g))
    assert out["file_url"] == "https://custom.web/url.xlsx"


def test_setup_merge_control_workbook_handles_existing_valid_file(mock_resolve):
    g = MockGraphSetup()
    asyncio.run(setup_merge_control_workbook(g))
    puts_after_create = g.put_count
    out = asyncio.run(setup_merge_control_workbook(g))
    assert out["created"] is False
    assert out["excel_protection_applied"] is True
    assert g.put_count == puts_after_create


def test_setup_merge_control_workbook_reports_security_warning_when_sharepoint_permissions_not_supported(
    mock_resolve,
):
    g = MockGraphSetup()
    out = asyncio.run(setup_merge_control_workbook(g))
    assert out["sharepoint_permission_applied"] is False
    assert out["security_warning"] == SECURITY_WARNING


def test_setup_merge_control_workbook_repairs_missing_table(mock_resolve):
    g = MockGraphSetup()
    base = _build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(base), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        if TABLE_DISPLAY_NAME in ws.tables:
            del ws.tables[TABLE_DISPLAY_NAME]
        buf = io.BytesIO()
        wb.save(buf)
        g.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH] = buf.getvalue()
    finally:
        wb.close()

    g.put_count = 0
    out = asyncio.run(setup_merge_control_workbook(g))
    assert out["created"] is False
    assert g.put_count == 1
    assert any(str(w).startswith("repaired:") for w in out["warnings"])
    wb2 = openpyxl.load_workbook(io.BytesIO(g.files[MERGE_CONTROL_WORKBOOK_RELATIVE_PATH]), data_only=False)
    try:
        assert TABLE_DISPLAY_NAME in wb2[SHEET_NAME].tables
    finally:
        wb2.close()


def build_app() -> FastAPI:
    app = FastAPI()
    init_graph_client(MockGraphSetup())
    app.include_router(router)
    return app


def test_setup_merge_control_workbook_handles_graph_403_with_clear_error(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST_SITE")

    async def fake_resolve(client, site_search, drive_name, path):
        return {"site_id": "s", "drive_id": "d", "path_encoded": "e", "file_path": path}

    monkeypatch.setattr(smcw, "resolve_sharepoint_path", fake_resolve)

    class G403(MockGraphSetup):
        async def get_bytes(self, endpoint: str, params=None):
            if self._is_file_content(endpoint):
                raise _http_error(403)
            raise _http_error(404)

    init_graph_client(G403())
    app = FastAPI()
    app.include_router(router)
    client = TestClient(app, raise_server_exceptions=False)
    res = client.post("/graph/sharepoint/payment-validation/setup/merge-control-workbook")
    assert res.status_code == 403
    body = res.json()
    assert "detail" in body
    d = body["detail"]
    assert "user_message" in d
    assert "No se pudo crear el archivo de control" in d["user_message"]
