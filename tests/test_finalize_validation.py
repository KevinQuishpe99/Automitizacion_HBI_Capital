import asyncio
import httpx
import io
import os
from datetime import date
from pathlib import Path
from urllib.parse import unquote

import openpyxl
import pytest

from app.application.services.review_schema import (
    CasosPagoCols,
    ControlCols,
    DISTRIB_LEGACY_HEADER_ALIASES,
    DistribucionCols,
    EstadoLinea,
    EstadoPago,
    ReviewSheets,
    ValidarPago,
)
from app.application.use_cases.payment_validation_finalize import (
    SECRETARY_FIRST_DATA_ROW,
    SECRETARY_HEADER_ROW,
    SECRETARY_HEADERS,
    SECRETARY_INSTRUCTION,
    SECRETARY_SHEET,
    SECRETARY_TITLE,
    _parse_extract_date_from_filename,
    _secretary_link_visible_text,
    finalize_payment_validation,
)


def _norm_fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    rgb = fill.fgColor.rgb
    if rgb is None:
        return None
    s = str(rgb).upper()
    return s[-6:] if len(s) > 6 else s


def _cell_has_hyperlink(cell) -> bool:
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return False
    return bool(getattr(hl, "target", None))

# Formato real HBI de ruta interna al drive (adjuntos por correo)
HBI_CLIENTS_ROOT = (
    "INFORMACION CREDITOS-CLIENTES/01 COMWARE AUTOMATIZACION - INFORMACION CREDITOS CLIENTES"
)
HBI_EXPECTED_RUTA = (
    "INFORMACION CREDITOS-CLIENTES/01 COMWARE AUTOMATIZACION - INFORMACION CREDITOS CLIENTES/"
    "EQUINORTE/CREDITO # 264/Extracto 2026-01-01 CREDITO # 264.pdf"
)
HBI_EXTRACT_PDF = "Extracto 2026-01-01 CREDITO # 264.pdf"


class MockGraphClient:
    def __init__(self):
        self.children: list = []
        self.downloaded_files: dict[str, bytes] = {}
        self.uploaded_files: dict[str, bytes] = {}
        self.fail_upload_path: str | None = None
        self.requested_file_paths: list[str] = []
        self.requested_endpoints: list[str] = []
        self.put_calls: list[tuple[str, str]] = []
        self.children_overrides: dict[str, list] | None = None
        self.children_get_paths: list[str] = []
        self.drive_item_responses: dict[str, dict] = {}
        # Si se define, get_bytes lanza HTTP 404 cuando esta subcadena aparece en la ruta (simula Graph).
        self.get_bytes_404_substring: str | None = None
        self.dynamic_folder_children: dict[str, list[dict]] = {}

    async def post_json(self, endpoint: str, body: dict):
        if "children" not in endpoint or "/root:/" not in endpoint:
            return {}, 201
        raw = endpoint.split("/root:/", 1)[1].rsplit(":/children", 1)[0]
        path_decoded = unquote(raw)
        name = str(body.get("name", "")).strip()
        item = {
            "name": name,
            "folder": {},
            "webUrl": f"https://mock.invalid/asientos/{path_decoded}/{name}",
        }
        self.dynamic_folder_children.setdefault(path_decoded, []).append(item)
        return item, 201

    async def get(self, endpoint, params=None):
        ep_base = endpoint.split("?", 1)[0]
        if ep_base in self.drive_item_responses:
            return self.drive_item_responses[ep_base]
        if "children" in endpoint and "/root:/" in endpoint:
            raw = endpoint.split("/root:/", 1)[1].rsplit(":/children", 1)[0]
            path_decoded = unquote(raw)
            self.children_get_paths.append(path_decoded)
            if self.children_overrides and path_decoded in self.children_overrides:
                return {"value": self.children_overrides[path_decoded]}
            builtin = {
                "clientes/CLI": [
                    {"name": "CRED", "folder": {}},
                    {"name": "CRED_A", "folder": {}},
                    {"name": "CRED_B", "folder": {}},
                ],
                "clientes/CLI/CRED": [
                    {"name": "ASIENTOS CONTABLES", "folder": {}, "webUrl": "https://mock.invalid/asientos/CRED"},
                ],
                "clientes/CLI/CRED_A": [],
                "clientes/CLI/CRED_B": [],
            }
            extra = self.dynamic_folder_children.get(path_decoded, [])
            if path_decoded in builtin:
                return {"value": list(builtin[path_decoded]) + extra}
            if not path_decoded.startswith("clientes/"):
                return {"value": self.children}
            return {"value": extra}
        if "/drives" in endpoint:
            return {"value": [{"id": "dummy_drive", "name": "DRIVE"}]}
        if "/sites" in endpoint:
            return {"value": [{"id": "dummy_site"}]}
        return {}

    async def get_bytes(self, endpoint, params=None):
        self.requested_endpoints.append(endpoint)
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        self.requested_file_paths.append(file_path)
        if self.get_bytes_404_substring and self.get_bytes_404_substring in file_path:
            req = httpx.Request("GET", "https://graph.microsoft.com/v1.0/mock")
            resp = httpx.Response(404, request=req)
            raise httpx.HTTPStatusError("404", request=req, response=resp)
        if file_path in self.downloaded_files:
            return self.downloaded_files[file_path]
        return b""

    async def put_bytes(self, endpoint, content, content_type):
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        if self.fail_upload_path and self.fail_upload_path in file_path:
            raise Exception("Mock Upload Error")
        self.put_calls.append((endpoint, content_type))
        self.uploaded_files[file_path] = content
        return {"id": "new_file_id", "webUrl": f"https://mock.invalid/web?path={file_path}"}

    async def delete(self, endpoint: str) -> None:
        return None


def set_env_vars():
    os.environ["GRAPH_BANK_PAYMENTS_FILE_PATH"] = "banco.xlsx"
    os.environ["GRAPH_CLIENTS_BASE_PATH"] = "clientes"
    os.environ["GRAPH_SHAREPOINT_SITE_SEARCH"] = "SITIO"
    os.environ["GRAPH_SHAREPOINT_DRIVE_NAME"] = "DRIVE"
    os.environ["GRAPH_PAYMENT_VALIDATION_REVIEW_PATH"] = "revision"
    os.environ["GRAPH_VALIDATION_FILE_PREFIX"] = "val"
    os.environ["GRAPH_PAYMENT_VALIDATION_HISTORY_PATH"] = "history"


def make_distrib_row(
    id_pago="ID1",
    cliente="CLI",
    credito="CRED",
    fecha_banco=None,
    fecha_limite=None,
    valor_int=70,
    abono_k=20,
    mora=10,
    monto_banco=100,
    estado=EstadoLinea.VALIDAR,
    validar_pago=ValidarPago.SI,
    observacion="",
    extract_route="clientes/CLI/CRED/extractos/e1.pdf",
    link_tabla_display="Ver tabla",
    tabla_hyperlink_target: str | None = None,
    ruta_pdf_internal: str = "",
    ruta_unidad_credito: str = "",
):
    fecha_banco = fecha_banco or date(2026, 5, 10)
    fecha_limite = fecha_limite or date(2026, 5, 15)

    def _default_ruta_unidad_credito() -> str:
        r = str(ruta_pdf_internal or extract_route or "").strip().replace("\\", "/")
        if not r or "/" not in r:
            return f"clientes/{cliente}/{credito}"
        parent = r.rsplit("/", 1)[0]
        if parent.lower().endswith("/extractos"):
            parent = parent.rsplit("/", 1)[0]
        return parent

    def _n(x):
        if isinstance(x, (int, float)):
            return float(x)
        return 0.0

    total = _n(valor_int) + _n(abono_k) + _n(mora)
    vals = {
        DistribucionCols.ID_PAGO: id_pago,
        DistribucionCols.CLIENTE: cliente,
        DistribucionCols.CREDITO: credito,
        DistribucionCols.MONTO_BANCO: monto_banco,
        DistribucionCols.FECHA_BANCO: fecha_banco,
        DistribucionCols.FECHA_LIMITE: fecha_limite,
        DistribucionCols.DIAS_MORA: 0,
        DistribucionCols.VALOR_EXTRACTO: 100,
        DistribucionCols.APLICAR_A_EXTRACTO: valor_int,
        DistribucionCols.MORA_A_APLICAR: abono_k,
        DistribucionCols.OTROS_VALORES: mora,
        DistribucionCols.TOTAL_APLICADO: total,
        DistribucionCols.SALDO_POR_ASIGNAR: 0,
        DistribucionCols.ESTADO_PAGO: estado,
        DistribucionCols.VALIDAR_PAGO: validar_pago,
        DistribucionCols.OBSERVACION: observacion,
        DistribucionCols.LINK_EXTRACTO: extract_route,
        DistribucionCols.LINK_TABLA: link_tabla_display,
        DistribucionCols.LINK_CARPETA_CREDITO: "",
        DistribucionCols.RUTA: ruta_pdf_internal,
        DistribucionCols.RUTA_UNIDAD_CREDITO: ruta_unidad_credito or _default_ruta_unidad_credito(),
        DistribucionCols.RUTA_TABLA_AMORTIZACION: "",
        DistribucionCols.CREDITO_NORMALIZADO: "",
    }
    row = [vals[c] for c in DistribucionCols.HEADERS]
    return row, tabla_hyperlink_target


def create_review_workbook(
    procesar="SI",
    estado="EN_REVISION",
    casos_data=None,
    distrib_specs=None,
    include_control=True,
    include_distrib=True,
):
    """
    distrib_specs: lista de tuplas (row_list_distrib_cols, optional_tabla_hyperlink_url)
    o lista de filas legacy de 14 columnas (sin links) para tests viejos.
    """
    wb = openpyxl.Workbook()
    ws_ctrl = wb.active
    ws_ctrl.title = ReviewSheets.CONTROL if include_control else "Otro"
    if include_control:
        ws_ctrl.append([ControlCols.CAMPO, ControlCols.VALOR])
        ws_ctrl.append([ControlCols.ROW_PROCESAR, procesar])
        ws_ctrl.append([ControlCols.ROW_ESTADO, estado])

    ws_casos = wb.create_sheet(ReviewSheets.CASOS_PAGO)
    ws_casos.append(CasosPagoCols.HEADERS)
    if casos_data:
        for row in casos_data:
            if len(row) == len(CasosPagoCols.HEADERS):
                ws_casos.append(row)
            else:
                ws_casos.append(
                    [
                        row[0] if len(row) > 0 else "",
                        row[1] if len(row) > 1 else None,
                        row[2] if len(row) > 2 else "",
                        row[3] if len(row) > 3 else "",
                        row[4] if len(row) > 4 else 0,
                        row[5] if len(row) > 5 else "",
                    ]
                )
    else:
        ws_casos.append(["ID1", None, "CLI", "concepto", 100, ""])

    ws_dist = wb.create_sheet(ReviewSheets.DISTRIBUCION if include_distrib else "OtroDist")
    if include_distrib:
        ws_dist.append(DistribucionCols.HEADERS)
        specs = distrib_specs or []
        for spec in specs:
            tabla_hl = None
            if isinstance(spec, tuple) and len(spec) == 2:
                row_vals, tabla_hl = spec
            else:
                row_vals = spec
            if len(row_vals) == len(DistribucionCols.HEADERS) - 1:
                row_vals = list(row_vals[:14]) + [ValidarPago.SI] + list(row_vals[14:])
            elif len(row_vals) != len(DistribucionCols.HEADERS):
                legacy = list(row_vals)
                row_vals = [
                    legacy[0] if len(legacy) > 0 else "",
                    legacy[1] if len(legacy) > 1 else "",
                    legacy[4] if len(legacy) > 4 else "",
                    legacy[2] if len(legacy) > 2 else 100,
                    legacy[3] if len(legacy) > 3 else None,
                    legacy[5] if len(legacy) > 5 else None,
                    legacy[6] if len(legacy) > 6 else 0,
                    legacy[7] if len(legacy) > 7 else 0,
                    legacy[8] if len(legacy) > 8 else 0,
                    legacy[9] if len(legacy) > 9 else 0,
                    legacy[10] if len(legacy) > 10 else 0,
                    legacy[11] if len(legacy) > 11 else 0,
                    legacy[12] if len(legacy) > 12 else 0,
                    legacy[13] if len(legacy) > 13 else "",
                    legacy[14] if len(legacy) > 14 else "",
                    legacy[15] if len(legacy) > 15 else "",
                    legacy[16] if len(legacy) > 16 else "",
                    legacy[17] if len(legacy) > 17 else "",
                    legacy[18] if len(legacy) > 18 else (legacy[17] if len(legacy) > 17 else ""),
                ]
                row_vals = row_vals + [""] * (len(DistribucionCols.HEADERS) - len(row_vals))
            r = ws_dist.max_row + 1
            for c, v in enumerate(row_vals, start=1):
                ws_dist.cell(r, c, v)
            le_col = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
            lt_col = DistribucionCols.HEADERS.index(DistribucionCols.LINK_TABLA) + 1
            if tabla_hl:
                ctab = ws_dist.cell(r, lt_col)
                ctab.hyperlink = tabla_hl
                if not ctab.value:
                    ctab.value = "Tabla"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _append_distrib_ruta_column(wb_bytes: bytes, ruta_by_row: dict[int, str]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = wb[ReviewSheets.DISTRIBUCION]
    col = ws.max_column + 1
    ws.cell(1, col, DistribucionCols.RUTA)
    for r, val in ruta_by_row.items():
        ws.cell(r, col, val)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _run(coro):
    return asyncio.run(coro)


def test_finalize_mock_graph_contract_has_no_legacy_methods():
    client = MockGraphClient()
    assert not hasattr(client, "get_file_content")
    assert not hasattr(client, "upload_file")


def test_finalize_fails_if_no_control_sheet():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(include_control=False)
        with pytest.raises(ValueError, match="missing_control_sheet"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_fails_if_no_distribucion_sheet():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(include_distrib=False)
        with pytest.raises(ValueError, match="missing_distribucion_sheet"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_fails_if_procesar_no():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            procesar="NO", distrib_specs=[(r, None)]
        )
        with pytest.raises(ValueError, match="process_not_approved"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_fails_if_estado_not_en_revision():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            estado="PROCESADO", distrib_specs=[(r, None)]
        )
        with pytest.raises(ValueError, match="invalid_control_state"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_missing_control_state():
    async def run_test():
        set_env_vars()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ReviewSheets.CONTROL
        ws.append([ControlCols.CAMPO, ControlCols.VALOR])
        ws.append([ControlCols.ROW_PROCESAR, ControlCols.VAL_PROCESAR_SI])
        ws_c = wb.create_sheet(ReviewSheets.CASOS_PAGO)
        ws_c.append(CasosPagoCols.HEADERS)
        ws_c.append(["ID1", None, "CLI", "c", 100, ""])
        ws_d = wb.create_sheet(ReviewSheets.DISTRIBUCION)
        ws_d.append(DistribucionCols.HEADERS)
        r, _ = make_distrib_row()
        ws_d.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        client = MockGraphClient()
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        with pytest.raises(ValueError, match="missing_control_state"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_fails_on_forbidden_states():
    """Solo REVISION_MANUAL impide finalizar; ATRASADO y el resto pueden cerrarse si cumplen reglas."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(estado=EstadoPago.REVISION_MANUAL)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="estado_pago_no_finalizable"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_validar_requires_intereses_mora_filled():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(valor_int=100, abono_k=0, mora="")
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="missing_mora"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_validar_accepts_zero_intereses_mora():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(mora=0, valor_int=100, abono_k=0)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def _legacy_distrib_visible_headers() -> list[str]:
    headers = list(DistribucionCols.HEADERS)
    for i, label in enumerate(headers):
        for legacy, canonical in DISTRIB_LEGACY_HEADER_ALIASES.items():
            if label == canonical:
                headers[i] = legacy
    return headers


def test_finalize_accepts_legacy_estado_linea_header():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(mora=0, valor_int=100, abono_k=0)
        wb_bytes = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb[ReviewSheets.DISTRIBUCION]
        estado_col = DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_LINEA) + 1
        ws.cell(1, estado_col, "Estado línea")
        out = io.BytesIO()
        wb.save(out)
        client.downloaded_files["revision/val_latest.xlsx"] = out.getvalue()
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_accepts_legacy_distrib_column_headers():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(mora=0, valor_int=100, abono_k=0)
        wb_bytes = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb[ReviewSheets.DISTRIBUCION]
        for c, h in enumerate(_legacy_distrib_visible_headers(), start=1):
            ws.cell(1, c, h)
        out = io.BytesIO()
        wb.save(out)
        client.downloaded_files["revision/val_latest.xlsx"] = out.getvalue()
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_treats_empty_mora_as_invalid_for_validar():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r_bad, _ = make_distrib_row(mora="")
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r_bad, None)])
        with pytest.raises(ValueError, match="missing_mora"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_reprogramar_expired_allows_blank_mora():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            fecha_banco=date(2026, 5, 22),
            fecha_limite=date(2026, 5, 15),
            valor_int=70,
            abono_k=20,
            mora=10,
            monto_banco=100,
            estado=EstadoLinea.REPROGRAMAR,
            observacion="Rep",
            extract_route="clientes/CLI/CRED/extractos/e1.pdf",
        )
        client.children_overrides = {
            "clientes/CLI/CRED/extractos": [
                {"name": "e1.pdf", "file": {}, "webUrl": "https://mock/e1.pdf"},
            ],
        }
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_no_validar_expired_allows_blank_mora_with_observation():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            fecha_banco=date(2026, 5, 22),
            fecha_limite=date(2026, 5, 15),
            valor_int=0,
            abono_k=0,
            mora="",
            estado=EstadoLinea.NO_VALIDAR,
            observacion="Excluido",
            extract_route="",
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_validar_requires_valor_intereses():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(valor_int="", abono_k=0, mora=0)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="missing_valor_intereses"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_validar_requires_abono_k():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(abono_k="")
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="missing_abono_k"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_fails_on_amount_mismatch():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(valor_int=150, abono_k=0, mora=0)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="amount_mismatch"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_adelantado_positive_total_succeeds():
    """ADELANTADO + Validar Pago = SI con importes reales cuadra como pago aplicable."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            estado=EstadoPago.ADELANTADO,
            validar_pago=ValidarPago.SI,
            valor_int=49118143,
            abono_k=881857,
            mora=0,
            monto_banco=50000000,
            extract_route="clientes/CLI/CRED/extractos/e1.pdf",
        )
        client.children_overrides = {
            "clientes/CLI/CRED/extractos": [
                {"name": "e1.pdf", "file": {}, "webUrl": "https://mock/e1.pdf"},
            ],
        }
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            casos_data=[["ID1", None, "CLI", "concepto", 50000000, ""]],
            distrib_specs=[(r, None)],
        )
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        assert res["validated_rows"] >= 1

    _run(run_test())


def test_finalize_legacy_reprogramar_migrates_with_positive_total():
    """REPROGRAMAR migra a ADELANTADO + SI y ya no exige total en cero."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            valor_int=50,
            abono_k=30,
            mora=20,
            monto_banco=100,
            estado=EstadoLinea.REPROGRAMAR,
            observacion="Adelanto",
            extract_route="clientes/CLI/CRED/extractos/e1.pdf",
        )
        client.children_overrides = {
            "clientes/CLI/CRED/extractos": [
                {"name": "e1.pdf", "file": {}, "webUrl": "https://mock/e1.pdf"},
            ],
        }
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_adelantado_amount_mismatch():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            estado=EstadoPago.ADELANTADO,
            validar_pago=ValidarPago.SI,
            valor_int=49118143,
            abono_k=881857,
            mora=0,
            monto_banco=50000000,
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            casos_data=[["ID1", None, "CLI", "concepto", 100, ""]],
            distrib_specs=[(r, None)],
        )
        with pytest.raises(ValueError, match="amount_mismatch"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_adelantado_registers_in_pagos_adelantados():
    async def run_test():
        from app.application.use_cases.setup_merge_control_workbook import MERGE_CONTROL_FOLDER_RELATIVE_PATH
        from app.application.use_cases.setup_payment_followup_workbooks import (
            ADELANTADOS_COLUMNS,
            FILENAME_ADELANTADOS,
            SHEET_PENDIENTES,
            _build_followup_workbook_bytes,
        )

        set_env_vars()
        path_ad = f"{MERGE_CONTROL_FOLDER_RELATIVE_PATH}/{FILENAME_ADELANTADOS}"
        client = MockGraphClient()
        client.downloaded_files[path_ad] = _build_followup_workbook_bytes(columns=ADELANTADOS_COLUMNS)
        r, _ = make_distrib_row(
            estado=EstadoPago.ADELANTADO,
            validar_pago=ValidarPago.SI,
            valor_int=49118143,
            abono_k=881857,
            mora=0,
            monto_banco=50000000,
            extract_route="clientes/CLI/CRED/extractos/e1.pdf",
        )
        client.children_overrides = {
            "clientes/CLI/CRED/extractos": [
                {"name": "e1.pdf", "file": {}, "webUrl": "https://mock/e1.pdf"},
            ],
        }
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            casos_data=[["ID1", None, "CLI", "concepto", 50000000, ""]],
            distrib_specs=[(r, None)],
        )
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        assert path_ad in client.uploaded_files
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[path_ad]))
        ws = wb[SHEET_PENDIENTES]
        assert ws.cell(2, 1).value == "ID1"
        assert not res.get("payment_followup_warnings")

    _run(run_test())


def test_finalize_validar_requires_positive_total():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(valor_int=0, abono_k=0, mora=0)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="validar_requires_positive_total"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_no_validar_requires_observation():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            valor_int=0,
            abono_k=0,
            mora=0,
            estado=EstadoLinea.NO_VALIDAR,
            observacion="",
            extract_route="",
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        with pytest.raises(ValueError, match="no_validar_requires_observation"):
            await finalize_payment_validation(client, "val_latest.xlsx")

    _run(run_test())


def test_finalize_pendiente_mora_atrasado_can_complete():
    """PENDIENTE_MORA → ATRASADO: se puede finalizar si cuadra."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(estado=EstadoLinea.PENDIENTE_MORA)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        assert client.uploaded_files

    _run(run_test())


def test_finalize_validar_parcial_skips_amortization_and_cuadratura():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            valor_int=80,
            abono_k=0,
            mora=20,
            estado=EstadoLinea.VALIDAR_PARCIAL,
            observacion="Parcial",
            # VALIDAR_PARCIAL migra a INCOMPLETO + SI: misma regla de ruta/extracto que líneas validadas.
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        assert res["amortization_updated"] is False

    _run(run_test())


def test_finalize_multi_credito_same_pago_cuadra():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r1, h1 = make_distrib_row(
            id_pago="ID1",
            credito="CRED_A",
            valor_int=60,
            abono_k=0,
            mora=0,
            extract_route="clientes/CLI/CRED_A/e.pdf",
        )
        r2, h2 = make_distrib_row(
            id_pago="ID1",
            credito="CRED_B",
            valor_int=40,
            abono_k=0,
            mora=0,
            extract_route="clientes/CLI/CRED_B/e.pdf",
        )
        r1 = list(r1)
        r2 = list(r2)
        r1[2] = 100
        r2[2] = 100
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            distrib_specs=[(r1, h1), (r2, h2)]
        )
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        assert res["validated_rows"] == 2

    _run(run_test())


def test_finalize_respects_validation_file_path_when_provided():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        exact_path = "revision/subcarpeta/val_manual.xlsx"
        r, _ = make_distrib_row()
        client.downloaded_files[exact_path] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(
            client,
            validation_file_path=exact_path,
            process_date=date(2026, 5, 10),
        )
        assert res["status"] == "success"
        assert res["validation_file_path"] == exact_path
        assert "history/cartera_validada_2026-05-10.xlsx" in client.uploaded_files

    _run(run_test())


def test_finalize_recalculates_totals_from_editable_cells():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        tcol = DistribucionCols.HEADERS.index("Total aplicado") + 1
        ws.cell(2, tcol, "=9999")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_does_not_update_amortization_tables():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, th = make_distrib_row()
        tab_url = "https://x.sharepoint.com/u?u=https://host/root:/clientes/CLI/CRED/tabla.xlsx:/"
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, tab_url)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert not any("tabla" in k and k.endswith(".xlsx") for k in client.uploaded_files)

    _run(run_test())


def test_finalize_does_not_clean_banco_bogota():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["bank_cleaned"] is False
        assert "banco.xlsx" not in client.uploaded_files

    _run(run_test())


def test_finalize_valid_excel_no_longer_fails_on_amortization_row_errors():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"

    _run(run_test())


def test_finalize_historical_workbook_has_exact_ruta_column():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = wb[ReviewSheets.DISTRIBUCION]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert headers.count(DistribucionCols.RUTA) == 1

    _run(run_test())


def test_finalize_historical_workbook_populates_ruta_for_validar_rows():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        path = "clientes/CLI/CRED/docs/ex.pdf"
        r, _ = make_distrib_row(extract_route=path)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = wb[ReviewSheets.DISTRIBUCION]
        cmap = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        assert ws.cell(2, cmap[DistribucionCols.RUTA]).value == path

    _run(run_test())


def test_finalize_missing_ruta_for_validar_blocks():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(extract_route="")
        wb_b = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(wb_b))
        ws = wb[ReviewSheets.DISTRIBUCION]
        le = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, le, "")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        with pytest.raises(ValueError, match="missing_extract_route"):
            await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))

    _run(run_test())


def test_finalize_creates_secretary_support_workbook():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        assert SECRETARY_SHEET in wb.sheetnames

    _run(run_test())


def test_secretary_workbook_contains_only_validar_rows():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        rv, _ = make_distrib_row(id_pago="V1")
        rnv, _ = make_distrib_row(
            id_pago="NV",
            valor_int=0,
            abono_k=0,
            mora=0,
            estado=EstadoLinea.NO_VALIDAR,
            observacion="x",
            extract_route="",
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            distrib_specs=[(rv, None), (rnv, None)]
        )
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        # filas 1–2 banner, 3 headers, 4 datos V1, 5 total
        assert ws.max_row == 5
        assert ws.cell(SECRETARY_FIRST_DATA_ROW, SECRETARY_HEADERS.index("ID Pago") + 1).value == "V1"

    _run(run_test())


def test_secretary_workbook_has_expected_columns():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        headers = [ws.cell(SECRETARY_HEADER_ROW, c).value for c in range(1, len(SECRETARY_HEADERS) + 1)]
        assert headers == SECRETARY_HEADERS
        assert "Estado asiento contable" not in headers
        assert "Observación" not in headers
        tot = 70 + 20 + 10
        assert ws.cell(SECRETARY_FIRST_DATA_ROW, SECRETARY_HEADERS.index("Total validado") + 1).value == tot

    _run(run_test())


def test_secretary_workbook_includes_extract_link():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, tab = make_distrib_row()
        ext_url = "https://host/path/root:/clientes/CLI/CRED/x.pdf:/"
        wb0 = create_review_workbook(distrib_specs=[(r, tab)])
        wb = openpyxl.load_workbook(io.BytesIO(wb0))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        cell = ws.cell(2, c)
        cell.value = "Abrir"
        cell.hyperlink = ext_url
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wbs = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        wss = wbs[SECRETARY_SHEET]
        col = SECRETARY_HEADERS.index("Link extracto") + 1
        assert _cell_has_hyperlink(wss.cell(SECRETARY_FIRST_DATA_ROW, col))

    _run(run_test())


def test_secretary_workbook_includes_amortization_table_link():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        tab_url = "https://z/root:/clientes/CLI/CRED/tabla.xlsx:/"
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, tab_url)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wbs = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        col = SECRETARY_HEADERS.index("Link tabla amortización") + 1
        assert _cell_has_hyperlink(wbs[SECRETARY_SHEET].cell(SECRETARY_FIRST_DATA_ROW, col))

    _run(run_test())


def test_secretary_workbook_includes_asientos_folder_link_when_http_url():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/CLI": [{"name": "CRED", "folder": {}}],
            "clientes/CLI/CRED": [
                {
                    "name": "ASIENTOS CONTABLES",
                    "folder": {},
                    "webUrl": "https://sharepoint/asientos/CRED",
                }
            ],
        }
        r, tab = make_distrib_row()
        tab_url = "https://z/root:/clientes/CLI/CRED/tabla.xlsx:/"
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, tab_url)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wbs = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        wss = wbs[SECRETARY_SHEET]
        col = SECRETARY_HEADERS.index("Link carpeta asientos contables") + 1
        c = wss.cell(SECRETARY_FIRST_DATA_ROW, col)
        assert _cell_has_hyperlink(c)
        assert "CRED" in str(c.value)

    _run(run_test())


def test_secretary_workbook_asientos_link_after_finalize_provisions_folder():
    """Finalize crea ASIENTOS CONTABLES CRED {n} y deja enlace en soporte secretaría."""
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/CLI": [{"name": "CRED", "folder": {}}],
            "clientes/CLI/CRED": [],
        }
        r, tab = make_distrib_row()
        tab_url = "https://z/root:/clientes/CLI/CRED/tabla.xlsx:/"
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, tab_url)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert "clientes/CLI/CRED" in client.dynamic_folder_children
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wbs = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        wss = wbs[SECRETARY_SHEET]
        col = SECRETARY_HEADERS.index("Link carpeta asientos contables") + 1
        c = wss.cell(SECRETARY_FIRST_DATA_ROW, col)
        assert c.value
        assert _cell_has_hyperlink(c)
        assert "CRED" in str(c.value)

    _run(run_test())


def test_finalize_writes_ruta_asientos_contables_on_historical_distrib():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(mora=0, valor_int=100, abono_k=0)
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = wb[ReviewSheets.DISTRIBUCION]
        hdr = 1
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == DistribucionCols.ID_PAGO:
                hdr = r
                break
        cmap = {
            str(ws.cell(hdr, c).value or "").strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(hdr, c).value
        }
        assert DistribucionCols.RUTA_ASIENTOS_CONTABLES in cmap
        dr = hdr + 1
        rv = str(ws.cell(dr, cmap[DistribucionCols.RUTA_ASIENTOS_CONTABLES]).value or "")
        assert "ASIENTOS CONTABLES CRED" in rv
        assert rv.startswith("clientes/CLI/CRED/")

    _run(run_test())


def test_finalize_fails_when_ruta_unidad_credito_missing():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(mora=0, valor_int=100, abono_k=0, ruta_unidad_credito="")
        wb_bytes = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        ws = wb[ReviewSheets.DISTRIBUCION]
        col_uc = DistribucionCols.HEADERS.index(DistribucionCols.RUTA_UNIDAD_CREDITO) + 1
        ws.cell(2, col_uc, "")
        out = io.BytesIO()
        wb.save(out)
        client.downloaded_files["revision/val_latest.xlsx"] = out.getvalue()
        with pytest.raises(ValueError, match="missing_ruta_unidad_credito"):
            await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))

    _run(run_test())


def test_secretary_workbook_has_professional_title_and_instruction():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        assert ws.cell(1, 1).value == SECRETARY_TITLE
        assert ws.cell(2, 1).value == SECRETARY_INSTRUCTION
        assert ws.cell(2, 1).font.size == 12

    _run(run_test())


def test_secretary_workbook_headers_start_on_row_3():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        got = [ws.cell(SECRETARY_HEADER_ROW, c).value for c in range(1, len(SECRETARY_HEADERS) + 1)]
        assert got == SECRETARY_HEADERS

    _run(run_test())


def test_secretary_workbook_data_starts_on_row_4():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(id_pago="PID-ROW4")
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        assert ws.cell(SECRETARY_FIRST_DATA_ROW, SECRETARY_HEADERS.index("ID Pago") + 1).value == "PID-ROW4"

    _run(run_test())


def test_secretary_workbook_has_freeze_panes_after_credito():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        assert ws.freeze_panes == f"C{SECRETARY_FIRST_DATA_ROW}"

    _run(run_test())


def test_secretary_workbook_has_no_autofilter():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        assert ws.auto_filter.ref is None

    _run(run_test())


def test_secretary_workbook_formats_money_columns():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        r = SECRETARY_FIRST_DATA_ROW
        for name in ("Total validado",):
            c = SECRETARY_HEADERS.index(name) + 1
            assert ws.cell(r, c).number_format == "#,##0.00"

    _run(run_test())


def test_secretary_workbook_formats_date_columns():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        r = SECRETARY_FIRST_DATA_ROW
        for name in ("Fecha banco", "Fecha límite"):
            c = SECRETARY_HEADERS.index(name) + 1
            assert ws.cell(r, c).number_format == "yyyy-mm-dd"

    _run(run_test())


def test_secretary_workbook_preserves_clickable_links():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        tab_url = "https://z/root:/clientes/CLI/CRED/tabla.xlsx:/"
        ext_url = "https://host/path/root:/clientes/CLI/CRED/x.pdf:/"
        wb0 = create_review_workbook(distrib_specs=[(r, tab_url)])
        wb = openpyxl.load_workbook(io.BytesIO(wb0))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        cell = ws.cell(2, c)
        cell.value = "Abrir"
        cell.hyperlink = ext_url
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wbs = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        wss = wbs[SECRETARY_SHEET]
        r = SECRETARY_FIRST_DATA_ROW
        assert _cell_has_hyperlink(wss.cell(r, SECRETARY_HEADERS.index("Link extracto") + 1))
        assert _cell_has_hyperlink(wss.cell(r, SECRETARY_HEADERS.index("Link tabla amortización") + 1))
        carpeta = wss.cell(r, SECRETARY_HEADERS.index("Link carpeta asientos contables") + 1)
        assert _cell_has_hyperlink(carpeta)

    _run(run_test())


def test_secretary_workbook_uses_descriptive_link_text_with_credito():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(credito="264")
        tab_url = "https://z/root:/clientes/CLI/264/tabla.xlsx:/"
        ext_url = "https://host/path/root:/clientes/CLI/264/x.pdf:/"
        wb0 = create_review_workbook(distrib_specs=[(r, tab_url)])
        wb = openpyxl.load_workbook(io.BytesIO(wb0))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        cell = ws.cell(2, c)
        cell.value = "Abrir"
        cell.hyperlink = ext_url
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wss = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))[SECRETARY_SHEET]
        r = SECRETARY_FIRST_DATA_ROW
        assert wss.cell(r, SECRETARY_HEADERS.index("Link extracto") + 1).value == "Ver extracto crédito 264"
        assert wss.cell(r, SECRETARY_HEADERS.index("Link tabla amortización") + 1).value == "Ver tabla crédito 264"
        hl = wss.cell(r, SECRETARY_HEADERS.index("Link extracto") + 1).hyperlink
        assert hl is not None and str(hl.target).startswith("https://")

    _run(run_test())


def test_secretary_link_visible_text_non_numeric_credito():
    assert _secretary_link_visible_text("extracto", "ACIMOR", "CLI") == "Ver extracto ACIMOR"


def test_secretary_workbook_header_navy_style():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        hcell = ws.cell(SECRETARY_HEADER_ROW, 1)
        assert _norm_fill_rgb(hcell) == "002060"
        assert hcell.font.bold is True
        font_rgb = getattr(hcell.font.color, "rgb", None) if hcell.font.color else None
        assert font_rgb in ("FFFFFF", "00FFFFFF", None)

    _run(run_test())


def test_secretary_workbook_client_block_borders():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r1, _ = make_distrib_row(id_pago="A1", cliente="CLI-A", credito="100")
        r2, _ = make_distrib_row(id_pago="A2", cliente="CLI-A", credito="101")
        r3, _ = make_distrib_row(id_pago="B1", cliente="CLI-B", credito="200")
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            distrib_specs=[(r1, None), (r2, None), (r3, None)]
        )
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        r_first = SECRETARY_FIRST_DATA_ROW
        r_second = r_first + 1
        r_third = r_first + 2
        assert ws.cell(r_first, 1).border.top.style == "thin"
        assert ws.cell(r_second, 1).border.top.style == "thin"
        assert ws.cell(r_third, 1).border.top.style == "medium"
        assert ws.cell(r_second, 1).border.bottom.style == "medium"
        assert ws.cell(r_third, 1).border.bottom.style == "medium"

    _run(run_test())


def test_secretary_workbook_keeps_expected_columns():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        headers = [ws.cell(SECRETARY_HEADER_ROW, c).value for c in range(1, len(SECRETARY_HEADERS) + 1)]
        assert headers == SECRETARY_HEADERS

    _run(run_test())


def test_secretary_workbook_total_row_with_sum():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        sec_key = next(k for k in client.uploaded_files if "soporte_asientos_contables_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[sec_key]))
        ws = wb[SECRETARY_SHEET]
        tcol = SECRETARY_HEADERS.index("Total validado") + 1
        trow = SECRETARY_FIRST_DATA_ROW + 1
        assert ws.cell(trow, 1).value == "Total validado general"
        f = ws.cell(trow, tcol).value
        assert isinstance(f, str) and f.startswith("=SUM(")

    _run(run_test())


def test_finalize_result_includes_secretary_file_url():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["secretary_file_url"] and "soporte_asientos_contables" in res["secretary_file_url"]

    _run(run_test())


def test_finalize_result_flags_amortization_and_bank_false():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["amortization_updated"] is False
        assert res["bank_cleaned"] is False

    _run(run_test())


def test_finalize_does_not_use_raw_credit_as_folder_path():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/EQUINORTE": [{"name": "CREDITO # 264", "folder": {}}],
            "clientes/EQUINORTE/CREDITO # 264": [],
        }
        r, _ = make_distrib_row(
            cliente="EQUINORTE",
            credito="264",
            extract_route="clientes/EQUINORTE/CREDITO # 264/e.pdf",
        )
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert "clientes/EQUINORTE/264" not in client.children_get_paths

    _run(run_test())


def test_finalize_falls_back_to_latest_prefixed_review_file_when_none_provided():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.children = [
            {"name": "~$val_temp.xlsx", "lastModifiedDateTime": "2026-05-10T09:00:00Z"},
            {"name": "otro_archivo.xlsx", "lastModifiedDateTime": "2026-05-10T09:30:00Z"},
            {"name": "val_old.xlsx", "lastModifiedDateTime": "2026-05-09T09:00:00Z"},
            {"name": "val_new.xlsx", "lastModifiedDateTime": "2026-05-10T10:00:00Z", "file": {}},
        ]
        client.downloaded_files["revision/val_new.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, process_date=date(2026, 5, 10))
        assert res["validation_file"] == "val_new.xlsx"

    _run(run_test())


def test_finalize_uses_existing_ruta_when_present():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(extract_route="")
        base = create_review_workbook(distrib_specs=[(r, None)])
        client.downloaded_files["revision/val_latest.xlsx"] = _append_distrib_ruta_column(
            base, {2: "INFORMACION/precargada/extracto.pdf"}
        )
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = wb[ReviewSheets.DISTRIBUCION]
        cmap = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        assert ws.cell(2, cmap[DistribucionCols.RUTA]).value == "INFORMACION/precargada/extracto.pdf"

    _run(run_test())


def test_finalize_extracts_ruta_from_hyperlink_root_path():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(extract_route="")
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = (
            "https://tenant.sharepoint.com/sites/x/_layouts/15/embed.aspx?"
            "u=https%3A%2F%2Fhost%2Froot%3A%2FINFORMACION%20CREDITOS%2FCLI%2FCRED%2Fex%2Epdf%3A%2F"
        )
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, c).value): c
            for c in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, c).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert rv.endswith("ex.pdf")
        assert "INFORMACION CREDITOS" in str(rv)

    _run(run_test())


def test_finalize_extracts_filename_from_sharepoint_doc_url():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/CLI": [{"name": "CRED", "folder": {}}],
            "clientes/CLI/CRED": [
                {"name": "ASIENTOS CONTABLES", "folder": {}},
                {"name": "Extracto Feb Exacto 1.pdf", "file": {}},
            ],
        }
        r, _ = make_distrib_row(extract_route="")
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = (
            "https://tenant.sharepoint.com/sites/acme/_layouts/15/Doc.aspx"
            "?sourcedoc=%7B111%7D&file=Extracto%20Feb%20Exacto%201.pdf"
        )
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert rv.endswith("Extracto Feb Exacto 1.pdf")
        assert rv.startswith("clientes/CLI/CRED/")

    _run(run_test())


def test_finalize_resolves_ruta_from_graph_drive_item_url():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.drive_item_responses["/drives/dummy_drive/items/ITEM99"] = {
            "name": "desde_graph.pdf",
            "parentReference": {"path": "/drives/dummy_drive/root:/clientes/CLI/CRED"},
        }
        r, _ = make_distrib_row(extract_route="")
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = "https://graph.microsoft.com/v1.0/drives/dummy_drive/items/ITEM99"
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert rv == "clientes/CLI/CRED/desde_graph.pdf"

    _run(run_test())


def test_finalize_resolves_extract_route_by_exact_filename_in_credit_folder():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/CLI": [{"name": "CRED", "folder": {}}],
            "clientes/CLI/CRED": [
                {"name": "Otro.pdf", "file": {}},
                {"name": "Extracto Match Exacto.pdf", "file": {}},
            ],
        }
        r, _ = make_distrib_row(extract_route="")
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = "https://x/Doc.aspx?file=Extracto%20Match%20Exacto.pdf"
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert str(rv).endswith("Extracto Match Exacto.pdf")

    _run(run_test())


def test_finalize_resolves_extract_route_by_legacy_style_pdf_scan():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/CLI": [{"name": "CRED", "folder": {}}],
            "clientes/CLI/CRED": [
                {"name": "Extracto 2026-05-01 Obligacion.pdf", "file": {}},
                {"name": "Extracto 2026-05-10 Obligacion.pdf", "file": {}},
            ],
        }
        r, _ = make_distrib_row(extract_route="", fecha_banco=date(2026, 5, 10))
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = "https://tenant.sharepoint.com/sites/x/_layouts/15/Doc.aspx?id=%2Fsites%2Fx"
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert "2026-05-10" in str(rv)

    _run(run_test())


def test_finalize_allows_blank_ruta_for_non_validar():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(
            valor_int=0,
            abono_k=0,
            mora=0,
            estado=EstadoLinea.NO_VALIDAR,
            validar_pago=ValidarPago.NO,
            observacion="rep",
            extract_route="",
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["status"] == "success"
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        if DistribucionCols.RUTA in cmap:
            assert w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value in (None, "")

    _run(run_test())


def test_finalize_preserves_link_extracto_and_link_tabla():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        ext_u = "https://keep-extract/root:/clientes/CLI/CRED/x.pdf:/"
        tab_u = "https://keep-tabla/root:/clientes/CLI/CRED/tabla.xlsx:/"
        r, _ = make_distrib_row(extract_route="")
        raw = create_review_workbook(distrib_specs=[(r, tab_u)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        ce = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ct = DistribucionCols.HEADERS.index(DistribucionCols.LINK_TABLA) + 1
        ws.cell(2, ce, "Ver extracto")
        ws.cell(2, ce).hyperlink = ext_u
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws2 = w[ReviewSheets.DISTRIBUCION]
        assert str(ws2.cell(2, ce).hyperlink.target) == ext_u
        assert str(ws2.cell(2, ct).hyperlink.target) == tab_u

    _run(run_test())


def test_finalize_does_not_import_or_depend_on_legacy_validate_payment_report():
    src = Path(__file__).resolve().parents[1] / "app" / "application" / "use_cases" / "payment_validation_finalize.py"
    assert "validate_payment_report" not in src.read_text(encoding="utf-8")


def test_finalize_does_not_reactivate_amortization_or_bank_cleanup():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row()
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(distrib_specs=[(r, None)])
        res = await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        assert res["amortization_updated"] is False
        assert res["bank_cleaned"] is False
        assert not any("tabla" in k and k.endswith(".xlsx") for k in client.uploaded_files)
        assert "banco.xlsx" not in client.uploaded_files

    _run(run_test())


def test_finalize_preserves_real_hbi_ruta_format_when_existing():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        r, _ = make_distrib_row(extract_route="")
        base = create_review_workbook(distrib_specs=[(r, None)])
        client.downloaded_files["revision/val_latest.xlsx"] = _append_distrib_ruta_column(base, {2: HBI_EXPECTED_RUTA})
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = w[ReviewSheets.DISTRIBUCION]
        cmap = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        assert ws.cell(2, cmap[DistribucionCols.RUTA]).value == HBI_EXPECTED_RUTA
        assert HBI_EXPECTED_RUTA.endswith(".pdf")
        assert "CREDITO #" in HBI_EXPECTED_RUTA
        assert "INFORMACION CREDITOS-CLIENTES/" in HBI_EXPECTED_RUTA

    _run(run_test())


def test_finalize_resolves_real_hbi_extract_path_from_credit_folder():
    async def run_test():
        set_env_vars()
        os.environ["GRAPH_CLIENTS_BASE_PATH"] = HBI_CLIENTS_ROOT
        client = MockGraphClient()
        client.children_overrides = {
            f"{HBI_CLIENTS_ROOT}/EQUINORTE": [{"name": "CREDITO # 264", "folder": {}}],
            f"{HBI_CLIENTS_ROOT}/EQUINORTE/CREDITO # 264": [
                {"name": HBI_EXTRACT_PDF, "file": {}},
            ],
        }
        r, _ = make_distrib_row(
            cliente="EQUINORTE",
            credito="264",
            extract_route="",
            fecha_banco=date(2026, 1, 1),
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = "https://tenant.sharepoint.com/sites/x/_layouts/15/Doc.aspx?id=onlyId"
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws2 = w[ReviewSheets.DISTRIBUCION]
        cmap = {str(ws2.cell(1, col).value): col for col in range(1, ws2.max_column + 1) if ws2.cell(1, col).value}
        rv = ws2.cell(2, cmap[DistribucionCols.RUTA]).value
        assert rv == HBI_EXPECTED_RUTA
        assert "EQUINORTE/" in str(rv)
        assert str(rv).endswith(HBI_EXTRACT_PDF)

    _run(run_test())


def test_finalize_ruta_is_drive_relative_not_web_url():
    async def run_test():
        set_env_vars()
        os.environ["GRAPH_CLIENTS_BASE_PATH"] = HBI_CLIENTS_ROOT
        client = MockGraphClient()
        client.children_overrides = {
            f"{HBI_CLIENTS_ROOT}/EQUINORTE": [{"name": "CREDITO # 264", "folder": {}}],
            f"{HBI_CLIENTS_ROOT}/EQUINORTE/CREDITO # 264": [
                {"name": HBI_EXTRACT_PDF, "file": {}},
            ],
        }
        r, _ = make_distrib_row(
            cliente="EQUINORTE",
            credito="264",
            extract_route="",
            fecha_banco=date(2026, 1, 1),
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "Ver extracto")
        ws.cell(2, c).hyperlink = "https://long.web.url/share?a=1"
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = str(w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value)
        assert not rv.lower().startswith("http")
        assert "://" not in rv
        assert rv.endswith(".pdf")

    _run(run_test())


def test_finalize_prefers_explicit_ruta_column_over_blank_link_hyperlink_resolve():
    """Con Ruta rellena, no hace falta hyperlink ni texto parseable en Link extracto."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/RUTP": [{"name": "CREDX", "folder": {}}],
            "clientes/RUTP/CREDX": [{"name": "Extracto rutp.pdf", "file": {}}],
        }
        r, _ = make_distrib_row(
            cliente="RUTP",
            credito="CREDX",
            extract_route="",
            ruta_pdf_internal="clientes/RUTP/CREDX/from_ruta_col.pdf",
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "solo texto amigable sin link")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value
        assert str(rv).replace("\\", "/") == "clientes/RUTP/CREDX/from_ruta_col.pdf"

    _run(run_test())


def test_finalize_fallback_extractos_folder_before_credit_root():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/XSUB": [{"name": "CR1", "folder": {}}],
            "clientes/XSUB/CR1": [
                {"name": "EXTRACTOS", "folder": {}},
                {"name": "solo_en_raiz_sin_extract_kw.pdf", "file": {}},
            ],
            "clientes/XSUB/CR1/EXTRACTOS": [
                {"name": "Extracto subcarpeta obligacion.pdf", "file": {}},
            ],
        }
        r, _ = make_distrib_row(
            cliente="XSUB",
            credito="CR1",
            extract_route="",
            ruta_pdf_internal="",
            link_tabla_display="tab",
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        c = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, c, "sin ruta tecnica previa")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        w = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        cmap = {
            str(w[ReviewSheets.DISTRIBUCION].cell(1, col).value): col
            for col in range(1, w[ReviewSheets.DISTRIBUCION].max_column + 1)
            if w[ReviewSheets.DISTRIBUCION].cell(1, col).value
        }
        rv = str(w[ReviewSheets.DISTRIBUCION].cell(2, cmap[DistribucionCols.RUTA]).value)
        assert "/EXTRACTOS/" in rv.replace("\\", "/")
        assert rv.endswith(".pdf")

    _run(run_test())


def test_finalize_missing_route_when_pdf_names_lack_extract_keyword():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children_overrides = {
            "clientes/YBAD": [{"name": "CY", "folder": {}}],
            "clientes/YBAD/CY": [
                {"name": "FACTURA_recibo_nomina_kw.pdf", "file": {}},
            ],
        }
        r, _ = make_distrib_row(
            cliente="YBAD",
            credito="CY",
            extract_route="",
            ruta_pdf_internal="",
            link_tabla_display="tab",
        )
        raw = create_review_workbook(distrib_specs=[(r, None)])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb[ReviewSheets.DISTRIBUCION]
        ce = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        ws.cell(2, ce, "x")
        buf = io.BytesIO()
        wb.save(buf)
        client.downloaded_files["revision/val_latest.xlsx"] = buf.getvalue()
        with pytest.raises(ValueError, match="missing_extract_route"):
            await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))

    _run(run_test())


def test_finalize_parse_extract_date_from_real_hbi_filename():
    assert _parse_extract_date_from_filename(HBI_EXTRACT_PDF) == date(2026, 1, 1)
    assert _parse_extract_date_from_filename("Extracto 2026-01-01 CREDITO # 264.pdf") == date(2026, 1, 1)
