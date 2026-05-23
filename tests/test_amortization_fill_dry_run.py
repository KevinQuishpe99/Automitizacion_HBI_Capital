"""Tests dry-run amortización (sin escritura en SharePoint)."""

import asyncio
import io
import json
from datetime import date
from urllib.parse import unquote

import openpyxl
import pytest
from pypdf import PdfWriter

from app.application.services.accounting_pdf_parser import ACCOUNT_VALOR_PAGADO_CLIENTE
from app.application.use_cases.amortization_fill_dry_run import (
    PDF_TEXT_NOT_EXTRACTABLE,
    TABLE_PATH_NOT_FOUND,
    run_amortization_fill_dry_run,
)


def _asiento_pdf_placeholder() -> bytes:
    """PDF mínimo; el texto del asiento se inyecta vía patch en tests."""
    return _empty_pdf_bytes()


def _empty_pdf_bytes() -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def _accounting_text() -> str:
    return f"""
    Comprobante 99 Fecha 22/05/2026
    {ACCOUNT_VALOR_PAGADO_CLIENTE} 50,000,000.00
    544113410519 49,118,143.00
    544113430501 0.00
    544141502030 881,857.00
    """


def _hist_bytes(id_pago: str, credito: str, tabla_path: str, fecha_limite: date) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(
        [
            "ID Pago",
            "Cliente",
            "Crédito",
            "Fecha límite",
            "Estado Pago",
            "Ruta",
            "Link tabla amortización",
        ]
    )
    link_val = tabla_path if tabla_path else "Ver tabla"
    ws.append([id_pago, "EQUINORTE", credito, fecha_limite, "ADELANTADO", "x.pdf", link_val])
    if tabla_path:
        cell = ws.cell(2, 7)
        cell.hyperlink = f"https://sharepoint/root:/{tabla_path.replace('/', '%2F')}:"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amort_table_bytes(fecha_limite: date) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amort"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "IBR +i",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "intereses mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([fecha_limite.day, fecha_limite.month, fecha_limite.year, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amort_table_with_existing_payment(fecha_limite: date) -> bytes:
    raw = _amort_table_bytes(fecha_limite)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    ws.cell(2, 9, value=50_000_000.0)
    ws.cell(2, 7, value=49_118_143.0)
    ws.cell(2, 6, value=0.0)
    ws.cell(2, 8, value=881_857.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amort_table_with_different_payment(fecha_limite: date) -> bytes:
    raw = _amort_table_bytes(fecha_limite)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    ws.cell(2, 9, value=1.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ibr_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IBR"
    ws.append(["Inicio", "Fin", "Valor"])
    ws.append([date(2026, 5, 1), date(2026, 5, 31), "10.580%"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class MockGraphDryRun:
    def __init__(self, files: dict[str, bytes]) -> None:
        self.files = files
        self.put_calls: list[tuple[str, bytes]] = []
        self.children: dict[str, list[dict]] = {}

    def _key(self, endpoint: str) -> str | None:
        if ":/content" not in endpoint:
            return None
        path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        return unquote(path)

    async def get(self, endpoint: str, params=None):
        if "children" in endpoint and "/root:/" in endpoint:
            folder = endpoint.split("/root:/", 1)[1].rsplit(":/children", 1)[0]
            folder = unquote(folder)
            return {"value": self.children.get(folder, [])}
        return {"value": [{"id": "site"}, {"id": "drive", "name": "Doc"}]}

    async def get_bytes(self, endpoint: str, params=None):
        key = self._key(endpoint)
        if key and key in self.files:
            return self.files[key]
        raise FileNotFoundError(key)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        self.put_calls.append((endpoint, content))
        return {"id": "new"}


@pytest.fixture(autouse=True)
def env_sharepoint(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST")
    monkeypatch.setenv("GRAPH_SHAREPOINT_DRIVE_NAME", "")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", "CTL")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "LOGS")
    monkeypatch.setenv("GRAPH_IBR_DIARIO_PATH", "CTL/IBR_DIARIO.xlsx")


def _base_files(
    *,
    hist: bytes,
    amort: bytes,
    asiento_pdf: bytes,
    ibr: bytes | None = None,
    fecha: date = date(2026, 5, 22),
) -> dict[str, bytes]:
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 258",
                "asiento_pdf_path": "clientes/EQUINORTE/asiento.pdf",
                "extracto_pdf_path": "clientes/EQUINORTE/extracto.pdf",
                "output_relative_path": "OUT/consolidado.pdf",
            }
        ],
        "skipped": [],
    }
    manifest_key = f"LOGS/merge_manifest_{fecha.isoformat()}.json"
    files = {
        "CTL/dummy.xlsx": b"x",
        manifest_key: json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort.xlsx": amort,
        "clientes/EQUINORTE/asiento.pdf": asiento_pdf,
    }
    if ibr is not None:
        files["CTL/IBR_DIARIO.xlsx"] = ibr
    return files


def test_dry_run_happy_path_would_apply_and_write_ibr(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(hist=hist, amort=_amort_table_bytes(fecha), asiento_pdf=_asiento_pdf_placeholder(), ibr=_ibr_bytes())
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )

    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    assert out["mode"] == "dry_run"
    assert g.put_calls == []
    item = out["items"][0]
    assert item["application_status"] == "WOULD_APPLY"
    assert item["target_row"] == 2
    assert item["ibr"]["status"] == "WOULD_WRITE_IBR"
    assert item["ibr"]["found"] is True
    assert item["payment_application"]["valor_pagado_cliente"] == 50_000_000.0


def test_dry_run_pending_ibr_without_ibr_file(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    files = _base_files(hist=hist, amort=_amort_table_bytes(fecha), asiento_pdf=_asiento_pdf_placeholder(), ibr=None)
    files.pop("CTL/IBR_DIARIO.xlsx", None)
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )

    out = asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert out["items"][0]["ibr"]["status"] == "PENDING_IBR"
    assert out["items"][0]["ibr"]["found"] is False


def test_dry_run_table_path_not_found(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "", fecha)
    g = MockGraphDryRun(
        _base_files(hist=hist, amort=_amort_table_bytes(fecha), asiento_pdf=_asiento_pdf_placeholder())
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert out["items"][0]["application_status"] == "ERROR"
    assert out["items"][0]["error_code"] == TABLE_PATH_NOT_FOUND


def test_dry_run_pdf_not_extractable():
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_amort_table_bytes(fecha),
            asiento_pdf=_empty_pdf_bytes(),
            ibr=_ibr_bytes(),
        )
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["items"][0]["error_code"] == PDF_TEXT_NOT_EXTRACTABLE


def test_dry_run_would_adopt_existing(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_amort_table_with_existing_payment(fecha),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert out["items"][0]["application_status"] == "WOULD_ADOPT_EXISTING"


def test_dry_run_revision_manual(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_amort_table_with_different_payment(fecha),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert out["items"][0]["application_status"] == "REVISION_MANUAL"


def test_dry_run_never_uploads_files(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(hist=hist, amort=_amort_table_bytes(fecha), asiento_pdf=_asiento_pdf_placeholder())
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert g.put_calls == []


_AMORT_HEADER_ROW = [
    "dia",
    "mes",
    "año",
    "IBR +i",
    "Fecha pago",
    "Valor intereses",
    "Abono a K",
    "intereses mora",
    "Valor pagado cliente",
    "Saldos Menores",
]


def _amort_table_date_at_row(fecha_limite: date, data_row: int, *, sheet_title: str = "EQUINORTE") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(_AMORT_HEADER_ROW)
    while (ws.max_row or 1) < data_row - 1:
        ws.append([None] * len(_AMORT_HEADER_ROW))
    ws.append(
        [
            fecha_limite.day,
            fecha_limite.month,
            fecha_limite.year,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amort_table_two_dates_at_rows(
    fecha_limite: date, row_a: int, row_b: int, *, sheet_title: str = "EQUINORTE"
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(_AMORT_HEADER_ROW)
    for target_row in (row_a, row_b):
        while (ws.max_row or 1) < target_row - 1:
            ws.append([None] * len(_AMORT_HEADER_ROW))
        ws.append(
            [
                fecha_limite.day,
                fecha_limite.month,
                fecha_limite.year,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hist_bytes_multi_credit(rows: list[tuple[str, str, str, date]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(
        [
            "ID Pago",
            "Cliente",
            "Crédito",
            "Fecha límite",
            "Estado Pago",
            "Ruta",
            "Link tabla amortización",
        ]
    )
    for id_pago, cliente, credito, tabla_path, fecha_limite in rows:
        ws.append([id_pago, cliente, credito, fecha_limite, "ADELANTADO", "x.pdf", tabla_path])
        cell = ws.cell(ws.max_row, 7)
        cell.hyperlink = f"https://sharepoint/root:/{tabla_path.replace('/', '%2F')}:"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _amort_table_two_rows_same_date(fecha_limite: date) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amort"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "IBR +i",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "intereses mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([fecha_limite.day, fecha_limite.month, fecha_limite.year, None, None, None, None, None, None, None])
    ws.append([fecha_limite.day, fecha_limite.month, fecha_limite.year, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_dry_run_two_asientos_produce_two_events(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    asiento_a = "clientes/EQUINORTE/ASIENTOS/Asiento cuota credito 258.pdf"
    asiento_b = "clientes/EQUINORTE/ASIENTOS/Asiento abono capital credito 258.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 258",
                "asiento_pdf_paths": [asiento_a, asiento_b],
                "asiento_pdf_path": asiento_a,
                "extracto_pdf_path": "clientes/EQUINORTE/extracto.pdf",
                "output_relative_path": "OUT/consolidado.pdf",
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort.xlsx": _amort_table_two_rows_same_date(fecha),
        asiento_a: _asiento_pdf_placeholder(),
        asiento_b: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    texts = iter([_accounting_text(), _accounting_text().replace("50,000,000", "10,000,000")])

    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: next(texts),
    )

    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["summary"]["total_events"] == 2
    assert len(out["items"]) == 2
    assert out["items"][0]["event_index"] == 1
    assert out["items"][1]["event_index"] == 2
    assert out["items"][0]["asiento_pdf_path"] != out["items"][1]["asiento_pdf_path"]
    assert out["items"][0]["due_date_row"] == 2
    assert out["items"][1]["due_date_row"] == 2
    assert out["items"][0]["application_row"] == 2
    assert out["items"][1]["application_row"] == 3


def test_dry_run_legacy_manifest_single_asiento_path(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(hist=hist, amort=_amort_table_bytes(fecha), asiento_pdf=_asiento_pdf_placeholder(), ibr=_ibr_bytes())
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(run_amortization_fill_dry_run(g, report_date_iso="2026-05-22"))
    assert len(out["items"]) == 1
    assert out["items"][0]["event_index"] == 1


def test_dry_run_credit_items_resolves_hist_per_individual_credit(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("8326b91b", "CREDITO # 258", "TABLAS/amort_258.xlsx", fecha)
    wb = openpyxl.load_workbook(io.BytesIO(hist))
    ws = wb.active
    ws.append(
        [
            "8326b91b",
            "EQUINORTE",
            "CREDITO # 265",
            fecha,
            "ADELANTADO",
            "x.pdf",
            "TABLAS/amort_265.xlsx",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    hist = buf.getvalue()

    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "8326b91b",
                "cliente": "EQUINORTE",
                "credito": "258, 265",
                "credit_items": [
                    {
                        "credito": "258",
                        "asiento_pdf_paths": ["clientes/E/asiento_258.pdf"],
                        "extracto_pdf_paths": ["clientes/E/extracto_258.pdf"],
                        "extracto_pdf_path": "clientes/E/extracto_258.pdf",
                    },
                    {
                        "credito": "265",
                        "asiento_pdf_paths": [
                            "clientes/E/asiento_265_1.pdf",
                            "clientes/E/asiento_265_2.pdf",
                        ],
                        "extracto_pdf_paths": ["clientes/E/extracto_265.pdf"],
                        "extracto_pdf_path": "clientes/E/extracto_265.pdf",
                    },
                ],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_258.xlsx": _amort_table_bytes(fecha),
        "TABLAS/amort_265.xlsx": _amort_table_two_rows_same_date(fecha),
        "clientes/E/asiento_258.pdf": _asiento_pdf_placeholder(),
        "clientes/E/asiento_265_1.pdf": _asiento_pdf_placeholder(),
        "clientes/E/asiento_265_2.pdf": _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["summary"]["total_events"] == 3
    creditos = [it["credito"] for it in out["items"]]
    assert creditos == ["258", "265", "265"]
    assert all(it["application_status"] != "ERROR" or it.get("error_code") != "TABLE_PATH_NOT_FOUND" for it in out["items"])
    assert out["items"][0]["tabla_amortizacion_path"] == "TABLAS/amort_258.xlsx"
    assert out["items"][1]["tabla_amortizacion_path"] == "TABLAS/amort_265.xlsx"
    assert out["items"][0]["due_date_row"] == 2
    assert out["items"][1]["due_date_row"] == 2
    assert out["items"][2]["due_date_row"] == 2
    assert out["items"][0]["application_row"] == 2
    assert out["items"][1]["application_row"] == 2
    assert out["items"][2]["application_row"] == 3
    assert out["items"][0]["application_status"] in (
        "WOULD_APPLY",
        "WOULD_ADOPT_EXISTING",
        "REVISION_MANUAL",
    )
    assert out["items"][1]["application_status"] in (
        "WOULD_APPLY",
        "WOULD_ADOPT_EXISTING",
        "REVISION_MANUAL",
    )


def test_dry_run_two_credits_different_tables_same_target_row(monkeypatch):
    fecha = date(2026, 4, 22)
    id_pago = "8326b91b"
    hist = _hist_bytes_multi_credit(
        [
            (id_pago, "EQUINORTE", "CREDITO # 258", "TABLAS/amort_258.xlsx", fecha),
            (id_pago, "EQUINORTE", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha),
        ]
    )
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": id_pago,
                "cliente": "EQUINORTE",
                "credito": "258, 265",
                "credit_items": [
                    {
                        "credito": "258",
                        "asiento_pdf_paths": ["clientes/E/asiento_258.pdf"],
                    },
                    {
                        "credito": "265",
                        "asiento_pdf_paths": ["clientes/E/asiento_265.pdf"],
                    },
                ],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_258.xlsx": _amort_table_date_at_row(fecha, 8),
        "TABLAS/amort_265.xlsx": _amort_table_date_at_row(fecha, 8),
        "clientes/E/asiento_258.pdf": _asiento_pdf_placeholder(),
        "clientes/E/asiento_265.pdf": _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert len(out["items"]) == 2
    assert out["items"][0]["credito"] == "258"
    assert out["items"][1]["credito"] == "265"
    assert out["items"][0]["due_date_row"] == 8
    assert out["items"][1]["due_date_row"] == 8
    assert out["items"][0]["application_row"] == 8
    assert out["items"][1]["application_row"] == 8
    assert out["items"][0]["target_row"] == 8
    assert out["items"][1]["target_row"] == 8
    assert out["items"][0]["error_code"] is None
    assert out["items"][1]["error_code"] is None


def test_dry_run_same_table_second_asiento_excludes_used_row(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha)
    asiento_a = "clientes/E/asiento_265_1.pdf"
    asiento_b = "clientes/E/asiento_265_2.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 265",
                "asiento_pdf_paths": [asiento_a, asiento_b],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_265.xlsx": _amort_table_two_dates_at_rows(fecha, 8, 9),
        asiento_a: _asiento_pdf_placeholder(),
        asiento_b: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["items"][0]["due_date_row"] == 8
    assert out["items"][1]["due_date_row"] == 8
    assert out["items"][0]["application_row"] == 8
    assert out["items"][1]["application_row"] == 9
    assert out["items"][0]["target_row"] == 8
    assert out["items"][1]["target_row"] == 9


def test_dry_run_row_already_assigned_warning_only_same_table(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha)
    asiento_a = "clientes/E/a1.pdf"
    asiento_b = "clientes/E/a2.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 265",
                "asiento_pdf_paths": [asiento_a, asiento_b],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_265.xlsx": _amort_table_date_at_row(fecha, 8),
        asiento_a: _asiento_pdf_placeholder(),
        asiento_b: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["items"][0]["target_row"] == 8
    second = out["items"][1]
    assert second["error_code"] == "REQUIRES_APPLICATION_ROW"
    assert second["due_date_row"] == 8
    assert second["application_row"] is None
    assert any(
        "filas de aplicación ya reservadas" in w.lower() for w in second.get("warnings", [])
    )


def test_dry_run_multi_credit_real_row_8_scenario(monkeypatch):
    fecha = date(2026, 4, 22)
    id_pago = "8326b91b"
    hist = _hist_bytes_multi_credit(
        [
            (id_pago, "EQUINORTE", "CREDITO # 258", "TABLAS/amort_258.xlsx", fecha),
            (id_pago, "EQUINORTE", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha),
        ]
    )
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": id_pago,
                "cliente": "EQUINORTE",
                "credito": "258, 265",
                "credit_items": [
                    {
                        "credito": "258",
                        "asiento_pdf_paths": ["clientes/E/asiento_258.pdf"],
                    },
                    {
                        "credito": "265",
                        "asiento_pdf_paths": [
                            "clientes/E/asiento_265_1.pdf",
                            "clientes/E/asiento_265_2.pdf",
                        ],
                    },
                ],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_258.xlsx": _amort_table_date_at_row(fecha, 8),
        "TABLAS/amort_265.xlsx": _amort_table_two_dates_at_rows(fecha, 8, 9),
        "clientes/E/asiento_258.pdf": _asiento_pdf_placeholder(),
        "clientes/E/asiento_265_1.pdf": _asiento_pdf_placeholder(),
        "clientes/E/asiento_265_2.pdf": _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert len(out["items"]) == 3
    assert out["items"][0]["credito"] == "258"
    assert out["items"][0]["due_date_row"] == 8
    assert out["items"][0]["application_row"] == 8
    assert out["items"][1]["credito"] == "265"
    assert out["items"][1]["due_date_row"] == 8
    assert out["items"][1]["application_row"] == 8
    assert out["items"][2]["credito"] == "265"
    if out["items"][1]["ibr"]["status"] == "WOULD_WRITE_IBR":
        assert out["items"][2]["ibr"]["status"] == "WOULD_SKIP_IBR_ALREADY_PLANNED"
    assert out["items"][2]["due_date_row"] == 8
    assert out["items"][2]["application_row"] == 9
    assert out["items"][2]["error_code"] is None


def _fill_application_cells(ws, row: int, *, amount: float = 1.0) -> None:
    ws.cell(row, 6, amount)
    ws.cell(row, 7, amount)
    ws.cell(row, 9, amount)


def _amort_table_displaced_application(
    fecha_limite: date,
    due_row: int,
    occupied_application_rows: list[int],
    *,
    free_application_row: int | None = None,
) -> bytes:
    """Cronograma en due_row; aplicación ocupada en filas sin fecha de corte."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(list(_AMORT_HEADER_ROW))
    max_row_needed = max([due_row, *occupied_application_rows, free_application_row or 0])
    while (ws.max_row or 1) < max_row_needed:
        ws.append([None] * len(_AMORT_HEADER_ROW))
    ws.cell(due_row, 1, fecha_limite.day)
    ws.cell(due_row, 2, fecha_limite.month)
    ws.cell(due_row, 3, fecha_limite.year)
    for r in occupied_application_rows:
        _fill_application_cells(ws, r)
    if free_application_row is not None:
        while (ws.max_row or 1) < free_application_row:
            ws.append([None] * len(_AMORT_HEADER_ROW))
        ws.cell(free_application_row, 5, "")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_dry_run_due_and_application_row_same_when_no_displacement(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    item = out["items"][0]
    assert item["due_date_row"] == 8
    assert item["ibr_row"] == 8
    assert item["application_row"] == 8
    assert item["target_row"] == 8


def test_dry_run_application_row_below_due_when_block_displaced(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha)
    amort = _amort_table_displaced_application(
        fecha,
        due_row=8,
        occupied_application_rows=[8, 9],
        free_application_row=10,
    )
    asiento = "clientes/E/asiento_265.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 265",
                "asiento_pdf_path": asiento,
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_265.xlsx": amort,
        asiento: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    item = out["items"][0]
    assert item["due_date_row"] == 8
    assert item["ibr_row"] == 8
    assert item["application_row"] == 10
    assert item["target_row"] == 10


def test_dry_run_summary_counts_events_not_outputs(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    asiento_a = "clientes/EQUINORTE/a1.pdf"
    asiento_b = "clientes/EQUINORTE/a2.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 258",
                "asiento_pdf_paths": [asiento_a, asiento_b],
                "asiento_pdf_path": asiento_a,
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort.xlsx": _amort_table_two_rows_same_date(fecha),
        asiento_a: _asiento_pdf_placeholder(),
        asiento_b: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["manifest_outputs_count"] == 1
    assert out["summary"]["total"] == 2


def _invalid_amort_table_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Total pagado", "Observación"])
    ws.append([100, "sin encabezados de amortización"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_dry_run_monthly_schedule_finds_target_row(monkeypatch):
    fecha = date(2026, 4, 22)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "Valor pagado cliente",
        ]
    )
    ws.append([22, 12, 2025, None, None, None, None])
    for r in range(3, 8):
        ws.cell(r, 1, f'=IF(A{r}<>" ",1,1)')
        ws.cell(r, 2, f'=IF(A{r}<>" ",1,1)')
        ws.cell(r, 3, f'=IF(A{r}<>" ",1,1)')
    buf = io.BytesIO()
    wb.save(buf)
    amort = buf.getvalue()

    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=amort,
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-04-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    item = out["items"][0]
    assert item["application_status"] in ("WOULD_APPLY", "WOULD_ADOPT_EXISTING", "REVISION_MANUAL")
    assert item.get("error_code") is None
    assert item["due_date_row"] == 6
    assert item["ibr_row"] == 6
    assert item["application_row"] == 6
    assert item["target_row"] == 6
    assert item["fecha_limite_pago"] == "2026-04-22"


def test_dry_run_amortization_sheet_not_found_item_error_job_completes(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_invalid_amort_table_bytes(),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    assert out["mode"] == "dry_run"
    assert out["summary"]["errors"] == 1
    item = out["items"][0]
    assert item["application_status"] == "ERROR"
    assert item["error_code"] == "AMORTIZATION_SHEET_NOT_FOUND"
    assert any("encabezados" in w.lower() for w in item["warnings"])
