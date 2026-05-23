"""Tests registro operativo en bandejas Pendientes tras Finalize."""

import asyncio
import io
from datetime import date

import httpx
import openpyxl
import pytest

from app.application.services.payment_followup_finalize import (
    ESTADO_APLICACION_PENDIENTE_TABLA,
    ESTADO_FINAL_ABIERTO,
    ESTADO_IBR_PENDIENTE,
    register_payment_followups_after_finalize,
)
from app.application.services.review_schema import DistribucionCols, EstadoPago, ValidarPago
from app.application.use_cases.setup_payment_followup_workbooks import (
    ADELANTADOS_COLUMNS,
    FILENAME_ADELANTADOS,
    FILENAME_INCOMPLETOS,
    INCOMPLETOS_COLUMNS,
    SHEET_HISTORICO,
    SHEET_PENDIENTES,
    _build_followup_workbook_bytes,
)


def _http_error(status: int) -> httpx.HTTPStatusError:
    req = httpx.Request("GET", "https://graph.microsoft.com/test")
    resp = httpx.Response(status, request=req, text=f"HTTP {status} body")
    return httpx.HTTPStatusError("err", request=req, response=resp)


CONTROL_FOLDER = "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL"
PATH_AD = f"{CONTROL_FOLDER}/{FILENAME_ADELANTADOS}"
PATH_IN = f"{CONTROL_FOLDER}/{FILENAME_INCOMPLETOS}"


class MockGraphFollowup:
    def __init__(self) -> None:
        self.files: dict[str, bytes] = {}

    def _path_key(self, endpoint: str) -> str | None:
        low = endpoint.lower()
        if ":/content" not in low:
            return None
        if FILENAME_ADELANTADOS.lower() in low:
            return PATH_AD
        if FILENAME_INCOMPLETOS.lower() in low:
            return PATH_IN
        return None

    async def get_bytes(self, endpoint: str, params=None):
        key = self._path_key(endpoint)
        if key and key in self.files:
            return self.files[key]
        raise _http_error(404)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        key = self._path_key(endpoint)
        if key:
            self.files[key] = content
        return {"webUrl": f"https://example/{key}"}


@pytest.fixture(autouse=True)
def env_paths(monkeypatch):
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", CONTROL_FOLDER)
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST")


def _dist(estado: str, id_pago: str = "P1", credito: str = "C1") -> dict:
    return {
        DistribucionCols.ID_PAGO: id_pago,
        DistribucionCols.CLIENTE: "Cliente X",
        DistribucionCols.CREDITO: credito,
        DistribucionCols.FECHA_BANCO: "2026-05-10",
        DistribucionCols.FECHA_LIMITE: "2026-05-15",
        DistribucionCols.MONTO_BANCO: 100.0,
        DistribucionCols.VALOR_EXTRACTO: 100.0,
        DistribucionCols.APLICAR_A_EXTRACTO: 80.0,
        DistribucionCols.MORA_A_APLICAR: 10.0,
        DistribucionCols.OTROS_VALORES: 10.0,
        DistribucionCols.TOTAL_APLICADO: 100.0,
        DistribucionCols.SALDO_POR_ASIGNAR: 0.0,
        DistribucionCols.ESTADO_PAGO: estado,
        DistribucionCols.VALIDAR_PAGO: ValidarPago.SI,
        DistribucionCols.OBSERVACION: "",
        DistribucionCols.LINK_TABLA: "/tabla",
        DistribucionCols.RUTA_UNIDAD_CREDITO: "/credito",
        DistribucionCols.RUTA: "/extracto",
        DistribucionCols.RUTA_ASIENTOS_CONTABLES: "/asiento",
    }


def _seed_workbooks(g: MockGraphFollowup) -> None:
    g.files[PATH_AD] = _build_followup_workbook_bytes(columns=ADELANTADOS_COLUMNS)
    g.files[PATH_IN] = _build_followup_workbook_bytes(columns=INCOMPLETOS_COLUMNS)


def _pendientes_rows(path: str, files: dict[str, bytes]) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(io.BytesIO(files[path]))
    ws = wb[SHEET_PENDIENTES]
    id_col = None
    cred_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip()
        if h == "ID Pago":
            id_col = c
        if h == "Crédito":
            cred_col = c
    rows = []
    for r in range(2, ws.max_row + 1):
        if id_col and cred_col:
            id_p = str(ws.cell(r, id_col).value or "").strip()
            cred = str(ws.cell(r, cred_col).value or "").strip()
            if id_p:
                rows.append((id_p, cred))
    return rows


def test_adelantado_upsert_en_pagos_adelantados(env_paths):
    g = MockGraphFollowup()
    _seed_workbooks(g)
    asyncio.run(
        register_payment_followups_after_finalize(
            g,
            "site-1",
            "drive-1",
            process_date=date(2026, 5, 10),
            distributions=[_dist(EstadoPago.ADELANTADO)],
            historical_relative_path="hist/cartera.xlsx",
        )
    )
    wb = openpyxl.load_workbook(io.BytesIO(g.files[PATH_AD]))
    ws = wb[SHEET_PENDIENTES]
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "P1"
    estado_app_col = ADELANTADOS_COLUMNS.index("EstadoAplicacionPago") + 1
    assert ws.cell(2, estado_app_col).value == ESTADO_APLICACION_PENDIENTE_TABLA
    estado_ibr_col = ADELANTADOS_COLUMNS.index("EstadoIBR") + 1
    assert ws.cell(2, estado_ibr_col).value == ESTADO_IBR_PENDIENTE
    assert ws.cell(2, ADELANTADOS_COLUMNS.index("EstadoFinal") + 1).value == ESTADO_FINAL_ABIERTO
    assert wb[SHEET_HISTORICO].max_row == 1


def test_incompleto_upsert_en_pagos_incompletos(env_paths):
    g = MockGraphFollowup()
    _seed_workbooks(g)
    asyncio.run(
        register_payment_followups_after_finalize(
            g,
            "site-1",
            "drive-1",
            process_date=date(2026, 5, 10),
            distributions=[_dist(EstadoPago.INCOMPLETO, id_pago="P2")],
            historical_relative_path="hist/cartera.xlsx",
        )
    )
    assert ("P2", "C1") in _pendientes_rows(PATH_IN, g.files)
    wb = openpyxl.load_workbook(io.BytesIO(g.files[PATH_IN]))
    ws = wb[SHEET_PENDIENTES]
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    assert "EstadoIBR" not in headers
    assert "FechaIBRRequerida" not in headers


def test_no_duplica_id_pago_credito(env_paths):
    g = MockGraphFollowup()
    _seed_workbooks(g)
    dist = _dist(EstadoPago.ADELANTADO)
    for _ in range(2):
        asyncio.run(
            register_payment_followups_after_finalize(
                g,
                "site-1",
                "drive-1",
                process_date=date(2026, 5, 10),
                distributions=[dist],
                historical_relative_path="hist/cartera.xlsx",
            )
        )
    assert _pendientes_rows(PATH_AD, g.files).count(("P1", "C1")) == 1
    wb = openpyxl.load_workbook(io.BytesIO(g.files[PATH_AD]))
    assert wb[SHEET_PENDIENTES].max_row == 2


def test_historico_sin_filas_nuevas(env_paths):
    g = MockGraphFollowup()
    _seed_workbooks(g)
    asyncio.run(
        register_payment_followups_after_finalize(
            g,
            "site-1",
            "drive-1",
            process_date=date(2026, 5, 10),
            distributions=[_dist(EstadoPago.INCOMPLETO)],
            historical_relative_path="hist/cartera.xlsx",
        )
    )
    wb = openpyxl.load_workbook(io.BytesIO(g.files[PATH_IN]))
    assert wb[SHEET_HISTORICO].max_row == 1
