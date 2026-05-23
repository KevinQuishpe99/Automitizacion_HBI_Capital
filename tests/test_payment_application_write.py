"""Tests escritura de aplicación del pago (estilo secretaria)."""

import io
from datetime import date

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from app.application.services.accounting_pdf_parser import (
    ACCOUNT_CAPITAL,
    ACCOUNT_SALDOS_MENORES,
    ACCOUNT_VALOR_PAGADO_CLIENTE,
    WARNING_BANK_INFERRED,
    PaymentApplicationEvent,
)
from app.application.services.amortization_workbook import (
    PaymentApplicationWriteOptions,
    detect_amortization_sheet,
    find_row_by_due_date,
    write_payment_application,
    _is_formula_value,
)

_SECRETARY_HEADERS = [
    "dia",
    "mes",
    "año",
    "IBR +i",
    "Fecha pago",
    "Valor intereses",
    "Abono a K",
    "intereses mora",
    "Retenciones",
    "Valor pagado cliente",
    "Saldo a capital",
    "Saldos Menores",
]


def _build_secretary_table(
    fecha_limite: date,
    *,
    application_row: int,
    prev_saldo_formula_row: int | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(_SECRETARY_HEADERS)
    headers = {name: idx + 1 for idx, name in enumerate(_SECRETARY_HEADERS)}
    while (ws.max_row or 1) < application_row - 1:
        ws.append([None] * len(_SECRETARY_HEADERS))
    ws.append(
        [
            fecha_limite.day,
            fecha_limite.month,
            fecha_limite.year,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    if prev_saldo_formula_row is not None and prev_saldo_formula_row >= 2:
        saldo_col = headers["Saldo a capital"]
        abono_col = headers["Abono a K"]
        ws.cell(
            prev_saldo_formula_row,
            saldo_col,
            value=(
                f"=+{get_column_letter(saldo_col)}{prev_saldo_formula_row - 1}-"
                f"{get_column_letter(abono_col)}{prev_saldo_formula_row}"
            ),
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normal_event() -> PaymentApplicationEvent:
    return PaymentApplicationEvent(
        id_pago="P1",
        cliente="EQUINORTE",
        credito="CREDITO # 258",
        asiento_pdf_path="a.pdf",
        comprobante="1",
        fecha_asiento=date(2026, 4, 20),
        valor_pagado_cliente=50_000_000.0,
        capital=49_118_143.0,
        intereses=0.0,
        mora=881_857.0,
        retenciones=0.0,
        saldos_menores=0.0,
        raw_text="",
        detected_codes=(ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL),
    )


def test_credito_258_writes_fecha_pago_and_formulas():
    fecha = date(2026, 4, 22)
    payment_date = date(2026, 4, 23)
    wb = openpyxl.load_workbook(
        io.BytesIO(_build_secretary_table(fecha, application_row=8, prev_saldo_formula_row=7))
    )
    match = detect_amortization_sheet(wb)
    ws, headers, header_row = match.worksheet, match.headers, match.header_row
    row = find_row_by_due_date(ws, headers, fecha, header_row=header_row)
    assert row == 8
    plan = write_payment_application(
        ws,
        row,
        headers,
        _normal_event(),
        write_options=PaymentApplicationWriteOptions(
            payment_date=payment_date,
            detected_codes=frozenset({ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL}),
        ),
        header_row=header_row,
    )
    assert ws.cell(row, headers["fecha_pago"]).value == payment_date
    assert plan["fecha_pago"] == "value"
    assert _is_formula_value(ws.cell(row, headers["saldo_a_capital"]).value)
    assert _is_formula_value(ws.cell(row, headers["valor_pagado_cliente"]).value)
    assert ws.cell(row, headers["valor_intereses"]).value is None
    assert ws.cell(row, headers["abono_k"]).value == pytest.approx(49_118_143.0)
    assert ws.cell(row, headers["intereses_mora"]).value == pytest.approx(881_857.0)


def test_saldos_menores_skips_valor_pagado_and_fills_saldos():
    fecha = date(2026, 4, 22)
    payment_date = date(2026, 4, 23)
    wb = openpyxl.load_workbook(
        io.BytesIO(_build_secretary_table(fecha, application_row=10, prev_saldo_formula_row=9))
    )
    match = detect_amortization_sheet(wb)
    ws, headers, header_row = match.worksheet, match.headers, match.header_row
    row = 10
    event = PaymentApplicationEvent(
        id_pago="P1",
        cliente="EQUINORTE",
        credito="CREDITO # 265",
        asiento_pdf_path="a2.pdf",
        comprobante="2",
        fecha_asiento=date(2026, 4, 20),
        valor_pagado_cliente=285.0,
        capital=285.0,
        intereses=0.0,
        mora=0.0,
        retenciones=0.0,
        saldos_menores=285.0,
        raw_text="",
        detected_codes=(ACCOUNT_SALDOS_MENORES, ACCOUNT_CAPITAL),
        parse_warnings=(WARNING_BANK_INFERRED,),
    )
    plan = write_payment_application(
        ws,
        row,
        headers,
        event,
        write_options=PaymentApplicationWriteOptions(
            payment_date=payment_date,
            detected_codes=frozenset({ACCOUNT_SALDOS_MENORES, ACCOUNT_CAPITAL}),
            warnings=frozenset({WARNING_BANK_INFERRED}),
        ),
        header_row=header_row,
    )
    assert ws.cell(row, headers["fecha_pago"]).value == payment_date
    assert plan["valor_pagado_cliente"] == "empty"
    assert ws.cell(row, headers["valor_pagado_cliente"]).value is None
    assert ws.cell(row, headers["abono_k"]).value == pytest.approx(285.0)
    sm_val = ws.cell(row, headers["saldos_menores"]).value
    assert _is_formula_value(sm_val) or sm_val == pytest.approx(285.0)
    assert plan["saldo_a_capital"] == "formula"
