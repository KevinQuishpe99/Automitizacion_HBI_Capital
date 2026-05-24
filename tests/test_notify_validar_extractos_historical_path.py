"""Notify validar extractos: historical_file_path obligatorio y sin auto-resolve de histórico."""

import asyncio
import time
from datetime import date
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.adapters.primary.http.deps import init_graph_client
from app.adapters.primary.http.routers import sharepoint as sharepoint_mod
from app.adapters.primary.http.routers.sharepoint import router
from app.application.job_store_factory import reset_job_store_for_tests
from app.application.job_status_enrichment import enrich_job_for_http_response
from app.application.use_cases.merge_control_workbook_notify import MergeControlNotifyWriteOutcome
from app.application.use_cases.send_validar_extractos_notification import (
    _find_distribucion_header_row,
    send_validar_extractos_notification_email,
)

_MC_SKIP_OUTCOME = MergeControlNotifyWriteOutcome(
    merge_control_updated=False,
    merge_control_file_path=(
        "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL/control_merge_pdfs.xlsx"
    ),
    merge_control_status=None,
    merge_control_warning=None,
    merge_control_error_code=None,
)


def _minimal_bank_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Concepto", "Crédito", "Extra"])
    ws.append([date(2026, 5, 12), "abono", "C1", "x"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _minimal_historico_xlsx(
    estado_header: str = "Estado línea",
    ruta_header: str = "Ruta",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append([estado_header, ruta_header])
    ws.append(["VALIDAR", ""])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture  # type: ignore[name-defined]
def client():
    app = FastAPI()
    init_graph_client(MagicMock())
    app.include_router(router)
    reset_job_store_for_tests()
    yield TestClient(app, raise_server_exceptions=False)
    reset_job_store_for_tests()


def _poll_job(client: TestClient, jid: str, timeout: float = 5.0) -> dict:
    deadline = time.perf_counter() + timeout
    last: dict = {}
    while time.perf_counter() < deadline:
        res = client.get(f"/graph/sharepoint/notify-validar-extractos-email/jobs/{jid}")
        last = res.json()
        if last.get("status") in ("completed", "failed"):
            return last
        time.sleep(0.02)
    return last


def test_notify_requires_historical_file_path(client):
    r = client.post("/graph/sharepoint/notify-validar-extractos-email", json={})
    assert r.status_code == 202
    body = _poll_job(client, r.json()["job_id"])
    assert body.get("status") == "failed"
    err = body.get("error")
    assert isinstance(err, dict)
    enriched = enrich_job_for_http_response(body)
    e = enriched["error"]
    assert e["error_code"] == "missing_historical_file_path"
    assert "historical_file_path" in e["next_action"].lower() or "finalize" in e["next_action"].lower()


def test_notify_rejects_blank_historical_file_path(client):
    r = client.post(
        "/graph/sharepoint/notify-validar-extractos-email",
        json={"historical_file_path": "   "},
    )
    assert r.status_code == 202
    body = _poll_job(client, r.json()["job_id"])
    assert body.get("status") == "failed"
    enriched = enrich_job_for_http_response(body)
    assert enriched["error"]["error_code"] == "missing_historical_file_path"


def test_notify_failed_explicit_history_missing_returns_standard_error():
    """404 al descargar histórico → historical_file_not_found enriquecido."""

    async def fail_hist(_g, _s, _d, _path):
        req = httpx.Request("GET", "https://graph.microsoft.com/v1.0/x")
        resp = httpx.Response(404, request=req, text="not found")
        raise httpx.HTTPStatusError("404", request=req, response=resp)

    class _GraphOk:
        async def get_bytes(self, *_a, **_k):
            return _minimal_bank_xlsx()

        async def post_json(self, *_a, **_k):
            return {}, 202

        async def put_bytes(self, *_a, **_k):
            return {}

    async def run():
        with (
            patch(
                "app.application.use_cases.send_validar_extractos_notification.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value={
                    "site_id": "s1",
                    "drive_id": "d1",
                    "path_encoded": "bank/report.xlsx",
                },
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._load_sender_and_recipients_from_correos_xlsx",
                new_callable=AsyncMock,
                return_value=("sender@example.com", ["to@example.com"]),
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._graph_download_by_path",
                new_callable=AsyncMock,
                side_effect=fail_hist,
            ),
        ):
            g = _GraphOk()
            with pytest.raises(ValueError) as ei:
                await send_validar_extractos_notification_email(
                    g,
                    historical_file_path="path/missing.xlsx",
                )
            assert str(ei.value).startswith("historical_file_not_found")

        raw = {
            "job_id": "j",
            "type": "notify_validar_extractos",
            "status": "failed",
            "error": str(ei.value),
        }
        out = enrich_job_for_http_response(raw)
        e = out["error"]
        assert e["error_code"] == "historical_file_not_found"
        assert e["user_message"]
        assert "finalize" in e["next_action"].lower() or "sharepoint" in e["next_action"].lower()

    asyncio.run(run())


def test_notify_standard_job_response_fields_remain_backward_compatible(client):
    r = client.post("/graph/sharepoint/notify-validar-extractos-email", json={})
    assert r.status_code == 202
    jid = r.json()["job_id"]
    body = _poll_job(client, jid)
    assert body.get("job_id") == jid
    assert body.get("type") == "notify_validar_extractos"
    enriched = enrich_job_for_http_response(body)
    assert enriched.get("severity") == "error"
    err = enriched.get("error") or {}
    assert err.get("user_message")
    assert err.get("next_action")
    assert err.get("error_code") == "missing_historical_file_path"


def test_notify_completed_result_includes_historical_file_source_explicit():
    raw = {
        "job_id": "j",
        "type": "notify_validar_extractos",
        "status": "completed",
        "result": {
            "status": "ok",
            "historical_file_path": "HIST/x.xlsx",
            "historico_excel_path": "HIST/x.xlsx",
            "historical_file_source": "explicit",
            "attachments_count": 0,
        },
    }
    out = enrich_job_for_http_response(raw)
    assert out["result"]["historical_file_source"] == "explicit"
    assert out["result"]["historical_file_path"] == "HIST/x.xlsx"


def test_notify_enrichment_maps_missing_historical_file_path_string():
    raw = {
        "job_id": "j",
        "type": "notify_validar_extractos",
        "status": "failed",
        "error": "missing_historical_file_path",
    }
    out = enrich_job_for_http_response(raw)
    assert out["error"]["error_code"] == "missing_historical_file_path"
    assert "Power Automate" in out["error"]["user_message"]
    assert "histórico" in out["error"]["user_message"].lower()
    assert out["error"]["next_action"]


def test_find_distribucion_header_row_accepts_estado_nuevo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(["ID Pago", "Estado", "Ruta"])
    row, hmap = _find_distribucion_header_row(ws)
    assert row == 1
    assert hmap["ESTADO"] == 2
    assert hmap["RUTA"] == 3


def test_find_distribucion_header_row_accepts_estado_linea_legacy():
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(["Estado línea", "Rutas"])
    row, hmap = _find_distribucion_header_row(ws)
    assert row == 1
    assert hmap["ESTADO LINEA"] == 1
    assert hmap["RUTAS"] == 2


def test_find_distribucion_header_row_missing_status_column():
    wb = Workbook()
    ws = wb.active
    ws.append(["Ruta", "Cliente"])
    with pytest.raises(ValueError, match="missing_distribucion_status_column"):
        _find_distribucion_header_row(ws)


def test_find_distribucion_header_row_missing_route_column():
    wb = Workbook()
    ws = wb.active
    ws.append(["Estado", "Cliente"])
    with pytest.raises(ValueError, match="missing_distribucion_route_column"):
        _find_distribucion_header_row(ws)


def test_find_distribucion_header_row_missing_headers_entirely():
    wb = Workbook()
    ws = wb.active
    ws.append(["Cliente", "Crédito"])
    with pytest.raises(ValueError, match="missing_distribucion_headers"):
        _find_distribucion_header_row(ws)


def test_notify_enrichment_maps_missing_distribucion_status_column():
    raw = {
        "job_id": "j",
        "type": "notify_validar_extractos",
        "status": "failed",
        "error": "missing_distribucion_status_column",
    }
    out = enrich_job_for_http_response(raw)
    assert out["error"]["error_code"] == "missing_distribucion_status_column"
    assert out["error"]["error_code"] != "unknown_error"
    assert "Estado" in out["error"]["user_message"]


def test_notify_enrichment_maps_missing_distribucion_headers():
    raw = {
        "job_id": "j",
        "type": "notify_validar_extractos",
        "status": "failed",
        "error": "missing_distribucion_headers",
    }
    out = enrich_job_for_http_response(raw)
    assert out["error"]["error_code"] == "missing_distribucion_headers"
    assert out["error"]["error_code"] != "unknown_error"


def test_notify_accepts_historico_with_estado_header_nuevo():
    hist_path = "HIST/cartera_estado_nuevo.xlsx"

    async def fake_download(_graph, _site, _drive, path):
        assert path == hist_path
        return _minimal_historico_xlsx(estado_header="Estado", ruta_header="Ruta")

    class _GraphOk:
        async def get_bytes(self, *_a, **_k):
            return _minimal_bank_xlsx()

        async def post_json(self, *_a, **_k):
            return {}, 202

        async def put_bytes(self, *_a, **_k):
            return {}

    async def run():
        with (
            patch(
                "app.application.use_cases.send_validar_extractos_notification.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value={
                    "site_id": "s1",
                    "drive_id": "d1",
                    "path_encoded": "bank/report.xlsx",
                },
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._load_sender_and_recipients_from_correos_xlsx",
                new_callable=AsyncMock,
                return_value=("sender@example.com", ["other@example.com"]),
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._graph_download_by_path",
                new_callable=AsyncMock,
                side_effect=fake_download,
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification.update_merge_control_workbook_after_notify",
                new_callable=AsyncMock,
                return_value=_MC_SKIP_OUTCOME,
            ),
        ):
            g = _GraphOk()
            r = await send_validar_extractos_notification_email(
                g,
                historical_file_path=hist_path,
            )
        assert r.historical_file_path == hist_path

    asyncio.run(run())


def test_notify_enrichment_maps_historical_file_not_found_prefix():
    raw = {
        "job_id": "j",
        "type": "notify_validar_extractos",
        "status": "failed",
        "error": "historical_file_not_found|HTTP 404 url=x detail='y'",
    }
    out = enrich_job_for_http_response(raw)
    assert out["error"]["error_code"] == "historical_file_not_found"


def test_notify_use_case_requires_historical_file_path():
    async def run():
        g = MagicMock()
        with pytest.raises(ValueError, match="missing_historical_file_path"):
            await send_validar_extractos_notification_email(g, historical_file_path=None)

    asyncio.run(run())


def test_notify_use_case_rejects_whitespace_only_path():
    async def run():
        g = MagicMock()
        with pytest.raises(ValueError, match="missing_historical_file_path"):
            await send_validar_extractos_notification_email(g, historical_file_path="  \t  ")

    asyncio.run(run())


def test_notify_uses_explicit_historical_file_path_and_skips_auto_resolve():
    hist_path = "ONLY/EXPLICIT/cartera_validada_2026-05-12.xlsx"
    downloaded: list[str] = []

    async def fake_download(_graph, _site, _drive, path):
        downloaded.append(path)
        return _minimal_historico_xlsx()

    resolve_auto = AsyncMock(side_effect=RuntimeError("auto-resolve must not run"))

    class _GraphOk:
        async def get_bytes(self, *_a, **_k):
            return _minimal_bank_xlsx()

        async def post_json(self, *_a, **_k):
            return {}, 202

        async def put_bytes(self, *_a, **_k):
            return {}

    async def run():
        with (
            patch(
                "app.application.use_cases.send_validar_extractos_notification.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value={
                    "site_id": "s1",
                    "drive_id": "d1",
                    "path_encoded": "bank/report.xlsx",
                },
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._load_sender_and_recipients_from_correos_xlsx",
                new_callable=AsyncMock,
                return_value=("sender@example.com", ["other@example.com"]),
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._graph_download_by_path",
                new_callable=AsyncMock,
                side_effect=fake_download,
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._resolve_historico_excel_path",
                resolve_auto,
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification.update_merge_control_workbook_after_notify",
                new_callable=AsyncMock,
                return_value=_MC_SKIP_OUTCOME,
            ),
        ):
            g = _GraphOk()
            r = await send_validar_extractos_notification_email(
                g,
                historical_file_path=hist_path,
            )
        assert downloaded == [hist_path.strip().strip("/")]
        assert r.historical_file_path == hist_path.strip().strip("/")
        assert r.historical_file_source == "explicit"
        assert r.historico_excel_path == r.historical_file_path
        resolve_auto.assert_not_awaited()

    asyncio.run(run())


def test_notify_to_cc_overrides_still_supported():
    hist_path = "HIST/a.xlsx"
    downloaded: list[str] = []

    async def fake_download(_graph, _site, _drive, path):
        downloaded.append(path)
        return _minimal_historico_xlsx()

    class _GraphOk:
        async def get_bytes(self, *_a, **_k):
            return _minimal_bank_xlsx()

        async def post_json(self, endpoint, body):
            self.last_body = body
            return {}, 202

        async def put_bytes(self, *_a, **_k):
            return {}

    async def run():
        with (
            patch(
                "app.application.use_cases.send_validar_extractos_notification.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value={
                    "site_id": "s1",
                    "drive_id": "d1",
                    "path_encoded": "bank/report.xlsx",
                },
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._load_sender_and_recipients_from_correos_xlsx",
                new_callable=AsyncMock,
                return_value=("sender@example.com", ["ignored@example.com"]),
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._graph_download_by_path",
                new_callable=AsyncMock,
                side_effect=fake_download,
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification.update_merge_control_workbook_after_notify",
                new_callable=AsyncMock,
                return_value=_MC_SKIP_OUTCOME,
            ),
        ):
            g = _GraphOk()
            await send_validar_extractos_notification_email(
                g,
                historical_file_path=hist_path,
                to_override="override1@example.com; override2@example.com",
                cc_override="cc@example.com",
            )
            msg = g.last_body["message"]
            to_addrs = {x["emailAddress"]["address"] for x in msg["toRecipients"]}
            assert "override1@example.com" in to_addrs
            assert "override2@example.com" in to_addrs
            cc_addrs = {x["emailAddress"]["address"] for x in msg.get("ccRecipients", [])}
            assert "cc@example.com" in cc_addrs

    asyncio.run(run())


def test_existing_notify_recipient_resolution_from_excel_remains_unchanged():
    """Sin to_override se usan destinatarios del mock de CORREOS."""

    hist_path = "HIST/a.xlsx"

    async def fake_download(_graph, _site, _drive, path):
        return _minimal_historico_xlsx()

    class _GraphOk:
        async def get_bytes(self, *_a, **_k):
            return _minimal_bank_xlsx()

        async def post_json(self, endpoint, body):
            self.last_body = body
            return {}, 202

        async def put_bytes(self, *_a, **_k):
            return {}

    async def run():
        with (
            patch(
                "app.application.use_cases.send_validar_extractos_notification.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value={
                    "site_id": "s1",
                    "drive_id": "d1",
                    "path_encoded": "bank/report.xlsx",
                },
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._load_sender_and_recipients_from_correos_xlsx",
                new_callable=AsyncMock,
                return_value=("sender@example.com", ["fromexcel@example.com"]),
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification._graph_download_by_path",
                new_callable=AsyncMock,
                side_effect=fake_download,
            ),
            patch(
                "app.application.use_cases.send_validar_extractos_notification.update_merge_control_workbook_after_notify",
                new_callable=AsyncMock,
                return_value=_MC_SKIP_OUTCOME,
            ),
        ):
            g = _GraphOk()
            await send_validar_extractos_notification_email(
                g,
                historical_file_path=hist_path,
            )
            to_addrs = [x["emailAddress"]["address"] for x in g.last_body["message"]["toRecipients"]]
            assert to_addrs == ["fromexcel@example.com"]

    asyncio.run(run())
