"""Tests utilidades tabla de amortización."""

import io
from datetime import date

import openpyxl
import pytest

from app.application.services.accounting_pdf_parser import (
    ACCOUNT_CAPITAL,
    ACCOUNT_VALOR_PAGADO_CLIENTE,
    PaymentApplicationEvent,
)
from app.application.services.amortization_workbook import (
    ADOPTADO_EXISTENTE,
    APLICADO,
    AUTOMATION_LOG_SHEET,
    REVISION_MANUAL,
    AmortizationSheetNotFoundError,
    append_automation_log,
    build_target_row_search_debug,
    compare_existing_application,
    detect_amortization_sheet,
    detect_headers,
    ensure_automation_log,
    build_amortization_idempotency_key,
    find_application_row_detailed,
    find_row_by_due_date,
    find_row_by_due_date_detailed,
    is_payment_application_empty,
    load_automation_log_idempotency_keys,
    _row_due_date,
    write_ibr,
    PaymentApplicationWriteOptions,
    write_payment_application,
    _is_formula_value,
)


def _sample_event(**kwargs) -> PaymentApplicationEvent:
    base = dict(
        id_pago="P1",
        cliente="EQUINORTE",
        credito="258",
        asiento_pdf_path="a.pdf",
        comprobante="1",
        fecha_asiento=date(2026, 5, 10),
        valor_pagado_cliente=50_000_000.0,
        capital=49_118_143.0,
        intereses=0.0,
        mora=881_857.0,
        retenciones=0.0,
        saldos_menores=0.0,
        raw_text="",
    )
    base.update(kwargs)
    return PaymentApplicationEvent(**base)


def _amort_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabla"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "IBR +i",
            "Spread",
            "Cuota",
            "Cuota + i",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "intereses mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([15, 5, 2026, None, None, 1000, 1100, None, None, None, None, None, None])
    ws.append([22, 5, 2026, None, None, 1000, 1100, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _equinorte_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "IBR +i",
            "Spread",
            "Cuota + i",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "Intereses de mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([22, 5, 2026, None, None, 1000, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _headers_on_row_workbook_bytes(header_row: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    for _ in range(header_row - 1):
        ws.append([None] * 12)
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "IBR +i",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "intereses mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([22, 5, 2026, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _duplicate_dia_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "dia",
            "Causac Inter Mes",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "intereses mora",
            "Valor pagado cliente",
            "Saldos Menores",
        ]
    )
    ws.append([22, 4, 2026, 99, None, None, None, None, None, None, None])
    ws.append([15, 4, 2026, 88, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _float_date_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabla"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "Valor pagado cliente",
        ]
    )
    ws.append([22.0, 4.0, 2026.0, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_duplicate_dia_uses_first_contiguous_group():
    wb = openpyxl.load_workbook(io.BytesIO(_duplicate_dia_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    assert match.headers["dia"] == 1
    assert match.headers["mes"] == 2
    assert match.headers["anio"] == 3
    row = find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 4, 22),
        header_row=match.header_row,
    )
    assert row == 2
    assert match.worksheet.cell(row, match.headers["dia"]).value == 22


def test_find_row_april_22_2026():
    wb = openpyxl.load_workbook(io.BytesIO(_duplicate_dia_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    assert find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 4, 22),
        header_row=match.header_row,
    ) == 2


def test_find_row_float_date_components():
    wb = openpyxl.load_workbook(io.BytesIO(_float_date_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    row = find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 4, 22),
        header_row=match.header_row,
    )
    assert row == 2


def _monthly_schedule_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EQUINORTE"
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "Valor pagado cliente",
        ]
    )
    ws.append([22, 12, 2025, None, None, None, None])
    for r in range(3, 8):
        ws.cell(r, 1, f'=IF(A{r}<>" ",1,1)')
        ws.cell(r, 2, f'=IF(A{r}<>" ",1,1)')
        ws.cell(r, 3, f'=IF(A{r}<>" ",1,1)')
        ws.cell(r, 7, 1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_formula_cells_do_not_produce_false_dates():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "dia",
            "mes",
            "año",
            "Fecha pago",
            "Valor intereses",
            "Abono a K",
            "Valor pagado cliente",
        ]
    )
    ws.cell(2, 1, '=IF(A5<>" ",+L4," ")')
    ws.cell(2, 2, "=IF(AND(A5<>\" \")*(M4=12),1,1+M4)")
    ws.cell(2, 3, "=IF(AND(A5<>\" \")*(M5=1),1,1+N4,N4)")
    h = detect_headers(ws, header_row=1)
    assert _row_due_date(ws, 2, h) is None
    assert find_row_by_due_date(ws, h, date(2005, 5, 5), header_row=1) is None


def test_monthly_schedule_finds_april_22_2026():
    wb = openpyxl.load_workbook(
        io.BytesIO(_monthly_schedule_workbook_bytes()), data_only=True
    )
    match = detect_amortization_sheet(wb)
    result = find_row_by_due_date_detailed(
        match.worksheet,
        match.headers,
        date(2026, 4, 22),
        header_row=match.header_row,
    )
    assert result.monthly_schedule_fallback_used is True
    assert result.row == 6
    assert result.monthly_schedule_base_date == "2025-12-22"


def test_monthly_schedule_out_of_range_returns_none():
    wb = openpyxl.load_workbook(
        io.BytesIO(_monthly_schedule_workbook_bytes()), data_only=True
    )
    match = detect_amortization_sheet(wb)
    result = find_row_by_due_date_detailed(
        match.worksheet,
        match.headers,
        date(2030, 1, 22),
        header_row=match.header_row,
    )
    assert result.row is None


def test_target_row_debug_includes_date_columns():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    debug = build_target_row_search_debug(
        match.worksheet,
        match.headers,
        date(2099, 1, 1),
        header_row=match.header_row,
        sheet_name=match.worksheet.title,
        tabla_amortizacion_path="TABLAS/x.xlsx",
    )
    assert debug["target_date"] == "2099-01-01"
    assert debug["date_header_columns"] == {"dia": 1, "mes": 2, "anio": 3}
    assert debug["header_row"] == match.header_row
    assert debug["sheet_name"] == "Tabla"
    assert debug["tabla_amortizacion_path"] == "TABLAS/x.xlsx"
    assert any(c["date"] == "2026-05-22" for c in debug["candidate_date_rows"])


def test_detect_headers_and_sheet():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws = match.worksheet
    h = match.headers
    assert h["dia"] == 1
    assert h["mes"] == 2
    assert h["anio"] == 3
    assert h["ibr_i"] == 4
    assert "cuota" in h


def test_detect_sheet_named_equinorte_by_headers():
    wb = openpyxl.load_workbook(io.BytesIO(_equinorte_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    assert match.worksheet.title == "EQUINORTE"
    assert "dia" in match.headers
    assert match.headers["intereses_mora"] is not None
    assert "ibr_i" in match.headers


def test_detect_headers_on_row_4():
    wb = openpyxl.load_workbook(io.BytesIO(_headers_on_row_workbook_bytes(4)))
    match = detect_amortization_sheet(wb)
    assert match.header_row == 4
    assert match.headers["valor_pagado_cliente"] > 0
    row = find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 5, 22),
        header_row=match.header_row,
    )
    assert row == 5


def test_detect_intereses_de_mora_and_ibr_plus_i_aliases():
    wb = openpyxl.load_workbook(io.BytesIO(_equinorte_workbook_bytes()))
    h = detect_headers(wb["EQUINORTE"])
    assert "intereses_mora" in h
    assert "ibr_i" in h


def test_detect_amortization_sheet_raises_with_debug():
    wb = openpyxl.Workbook()
    wb.active.title = "SoloResumen"
    wb.active.append(["Total", "Monto"])
    wb.active.append([1, 2])
    with pytest.raises(AmortizationSheetNotFoundError) as excinfo:
        detect_amortization_sheet(wb, tabla_amortizacion_path="TABLAS/x.xlsx")
    exc = excinfo.value
    assert "SoloResumen" in exc.workbook_sheets
    assert exc.tabla_amortizacion_path == "TABLAS/x.xlsx"
    assert exc.required_headers_missing


def test_find_row_by_due_date():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    row = find_row_by_due_date(
        match.worksheet, match.headers, date(2026, 5, 22), header_row=match.header_row
    )
    assert row == 3


def test_find_row_by_due_date_excludes_used_rows():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    first = find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 5, 22),
        header_row=match.header_row,
    )
    second = find_row_by_due_date(
        match.worksheet,
        match.headers,
        date(2026, 5, 22),
        exclude_rows=frozenset({first}),
        header_row=match.header_row,
    )
    assert first == 3
    assert second is None


def test_build_amortization_idempotency_key_includes_asiento_and_comprobante():
    key = build_amortization_idempotency_key(
        "7785e37e",
        "CREDITO # 258",
        "clientes/E/asiento.pdf",
        "99-1",
        pdf_hash="abc",
    )
    assert "7785e37e" in key
    assert "asiento.pdf" in key
    assert "99-1" in key
    assert "abc" in key


def test_ensure_automation_log_created():
    wb = openpyxl.Workbook()
    ws = ensure_automation_log(wb)
    assert ws.title == AUTOMATION_LOG_SHEET
    assert ws.cell(1, 1).value == "Timestamp"


def test_write_empty_row_aplicado():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.headers
    row = find_row_by_due_date(ws, h, date(2026, 5, 22), header_row=match.header_row)
    assert row is not None
    ev = _sample_event()
    assert compare_existing_application(ws, row, h, ev) == APLICADO
    opts = PaymentApplicationWriteOptions(
        payment_date=date(2026, 5, 10),
        detected_codes=frozenset({ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL}),
    )
    write_payment_application(ws, row, h, ev, write_options=opts, header_row=match.header_row)
    assert not is_payment_application_empty(ws, row, h)


def test_existing_matching_adoptado():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.headers
    row = find_row_by_due_date(ws, h, date(2026, 5, 22), header_row=match.header_row)
    ev = _sample_event()
    opts = PaymentApplicationWriteOptions(
        payment_date=date(2026, 5, 10),
        detected_codes=frozenset({ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL}),
    )
    write_payment_application(ws, row, h, ev, write_options=opts, header_row=match.header_row)
    assert compare_existing_application(
        ws, row, h, ev, payment_date=date(2026, 5, 10), detected_codes=opts.detected_codes
    ) == ADOPTADO_EXISTENTE


def test_existing_mismatch_revision_manual():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.headers
    row = find_row_by_due_date(ws, h, date(2026, 5, 22), header_row=match.header_row)
    ev = _sample_event()
    opts = PaymentApplicationWriteOptions(
        payment_date=date(2026, 5, 10),
        detected_codes=frozenset({ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL}),
    )
    write_payment_application(ws, row, h, ev, write_options=opts, header_row=match.header_row)
    other = _sample_event(capital=1.0)
    assert compare_existing_application(ws, row, h, other) == REVISION_MANUAL


def test_write_ibr_on_corte_row():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.headers
    row = find_row_by_due_date(ws, h, date(2026, 5, 22), header_row=match.header_row)
    write_ibr(ws, row, h, 0.10580)
    assert ws.cell(row, h["ibr_i"]).value == pytest.approx(0.10580)


def test_append_automation_log():
    wb = openpyxl.Workbook()
    ensure_automation_log(wb)
    append_automation_log(
        wb,
        {
            "id_pago": "P1",
            "cliente": "X",
            "credito": "258",
            "fila": 3,
            "accion": APLICADO,
            "detalle": "ok",
            "idempotency_key": "k1",
            "asiento_pdf_path": "a.pdf",
            "application_row": 3,
            "ibr_row": 3,
        },
    )
    ws = wb[AUTOMATION_LOG_SHEET]
    assert ws.cell(2, 2).value == "P1"
    assert ws.cell(2, 6).value == APLICADO
    assert ws.cell(2, 8).value == "k1"
    assert load_automation_log_idempotency_keys(wb) == {"k1"}


def test_find_application_row_skips_occupied_and_uses_next_free():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.header_row
    headers = match.headers
    due = find_row_by_due_date(ws, headers, date(2026, 5, 22), header_row=h)
    assert due == 3
    ws.cell(3, headers["valor_pagado_cliente"], 99.0)
    ws.cell(4, headers["valor_pagado_cliente"], None)
    ev = _sample_event()
    result = find_application_row_detailed(
        ws, headers, ev, due_date_row=due, header_row=h
    )
    assert result.row == 4
    assert result.compare_status == APLICADO


def test_find_application_row_adopts_matching_row_below_due():
    wb = openpyxl.load_workbook(io.BytesIO(_amort_workbook_bytes()))
    match = detect_amortization_sheet(wb)
    ws, h = match.worksheet, match.headers
    due = find_row_by_due_date(ws, match.headers, date(2026, 5, 22), header_row=match.header_row)
    ev = _sample_event()
    opts = PaymentApplicationWriteOptions(
        payment_date=date(2026, 5, 10),
        detected_codes=frozenset({ACCOUNT_VALOR_PAGADO_CLIENTE, ACCOUNT_CAPITAL}),
    )
    write_payment_application(
        ws, 4, match.headers, ev, write_options=opts, header_row=match.header_row
    )
    result = find_application_row_detailed(
        ws, match.headers, ev, due_date_row=due, header_row=match.header_row
    )
    assert result.row == 4
    assert result.compare_status == ADOPTADO_EXISTENTE
