"""Merge composite: control_merge_pdfs.xlsx, asientos y actualización de estado."""

from __future__ import annotations

import asyncio
import inspect
import unicodedata
from datetime import date
from io import BytesIO
from unittest.mock import AsyncMock, patch
from urllib.parse import unquote

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

from app.application.job_status_enrichment import enrich_job_for_http_response
from app.application.sharepoint_resolution import encode_graph_drive_path
from app.application.use_cases.merge_composite_validado_pdfs import (
    _classify_asiento_pdf_names,
    _credit_number_from_extract_parent,
    _credit_number_from_folder_segment,
    _merge_skip_line,
    _pick_single_asiento_pdf,
    merge_composite_validado_pdfs,
)
from app.application.use_cases.merge_control_workbook_merge import (
    MergeControlValidationError,
    merge_control_parse_row2_validate_for_merge,
    merge_control_upload_row2_updates,
)
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_COLUMNS,
    SHEET_NAME,
    apply_merge_control_worksheet_protection,
)


def _tiny_pdf() -> bytes:
    w = PdfWriter()
    w.add_blank_page(width=72, height=72)
    bio = BytesIO()
    w.write(bio)
    return bio.getvalue()


def _control_row_bytes(
    *,
    estado: str,
    is_active: bool | str,
    hist: str,
    email: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(list(MERGE_CONTROL_COLUMNS))
    row = {c: "" for c in MERGE_CONTROL_COLUMNS}
    row["Title"] = "T"
    row["EstadoProceso"] = estado
    row["IsActive"] = is_active
    row["HistoricalFilePath"] = hist
    row["EmailPdfPath"] = email
    row["MergeOutputCount"] = 0
    row["MergeSkippedCount"] = 0
    row["LastErrorUserMessage"] = ""
    row["LastErrorNextAction"] = ""
    row["CreatedAtProceso"] = ""
    row["LastUpdatedAtProceso"] = ""
    ws.append([row[c] for c in MERGE_CONTROL_COLUMNS])
    apply_merge_control_worksheet_protection(ws)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _bank_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Concepto", "Crédito", "Monto"])
    ws.append([date(2026, 5, 12), "abono", "264", "1000"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


_HIST_DIST_HEADERS = [
    "Estado línea",
    "Ruta",
    "ID Pago",
    "Cliente",
    "Crédito",
    "RutaAsientosContables",
]


def _hist_workbook_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(_HIST_DIST_HEADERS)
    for row in rows:
        ws.append(row)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _hist_bytes() -> bytes:
    return _hist_workbook_bytes(
        [
            [
                "VALIDAR",
                "",
                "P1",
                "Acme",
                "264",
                "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264",
            ]
        ]
    )


class _MergeGraph:
    def __init__(self) -> None:
        self.initial: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}
        self.children: dict[str, list] = {}
        self.deleted: list[str] = []
        self.put_fail_substr: str | None = None

    def _path_from_content_ep(self, endpoint: str) -> str:
        return unquote(endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0])

    def _path_from_delete_ep(self, endpoint: str) -> str:
        return unquote(endpoint.split("/root:/", 1)[1].rstrip(":"))

    async def get(self, endpoint: str, params=None):
        if ":/children" in endpoint and "/root:/" in endpoint:
            raw = endpoint.split("/root:/", 1)[1].rsplit(":/children", 1)[0]
            path = unquote(raw)
            return {"value": self.children.get(path, [])}
        return {"value": []}

    async def get_bytes(self, endpoint: str, params=None):
        path = self._path_from_content_ep(endpoint)
        if path in self.uploaded:
            return self.uploaded[path]
        if path in self.initial:
            return self.initial[path]
        raise httpx.HTTPStatusError(
            "not found",
            request=httpx.Request("GET", endpoint),
            response=httpx.Response(404),
        )

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str | None = None):
        path = self._path_from_content_ep(endpoint)
        if self.put_fail_substr and self.put_fail_substr in path:
            raise RuntimeError("mock put failure")
        self.uploaded[path] = content
        return {}

    async def delete(self, endpoint: str) -> None:
        self.deleted.append(self._path_from_delete_ep(endpoint))

    async def post_json(self, endpoint: str, body: dict):
        return {}, 202


def test_merge_fails_when_historical_missing_ruta_asientos_contables_column(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    hist = "HIST/no_asientos_col.xlsx"
    email = "EMAIL/mail.pdf"
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(["Estado línea", "Ruta", "ID Pago", "Cliente", "Crédito"])
    ws.append(["VALIDAR", "", "P1", "Acme", "264"])
    bio = BytesIO()
    wb.save(bio)
    g.initial[hist] = bio.getvalue()
    g.initial[email] = _tiny_pdf()

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def run():
        with patch(
            "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
            new_callable=AsyncMock,
            return_value=ctx,
        ):
            with pytest.raises(ValueError, match="RutaAsientosContables"):
                await merge_composite_validado_pdfs(g)

    asyncio.run(run())


def test_merge_reads_historical_and_email_paths_from_control_workbook():
    hist = "HIST/cartera.xlsx"
    email = "EMAIL/correo.pdf"
    raw = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    snap = merge_control_parse_row2_validate_for_merge(raw)
    assert snap.historical_file_path == hist
    assert snap.email_pdf_path == email


def test_merge_fails_when_no_pending_control_process():
    raw = _control_row_bytes(
        estado="CONSOLIDADO",
        is_active=False,
        hist="H/a.xlsx",
        email="E/m.pdf",
    )
    with pytest.raises(MergeControlValidationError) as ei:
        merge_control_parse_row2_validate_for_merge(raw)
    assert ei.value.error_code == "merge_control_no_pending_process"


def test_merge_fails_when_control_missing_historical_file_path():
    raw = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist="",
        email="E/m.pdf",
    )
    with pytest.raises(MergeControlValidationError) as ei:
        merge_control_parse_row2_validate_for_merge(raw)
    assert ei.value.error_code == "missing_historical_file_path"


def test_merge_fails_when_control_missing_email_pdf_path():
    raw = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist="H/a.xlsx",
        email="",
    )
    with pytest.raises(MergeControlValidationError) as ei:
        merge_control_parse_row2_validate_for_merge(raw)
    assert ei.value.error_code == "missing_email_pdf_path"


def test_merge_sets_control_status_consolidando_before_work(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    raw = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist="H/h.xlsx",
        email="E/e.pdf",
    )
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = raw

    from app.application.use_cases.merge_control_workbook_merge import merge_control_set_consolidando

    async def _run():
        await merge_control_set_consolidando(g, "s1", "d1", "CTL/control.xlsx")

    asyncio.run(_run())
    assert "CTL/control.xlsx" in g.uploaded
    wb = load_workbook(BytesIO(g.uploaded["CTL/control.xlsx"]), data_only=True)
    try:
        ws = wb[SHEET_NAME]
        c_est = MERGE_CONTROL_COLUMNS.index("EstadoProceso") + 1
        assert ws.cell(2, c_est).value == "CONSOLIDANDO"
    finally:
        wb.close()


def test_merge_updates_control_status_consolidado_on_success(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    raw = _control_row_bytes(
        estado="CONSOLIDANDO",
        is_active=True,
        hist="H/h.xlsx",
        email="E/e.pdf",
    )
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = raw
    from app.application.use_cases.merge_control_workbook_merge import merge_control_set_consolidado_success

    async def _run():
        await merge_control_set_consolidado_success(g, "s1", "d1", "CTL/control.xlsx", outputs_count=3)

    asyncio.run(_run())
    wb = load_workbook(BytesIO(g.uploaded["CTL/control.xlsx"]), data_only=True)
    try:
        ws = wb[SHEET_NAME]
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("EstadoProceso") + 1).value == "CONSOLIDADO"
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("IsActive") + 1).value is False
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("MergeOutputCount") + 1).value == 3
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("MergeSkippedCount") + 1).value == 0
    finally:
        wb.close()


def test_merge_updates_control_status_merge_parcial_on_skipped(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    raw = _control_row_bytes(
        estado="CONSOLIDANDO",
        is_active=True,
        hist="H/h.xlsx",
        email="E/e.pdf",
    )
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = raw
    from app.application.use_cases.merge_control_workbook_merge import merge_control_set_merge_parcial

    async def _run():
        await merge_control_set_merge_parcial(
            g,
            "s1",
            "d1",
            "CTL/control.xlsx",
            outputs_count=1,
            skipped_count=2,
            user_message="Hubo omisiones.",
            next_action="Revise skipped.",
        )

    asyncio.run(_run())
    wb = load_workbook(BytesIO(g.uploaded["CTL/control.xlsx"]), data_only=True)
    try:
        ws = wb[SHEET_NAME]
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("EstadoProceso") + 1).value == "MERGE_PARCIAL"
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("MergeSkippedCount") + 1).value == 2
    finally:
        wb.close()


def test_merge_updates_control_status_error_merge_on_failure(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    raw = _control_row_bytes(
        estado="CONSOLIDANDO",
        is_active=True,
        hist="H/h.xlsx",
        email="E/e.pdf",
    )
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = raw
    from app.application.use_cases.merge_control_workbook_merge import merge_control_set_error_merge

    async def _run():
        await merge_control_set_error_merge(
            g,
            "s1",
            "d1",
            "CTL/control.xlsx",
            user_message="Falló algo.",
            next_action="Reintente.",
        )

    asyncio.run(_run())
    wb = load_workbook(BytesIO(g.uploaded["CTL/control.xlsx"]), data_only=True)
    try:
        ws = wb[SHEET_NAME]
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("EstadoProceso") + 1).value == "ERROR_MERGE"
        assert ws.cell(2, MERGE_CONTROL_COLUMNS.index("IsActive") + 1).value is True
    finally:
        wb.close()


def test_merge_does_not_call_auto_historical_resolver():
    src = inspect.getsource(merge_composite_validado_pdfs)
    assert "_resolve_historico_excel_path" not in src


def test_merge_does_not_call_email_pdf_picker():
    src = inspect.getsource(merge_composite_validado_pdfs)
    assert "_pick_email_export_pdf" not in src


def test_credit_number_from_extract_parent_uses_immediate_folder():
    ep = "clientes/GEOEXCON/CREDITO # 254/Extracto 2025-12-23 CREDITO # 254.pdf"
    assert _credit_number_from_extract_parent(ep) == "254"


def test_credit_number_from_folder_segment_accepts_credito_accent():
    seg = unicodedata.normalize("NFC", "CRÉDITO # 254")
    assert _credit_number_from_folder_segment(seg) == "254"


def test_credit_number_from_folder_segment_fullwidth_hash():
    seg = "CREDITO" + "\uff03" + " 231"
    assert _credit_number_from_folder_segment(seg) == "231"


@pytest.mark.parametrize(
    "suffix",
    [
        "TERMINADO",
        "FINALIZADO",
        "CANCELADO",
        "PAGADO",
        "LIQUIDADO",
        "VIGENTE",
        "REPUESTOS",
    ],
)
def test_credit_number_from_folder_segment_accepts_status_suffixes(suffix: str):
    assert _credit_number_from_folder_segment(f"CREDITO # 254 {suffix}") == "254"


def test_credit_number_from_folder_segment_credito_without_hash():
    assert _credit_number_from_folder_segment("CREDITO 254") == "254"


@pytest.mark.parametrize(
    "seg",
    [
        "DOCUMENTOS",
        "ASIENTOS CONTABLES",
        "GEOEXCON",
        "254",
        "Extracto 2025",
    ],
)
def test_credit_number_from_folder_segment_rejects_non_credit_folders(seg: str):
    assert _credit_number_from_folder_segment(seg) == ""


def test_merge_skip_line_includes_fields():
    s = _merge_skip_line(
        "id1",
        "asiento_contable_credit_mismatch",
        credit_number_expected="264",
        asiento_pdf_found="Asiento 254.pdf",
        asiento_folder_path="c/a/ASIENTOS",
        extracto_path="c/a/ex.pdf",
        names_seen="a.pdf, b.pdf",
    )
    assert "id_pago=id1" in s
    assert "reason=asiento_contable_credit_mismatch" in s
    assert "credit_number_expected=264" in s
    assert "asiento_pdf_found=Asiento 254.pdf" in s


def test_merge_multi_credit_two_extracts_validates_each_credit(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    p231 = "clientes/GEO/CREDITO # 231/Extracto 231.pdf"
    p254 = "clientes/GEO/CREDITO # 254/Extracto 254.pdf"
    d231 = "clientes/GEO/CREDITO # 231/ASIENTOS CONTABLES CRED 231"
    d254 = "clientes/GEO/CREDITO # 254/ASIENTOS CONTABLES CRED 254"
    a231 = f"{d231}/Asiento 2025-12-23 CREDITO # 231.pdf"
    a254 = f"{d254}/Asiento 2025-12-23 CREDITO # 254.pdf"

    pdf = _tiny_pdf()
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_workbook_bytes(
        [
            ["VALIDAR", "", "G1", "GEO", "231", d231],
            ["VALIDAR", "", "G1", "GEO", "254", d254],
        ]
    )
    g.initial[email] = pdf
    g.initial[p231] = pdf
    g.initial[p254] = pdf
    g.initial[a231] = pdf
    g.initial[a254] = pdf
    g.children[d231] = [{"name": "Asiento 2025-12-23 CREDITO # 231.pdf", "file": {}}]
    g.children[d254] = [{"name": "Asiento 2025-12-23 CREDITO # 254.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }
    ret = iter([[p231], [p254]])

    async def fake_collect(_gr, _si, _dr, _cell):
        return next(ret)

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.skipped_count == 0
    assert g.deleted == []
    assert r.merge_control_status == "CONSOLIDADO"


def test_merge_parcial_no_delete_when_second_id_fails(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    p_ok = "clientes/ACME/CREDITO # 264/Extracto.pdf"
    d_ok = "clientes/ACME/CREDITO # 264/ASIENTOS CONTABLES CRED 264"
    a_ok = f"{d_ok}/asiento_264.pdf"
    p_bad = "clientes/BETA/CREDITO # 999/Extracto.pdf"
    d_bad = "clientes/BETA/CREDITO # 999/ASIENTOS CONTABLES CRED 999"

    pdf = _tiny_pdf()
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_workbook_bytes(
        [
            ["VALIDAR", "", "PAY_A", "Acme", "264", d_ok],
            ["VALIDAR", "", "PAY_B", "Beta", "999", d_bad],
        ]
    )
    g.initial[email] = pdf
    g.initial[p_ok] = pdf
    g.initial[a_ok] = pdf
    g.initial[p_bad] = pdf
    g.children[d_ok] = [{"name": "asiento_264.pdf", "file": {}}]
    g.children[d_bad] = []

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }
    ret = iter([[p_ok], [p_bad]])

    async def fake_collect(_gr, _si, _dr, _cell):
        return next(ret)

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.skipped_count >= 1
    assert r.merge_control_status == "MERGE_PARCIAL"
    assert g.deleted == []


def test_merge_multi_credit_credit_number_not_resolved_when_no_folder_nor_row_credit(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    p1 = "clientes/GEO/SIN_CREDITO/ex1.pdf"
    p2 = "clientes/GEO/SIN_CREDITO/ex2.pdf"

    pdf = _tiny_pdf()
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    d_sin = "clientes/GEO/SIN_CREDITO/ASIENTOS CONTABLES CRED 0"
    g.initial[hist] = _hist_workbook_bytes(
        [
            ["VALIDAR", "", "G1", "GEO", "", d_sin],
            ["VALIDAR", "", "G1", "GEO", "", d_sin],
        ]
    )
    g.initial[email] = pdf
    g.initial[p1] = pdf
    g.initial[p2] = pdf
    g.children[d_sin] = []

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }
    ret = iter([[p1], [p2]])

    async def fake_collect(_gr, _si, _dr, _cell):
        return next(ret)

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 0
    assert r.skipped_count >= 1
    assert any("credit_number_not_resolved" in s for s in r.skipped)
    assert g.deleted == []


def test_merge_id_pago_two_credits_missing_one_asiento_partial_merge(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    p231 = "clientes/GEO/CREDITO # 231/Extracto 231.pdf"
    p254 = "clientes/GEO/CREDITO # 254/Extracto 254.pdf"
    d231 = "clientes/GEO/CREDITO # 231/ASIENTOS CONTABLES"
    a231 = f"{d231}/Asiento 231.pdf"
    d254 = "clientes/GEO/CREDITO # 254/ASIENTOS CONTABLES CRED 254"

    pdf = _tiny_pdf()
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_workbook_bytes(
        [
            ["VALIDAR", "", "G1", "GEO", "231", d231],
            ["VALIDAR", "", "G1", "GEO", "254", d254],
        ]
    )
    g.initial[email] = pdf
    g.initial[p231] = pdf
    g.initial[p254] = pdf
    g.initial[a231] = pdf
    g.children[d231] = [{"name": "Asiento 231.pdf", "file": {}}]
    g.children[d254] = []

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }
    ret = iter([[p231], [p254]])

    async def fake_collect(_gr, _si, _dr, _cell):
        return next(ret)

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.merge_control_status == "MERGE_PARCIAL"
    assert any("asiento_contable_not_found" in s for s in r.skipped)
    assert r.outputs[0].asiento_pdf_paths == (a231,)
    assert any(p.startswith("OUT/PDFS/") for p in g.uploaded)
    assert g.deleted == []


def test_merge_terminado_credit_folder_consolidates(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/GEOEXCON/CREDITO # 254 TERMINADO/Extracto 254.pdf"
    asiento_dir = "clientes/GEOEXCON/CREDITO # 254 TERMINADO/ASIENTOS CONTABLES CRED 254"
    asiento_rel = f"{asiento_dir}/Asiento 254.pdf"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_workbook_bytes(
        [["VALIDAR", "", "P254", "GEOEXCON", "254", asiento_dir]]
    )
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_rel] = _tiny_pdf()
    g.children[asiento_dir] = [{"name": "Asiento 254.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_gr, _si, _dr, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.skipped_count == 0
    assert not any("credit_number_not_resolved" in s for s in r.skipped)
    assert g.deleted == []
    assert r.outputs[0].asiento_pdf_path == asiento_rel


def test_merge_terminado_extract_wrong_asiento_folder_not_credit_unresolved(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/GEOEXCON/CREDITO # 254 TERMINADO/Extracto 254.pdf"
    asiento_dir_254 = "clientes/GEOEXCON/CREDITO # 254 TERMINADO/ASIENTOS CONTABLES CRED 254"
    asiento_dir_231 = "clientes/GEOEXCON/CREDITO # 231/ASIENTOS CONTABLES CRED 231"
    asiento_wrong = f"{asiento_dir_231}/Asiento 231.pdf"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_workbook_bytes(
        [["VALIDAR", "", "P254", "GEOEXCON", "254", asiento_dir_254]]
    )
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_wrong] = _tiny_pdf()
    g.children[asiento_dir_254] = []
    g.children[asiento_dir_231] = [{"name": "Asiento 231.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_gr, _si, _dr, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 0
    assert any("asiento_contable_not_found" in s for s in r.skipped)
    assert not any("credit_number_not_resolved" in s for s in r.skipped)
    assert g.deleted == []


def test_merge_parcial_retry_does_not_duplicate_existing_output(monkeypatch):
    from app.application.use_cases.merge_composite_validado_pdfs import (
        _merge_composite_output_basename,
    )

    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/ACME/CREDITO# 264/Extracto.pdf"
    asiento_dir = "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264"
    asiento_rel = f"{asiento_dir}/asiento_264.pdf"
    out_base = _merge_composite_output_basename(date(2026, 5, 12), "Acme", "264")
    out_rel = f"OUT/PDFS/{out_base}"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_bytes()
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_rel] = _tiny_pdf()
    g.initial[out_rel] = _tiny_pdf()
    g.children[asiento_dir] = [{"name": "asiento_264.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_gr, _si, _dr, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert len([p for p in g.uploaded if p.startswith("OUT/PDFS/")]) == 0
    assert g.deleted == []
    assert any(
        str(o.sources_summary).startswith("already_consolidated") for o in r.outputs
    )


def test_merge_finds_single_asiento_pdf_in_credit_folder():
    name, reason = _pick_single_asiento_pdf(["asiento_264.pdf"], "264")
    assert name == "asiento_264.pdf"
    assert reason is None


def test_merge_skips_when_no_asiento_pdf():
    name, reason = _pick_single_asiento_pdf([], "264")
    assert name is None
    assert reason == "asiento_contable_not_found"


def test_merge_accepts_multiple_valid_asiento_pdfs_for_same_credit():
    valid, rejected = _classify_asiento_pdf_names(
        ["Asiento cuota credito 264.pdf", "Asiento abono capital credito 264.pdf"],
        "264",
    )
    assert len(valid) == 2
    assert rejected == []
    name, reason = _pick_single_asiento_pdf(valid, "264")
    assert name == valid[0]
    assert reason is None


def test_merge_skips_when_asiento_filename_does_not_match_credit_number():
    name, reason = _pick_single_asiento_pdf(["asiento_1264.pdf"], "264")
    assert name is None
    assert reason == "asiento_contable_credit_mismatch"


def test_merge_success_does_not_delete_asientos(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/ACME/CREDITO# 264/Extracto.pdf"
    asiento_dir = "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264"
    asiento_rel = f"{asiento_dir}/asiento_264.pdf"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_bytes()
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_rel] = _tiny_pdf()
    g.children[asiento_dir] = [{"name": "asiento_264.pdf", "file": {}}]

    bank_enc = encode_graph_drive_path("bank/report.xlsx")
    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": bank_enc,
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_g, _s, _d, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.merge_control_status == "CONSOLIDADO"
    assert g.deleted == []
    assert r.outputs[0].asiento_pdf_path == asiento_rel
    assert r.outputs[0].extracto_pdf_path == extract
    assert r.outputs[0].email_pdf_path == email


def test_merge_does_not_delete_asiento_when_upload_fails(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/ACME/CREDITO# 264/Extracto.pdf"
    asiento_dir = "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264"
    asiento_rel = f"{asiento_dir}/asiento_264.pdf"

    g = _MergeGraph()
    g.put_fail_substr = "OUT/PDFS"
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_bytes()
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_rel] = _tiny_pdf()
    g.children[asiento_dir] = [{"name": "asiento_264.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_g, _s, _d, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 0
    assert r.skipped_count >= 1
    assert g.deleted == []


def test_merge_result_includes_outputs_count_and_skipped_count():
    from app.application.use_cases.merge_composite_validado_pdfs import MergeCompositeValidadoPdfsResult

    r = MergeCompositeValidadoPdfsResult(
        report_date_iso="2026-01-01",
        historico_excel_path="h.xlsx",
        estado_linea_contains="VALIDAR",
        email_pdf_used="e.pdf",
        outputs=(),
        skipped=("a: skip",),
        merge_control_file_path="CTL/x.xlsx",
        merge_control_updated=True,
        merge_control_status="MERGE_PARCIAL",
        outputs_count=0,
        skipped_count=1,
    )
    assert r.outputs_count == 0
    assert r.skipped_count == 1


def test_merge_preserves_control_workbook_protection(monkeypatch):
    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    raw = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist="H/h.xlsx",
        email="E/e.pdf",
    )
    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = raw

    async def _run():
        await merge_control_upload_row2_updates(
            g,
            "s1",
            "d1",
            "CTL/control.xlsx",
            updates={"EstadoProceso": "CONSOLIDANDO"},
        )

    asyncio.run(_run())
    wb = load_workbook(BytesIO(g.uploaded["CTL/control.xlsx"]), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        assert ws.protection.sheet is True
        assert ws.cell(1, 1).protection.locked is True
    finally:
        wb.close()


def test_merge_failed_job_enrichment_merge_control_no_pending():
    raw = {
        "job_id": "j",
        "type": "merge_composite_validado_pdfs",
        "status": "failed",
        "error": "merge_control_no_pending_process",
    }
    out = enrich_job_for_http_response(raw)
    assert out["error"]["error_code"] == "merge_control_no_pending_process"
    msg = out["error"]["user_message"].lower()
    assert "unir pdfs" in msg
    assert "proceso activo" in msg
    assert out["error"]["next_action"]


def test_merge_completed_example_success_payload():
    body = {
        "status": "completed",
        "result": {
            "status": "ok",
            "merge_control_updated": True,
            "merge_control_status": "CONSOLIDADO",
            "merge_control_file_path": "CTL/control_merge_pdfs.xlsx",
            "outputs_count": 2,
            "skipped_count": 0,
            "outputs": [
                {
                    "id_pago": "P1",
                    "output_relative_path": "OUT/x.pdf",
                    "bytes_written": 100,
                    "asiento_pdf_path": "clientes/X/asiento.pdf",
                }
            ],
            "skipped": [],
        },
    }
    out = enrich_job_for_http_response({**body, "job_id": "x", "type": "merge_composite_validado_pdfs"})
    assert out["severity"] == "success"
    assert "amortiz" in out["next_action"].lower()
    assert "cerrado en este paso" not in out["next_action"].lower()


def test_merge_writes_manifest_json(monkeypatch):
    import json

    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "LOGS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/ACME/CREDITO# 264/Extracto.pdf"
    asiento_dir = "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264"
    asiento_rel = f"{asiento_dir}/asiento_264.pdf"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_bytes()
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_rel] = _tiny_pdf()
    g.children[asiento_dir] = [{"name": "asiento_264.pdf", "file": {}}]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_g, _s, _d, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.merge_manifest_path.startswith("LOGS/merge_manifest_")
    manifest_key = next(k for k in g.uploaded if k.endswith(".json"))
    data = json.loads(g.uploaded[manifest_key].decode("utf-8"))
    assert data["historico_excel_path"] == hist
    assert len(data["outputs"]) == 1
    assert data["outputs"][0]["asiento_pdf_path"] == asiento_rel
    assert data["outputs"][0]["extracto_pdf_path"] == extract


def test_merge_two_asientos_same_credit_consolidates_both_and_manifest_paths(monkeypatch):
    import json

    monkeypatch.setenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "CTL/control.xlsx")
    monkeypatch.setenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "OUT/PDFS")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "LOGS")
    hist = "HIST/hist.xlsx"
    email = "EMAIL/mail.pdf"
    extract = "clientes/ACME/CREDITO# 264/Extracto.pdf"
    asiento_dir = "clientes/ACME/CREDITO# 264/ASIENTOS CONTABLES CRED 264"
    asiento_abono = f"{asiento_dir}/Asiento abono capital credito 264.pdf"
    asiento_cuota = f"{asiento_dir}/Asiento cuota credito 264.pdf"

    g = _MergeGraph()
    g.initial["CTL/control.xlsx"] = _control_row_bytes(
        estado="PENDIENTE_ASIENTOS",
        is_active=True,
        hist=hist,
        email=email,
    )
    g.initial["bank/report.xlsx"] = _bank_bytes()
    g.initial[hist] = _hist_bytes()
    g.initial[email] = _tiny_pdf()
    g.initial[extract] = _tiny_pdf()
    g.initial[asiento_abono] = _tiny_pdf()
    g.initial[asiento_cuota] = _tiny_pdf()
    g.children[asiento_dir] = [
        {"name": "Asiento abono capital credito 264.pdf", "file": {}},
        {"name": "Asiento cuota credito 264.pdf", "file": {}},
        {"folder": {"name": "PROCESADOS"}},
    ]

    ctx = {
        "site_id": "s1",
        "drive_id": "d1",
        "path_encoded": encode_graph_drive_path("bank/report.xlsx"),
        "file_path": "bank/report.xlsx",
    }

    async def fake_collect(_g, _s, _d, _cell):
        return [extract]

    async def run():
        with (
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs.resolve_sharepoint_from_env",
                new_callable=AsyncMock,
                return_value=ctx,
            ),
            patch(
                "app.application.use_cases.merge_composite_validado_pdfs._collect_pdf_paths_from_ruta_cell",
                new_callable=AsyncMock,
                side_effect=fake_collect,
            ),
        ):
            return await merge_composite_validado_pdfs(g)

    r = asyncio.run(run())
    assert r.outputs_count == 1
    assert r.merge_control_status == "CONSOLIDADO"
    assert g.deleted == []
    out = r.outputs[0]
    assert out.asiento_pdf_paths == (asiento_abono, asiento_cuota)
    assert out.asiento_pdf_path == asiento_abono
    assert "asiento:" in out.sources_summary
    manifest_key = next(k for k in g.uploaded if k.endswith(".json"))
    data = json.loads(g.uploaded[manifest_key].decode("utf-8"))
    assert data["outputs"][0]["asiento_pdf_paths"] == [asiento_abono, asiento_cuota]
    assert data["outputs"][0]["asiento_pdf_path"] == asiento_abono


def test_merge_completed_example_warning_payload():
    body = {
        "status": "completed",
        "result": {
            "status": "ok",
            "merge_control_updated": True,
            "merge_control_status": "MERGE_PARCIAL",
            "outputs_count": 1,
            "skipped_count": 1,
            "outputs": [{"id_pago": "1"}],
            "skipped": ["1: x"],
        },
    }
    out = enrich_job_for_http_response({**body, "job_id": "x", "type": "merge_composite_validado_pdfs"})
    assert out["severity"] == "warning"
