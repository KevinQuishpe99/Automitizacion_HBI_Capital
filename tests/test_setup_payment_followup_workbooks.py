"""Tests setup bandejas operativas pagos_adelantados / pagos_incompletos."""

import asyncio
import io

import httpx
import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from app.adapters.primary.http.deps import get_graph_client
from app.adapters.primary.http.routers.payment_validation import router
from app.application.use_cases import setup_payment_followup_workbooks as spfw
from app.application.use_cases.setup_payment_followup_workbooks import (
    ADELANTADOS_COLUMNS,
    FILENAME_ADELANTADOS,
    FILENAME_INCOMPLETOS,
    SHEET_HISTORICO,
    SHEET_PENDIENTES,
    _build_followup_workbook_bytes,
    setup_payment_followup_workbooks,
)


def _http_error(status: int) -> httpx.HTTPStatusError:
    req = httpx.Request("GET", "https://graph.microsoft.com/test")
    resp = httpx.Response(status, request=req, text=f"HTTP {status} body")
    return httpx.HTTPStatusError("err", request=req, response=resp)


CONTROL_FOLDER = "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL"
PATH_AD = f"{CONTROL_FOLDER}/{FILENAME_ADELANTADOS}"
PATH_IN = f"{CONTROL_FOLDER}/{FILENAME_INCOMPLETOS}"


class MockGraphFollowup:
    def __init__(self) -> None:
        self.files: dict[str, bytes] = {}
        self.put_count = 0

    def _path_key(self, endpoint: str) -> str | None:
        low = endpoint.lower()
        if ":/content" not in low:
            return None
        if FILENAME_ADELANTADOS.lower() in low:
            return PATH_AD
        if FILENAME_INCOMPLETOS.lower() in low:
            return PATH_IN
        return None

    async def get(self, endpoint: str, params=None):
        return {"value": []}

    async def get_bytes(self, endpoint: str, params=None):
        key = self._path_key(endpoint)
        if key and key in self.files:
            return self.files[key]
        raise _http_error(404)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        self.put_count += 1
        key = self._path_key(endpoint)
        if key:
            self.files[key] = content
        return {"webUrl": f"https://example.sharepoint.com/{key or 'file'}"}

    async def post_json(self, endpoint: str, body: dict):
        return {}, 201


@pytest.fixture
def mock_resolve(monkeypatch):
    async def _fake_resolve(client, site_search, drive_name, path):
        return {"site_id": "site-1", "drive_id": "drive-1", "path_encoded": "enc", "file_path": path}

    monkeypatch.setattr(spfw, "resolve_sharepoint_path", _fake_resolve)


@pytest.fixture(autouse=True)
def env_site(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST_SITE")
    monkeypatch.setenv("GRAPH_SHAREPOINT_DRIVE_NAME", "")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", CONTROL_FOLDER)


def _assert_sheet_fully_locked(ws) -> None:
    assert ws.protection.sheet is True
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 1), max_col=len(ADELANTADOS_COLUMNS)):
        for cell in row:
            assert cell.protection.locked is True


def test_build_workbook_has_pendientes_historico_no_registros_no_tables():
    raw = _build_followup_workbook_bytes(columns=ADELANTADOS_COLUMNS)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert SHEET_PENDIENTES in wb.sheetnames
    assert SHEET_HISTORICO in wb.sheetnames
    assert "Registros" not in wb.sheetnames
    ws_p = wb[SHEET_PENDIENTES]
    ws_h = wb[SHEET_HISTORICO]
    assert not ws_p.tables
    assert not ws_h.tables
    assert ws_p.cell(1, 1).value == ADELANTADOS_COLUMNS[0]
    _assert_sheet_fully_locked(ws_p)
    _assert_sheet_fully_locked(ws_h)


def test_setup_creates_both_when_missing(mock_resolve):
    g = MockGraphFollowup()
    out = asyncio.run(setup_payment_followup_workbooks(g))
    assert out["status"] == "success"
    assert out["pagos_adelantados"]["created"] is True
    assert out["pagos_incompletos"]["created"] is True
    assert g.put_count == 2


def test_setup_idempotent_no_overwrite(mock_resolve):
    g = MockGraphFollowup()
    asyncio.run(setup_payment_followup_workbooks(g))
    first_puts = g.put_count
    out2 = asyncio.run(setup_payment_followup_workbooks(g, force_recreate=False))
    assert g.put_count == first_puts
    assert out2["pagos_adelantados"]["created"] is False
    assert out2["pagos_incompletos"]["created"] is False


def test_force_recreate_replaces(mock_resolve):
    g = MockGraphFollowup()
    asyncio.run(setup_payment_followup_workbooks(g))
    puts_before = g.put_count
    out = asyncio.run(setup_payment_followup_workbooks(g, force_recreate=True))
    assert g.put_count == puts_before + 2
    assert out["pagos_adelantados"]["recreated"] is True
    assert out["pagos_incompletos"]["recreated"] is True


def test_incompletos_workbook_structure():
    from app.application.use_cases.setup_payment_followup_workbooks import INCOMPLETOS_COLUMNS

    raw = _build_followup_workbook_bytes(columns=INCOMPLETOS_COLUMNS)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert not wb[SHEET_PENDIENTES].tables
    assert wb[SHEET_PENDIENTES].cell(1, 1).value == "ID Pago"


def test_router_payment_followup_setup(mock_resolve):
    g = MockGraphFollowup()
    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[get_graph_client] = lambda: g
    client = TestClient(app)

    r = client.post("/graph/sharepoint/payment-validation/setup/payment-followup-workbooks")
    assert r.status_code == 200
    assert r.json()["pagos_adelantados"]["created"] is True

    r2 = client.post(
        "/graph/sharepoint/payment-validation/setup/payment-followup-workbooks",
        json={"force_recreate": True},
    )
    assert r2.status_code == 200
    assert r2.json()["force_recreate"] is True

    r_old = client.post("/graph/sharepoint/payment-validation/setup/audit-tracking-workbooks")
    assert r_old.status_code == 404
