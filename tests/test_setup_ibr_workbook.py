"""Tests setup idempotente IBR_DIARIO.xlsx."""

import asyncio
import io

import httpx
import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.adapters.primary.http.deps import get_graph_client
from app.adapters.primary.http.routers.payment_validation import router
from app.application.use_cases import setup_ibr_workbook as sib
from app.application.use_cases.setup_ibr_workbook import (
    FILENAME_IBR,
    IBR_COLUMNS,
    IBR_SHEET_NAME,
    _build_new_ibr_workbook_bytes,
    _repair_existing_workbook,
    setup_ibr_workbook,
)


def _http_error(status: int) -> httpx.HTTPStatusError:
    req = httpx.Request("GET", "https://graph.microsoft.com/test")
    resp = httpx.Response(status, request=req, text=f"HTTP {status} body")
    return httpx.HTTPStatusError("err", request=req, response=resp)


CONTROL_FOLDER = "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL"
PATH_IBR = f"{CONTROL_FOLDER}/{FILENAME_IBR}"


class MockGraphIbr:
    def __init__(self) -> None:
        self.files: dict[str, bytes] = {}
        self.put_count = 0

    def _path_key(self, endpoint: str) -> str | None:
        if ":/content" not in endpoint.lower():
            return None
        if FILENAME_IBR.lower() in endpoint.lower():
            return PATH_IBR
        return None

    async def get(self, endpoint: str, params=None):
        return {"webUrl": "https://example/ibr.xlsx"}

    async def get_bytes(self, endpoint: str, params=None):
        key = self._path_key(endpoint)
        if key and key in self.files:
            return self.files[key]
        raise _http_error(404)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        self.put_count += 1
        key = self._path_key(endpoint)
        if key:
            self.files[key] = content
        return {"webUrl": "https://example/ibr.xlsx"}

    async def post_json(self, endpoint: str, body: dict):
        return {}, 201


@pytest.fixture
def mock_resolve(monkeypatch):
    async def _fake_resolve(client, site_search, drive_name, path):
        return {"site_id": "site-1", "drive_id": "drive-1", "path_encoded": "enc", "file_path": path}

    monkeypatch.setattr(sib, "resolve_sharepoint_path", _fake_resolve)


@pytest.fixture(autouse=True)
def env_site(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST_SITE")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", CONTROL_FOLDER)
    monkeypatch.setenv("GRAPH_IBR_DIARIO_PATH", PATH_IBR)


def _assert_ibr_protection(ws) -> None:
    assert ws.protection.sheet is True
    for c in range(1, 4):
        assert ws.cell(1, c).protection.locked is True
        assert ws.cell(2, c).protection.locked is False


def test_build_ibr_has_sheet_and_columns_no_tables():
    raw = _build_new_ibr_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[IBR_SHEET_NAME]
    assert [ws.cell(1, i).value for i in range(1, 4)] == list(IBR_COLUMNS)
    assert not ws.tables
    _assert_ibr_protection(ws)


def test_setup_creates_when_missing(mock_resolve):
    g = MockGraphIbr()
    out = asyncio.run(setup_ibr_workbook(g))
    assert out["created"] is True
    assert PATH_IBR in g.files


def test_setup_idempotent_second_call(mock_resolve):
    g = MockGraphIbr()
    asyncio.run(setup_ibr_workbook(g))
    puts = g.put_count
    out2 = asyncio.run(setup_ibr_workbook(g))
    assert out2["created"] is False
    assert g.put_count == puts


def test_force_recreate_replaces(mock_resolve):
    g = MockGraphIbr()
    asyncio.run(setup_ibr_workbook(g))
    puts = g.put_count
    out = asyncio.run(setup_ibr_workbook(g, force_recreate=True))
    assert g.put_count == puts + 1
    assert out["recreated"] is True


def test_repair_preserves_data_rows(mock_resolve):
    g = MockGraphIbr()
    raw = _build_new_ibr_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[IBR_SHEET_NAME]
    ws.cell(3, 1, value="2026-01-01")
    ws.cell(3, 2, value="2026-01-31")
    ws.cell(3, 3, value=0.12)
    buf = io.BytesIO()
    wb.save(buf)
    g.files[PATH_IBR] = buf.getvalue()

    payload, warnings, modified = _repair_existing_workbook(g.files[PATH_IBR])
    wb2 = openpyxl.load_workbook(io.BytesIO(payload))
    assert wb2[IBR_SHEET_NAME].cell(3, 1).value == "2026-01-01"
    assert wb2[IBR_SHEET_NAME].cell(3, 3).value == 0.12
    assert modified is False
    assert not warnings


def test_repair_adds_missing_ibr_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Otro"
    ws.cell(1, 1, value="x")
    buf = io.BytesIO()
    wb.save(buf)
    payload, warnings, modified = _repair_existing_workbook(buf.getvalue())
    wb2 = openpyxl.load_workbook(io.BytesIO(payload))
    assert IBR_SHEET_NAME in wb2.sheetnames
    assert modified is True
    assert any("sheet_added" in w for w in warnings)


def test_existing_with_data_force_recreate_false_not_destroyed(mock_resolve):
    g = MockGraphIbr()
    raw = _build_new_ibr_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[IBR_SHEET_NAME]
    ws.cell(3, 1, value="2026-02-01")
    ws.cell(3, 2, value="2026-02-28")
    ws.cell(3, 3, value=0.15)
    buf = io.BytesIO()
    wb.save(buf)
    g.files[PATH_IBR] = buf.getvalue()

    out = asyncio.run(setup_ibr_workbook(g, force_recreate=False))
    assert out["created"] is False
    wb2 = openpyxl.load_workbook(io.BytesIO(g.files[PATH_IBR]))
    assert wb2[IBR_SHEET_NAME].cell(3, 1).value == "2026-02-01"


def test_router_setup_ibr_empty_body_and_force_recreate(mock_resolve):
    g = MockGraphIbr()
    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[get_graph_client] = lambda: g
    client = TestClient(app)
    r = client.post("/graph/sharepoint/payment-validation/setup/ibr-workbook")
    assert r.status_code == 200
    assert r.json()["created"] is True

    r2 = client.post(
        "/graph/sharepoint/payment-validation/setup/ibr-workbook",
        json={"force_recreate": True},
    )
    assert r2.status_code == 200
    assert r2.json()["force_recreate"] is True
    assert r2.json()["recreated"] is True
