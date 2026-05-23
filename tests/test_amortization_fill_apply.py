"""Tests apply real de amortización (mock SharePoint con put_bytes)."""

import asyncio
import hashlib
import io
import json
from datetime import date
from urllib.parse import unquote

import httpx
import openpyxl
import pytest

from app.application.services.accounting_pdf_parser import (
    ACCOUNT_CAPITAL,
    ACCOUNT_SALDOS_MENORES,
)
from app.application.services.amortization_workbook import (
    AUTOMATION_LOG_SHEET,
    load_automation_log_idempotency_keys,
)
from app.application.use_cases.amortization_fill_apply import (
    AmortizationPreflightError,
    run_amortization_fill_apply,
    validate_amortization_preflight,
)
from app.application.use_cases.amortization_fill_dry_run import run_amortization_fill_dry_run
from tests.test_amortization_fill_dry_run import (
    MockGraphDryRun,
    _accounting_text,
    _amort_table_date_at_row,
    _amort_table_displaced_application,
    _amort_table_two_dates_at_rows,
    _asiento_pdf_placeholder,
    _base_files,
    _hist_bytes,
    _hist_bytes_multi_credit,
    _ibr_bytes,
)


class MockGraphApply(MockGraphDryRun):
    """Mock Graph con almacén mutable para simular round-trip de tablas."""

    def __init__(self, files: dict[str, bytes], *, fail_upload_423: bool = False) -> None:
        super().__init__(files)
        self.uploaded: dict[str, bytes] = {}
        self.fail_upload_423 = fail_upload_423

    async def get(self, endpoint: str, params=None):
        if "/root:/" in endpoint and ":/content" not in endpoint and "children" not in endpoint:
            path = unquote(endpoint.split("/root:/", 1)[1].rstrip(":"))
            blob = self.files.get(path, b"")
            digest = hashlib.sha256(blob).hexdigest()[:16]
            return {
                "eTag": f'"{digest}"',
                "size": len(blob),
                "lastModifiedDateTime": "2026-04-23T12:00:00Z",
            }
        return await super().get(endpoint, params)

    async def get_bytes(self, endpoint: str, params=None):
        key = self._key(endpoint)
        if key and key in self.uploaded:
            return self.uploaded[key]
        return await super().get_bytes(endpoint, params)

    async def put_bytes(self, endpoint: str, content: bytes, content_type: str = ""):
        if self.fail_upload_423:
            request = httpx.Request("PUT", "https://graph.test/upload")
            response = httpx.Response(423, request=request, text="Locked")
            raise httpx.HTTPStatusError("423 Locked", request=request, response=response)
        key = self._key(endpoint)
        if key:
            self.uploaded[key] = content
            self.files[key] = content
        self.put_calls.append((endpoint, content))
        return {"id": "uploaded"}


def _saldos_menores_accounting_text() -> str:
    return f"""
    Comprobante 1 Fecha 22/04/2026
    {ACCOUNT_SALDOS_MENORES} 100.00
    {ACCOUNT_CAPITAL} 100.00
    """


@pytest.fixture(autouse=True)
def env_sharepoint(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST")
    monkeypatch.setenv("GRAPH_SHAREPOINT_DRIVE_NAME", "")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", "CTL")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "LOGS")
    monkeypatch.setenv("GRAPH_IBR_DIARIO_PATH", "CTL/IBR_DIARIO.xlsx")


def test_preflight_rejects_errors():
    dry = {"summary": {"errors": 1, "revision_manual": 0}}
    with pytest.raises(AmortizationPreflightError) as excinfo:
        validate_amortization_preflight(dry)
    assert excinfo.value.error_code == "preflight_errors"


def test_apply_single_table_writes_and_logs(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    assert out["summary"]["applied"] == 1
    assert len(g.put_calls) == 1
    item = out["items"][0]
    assert item["apply_status"] == "APPLIED"
    assert item["due_date_row"] == 8
    assert item["application_row"] == 8
    assert item.get("payment_date_iso") == fecha.isoformat()
    wb = openpyxl.load_workbook(io.BytesIO(g.uploaded["TABLAS/amort.xlsx"]), data_only=False)
    ws = wb["EQUINORTE"]
    from app.application.services.amortization_workbook import detect_headers, find_header_row, _is_formula_value

    headers = detect_headers(ws, header_row=find_header_row(ws))
    fecha_cell = ws.cell(8, headers["fecha_pago"]).value
    if hasattr(fecha_cell, "date"):
        fecha_cell = fecha_cell.date()
    assert fecha_cell == fecha
    assert _is_formula_value(ws.cell(8, headers["valor_pagado_cliente"]).value)
    assert item["write_plan"].get("valor_pagado_cliente") == "formula"
    log_ws = wb[AUTOMATION_LOG_SHEET]
    assert log_ws.max_row >= 2
    assert load_automation_log_idempotency_keys(wb)


def test_apply_multiple_asientos_different_application_rows(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha)
    asiento_a = "clientes/E/a1.pdf"
    asiento_b = "clientes/E/a2.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "historico_excel_path": "HIST/cartera.xlsx",
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 265",
                "asiento_pdf_paths": [asiento_a, asiento_b],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_265.xlsx": _amort_table_two_dates_at_rows(fecha, 8, 9),
        asiento_a: _asiento_pdf_placeholder(),
        asiento_b: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphApply(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    assert out["summary"]["applied"] == 2
    rows = sorted(it["application_row"] for it in out["items"])
    assert rows == [8, 9]
    assert all(it["due_date_row"] == 8 for it in out["items"])
    assert all(it["ibr_row"] == 8 for it in out["items"])


def test_apply_displaced_application_row(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 265", "TABLAS/amort_265.xlsx", fecha)
    amort = _amort_table_displaced_application(
        fecha, due_row=8, occupied_application_rows=[8, 9], free_application_row=10
    )
    asiento = "clientes/E/asiento.pdf"
    manifest = {
        "report_date_iso": fecha.isoformat(),
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "CREDITO # 265",
                "asiento_pdf_path": asiento,
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort_265.xlsx": amort,
        asiento: _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphApply(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g,
            report_date_iso=fecha.isoformat(),
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["items"][0]["due_date_row"] == 8
    assert out["items"][0]["application_row"] == 10


def test_apply_idempotent_on_second_run(monkeypatch):
    fecha = date(2026, 4, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out1 = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso=fecha.isoformat(), historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out1["summary"]["applied"] == 1
    out2 = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso=fecha.isoformat(), historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out2["summary"]["skipped_idempotent"] == 1
    assert out2["summary"]["applied"] == 0


def test_apply_allows_bank_inferred_saldos_menores(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _saldos_menores_accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g,
            report_date_iso="2026-04-23",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "ok"
    assert out["summary"]["applied"] == 1
    item = out["items"][0]
    assert "BANK_VALUE_INFERRED_OR_MISSING" in (item.get("warnings") or [])


def test_apply_blocks_disallowed_warning_preflight(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )

    def _bad_text(_b):
        return """
        Comprobante 1 Fecha 22/04/2026
        544113410519 100.00
        544113430501 50.00
        544141502030 10.00
        """

    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        _bad_text,
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g,
            report_date_iso="2026-04-23",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["status"] == "preflight_failed"
    assert out["preflight_error_code"] == "preflight_warnings_not_allowed"
    assert g.put_calls == []


def test_apply_pdf_changed_same_path_not_idempotent(monkeypatch):
    fecha = date(2026, 4, 22)
    asiento = "clientes/EQUINORTE/asiento.pdf"
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    files = _base_files(
        hist=hist,
        amort=_amort_table_date_at_row(fecha, 8),
        asiento_pdf=b"pdf-v1",
        ibr=_ibr_bytes(),
        fecha=fecha,
    )
    files[asiento] = b"pdf-v1"
    g = MockGraphApply(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out1 = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso=fecha.isoformat(), historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out1["summary"]["applied"] == 1
    g.files[asiento] = b"pdf-v2-changed"
    out2 = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso=fecha.isoformat(), historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out2["items"][0]["apply_status"] == "ERROR"
    assert out2["items"][0]["apply_error_code"] == "PDF_CHANGED_SAME_PATH"


def test_apply_post_upload_verification_ok_includes_tables_summary(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso="2026-04-23", historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out["tables_summary"]
    ts = out["tables_summary"][0]
    assert ts["verification_status"] == "ok"
    assert ts["upload_status"] == "uploaded"
    assert ts["eventos_aplicados"] == 1


def test_apply_post_upload_verification_formula_mismatch(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )

    def _fake_write(ws, row, headers, event, **kwargs):
        from app.application.services.amortization_workbook import write_payment_application as real_write

        plan = real_write(ws, row, headers, event, **kwargs)
        vp_col = headers.get("valor_pagado_cliente")
        if vp_col is not None and plan.get("valor_pagado_cliente") == "formula":
            ws.cell(row, vp_col, value=50_000_000.0)
        return plan

    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_apply.write_payment_application",
        _fake_write,
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso="2026-04-23", historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out["items"][0]["apply_error_code"] == "POST_UPLOAD_VERIFICATION_FAILED_FORMULA_MISMATCH"


def test_apply_post_upload_verification_failed(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_apply.verify_uploaded_table",
        lambda *_a, **_k: ["IBR no coincide"],
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso="2026-04-23", historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out["items"][0]["apply_error_code"] == "POST_UPLOAD_VERIFICATION_FAILED"
    assert out["tables_summary"][0]["verification_status"] == "POST_UPLOAD_VERIFICATION_FAILED"


def test_apply_excel_locked_does_not_count_as_applied(monkeypatch):
    fecha = date(2026, 4, 23)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphApply(
        _base_files(
            hist=hist,
            amort=_amort_table_date_at_row(fecha, 8),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        ),
        fail_upload_423=True,
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_apply(
            g, report_date_iso="2026-04-23", historical_file_path="HIST/cartera.xlsx"
        )
    )
    assert out["summary"]["applied"] == 0
    assert out["summary"]["errors"] == 1
    assert out["items"][0]["apply_error_code"] == "EXCEL_LOCKED"
    assert out["tables_uploaded"] == []
