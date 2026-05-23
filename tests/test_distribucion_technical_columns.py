"""Columnas técnicas RutaTablaAmortizacion / CreditoNormalizado en Generate, Finalize y dry-run."""

import asyncio
import io
import json
import os
from datetime import date, datetime

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from app.application.services.review_schema import (
    DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS,
    DistribucionCols,
    EstadoPago,
    ReviewSheets,
    ValidarPago,
    normalize_credito_digits,
)
from app.application.use_cases.amortization_fill_dry_run import (
    _load_historical_index,
    _resolve_tabla_path_for_row,
    run_amortization_fill_dry_run,
)
from app.application.use_cases.payment_validation_generate import _configure_distrib_technical_path_columns
from app.application.use_cases.payment_validation_finalize import _configure_hist_distrib_technical_path_columns
from tests.test_amortization_fill_dry_run import (
    MockGraphDryRun,
    _accounting_text,
    _amort_table_bytes,
    _asiento_pdf_placeholder,
    _base_files,
    _hist_bytes,
    _ibr_bytes,
)
from tests.test_finalize_validation import (
    MockGraphClient,
    create_review_workbook,
    make_distrib_row,
    set_env_vars as finalize_set_env_vars,
)
from tests.test_generate_validation import run_generate, sheet_to_dicts

# Misma configuración Graph que tests/test_amortization_fill_dry_run.py
@pytest.fixture(autouse=True)
def env_sharepoint(monkeypatch):
    monkeypatch.setenv("GRAPH_SHAREPOINT_SITE_SEARCH", "TEST")
    monkeypatch.setenv("GRAPH_SHAREPOINT_DRIVE_NAME", "")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", "CTL")
    monkeypatch.setenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "LOGS")
    monkeypatch.setenv("GRAPH_IBR_DIARIO_PATH", "CTL/IBR_DIARIO.xlsx")


def test_schema_has_technical_columns_in_headers():
    assert DistribucionCols.RUTA_TABLA_AMORTIZACION in DistribucionCols.HEADERS
    assert DistribucionCols.CREDITO_NORMALIZADO in DistribucionCols.HEADERS
    assert DistribucionCols.RUTA_TABLA_AMORTIZACION in DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS


def test_normalize_credito_digits():
    assert normalize_credito_digits("CREDITO # 258") == "258"
    assert normalize_credito_digits("258") == "258"


def test_generate_writes_ruta_tabla_and_credito_normalizado():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    row = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])[0]
    assert (
        row[DistribucionCols.RUTA_TABLA_AMORTIZACION]
        == "clientes/GEOEXCON/254/Tabla amortizacion GEOEXCON 254.xlsx"
    )
    assert row[DistribucionCols.CREDITO_NORMALIZADO] == "254"
    ws = wb[ReviewSheets.DISTRIBUCION]
    for col_name in DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS:
        cidx = DistribucionCols.HEADERS.index(col_name) + 1
        assert ws.column_dimensions[get_column_letter(cidx)].hidden is True


def test_generate_hides_technical_columns_only():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ReviewSheets.DISTRIBUCION
    for col, name in enumerate(DistribucionCols.HEADERS, start=1):
        ws.cell(3, col, name)
    _configure_distrib_technical_path_columns(ws)
    for col_name in DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS:
        cidx = DistribucionCols.HEADERS.index(col_name) + 1
        assert ws.column_dimensions[get_column_letter(cidx)].hidden is True
    ruta_c = DistribucionCols.HEADERS.index(DistribucionCols.RUTA) + 1
    assert ws.column_dimensions[get_column_letter(ruta_c)].hidden is not True


def test_finalize_hides_technical_columns_on_hist():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ReviewSheets.DISTRIBUCION
    headers = list(DistribucionCols.HEADERS) + [DistribucionCols.RUTA_ASIENTOS_CONTABLES]
    for col, name in enumerate(headers, start=1):
        ws.cell(1, col, name)
    _configure_hist_distrib_technical_path_columns(ws, 1)
    rt = headers.index(DistribucionCols.RUTA_TABLA_AMORTIZACION) + 1
    cn = headers.index(DistribucionCols.CREDITO_NORMALIZADO) + 1
    assert ws.column_dimensions[get_column_letter(rt)].hidden is True
    assert ws.column_dimensions[get_column_letter(cn)].hidden is True


def test_resolve_tabla_path_prefers_ruta_tabla_amortizacion():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "ID Pago",
            "Crédito",
            "Link tabla amortización",
            "RutaTablaAmortizacion",
            "RutaUnidadCredito",
        ]
    )
    ws.append(
        [
            "P1",
            "CREDITO # 258",
            "Ver tabla",
            "TABLAS/amort_258.xlsx",
            "clientes/E/CREDITO # 258",
        ]
    )
    path = _resolve_tabla_path_for_row(
        ws,
        2,
        col_ruta_tabla=4,
        col_link_tabla=3,
        col_ruta_unidad=5,
    )
    assert path == "TABLAS/amort_258.xlsx"


def test_load_historical_index_uses_credito_normalizado_key():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(
        [
            "ID Pago",
            "Cliente",
            "Crédito",
            "Fecha límite",
            "Estado Pago",
            "Ruta",
            "Link tabla amortización",
            "RutaTablaAmortizacion",
            "CreditoNormalizado",
        ]
    )
    ws.append(
        [
            "8326b91b",
            "EQUINORTE",
            "CREDITO # 258",
            date(2026, 5, 22),
            "ADELANTADO",
            "x.pdf",
            "Ver tabla",
            "TABLAS/amort_258.xlsx",
            "258",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    index = _load_historical_index(buf.getvalue())
    assert ("8326b91b", "258") in index
    assert index[("8326b91b", "258")]["tabla_amortizacion_path"] == "TABLAS/amort_258.xlsx"


def test_dry_run_uses_ruta_tabla_before_link_fallback(monkeypatch):
    fecha = date(2026, 5, 22)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distribución"
    ws.append(
        [
            "ID Pago",
            "Cliente",
            "Crédito",
            "Fecha límite",
            "Estado Pago",
            "Ruta",
            "Link tabla amortización",
            "RutaTablaAmortizacion",
            "CreditoNormalizado",
        ]
    )
    ws.append(
        [
            "7785e37e",
            "EQUINORTE",
            "CREDITO # 258",
            fecha,
            "ADELANTADO",
            "x.pdf",
            "Ver tabla sin ruta util",
            "TABLAS/amort.xlsx",
            "258",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    hist = buf.getvalue()

    manifest = {
        "report_date_iso": fecha.isoformat(),
        "outputs": [
            {
                "id_pago": "7785e37e",
                "cliente": "EQUINORTE",
                "credito": "258, 265",
                "credit_items": [
                    {
                        "credito": "258",
                        "asiento_pdf_paths": ["clientes/E/asiento.pdf"],
                        "extracto_pdf_path": "clientes/E/extracto.pdf",
                    }
                ],
            }
        ],
    }
    files = {
        "CTL/dummy.xlsx": b"x",
        f"LOGS/merge_manifest_{fecha.isoformat()}.json": json.dumps(manifest).encode("utf-8"),
        "HIST/cartera.xlsx": hist,
        "TABLAS/amort.xlsx": _amort_table_bytes(fecha),
        "clientes/E/asiento.pdf": _asiento_pdf_placeholder(),
        "CTL/IBR_DIARIO.xlsx": _ibr_bytes(),
    }
    g = MockGraphDryRun(files)
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    item = out["items"][0]
    assert item["credito"] == "258"
    assert item["tabla_amortizacion_path"] == "TABLAS/amort.xlsx"
    assert item.get("error_code") != "TABLE_PATH_NOT_FOUND"


def test_finalize_preserves_ruta_tabla_amortizacion_in_historical():
    async def run_test():
        from app.application.use_cases.payment_validation_finalize import finalize_payment_validation

        finalize_set_env_vars()
        os.environ["GRAPH_PAYMENT_VALIDATION_HISTORY_PATH"] = "history"
        client = MockGraphClient()
        ruta_tabla = "clientes/CLI/CRED/Tabla amortizacion CLI CRED.xlsx"
        row, _ = make_distrib_row(
            credito="CREDITO # 264",
            estado=EstadoPago.ADELANTADO,
            validar_pago=ValidarPago.SI,
            extract_route="clientes/CLI/CRED/extractos/e1.pdf",
        )
        vals = dict(zip(DistribucionCols.HEADERS, row))
        vals[DistribucionCols.RUTA_TABLA_AMORTIZACION] = ruta_tabla
        vals[DistribucionCols.CREDITO_NORMALIZADO] = "264"
        row = [vals[c] for c in DistribucionCols.HEADERS]
        client.downloaded_files["revision/val_latest.xlsx"] = create_review_workbook(
            distrib_specs=[(row, None)]
        )
        await finalize_payment_validation(client, "val_latest.xlsx", process_date=date(2026, 5, 10))
        hist_key = next(k for k in client.uploaded_files if "cartera_validada_" in k)
        wb = openpyxl.load_workbook(io.BytesIO(client.uploaded_files[hist_key]))
        ws = wb[ReviewSheets.DISTRIBUCION]
        cmap = {
            str(ws.cell(1, c).value or "").strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value
        }
        assert ws.cell(2, cmap[DistribucionCols.RUTA_TABLA_AMORTIZACION]).value == ruta_tabla
        assert ws.cell(2, cmap[DistribucionCols.CREDITO_NORMALIZADO]).value == "264"

    asyncio.run(run_test())


def test_dry_run_legacy_hist_without_technical_columns(monkeypatch):
    fecha = date(2026, 5, 22)
    hist = _hist_bytes("7785e37e", "CREDITO # 258", "TABLAS/amort.xlsx", fecha)
    g = MockGraphDryRun(
        _base_files(
            hist=hist,
            amort=_amort_table_bytes(fecha),
            asiento_pdf=_asiento_pdf_placeholder(),
            ibr=_ibr_bytes(),
            fecha=fecha,
        )
    )
    monkeypatch.setattr(
        "app.application.use_cases.amortization_fill_dry_run.extract_text_from_pdf",
        lambda _b: _accounting_text(),
    )
    out = asyncio.run(
        run_amortization_fill_dry_run(
            g,
            report_date_iso="2026-05-22",
            historical_file_path="HIST/cartera.xlsx",
        )
    )
    assert out["items"][0]["application_status"] in (
        "WOULD_APPLY",
        "WOULD_ADOPT_EXISTING",
        "REVISION_MANUAL",
    )
    assert out["items"][0].get("error_code") != "TABLE_PATH_NOT_FOUND"
