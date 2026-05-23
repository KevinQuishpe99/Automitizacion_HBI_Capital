import asyncio
import contextlib
import io
import os
import re
from datetime import date, datetime
from unittest import mock
from urllib.parse import unquote

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from app.application.services.review_schema import (
    CasosPagoCols,
    ControlCols,
    DistribucionCols,
    ErroresCols,
    EstadoPago,
    ReviewSheets,
    ValidarPago,
)

try:
    from app.application.use_cases.payment_validation_generate import (
        CASOS_OBS_COL_WIDTH,
        CASOS_SUBTITLE,
        CONTROL_HELP,
        CONTROL_TITLE,
        DISTRIB_HELP,
        DIST_MONEY_COL_WIDTH,
        ERRORES_HELP,
        _ESTADO_PAGO_FILLS,
        _FILL_EDITABLE_COL,
        _FILL_PROCESAR_SI,
        _FILL_SALDO_COL,
        _distrib_freeze_panes_cell,
        generate_payment_validation,
        _extract_pending_installment,
    )
except ImportError:
    generate_payment_validation = None
    _extract_pending_installment = None  # type: ignore
    _distrib_freeze_panes_cell = None  # type: ignore
    CONTROL_TITLE = ""
    CONTROL_HELP = ""
    _ESTADO_PAGO_FILLS = {}
    _FILL_EDITABLE_COL = None
    _FILL_PROCESAR_SI = None
    _FILL_SALDO_COL = None
    CASOS_SUBTITLE = ""
    CASOS_OBS_COL_WIDTH = 48.0
    DISTRIB_HELP = ""
    ERRORES_HELP = ""
    DIST_MONEY_COL_WIDTH = 16.0


class MockGraphClient:
    def __init__(self):
        self.children = []
        self.downloaded_files = {}
        self.folder_children = {}
        self.folder_web_urls: dict[str, str] = {}
        self.uploaded_files = {}
        self.requested_endpoints = []
        self.put_calls = []

    async def get(self, endpoint, params=None):
        if endpoint == "/sites":
            return {"value": [{"id": "dummy_site"}]}
        if endpoint == "/sites/dummy_site/drives":
            return {"value": [{"id": "dummy_drive", "name": "DRIVE"}]}
        if endpoint.endswith(":/children"):
            path = endpoint.split("/root:/", 1)[1].rsplit(":/children", 1)[0]
            path = unquote(path)
            if path == "revision":
                return {"value": self.children}
            return {"value": self.folder_children.get(path, [])}
        if "/root:/" in endpoint and ":/children" not in endpoint and ":/content" not in endpoint:
            path = unquote(endpoint.split("/root:/", 1)[1].split(":", 1)[0])
            web = self.folder_web_urls.get(path)
            if web:
                return {"webUrl": web}
            return {}
        return {}

    async def get_bytes(self, endpoint, params=None):
        self.requested_endpoints.append(endpoint)
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        return self.downloaded_files.get(file_path, b"")

    async def put_bytes(self, endpoint, content, content_type):
        self.put_calls.append((endpoint, content_type))
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        self.uploaded_files[file_path] = content
        return {"id": "new_file_id"}

    async def delete(self, endpoint: str) -> None:
        return None


class MockGraphClientPutReturnsWebUrl(MockGraphClient):
    """Simula Graph: put_bytes devuelve driveItem con webUrl."""

    def __init__(self, web_url: str = "https://comwareec.sharepoint.com/sites/x/file.xlsx"):
        super().__init__()
        self._web_url = web_url

    async def put_bytes(self, endpoint, content, content_type):
        self.put_calls.append((endpoint, content_type))
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        self.uploaded_files[file_path] = content
        return {"id": "new_file_id", "webUrl": self._web_url}


class MockGraphClientPutCustomReturn(MockGraphClient):
    """put_bytes devuelve un valor configurable (dict sin webUrl, None, etc.)."""

    def __init__(self, put_return):
        super().__init__()
        self._put_return = put_return

    async def put_bytes(self, endpoint, content, content_type):
        self.put_calls.append((endpoint, content_type))
        file_path = endpoint.split("/root:/", 1)[1].rsplit(":/content", 1)[0]
        file_path = unquote(file_path)
        self.uploaded_files[file_path] = content
        return self._put_return


def make_item(name, is_folder=False, web_url=None):
    item = {"name": name}
    if is_folder:
        item["folder"] = {}
    if web_url:
        item["webUrl"] = web_url
    return item


def create_excel(headers, data, sheet_name="Sheet"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in data:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def create_amortization_excel(rows):
    return create_excel(
        ["Fecha límite", "Fecha pago", "Total pagado", "Cuota", "Intereses de mora"],
        rows,
    )


# Mapa global: path_del_pdf -> valor TOTAL A PAGAR (usado por el mock de extract_total_a_pagar_from_pdf)
_PDF_AMOUNT_REGISTRY: dict[str, float] = {}


def create_pdf_bytes(
    total_a_pagar: str,
    fecha_limite: date | None = None,
    *,
    skip_fecha_fake: bool = False,
) -> bytes:
    """
    Genera bytes marcadores para un PDF de test. Estos bytes NO son un PDF real.
    El mock de extract_total_a_pagar_from_pdf lee TOTAL A PAGAR;
    el mock de extract_fecha_limite_pago_from_pdf lee FECHA_LIMITE=YYYY-MM-DD si está presente.
    Si no hay marcador de fecha, el fake de tests usa una fecha fija (un solo extracto).
    skip_fecha_fake: el fake de fecha en tests devuelve None (simula PDF sin fecha legible).
    """
    sk = "SKIP_FECHA;" if skip_fecha_fake else ""
    fl = ""
    if fecha_limite is not None:
        fl = f"FECHA_LIMITE={fecha_limite.isoformat()};"
    if not total_a_pagar:
        return f"PDF_MOCK:{sk}{fl}sin_total".encode()
    return f"PDF_MOCK:{sk}{fl}TOTAL A PAGAR {total_a_pagar}".encode()


def create_pdf_no_total() -> bytes:
    """PDF sin TOTAL A PAGAR (marcador)."""
    return b"PDF_MOCK:sin_total"


def make_pdf_extractor_mock(client: "MockGraphClient"):
    """
    Devuelve una función que reemplaza extract_total_a_pagar_from_pdf.
    Lee los bytes que el mock guardó y parsea el marcador PDF_MOCK.
    """
    def _fake_extract(pdf_bytes: bytes) -> float:
        text = pdf_bytes.decode(errors="ignore")
        if not text.startswith("PDF_MOCK:"):
            raise ValueError("extract_amount_not_found")
        # Formato: "PDF_MOCK:TOTAL A PAGAR 18.826.879" o "PDF_MOCK:sin_total"
        if "TOTAL A PAGAR" not in text:
            raise ValueError("extract_amount_not_found")
        raw = text.split("TOTAL A PAGAR", 1)[1].strip()
        if not raw:
            raise ValueError("extract_amount_not_found")
        # Parsear formato latino
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(".", "")
        try:
            return float(raw)
        except ValueError:
            raise ValueError("extract_amount_not_found")
    return _fake_extract


def set_env_vars():
    os.environ["GRAPH_BANK_PAYMENTS_FILE_PATH"] = "banco.xlsx"
    os.environ["GRAPH_CLIENTS_BASE_PATH"] = "clientes"
    os.environ["GRAPH_SHAREPOINT_SITE_SEARCH"] = "SITIO"
    os.environ["GRAPH_SHAREPOINT_DRIVE_NAME"] = "DRIVE"
    os.environ["GRAPH_PAYMENT_VALIDATION_REVIEW_PATH"] = "revision"
    os.environ["GRAPH_VALIDATION_FILE_PREFIX"] = "val"


def setup_client_structure(client, include_web_urls=False):
    client.folder_children["clientes"] = [
        make_item("GEOEXCON", is_folder=True),
        make_item("EQUINORTE", is_folder=True),
    ]
    folder_url_254 = "https://contoso/folder/254" if include_web_urls else None
    folder_url_231 = "https://contoso/folder/231" if include_web_urls else None
    folder_url_900 = "https://contoso/folder/900" if include_web_urls else None
    client.folder_children["clientes/GEOEXCON"] = [
        make_item("254", is_folder=True, web_url=folder_url_254),
        make_item("231", is_folder=True, web_url=folder_url_231),
    ]
    client.folder_children["clientes/EQUINORTE"] = [
        make_item("900", is_folder=True, web_url=folder_url_900),
    ]

    geo_254_pdf = "clientes/GEOEXCON/254/Extracto 2025-12-23 CREDITO # 254.pdf"
    geo_254_tabla = "clientes/GEOEXCON/254/Tabla amortizacion GEOEXCON 254.xlsx"
    geo_231_pdf = "clientes/GEOEXCON/231/Extracto 2025-12-23 CREDITO # 231.pdf"
    geo_231_tabla = "clientes/GEOEXCON/231/Tabla amortizacion GEOEXCON 231.xlsx"
    eq_900_pdf = "clientes/EQUINORTE/900/Extracto 2025-12-30 CREDITO # 900.pdf"
    eq_900_tabla = "clientes/EQUINORTE/900/Tabla amortizacion EQUINORTE 900.xlsx"

    client.folder_children["clientes/GEOEXCON/254"] = [
        make_item(
            "Extracto 2025-12-23 CREDITO # 254.pdf",
            web_url="https://contoso/254.pdf" if include_web_urls else None,
        ),
        make_item(
            "Tabla amortizacion GEOEXCON 254.xlsx",
            web_url="https://contoso/254.xlsx" if include_web_urls else None,
        ),
    ]
    client.folder_children["clientes/GEOEXCON/231"] = [
        make_item(
            "Extracto 2025-12-23 CREDITO # 231.pdf",
            web_url="https://contoso/231.pdf" if include_web_urls else None,
        ),
        make_item(
            "Tabla amortizacion GEOEXCON 231.xlsx",
            web_url="https://contoso/231.xlsx" if include_web_urls else None,
        ),
    ]
    client.folder_children["clientes/EQUINORTE/900"] = [
        make_item(
            "Extracto 2025-12-30 CREDITO # 900.pdf",
            web_url="https://contoso/900.pdf" if include_web_urls else None,
        ),
        make_item(
            "Tabla amortizacion EQUINORTE 900.xlsx",
            web_url="https://contoso/900.xlsx" if include_web_urls else None,
        ),
    ]

    client.downloaded_files[geo_254_tabla] = create_amortization_excel([
        [date(2025, 12, 23), None, None, 6734920.07, None],
    ])
    client.downloaded_files[geo_231_tabla] = create_amortization_excel([
        [date(2025, 12, 23), None, None, 19540684.91, None],
    ])
    client.downloaded_files[eq_900_tabla] = create_amortization_excel([
        [date(2025, 12, 30), None, None, 5000000, None],
    ])
    # PDFs con TOTAL A PAGAR real (valores correctos, distintos de Cuota en tabla)
    client.downloaded_files[geo_254_pdf] = create_pdf_bytes("6.500.739", date(2025, 12, 23))
    client.downloaded_files[geo_231_pdf] = create_pdf_bytes("18.826.879", date(2025, 12, 23))
    client.downloaded_files[eq_900_pdf] = create_pdf_bytes("5.000.000", date(2025, 12, 30))

    return {
        "geo_254_pdf": geo_254_pdf,
        "geo_254_tabla": geo_254_tabla,
        "geo_231_pdf": geo_231_pdf,
        "geo_231_tabla": geo_231_tabla,
        "eq_900_pdf": eq_900_pdf,
        "eq_900_tabla": eq_900_tabla,
    }


def load_generated_workbook(client):
    assert client.uploaded_files, "Generate no subió ningún workbook"
    uploaded_key = list(client.uploaded_files.keys())[0]
    return openpyxl.load_workbook(io.BytesIO(client.uploaded_files[uploaded_key]))


def _header_row_index(ws, marker: str = "ID Pago") -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and str(row[0]).strip() == marker:
            return i
    return 1


def _first_data_row(ws, marker: str = "ID Pago") -> int:
    return _header_row_index(ws, marker) + 1


def sheet_to_dicts(ws, marker: str | None = None):
    m = marker or DistribucionCols.ID_PAGO
    rows = list(ws.iter_rows(values_only=True))
    hidx = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == m), None)
    if hidx is None:
        raise AssertionError(f"No se encontró fila de encabezados con primera columna {m!r}")
    headers = list(rows[hidx])
    width = len(headers)
    out = []
    for row in rows[hidx + 1 :]:
        if not any(row):
            continue
        d = {}
        for idx in range(min(width, len(row) if row else 0)):
            h = headers[idx]
            if h is None or str(h).strip() == "":
                continue
            d[h] = row[idx] if idx < len(row) else None
        out.append(d)
    return out


def run_generate(bank_rows, process_date, include_web_urls=False):
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    async def _run():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        setup_client_structure(client, include_web_urls=include_web_urls)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            bank_rows,
        )
        extractor = make_pdf_extractor_mock(client)
        fake_fecha = _fake_fecha_limite_from_marker_bytes()
        with mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_total_a_pagar_from_pdf",
            side_effect=extractor,
        ), mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_fecha_limite_pago_from_pdf",
            side_effect=fake_fecha,
        ):
            result = await generate_payment_validation(client, process_date)
        workbook = load_generated_workbook(client)
        return client, result, workbook

    return asyncio.run(_run())


def test_generate_review_folder_not_empty():
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = [{"name": "archivo_viejo.xlsx"}]

        with pytest.raises(ValueError, match="review_folder_not_empty"):
            await generate_payment_validation(client, date(2026, 5, 10))

    asyncio.run(run_test())


def test_generate_mock_graph_contract_has_no_legacy_methods():
    client = MockGraphClient()
    assert not hasattr(client, "get_file_content")
    assert not hasattr(client, "upload_file")


def test_generate_ignores_temp_files():
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = [{"name": "~$archivo.xlsx"}]
        setup_client_structure(client)
        client.downloaded_files["banco.xlsx"] = create_excel(["Fecha", "Crédito", "Concepto", "Transacción"], [])

        result = await generate_payment_validation(client, date(2026, 5, 10))
        assert result["process_id"] is not None
        assert any(endpoint.endswith("/banco.xlsx:/content") for endpoint in client.requested_endpoints)
        assert any(
            endpoint.endswith("/revision/val_2026-05-10.xlsx:/content")
            for endpoint, _ in client.put_calls
        )

    asyncio.run(run_test())


def test_generate_multiple_credits_per_payment():
    client, result, workbook = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_dist = workbook[ReviewSheets.DISTRIBUCION]
    ws_cp = workbook[ReviewSheets.CASOS_PAGO]
    dist_rows = sheet_to_dicts(ws_dist)
    case_rows = sheet_to_dicts(ws_cp)

    ws_err = workbook[ReviewSheets.ERRORES]
    err_dr = _first_data_row(ws_err)
    err_rows = list(ws_err.iter_rows(min_row=err_dr, values_only=True))
    assert result["summary"]["errores"] == 0, f"Errores inesperados: {err_rows}"
    assert len(dist_rows) == 2
    assert len(case_rows) == 1
    assert {row[DistribucionCols.CREDITO] for row in dist_rows} == {"254", "231"}
    assert len({row[DistribucionCols.ID_PAGO] for row in dist_rows}) == 1
    assert case_rows[0][CasosPagoCols.ID_PAGO] == dist_rows[0][DistribucionCols.ID_PAGO]
    assert case_rows[0][CasosPagoCols.CLIENTE] == "GEOEXCON"
    assert any(endpoint.endswith("/banco.xlsx:/content") for endpoint in client.requested_endpoints)


def test_generate_mora_fields_are_empty():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 30), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 30),
    )
    row = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])[0]

    fl = row[DistribucionCols.FECHA_LIMITE]
    assert fl == "2025-12-23" or (hasattr(fl, "isoformat") and fl.isoformat()[:10] == "2025-12-23")
    assert row[DistribucionCols.DIAS_MORA] == 7
    assert row[DistribucionCols.ESTADO_LINEA] == EstadoPago.ATRASADO
    assert row[DistribucionCols.INTERESES_MORA] in (None, "")
    assert isinstance(row[DistribucionCols.TOTAL_APLICADO], str)
    assert row[DistribucionCols.TOTAL_APLICADO].startswith("=")


def test_generate_does_not_calculate_mora():
    client, _, workbook = run_generate(
        [[datetime(2025, 12, 30), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 30),
    )
    row = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])[0]

    assert row[DistribucionCols.VALOR_EXTRACTO] == 6500739.0  # viene del PDF, no de la tabla
    assert row[DistribucionCols.VALOR_INTERESES] in (None, "")
    assert row[DistribucionCols.ABONO_K] in (None, "")
    assert row[DistribucionCols.INTERESES_MORA] in (None, "")
    assert isinstance(row[DistribucionCols.TOTAL_APLICADO], str)
    assert row[DistribucionCols.TOTAL_APLICADO].startswith("=")


def test_generate_reprogramar_fields():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 20), 5000000, "EQUINORTE", ""]],
        date(2025, 12, 20),
    )
    row = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])[0]

    assert row[DistribucionCols.ESTADO_LINEA] == EstadoPago.ADELANTADO
    assert isinstance(row[DistribucionCols.TOTAL_APLICADO], str)
    assert row[DistribucionCols.TOTAL_APLICADO].startswith("=")
    assert row[DistribucionCols.INTERESES_MORA] in (None, "")


def test_generate_schema_alignment():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", "trx-geo"]],
        date(2025, 12, 23),
    )
    ws_cp = workbook[ReviewSheets.CASOS_PAGO]
    ws_dist = workbook[ReviewSheets.DISTRIBUCION]
    cp_hr = _header_row_index(ws_cp, CasosPagoCols.ID_PAGO)
    dist_hr = _header_row_index(ws_dist, DistribucionCols.ID_PAGO)
    cp_headers = [cell.value for cell in ws_cp[cp_hr]]
    dist_headers = [cell.value for cell in ws_dist[dist_hr]]
    case_row = sheet_to_dicts(ws_cp)[0]
    dist_row = sheet_to_dicts(ws_dist)[0]

    assert cp_headers == CasosPagoCols.HEADERS
    assert dist_headers == DistribucionCols.HEADERS
    assert case_row[CasosPagoCols.CLIENTE] == "GEOEXCON"
    assert case_row[CasosPagoCols.MONTO_BANCO] == 25443565
    assert isinstance(case_row.get(CasosPagoCols.OBSERVACION), str)
    assert dist_row[DistribucionCols.ESTADO_LINEA] == EstadoPago.NORMAL


def test_distribucion_estado_linea_column():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 30), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 30),
    )
    row = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])[0]

    assert row[DistribucionCols.ESTADO_LINEA] == EstadoPago.ATRASADO
    assert row[DistribucionCols.TOTAL_APLICADO] not in {
        EstadoPago.NORMAL,
        EstadoPago.ATRASADO,
    }


def test_generate_uses_weburl():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=True,
    )
    ws = workbook[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    c_cred = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    credito = ws.cell(row=dr, column=c_cred).value
    c_ext = ws.cell(row=dr, column=DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1)
    c_tab = ws.cell(row=dr, column=DistribucionCols.HEADERS.index(DistribucionCols.LINK_TABLA) + 1)
    c_fold = ws.cell(row=dr, column=DistribucionCols.HEADERS.index(DistribucionCols.LINK_CARPETA_CREDITO) + 1)
    assert c_ext.value == f"📄 Ver extracto crédito {credito}"
    assert c_tab.value == f"📄 Ver tabla crédito {credito}"
    assert c_fold.value == f"Ver carpeta crédito {credito}"
    assert c_ext.hyperlink and str(getattr(c_ext.hyperlink, "target", c_ext.hyperlink)).startswith("https://contoso/")
    assert c_tab.hyperlink and str(getattr(c_tab.hyperlink, "target", c_tab.hyperlink)).startswith("https://contoso/")
    assert c_fold.hyperlink and str(getattr(c_fold.hyperlink, "target", c_fold.hyperlink)).startswith("https://contoso/folder/")


def test_generate_output_is_secretary_usable_minimum():
    _, result, workbook = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 20), 5000000, "", "EQUINORTE pago"],
        ],
        date(2025, 12, 23),
    )
    case_rows = sheet_to_dicts(workbook[ReviewSheets.CASOS_PAGO])
    dist_rows = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])

    assert result["summary"]["errores"] == 0
    assert len(case_rows) == 2
    assert len(dist_rows) == 3
    assert case_rows[0][CasosPagoCols.CLIENTE] == "GEOEXCON"
    assert case_rows[0][CasosPagoCols.MONTO_BANCO] == 25443565
    assert {row[DistribucionCols.CLIENTE] for row in dist_rows} == {"GEOEXCON", "EQUINORTE"}
    for row in dist_rows:
        assert row[DistribucionCols.CLIENTE]
        assert row[DistribucionCols.CREDITO]
        assert row[DistribucionCols.FECHA_LIMITE]
        assert row[DistribucionCols.DIAS_MORA] is not None
        assert row[DistribucionCols.VALOR_EXTRACTO] is not None
        assert row[DistribucionCols.ESTADO_LINEA] in {
            EstadoPago.NORMAL,
            EstadoPago.ADELANTADO,
            EstadoPago.ATRASADO,
            EstadoPago.INCOMPLETO,
            EstadoPago.REVISION_MANUAL,
        }


def test_generate_workbook_headers_compatible_with_finalize():
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        setup_client_structure(client)
        client.downloaded_files["banco.xlsx"] = create_excel(["Fecha", "Crédito", "Concepto", "Transacción"], [])

        await generate_payment_validation(client, date(2026, 5, 10))
        workbook = load_generated_workbook(client)

        assert ReviewSheets.CONTROL in workbook.sheetnames
        assert ReviewSheets.CASOS_PAGO in workbook.sheetnames
        assert ReviewSheets.DISTRIBUCION in workbook.sheetnames
        assert ReviewSheets.LISTAS in workbook.sheetnames
        assert ReviewSheets.ERRORES in workbook.sheetnames
        assert "Adelantados" not in workbook.sheetnames

        ws_cp = workbook[ReviewSheets.CASOS_PAGO]
        ws_dist = workbook[ReviewSheets.DISTRIBUCION]
        cp_hr = _header_row_index(ws_cp, CasosPagoCols.ID_PAGO)
        dist_hr = _header_row_index(ws_dist, DistribucionCols.ID_PAGO)
        cp_headers = [cell.value for cell in ws_cp[cp_hr]]
        dist_headers = [cell.value for cell in ws_dist[dist_hr]]
        assert cp_headers == CasosPagoCols.HEADERS
        assert dist_headers == DistribucionCols.HEADERS

    asyncio.run(run_test())


def test_extract_pending_installment_detects_hbi_header_row_3():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["INFO CORPORATIVA", ""])
    ws.append(["Mas info", ""])
    ws.append(["Día", "Mes", "Año", "Cuota", "Fecha de pago", "Total pagado"])
    ws.append([23, 12, 2025, 5000, "2025-12-20", 5000])
    ws.append([23, 1, 2026, 5000, "", ""])
    
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue())
    assert res["header_row"] == 3
    assert res["fecha_limite"] == date(2026, 1, 23)
    assert res["valor_extracto"] == 5000

def test_extract_pending_installment_with_split_due_date():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Día", "Mes", "Año", "Cuota", "Fecha de pago", "Total pagado"])
    ws.append([15, 6, 2026, 1200, "", ""])
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue())
    assert res["fecha_limite"] == date(2026, 6, 15)
    assert res["valor_extracto"] == 1200

def test_extract_pending_installment_prefers_first_unpaid_fecha_pago():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fecha límite", "Cuota", "Fecha de pago"])
    ws.append([date(2025, 11, 1), 100, date(2025, 10, 30)])
    ws.append([date(2025, 12, 1), 100, None])
    ws.append([date(2026, 1, 1), 100, None])
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue())
    assert res["fecha_limite"] == date(2025, 12, 1)

def test_extract_pending_installment_uses_cuota_as_base_amount():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fecha límite", "Cuota", "Fecha de pago", "Total a pagar"])
    ws.append([date(2025, 12, 1), 999, None, 1500])
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue())
    assert res["valor_extracto"] == 999

def test_extract_pending_installment_ignores_intereses_mora():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Fecha límite", "Cuota", "Intereses de mora", "Fecha de pago"])
    ws.append([date(2025, 12, 1), 500, 50, None])
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue())
    assert res["valor_extracto"] == 500

def test_extract_pending_installment_selects_client_named_sheet():
    wb = openpyxl.Workbook()
    ws_tasa = wb.active
    ws_tasa.title = "TasaImplicita"
    ws_tasa.append(["Fecha límite", "Cuota", "Fecha de pago"])
    ws_tasa.append([date(2020, 1, 1), 1, None])
    
    ws_geo = wb.create_sheet("GEOEXCON")
    ws_geo.append(["Fecha límite", "Cuota", "Fecha de pago"])
    ws_geo.append([date(2025, 12, 1), 500, None])
    
    out = io.BytesIO()
    wb.save(out)
    
    res = _extract_pending_installment(out.getvalue(), "GEOEXCON")
    assert res["sheet_name"] == "GEOEXCON"
    assert res["fecha_limite"] == date(2025, 12, 1)

def test_generate_proposes_all_pending_installments_geoexcon():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        setup_client_structure(client, include_web_urls=True)
        # El pago es el 30/12/2025
        headers = ["Fecha", "Crédito", "Concepto", "Transacción"]
        data = [
            [date(2025, 12, 30), 500000, "GEOEXCON", ""],
        ]
        client.downloaded_files["banco.xlsx"] = create_excel(headers, data)
        
        client.downloaded_files["clientes/GEOEXCON/254/Tabla amortizacion GEOEXCON 254.xlsx"] = create_amortization_excel([
            [date(2025, 12, 23), None, None, 12000000, None],
        ])
        client.downloaded_files["clientes/GEOEXCON/231/Tabla amortizacion GEOEXCON 231.xlsx"] = create_amortization_excel([
            [date(2025, 12, 23), None, None, 12000000, None],
        ])
        
        # En el setup, geoexcon tiene creditos 254 y 231 (fechas limite 23/12/2025)
        # por lo que ambas deberian ser PENDIENTE_MORA con mora=7
        with _run_with_pdf_mock(client):
            res = await generate_payment_validation(client, date(2026, 1, 1))
        
        wb = load_generated_workbook(client)
        ws_dist = wb[ReviewSheets.DISTRIBUCION]
        dist_dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
        dist_rows = list(ws_dist.iter_rows(min_row=dist_dr, values_only=True))
        ws_errores = wb[ReviewSheets.ERRORES]
        err_dr = _first_data_row(ws_errores)
        errores_rows = list(ws_errores.iter_rows(min_row=err_dr, values_only=True))
        assert len(dist_rows) == 2, f"Errores encontrados: {errores_rows}"

        dist_hr = _header_row_index(ws_dist, DistribucionCols.ID_PAGO)
        headers = [c.value for c in ws_dist[dist_hr]]
        estado_idx = headers.index(DistribucionCols.ESTADO_LINEA)
        mora_idx = headers.index(DistribucionCols.DIAS_MORA)
        total_idx = headers.index(DistribucionCols.TOTAL_APLICADO)
        mora_apl_idx = headers.index(DistribucionCols.INTERESES_MORA)
        cred_idx = headers.index(DistribucionCols.CREDITO)
        
        creditos = [r[cred_idx] for r in dist_rows]
        assert "254" in creditos and "231" in creditos
        
        for idx, r in enumerate(dist_rows):
            if r[estado_idx] != EstadoPago.ATRASADO:
                print("DEBUG GEOEXCON ROW:", r)
            assert r[estado_idx] == EstadoPago.ATRASADO
            assert r[mora_idx] == 7
            if idx == 0:
                assert isinstance(r[total_idx], str) and r[total_idx].startswith("=")
            else:
                assert r[total_idx] in ["", None]
            assert r[mora_apl_idx] in ["", None]

    asyncio.run(run_test())

def test_generate_proposes_all_future_installments_equinorte():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        setup_client_structure(client)
        
        # Add another credit to equinorte
        client.folder_children["clientes/EQUINORTE"].append(make_item("901", is_folder=True))
        client.folder_children["clientes/EQUINORTE/901"] = [
            make_item("Extracto 2026-01-22 CREDITO # 901.pdf"),
            make_item("Tabla amortizacion EQUINORTE 901.xlsx")
        ]
        client.downloaded_files["clientes/EQUINORTE/901/Tabla amortizacion EQUINORTE 901.xlsx"] = create_amortization_excel([
            [date(2026, 1, 22), None, None, 1000, None],
        ])
        
        headers = ["Fecha", "Crédito", "Concepto", "Transacción"]
        data = [
            [date(2025, 12, 30), 500000, "EQUINORTE", ""],
        ]
        client.downloaded_files["banco.xlsx"] = create_excel(headers, data)
        
        client.downloaded_files["clientes/EQUINORTE/900/Tabla amortizacion EQUINORTE 900.xlsx"] = create_amortization_excel([
            [date(2026, 1, 22), None, None, 1000, None],
        ])
        
        client.downloaded_files["clientes/EQUINORTE/901/Extracto 2026-01-22 CREDITO # 901.pdf"] = create_pdf_bytes(
            "4.500", date(2026, 1, 22)
        )
        client.downloaded_files["clientes/EQUINORTE/900/Extracto 2025-12-30 CREDITO # 900.pdf"] = create_pdf_bytes(
            "5.000.000", date(2026, 1, 22)
        )

        with _run_with_pdf_mock(client):
            res = await generate_payment_validation(client, date(2026, 1, 1))
        
        wb = load_generated_workbook(client)
        ws_dist = wb[ReviewSheets.DISTRIBUCION]
        dist_dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
        dist_rows = list(ws_dist.iter_rows(min_row=dist_dr, values_only=True))
        ws_errores = wb[ReviewSheets.ERRORES]
        err_dr = _first_data_row(ws_errores)
        errores_rows = list(ws_errores.iter_rows(min_row=err_dr, values_only=True))
        assert len(dist_rows) == 2, f"Errores encontrados: {errores_rows}"

        dist_hr = _header_row_index(ws_dist, DistribucionCols.ID_PAGO)
        headers = [c.value for c in ws_dist[dist_hr]]
        estado_idx = headers.index(DistribucionCols.ESTADO_LINEA)

        for r in dist_rows:
            assert r[estado_idx] == EstadoPago.ADELANTADO

    asyncio.run(run_test())


def test_generate_does_not_filter_by_amount_or_date():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(
            client,
            "777",
            tabla_cuota=99999.0,
            pdf_total_str="12.345",
            due_date=date(2025, 12, 23),
            banco_date=date(2025, 12, 30),
        )

        with _run_with_pdf_mock(client):
            result = await generate_payment_validation(client, date(2026, 1, 1))

        workbook = load_generated_workbook(client)
        dist_rows = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])
        err_dr = _first_data_row(workbook[ReviewSheets.ERRORES])
        error_rows = list(workbook[ReviewSheets.ERRORES].iter_rows(min_row=err_dr, values_only=True))

        assert result["summary"]["errores"] == 0, f"Errores inesperados: {error_rows}"
        assert len(dist_rows) == 1
        assert dist_rows[0][DistribucionCols.CREDITO] == "777"
        assert dist_rows[0][DistribucionCols.ESTADO_LINEA] == EstadoPago.ATRASADO
        assert dist_rows[0][DistribucionCols.VALOR_EXTRACTO] == 12345.0

    asyncio.run(run_test())


def test_generate_no_pending_installment_error_for_realistic_hbi_table():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("HBI", is_folder=True)]
        client.folder_children["clientes/HBI"] = [make_item("321", is_folder=True)]
        client.folder_children["clientes/HBI/321"] = [
            make_item("Extracto 2025-12-23 CREDITO # 321.pdf"),
            make_item("Tabla amortizacion HBI 321.xlsx"),
        ]
        client.downloaded_files["clientes/HBI/321/Tabla amortizacion HBI 321.xlsx"] = create_excel(
            ["Info general", ""],
            [
                ["Más contexto", ""],
                ["Día", "Mes", "Año", "Cuota", "Fecha de pago", "Total pagado"],
                [23, 12, 2025, 5000, None, None],
            ],
        )
        client.downloaded_files["clientes/HBI/321/Extracto 2025-12-23 CREDITO # 321.pdf"] = create_pdf_bytes(
            "5.000", date(2025, 12, 23)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 23), 5000, "HBI", ""]],
        )

        with _run_with_pdf_mock(client):
            result = await generate_payment_validation(client, date(2026, 1, 1))

        workbook = load_generated_workbook(client)
        dist_rows = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])
        err_dr = _first_data_row(workbook[ReviewSheets.ERRORES])
        error_rows = list(workbook[ReviewSheets.ERRORES].iter_rows(min_row=err_dr, values_only=True))

        assert result["summary"]["errores"] == 0, f"Errores inesperados: {error_rows}"
        assert len(dist_rows) == 1
        assert not any("pending_installment_not_found" in str(row) for row in error_rows)
        fl = dist_rows[0][DistribucionCols.FECHA_LIMITE]
        assert fl == "2025-12-23" or (hasattr(fl, "isoformat") and fl.isoformat()[:10] == "2025-12-23")

    asyncio.run(run_test())


def test_generate_does_not_use_mora_from_amortization_table():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(
            client,
            "300",
            tabla_cuota=10000.0,
            pdf_total_str="15.000",
            due_date=date(2025, 12, 23),
            banco_date=date(2025, 12, 30),
        )

        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))

        workbook = load_generated_workbook(client)
        row = sheet_to_dicts(workbook[ReviewSheets.DISTRIBUCION])[0]

        assert row[DistribucionCols.VALOR_EXTRACTO] == 15000.0
        assert row[DistribucionCols.INTERESES_MORA] in (None, "")
        assert isinstance(row[DistribucionCols.TOTAL_APLICADO], str)
        assert row[DistribucionCols.TOTAL_APLICADO].startswith("=")
        assert row[DistribucionCols.ESTADO_LINEA] == EstadoPago.ATRASADO

    asyncio.run(run_test())


# ─── Tests: Valor extracto debe venir exclusivamente del PDF ─────────────────

def _make_single_client_setup(client, credit_id: str, tabla_cuota: float, pdf_total_str: str, due_date: date, banco_date: date):
    """Helper: configura un cliente TESTCLIENT con un solo crédito."""
    client.folder_children["clientes"] = [make_item("TESTCLIENT", is_folder=True)]
    client.folder_children["clientes/TESTCLIENT"] = [make_item(credit_id, is_folder=True)]
    fecha_str = due_date.strftime("%Y-%m-%d")
    pdf_name = f"Extracto {fecha_str} CREDITO # {credit_id}.pdf"
    tabla_name = f"Tabla amortizacion TESTCLIENT {credit_id}.xlsx"
    client.folder_children[f"clientes/TESTCLIENT/{credit_id}"] = [
        make_item(pdf_name),
        make_item(tabla_name),
    ]
    client.downloaded_files[f"clientes/TESTCLIENT/{credit_id}/{tabla_name}"] = create_excel(
        ["Fecha límite", "Cuota", "Intereses de mora", "Fecha de pago"],
        [[due_date, tabla_cuota, tabla_cuota * 0.05, None]],
    )
    # Registrar PDF como marcador — el mock_extractor lo parseará
    client.downloaded_files[f"clientes/TESTCLIENT/{credit_id}/{pdf_name}"] = create_pdf_bytes(
        pdf_total_str, due_date
    )
    client.downloaded_files["banco.xlsx"] = create_excel(
        ["Fecha", "Crédito", "Concepto", "Transacción"],
        [[banco_date, 500000, "TESTCLIENT", ""]],
    )


def _fake_fecha_limite_from_marker_bytes():
    """Lee FECHA_LIMITE=YYYY-MM-DD del marcador PDF_MOCK; si falta, fecha fija para un solo extracto."""

    def _inner(pdf_bytes: bytes):
        text = pdf_bytes.decode(errors="ignore")
        if "SKIP_FECHA;" in text:
            return None
        m = re.search(r"FECHA_LIMITE=([0-9]{4}-[0-9]{2}-[0-9]{2})", text)
        if m:
            return date.fromisoformat(m.group(1))
        return date(2000, 1, 1)

    return _inner


@contextlib.contextmanager
def _run_with_pdf_mock(client):
    """Parchea extract_total_a_pagar_from_pdf y extract_fecha_limite_pago_from_pdf (marcadores PDF_MOCK)."""
    with mock.patch(
        "app.application.use_cases.payment_validation_generate.extract_total_a_pagar_from_pdf",
        side_effect=make_pdf_extractor_mock(client),
    ), mock.patch(
        "app.application.use_cases.payment_validation_generate.extract_fecha_limite_pago_from_pdf",
        side_effect=_fake_fecha_limite_from_marker_bytes(),
    ):
        yield


def _get_dist_col_idx(wb, col_name: str) -> int:
    ws = wb[ReviewSheets.DISTRIBUCION]
    hr = _header_row_index(ws, DistribucionCols.ID_PAGO)
    headers = [c.value for c in ws[hr]]
    return headers.index(col_name)


def test_generate_extract_value_from_pdf_total_a_pagar_credit_231():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "231", tabla_cuota=19540684.91, pdf_total_str="18.826.879",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        rows = list(ws_d.iter_rows(min_row=dr, values_only=True))
        assert len(rows) == 1, f"Expected 1 row, got {len(rows)}"
        val_idx = _get_dist_col_idx(wb, DistribucionCols.VALOR_EXTRACTO)
        assert rows[0][val_idx] == 18826879.0, f"Valor extracto incorrecto: {rows[0][val_idx]}"
    asyncio.run(run_test())


def test_generate_extract_value_from_pdf_total_a_pagar_credit_254():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "254", tabla_cuota=6734920.07, pdf_total_str="6.500.739",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        rows = list(ws_d.iter_rows(min_row=dr, values_only=True))
        assert len(rows) == 1
        val_idx = _get_dist_col_idx(wb, DistribucionCols.VALOR_EXTRACTO)
        assert rows[0][val_idx] == 6500739.0, f"Valor extracto incorrecto: {rows[0][val_idx]}"
    asyncio.run(run_test())


def test_generate_pdf_amount_overrides_table_amount():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "100", tabla_cuota=99999.0, pdf_total_str="12.345",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        rows = list(ws_d.iter_rows(min_row=dr, values_only=True))
        assert len(rows) == 1
        val_idx = _get_dist_col_idx(wb, DistribucionCols.VALOR_EXTRACTO)
        assert rows[0][val_idx] == 12345.0
        assert rows[0][val_idx] != 99999.0, "Usó tabla en vez de PDF"
    asyncio.run(run_test())


def test_generate_no_table_amount_fallback_when_pdf_amount_missing():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "555", tabla_cuota=50000.0, pdf_total_str="",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr_d = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        dist_rows = list(ws_d.iter_rows(min_row=dr_d, values_only=True))
        assert len(dist_rows) == 0, "Sin TOTAL A PAGAR en PDF no debe haber línea en Distribucion"
        ws_e = wb[ReviewSheets.ERRORES]
        dr_e = _first_data_row(ws_e)
        err_rows = list(ws_e.iter_rows(min_row=dr_e, values_only=True))
        assert any("extract_amount_not_found" in str(r) for r in err_rows), \
            f"Debe haber error extract_amount_not_found. Errores: {err_rows}"
    asyncio.run(run_test())


def test_generate_extract_not_found_when_pdf_missing():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("NOPDF", is_folder=True)]
        client.folder_children["clientes/NOPDF"] = [make_item("200", is_folder=True)]
        client.folder_children["clientes/NOPDF/200"] = [make_item("Tabla amortizacion NOPDF 200.xlsx")]
        client.downloaded_files["clientes/NOPDF/200/Tabla amortizacion NOPDF 200.xlsx"] = create_excel(
            ["Fecha límite", "Cuota", "Fecha de pago"], [[date(2025, 12, 23), 10000, None]])
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 30), 100000, "NOPDF", ""]])
        # Este test no necesita mock de PDF porque no hay PDF — testea extract_not_found
        await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr_d = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        dist_rows = list(ws_d.iter_rows(min_row=dr_d, values_only=True))
        assert len(dist_rows) == 0, "Sin PDF no debe haber línea en Distribucion"
        ws_e = wb[ReviewSheets.ERRORES]
        dr_e = _first_data_row(ws_e)
        err_rows = list(ws_e.iter_rows(min_row=dr_e, values_only=True))
        assert any("extract_not_found" in str(r) for r in err_rows), \
            f"Debe haber error extract_not_found. Errores: {err_rows}"
    asyncio.run(run_test())


def test_extract_fecha_limite_pago_from_pdf_text_iso_and_european():
    from app.application.services.payment_helpers import extract_fecha_limite_pago_from_pdf_text

    assert extract_fecha_limite_pago_from_pdf_text(
        "Texto Fecha Limite de pago 2026-04-15 más texto"
    ) == date(2026, 4, 15)
    assert extract_fecha_limite_pago_from_pdf_text(
        "Fecha Limite de pago 23/04/2026"
    ) == date(2026, 4, 23)
    assert extract_fecha_limite_pago_from_pdf_text(
        "Fecha Limite de pago 23-04-2026"
    ) == date(2026, 4, 23)
    assert extract_fecha_limite_pago_from_pdf_text(
        "Fecha desembolso 23/10/2025  ...  Fecha Limite de pago 23/04/2026  fin"
    ) == date(2026, 4, 23)
    assert extract_fecha_limite_pago_from_pdf_text(
        "TOTAL A PAGAR $ 100\nTOTAL PAGADO $ 0\nFecha Limite de pago 23/04/2026\n"
    ) == date(2026, 4, 23)


def test_phase1_extractos_subfolder_priority_over_parent_pdf():
    """A: Con EXTRACTOS con PDFs válidos, no se usa el extracto de la carpeta padre."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("MIXCLI", is_folder=True)]
        client.folder_children["clientes/MIXCLI"] = [make_item("100", is_folder=True)]
        client.folder_children["clientes/MIXCLI/100"] = [
            make_item("EXTRACTOS", is_folder=True),
            make_item("Extracto padre CREDITO # 100.pdf"),
            make_item("Tabla amortizacion MIXCLI 100.xlsx"),
        ]
        client.folder_children["clientes/MIXCLI/100/EXTRACTOS"] = [
            make_item("Extracto interno CREDITO # 100.pdf"),
        ]
        tab_path = "clientes/MIXCLI/100/Tabla amortizacion MIXCLI 100.xlsx"
        client.downloaded_files[tab_path] = create_amortization_excel([[date(2025, 12, 23), None, None, 1000, None]])
        client.downloaded_files["clientes/MIXCLI/100/Extracto padre CREDITO # 100.pdf"] = create_pdf_bytes(
            "1.111.111", date(2025, 1, 1)
        )
        client.downloaded_files["clientes/MIXCLI/100/EXTRACTOS/Extracto interno CREDITO # 100.pdf"] = create_pdf_bytes(
            "8.888.888", date(2026, 5, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 23), 100000, "MIXCLI", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.VALOR_EXTRACTO] == 8888888.0
        ws = wb[ReviewSheets.DISTRIBUCION]
        hdr = _header_row_index(ws, DistribucionCols.ID_PAGO)
        dr = hdr + 1
        c_ruta = DistribucionCols.HEADERS.index(DistribucionCols.RUTA) + 1
        expected = (
            "clientes/MIXCLI/100/EXTRACTOS/Extracto interno CREDITO # 100.pdf"
        ).replace("\\", "/")
        assert ws.cell(dr, c_ruta).value.replace("\\", "/") == expected
        c_le = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
        assert ws.cell(dr, c_le).value in ("", None)

    asyncio.run(run_test())


def test_phase1_empty_extractos_fallback_to_credit_folder():
    """B: EXTRACTOS vacía → candidatos en carpeta del crédito."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("FALLB", is_folder=True)]
        client.folder_children["clientes/FALLB"] = [make_item("200", is_folder=True)]
        client.folder_children["clientes/FALLB/200"] = [
            make_item("extractos", is_folder=True),
            make_item("Extracto root CREDITO # 200.pdf"),
            make_item("Tabla amortizacion FALLB 200.xlsx"),
        ]
        client.folder_children["clientes/FALLB/200/extractos"] = []
        client.downloaded_files["clientes/FALLB/200/Tabla amortizacion FALLB 200.xlsx"] = create_amortization_excel(
            [[date(2025, 6, 1), None, None, 2000, None]]
        )
        client.downloaded_files["clientes/FALLB/200/Extracto root CREDITO # 200.pdf"] = create_pdf_bytes(
            "3.333.333", date(2025, 6, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 6, 1), 500000, "FALLB", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.VALOR_EXTRACTO] == 3333333.0
        ws = wb[ReviewSheets.DISTRIBUCION]
        hdr = _header_row_index(ws, DistribucionCols.ID_PAGO)
        dr = hdr + 1
        c_ruta = DistribucionCols.HEADERS.index(DistribucionCols.RUTA) + 1
        exp = "clientes/FALLB/200/Extracto root CREDITO # 200.pdf".replace("\\", "/")
        assert ws.cell(dr, c_ruta).value.replace("\\", "/") == exp

    asyncio.run(run_test())


def test_generate_flat_acimor_style_writes_ruta_for_root_extract():
    """Cliente plano sin subcarpetas de crédito: Ruta al PDF en raíz."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("ACIMRU", is_folder=True)]
        client.folder_children["clientes/ACIMRU"] = [
            make_item("Extracto 2026-04-01 CREDITO # 801.pdf"),
            make_item("Tabla amortizacion ACIMRU ACIMRU.xlsx"),
        ]
        client.downloaded_files["clientes/ACIMRU/Tabla amortizacion ACIMRU ACIMRU.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 5000, None]]
        )
        client.downloaded_files["clientes/ACIMRU/Extracto 2026-04-01 CREDITO # 801.pdf"] = create_pdf_bytes(
            "4.000", date(2026, 4, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "ACIMRU", "pago"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        hdr = _header_row_index(wb[ReviewSheets.DISTRIBUCION], DistribucionCols.ID_PAGO)
        dr = hdr + 1
        c_ruta = DistribucionCols.HEADERS.index(DistribucionCols.RUTA) + 1
        exp = "clientes/ACIMRU/Extracto 2026-04-01 CREDITO # 801.pdf".replace("\\", "/")
        assert wb[ReviewSheets.DISTRIBUCION].cell(dr, c_ruta).value.replace("\\", "/") == exp

    asyncio.run(run_test())


def test_phase1_strict_name_rejects_pdf_without_extracto_in_name():
    """D: No seleccionar PDF que no contenga 'extracto' en el nombre."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("STRICT", is_folder=True)]
        client.folder_children["clientes/STRICT"] = [make_item("300", is_folder=True)]
        client.folder_children["clientes/STRICT/300"] = [
            make_item("Pagaré 300.pdf"),
            make_item("Tabla amortizacion STRICT 300.xlsx"),
        ]
        client.downloaded_files["clientes/STRICT/300/Tabla amortizacion STRICT 300.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 23), None, None, 5000, None]]
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 23), 100000, "STRICT", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        err_rows = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert not sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert any(
            "extract_not_found" in str(r.get(ErroresCols.CODIGO_TECNICO, "")) for r in err_rows
        )

    asyncio.run(run_test())


def test_phase1_selects_extract_with_max_fecha_limite():
    """F: Entre varios extractos, gana el de fecha límite máxima."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("MAXDT", is_folder=True)]
        client.folder_children["clientes/MAXDT"] = [make_item("400", is_folder=True)]
        client.folder_children["clientes/MAXDT/400"] = [
            make_item("Extracto viejo CREDITO # 400.pdf"),
            make_item("Extracto nuevo CREDITO # 400.pdf"),
            make_item("Tabla amortizacion MAXDT 400.xlsx"),
        ]
        client.downloaded_files["clientes/MAXDT/400/Tabla amortizacion MAXDT 400.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/MAXDT/400/Extracto viejo CREDITO # 400.pdf"] = create_pdf_bytes(
            "1.000", date(2025, 1, 1)
        )
        client.downloaded_files["clientes/MAXDT/400/Extracto nuevo CREDITO # 400.pdf"] = create_pdf_bytes(
            "9.999", date(2026, 8, 15)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "MAXDT", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.VALOR_EXTRACTO] == 9999.0

    asyncio.run(run_test())


def test_phase1_no_selection_when_fecha_limite_unreadable():
    """G: Sin fecha legible en ningún candidato → error, sin elegir por orden."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("NODTE", is_folder=True)]
        client.folder_children["clientes/NODTE"] = [make_item("500", is_folder=True)]
        client.folder_children["clientes/NODTE/500"] = [
            make_item("Extracto a CREDITO # 500.pdf"),
            make_item("Extracto b CREDITO # 500.pdf"),
            make_item("Tabla amortizacion NODTE 500.xlsx"),
        ]
        client.downloaded_files["clientes/NODTE/500/Tabla amortizacion NODTE 500.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/NODTE/500/Extracto a CREDITO # 500.pdf"] = create_pdf_bytes(
            "5.000", date(2025, 1, 1)
        )
        client.downloaded_files["clientes/NODTE/500/Extracto b CREDITO # 500.pdf"] = create_pdf_bytes(
            "9.000", date(2026, 1, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "NODTE", ""]],
        )
        extractor = make_pdf_extractor_mock(client)
        with mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_total_a_pagar_from_pdf",
            side_effect=extractor,
        ), mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_fecha_limite_pago_from_pdf",
            return_value=None,
        ):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert not sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err_rows = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert any(
            "fecha_limite_extracto_not_readable"
            in str(r.get(ErroresCols.CODIGO_TECNICO, ""))
            for r in err_rows
        )

    asyncio.run(run_test())


def test_phase1_tie_max_fecha_limite_does_not_pick_silently():
    """H: Empate en fecha máxima → error explícito."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("TIECL", is_folder=True)]
        client.folder_children["clientes/TIECL"] = [make_item("600", is_folder=True)]
        client.folder_children["clientes/TIECL/600"] = [
            make_item("Extracto uno CREDITO # 600.pdf"),
            make_item("Extracto dos CREDITO # 600.pdf"),
            make_item("Tabla amortizacion TIECL 600.xlsx"),
        ]
        client.downloaded_files["clientes/TIECL/600/Tabla amortizacion TIECL 600.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 1), None, None, 1000, None]]
        )
        same = date(2026, 3, 1)
        client.downloaded_files["clientes/TIECL/600/Extracto uno CREDITO # 600.pdf"] = create_pdf_bytes(
            "1.111", same
        )
        client.downloaded_files["clientes/TIECL/600/Extracto dos CREDITO # 600.pdf"] = create_pdf_bytes(
            "2.222", same
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "TIECL", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert not sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err_rows = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert any(
            "extract_tie_max_fecha_limite" in str(r.get(ErroresCols.CODIGO_TECNICO, "")) for r in err_rows
        )

    asyncio.run(run_test())


def test_phase1_possibly_finalized_observation_on_folder_name():
    """I: Carpeta TERMINADO añade observación entendible y se concatena con la observación base."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("PFCLI", is_folder=True)]
        client.folder_children["clientes/PFCLI"] = [make_item("700-TERMINADO", is_folder=True)]
        client.folder_children["clientes/PFCLI/700-TERMINADO"] = [
            make_item("Extracto 2025-12-23 CREDITO # 700.pdf"),
            make_item("Tabla amortizacion PFCLI 700.xlsx"),
        ]
        client.downloaded_files["clientes/PFCLI/700-TERMINADO/Tabla amortizacion PFCLI 700.xlsx"] = (
            create_amortization_excel([[date(2025, 12, 23), None, None, 10000, None]])
        )
        client.downloaded_files["clientes/PFCLI/700-TERMINADO/Extracto 2025-12-23 CREDITO # 700.pdf"] = create_pdf_bytes(
            "7.000", date(2025, 12, 23)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 1), 500000, "PFCLI", "concepto-trx"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert "Revisión" not in wb.sheetnames
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "TERMINADO/FINALIZADO/CANCELADO/PAGADO/LIQUIDADO" in obs
        assert "crédito sigue vigente" in obs
        assert "Pago adelantado; requiere reprogramación" in obs

    asyncio.run(run_test())


def test_phase11_flat_client_without_credit_subfolders_acimor_style():
    """E: Sin subcarpetas de crédito; extractos en raíz como una unidad operativa."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("ACIMOR", is_folder=True)]
        client.folder_children["clientes/ACIMOR"] = [
            make_item("Extracto 2026-04-01 CREDITO # 801.pdf"),
            make_item("Tabla amortizacion ACIMOR ACIMOR.xlsx"),
        ]
        client.downloaded_files["clientes/ACIMOR/Tabla amortizacion ACIMOR ACIMOR.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 5000, None]]
        )
        client.downloaded_files["clientes/ACIMOR/Extracto 2026-04-01 CREDITO # 801.pdf"] = create_pdf_bytes(
            "4.000", date(2026, 4, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "ACIMOR", "pago acimor"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert "Revisión" not in wb.sheetnames
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert not any(
            "credit_folder_not_found" in str(r.get(ErroresCols.CODIGO_TECNICO, "")) for r in err
        )
        assert len(dist) == 1
        assert dist[0][DistribucionCols.CREDITO] == "801"

    asyncio.run(run_test())


def test_phase11_mixed_candidates_one_illegible_fecha_still_selects_max():
    """Variación D: algunos PDFs ilegibles pero al menos uno con fecha válida."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("MIXRO", is_folder=True)]
        client.folder_children["clientes/MIXRO"] = [make_item("410", is_folder=True)]
        client.folder_children["clientes/MIXRO/410"] = [
            make_item("Extracto sin fecha CREDITO # 410.pdf"),
            make_item("Extracto con fecha CREDITO # 410.pdf"),
            make_item("Tabla amortizacion MIXRO 410.xlsx"),
        ]
        client.downloaded_files["clientes/MIXRO/410/Tabla amortizacion MIXRO 410.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/MIXRO/410/Extracto sin fecha CREDITO # 410.pdf"] = create_pdf_bytes(
            "500.000", date(2099, 1, 1), skip_fecha_fake=True
        )
        client.downloaded_files["clientes/MIXRO/410/Extracto con fecha CREDITO # 410.pdf"] = create_pdf_bytes(
            "123", date(2026, 4, 23)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "MIXRO", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.VALOR_EXTRACTO] == 123.0
        assert not any(
            "fecha_limite_extracto_not_readable" in str(r.get(ErroresCols.CODIGO_TECNICO, "")) for r in err
        )

    asyncio.run(run_test())


def test_phase11_fecha_limite_unreadable_in_errores_includes_payment_context():
    """C: fecha_limite_extracto_not_readable con guía legible; código técnico solo al final."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("NODTE2", is_folder=True)]
        client.folder_children["clientes/NODTE2"] = [make_item("501", is_folder=True)]
        client.folder_children["clientes/NODTE2/501"] = [
            make_item(
                "Extracto x CREDITO # 501.pdf",
                web_url="https://example.invalid/extracto501.pdf",
            ),
            make_item("Tabla amortizacion NODTE2 501.xlsx"),
        ]
        client.downloaded_files["clientes/NODTE2/501/Tabla amortizacion NODTE2 501.xlsx"] = create_amortization_excel(
            [[date(2025, 12, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/NODTE2/501/Extracto x CREDITO # 501.pdf"] = create_pdf_bytes(
            "12.345", date(2025, 12, 1), skip_fecha_fake=True
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "NODTE2", "ref-banco-123"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert "Revisión" not in wb.sheetnames
        assert not sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        errs = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert errs
        row = errs[0]
        assert row.get(ErroresCols.CREDITO) == "501"
        assert row.get(ErroresCols.TIPO_CASO) == "Extracto"
        desc = str(row.get(ErroresCols.DESCRIPCION, "")).lower()
        assert "fecha límite" in desc
        assert "fecha_limite_extracto_not_readable" not in desc
        assert row.get(ErroresCols.CODIGO_TECNICO) == "fecha_limite_extracto_not_readable"
        assert row.get(ErroresCols.CLIENTE) == "NODTE2"
        assert row.get(ErroresCols.ID_PAGO)
        ws_e = wb[ReviewSheets.ERRORES]
        dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
        c_link = ErroresCols.HEADERS.index(ErroresCols.LINK_EXTRACTO) + 1
        cell = ws_e.cell(row=dr, column=c_link)
        assert cell.value == "Ver extracto crédito 501"
        assert getattr(cell, "hyperlink", None) is not None

    asyncio.run(run_test())


def test_errores_sheet_operational_headers_no_bank_columns():
    esperado = [
        "ID Pago",
        "Cliente",
        "Crédito",
        "Tipo de caso",
        "Descripción para revisión",
        "Qué debe hacer",
        "Requiere soporte",
        "Link extracto",
        "Link carpeta crédito",
        "Código técnico",
    ]
    assert ErroresCols.HEADERS == esperado
    prohibidas = {
        "Monto banco",
        "Fecha banco",
        "Concepto banco",
        "Detalle técnico",
        "Error",
        "Acción recomendada",
    }
    assert not prohibidas.intersection(set(ErroresCols.HEADERS))


def test_errores_extract_not_found_friendly_message_and_code_column():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        credit = make_item(
            "200",
            is_folder=True,
            web_url="https://example.invalid/carpeta200",
        )
        client.folder_children["clientes"] = [
            make_item("NOPDF", is_folder=True),
        ]
        client.folder_children["clientes/NOPDF"] = [credit]
        client.folder_children["clientes/NOPDF/200"] = [make_item("Tabla amortizacion NOPDF 200.xlsx")]
        client.downloaded_files["clientes/NOPDF/200/Tabla amortizacion NOPDF 200.xlsx"] = create_excel(
            ["Fecha límite", "Cuota", "Fecha de pago"], [[date(2025, 12, 23), 10000, None]]
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 30), 100000, "NOPDF", ""]],
        )
        await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        errs = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert len(errs) >= 1
        row = next(r for r in errs if r.get(ErroresCols.CODIGO_TECNICO) == "extract_not_found")
        assert row[ErroresCols.TIPO_CASO] == "Extracto"
        assert "PDF" in row[ErroresCols.DESCRIPCION]
        assert "extract_not_found" not in str(row.get(ErroresCols.DESCRIPCION, ""))
        ws_e = wb[ReviewSheets.ERRORES]
        dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
        col_dir = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
        cdir = ws_e.cell(row=dr, column=col_dir)
        assert cdir.value == "Ver carpeta crédito 200"
        assert cdir.hyperlink is not None
        assert str(cdir.hyperlink.target).startswith("https://")

    asyncio.run(run_test())


def test_errores_flat_client_extract_not_found_carpeta_hyperlink():
    """Cliente plano: Link carpeta crédito apunta a webUrl de la carpeta del cliente."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client_url = "https://sharepoint.invalid/clientes/MINCIVIL"
        client.folder_web_urls["clientes/MINCIVIL"] = client_url
        client.folder_children["clientes"] = [
            make_item("MINCIVIL", is_folder=True, web_url=client_url),
        ]
        client.folder_children["clientes/MINCIVIL"] = [
            make_item("Tabla amortizacion MINCIVIL MINCIVIL.xlsx"),
        ]
        client.downloaded_files["clientes/MINCIVIL/Tabla amortizacion MINCIVIL MINCIVIL.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 1000, None]])
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "MINCIVIL", "pago"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        row = next(
            r
            for r in sheet_to_dicts(wb[ReviewSheets.ERRORES])
            if r.get(ErroresCols.CODIGO_TECNICO) == "extract_not_found"
        )
        assert row[ErroresCols.LINK_EXTRACTO] in ("", None) or not str(row[ErroresCols.LINK_EXTRACTO]).startswith("http")
        ws_e = wb[ReviewSheets.ERRORES]
        dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
        col_dir = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
        cdir = ws_e.cell(row=dr, column=col_dir)
        assert cdir.value == "Ver carpeta MINCIVIL"
        assert cdir.hyperlink is not None
        assert str(cdir.hyperlink.target) == client_url

    asyncio.run(run_test())


def test_errores_flat_client_illegible_extract_both_hyperlinks():
    """Cliente plano: PDF ilegible → Link extracto y Link carpeta crédito con hipervínculo."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client_url = "https://sharepoint.invalid/clientes/MINCIVIL2"
        pdf_url = "https://sharepoint.invalid/clientes/MINCIVIL2/Extracto sin fecha.pdf"
        client.folder_web_urls["clientes/MINCIVIL2"] = client_url
        client.folder_children["clientes"] = [
            make_item("MINCIVIL2", is_folder=True, web_url=client_url),
        ]
        client.folder_children["clientes/MINCIVIL2"] = [
            make_item("Extracto sin fecha CREDITO # 99.pdf", web_url=pdf_url),
            make_item("Tabla amortizacion MINCIVIL2 MINCIVIL2.xlsx"),
        ]
        client.downloaded_files["clientes/MINCIVIL2/Extracto sin fecha CREDITO # 99.pdf"] = (
            create_pdf_bytes("1.000", date(2099, 1, 1), skip_fecha_fake=True)
        )
        client.downloaded_files["clientes/MINCIVIL2/Tabla amortizacion MINCIVIL2 MINCIVIL2.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 1000, None]])
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "MINCIVIL2", "pago"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        assert any(
            r.get(ErroresCols.CODIGO_TECNICO) == "fecha_limite_extracto_not_readable"
            for r in sheet_to_dicts(wb[ReviewSheets.ERRORES])
        )
        ws_e = wb[ReviewSheets.ERRORES]
        dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
        col_ext = ErroresCols.HEADERS.index(ErroresCols.LINK_EXTRACTO) + 1
        col_dir = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
        cext = ws_e.cell(row=dr, column=col_ext)
        cdir = ws_e.cell(row=dr, column=col_dir)
        assert cext.value == "Ver extracto MINCIVIL2"
        assert cext.hyperlink is not None
        assert str(cext.hyperlink.target) == pdf_url
        assert cdir.value == "Ver carpeta MINCIVIL2"
        assert cdir.hyperlink is not None
        assert str(cdir.hyperlink.target) == client_url

    asyncio.run(run_test())


def _errores_link_cell(ws, row: int, col_name: str):
    col = ErroresCols.HEADERS.index(col_name) + 1
    return ws.cell(row=row, column=col)


def test_errores_sin_url_no_muestra_ver_carpeta_enganoso():
    """Sin webUrl real: celda Link carpeta crédito vacía (no texto clickeable falso)."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("NOURL", is_folder=True)]
        client.folder_children["clientes/NOURL"] = [
            make_item("Tabla amortizacion NOURL NOURL.xlsx"),
        ]
        client.downloaded_files["clientes/NOURL/Tabla amortizacion NOURL NOURL.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 1000, None]])
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "NOURL", "pago"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_e = wb[ReviewSheets.ERRORES]
        dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
        cdir = _errores_link_cell(ws_e, dr, ErroresCols.LINK_CARPETA_CREDITO)
        assert cdir.value in (None, "")
        assert cdir.hyperlink is None

    asyncio.run(run_test())


def test_errores_extract_tie_max_fecha_friendly_message():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("TIECLX", is_folder=True)]
        client.folder_children["clientes/TIECLX"] = [make_item("601", is_folder=True)]
        client.folder_children["clientes/TIECLX/601"] = [
            make_item("Extracto uno CREDITO # 601.pdf"),
            make_item("Extracto dos CREDITO # 601.pdf"),
            make_item("Tabla amortizacion TIECLX 601.xlsx"),
        ]
        client.downloaded_files["clientes/TIECLX/601/Tabla amortizacion TIECLX 601.xlsx"] = (
            create_amortization_excel([[date(2025, 12, 1), None, None, 1000, None]])
        )
        same = date(2026, 3, 15)
        client.downloaded_files["clientes/TIECLX/601/Extracto uno CREDITO # 601.pdf"] = create_pdf_bytes("1", same)
        client.downloaded_files["clientes/TIECLX/601/Extracto dos CREDITO # 601.pdf"] = create_pdf_bytes("2", same)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2025, 12, 15), 100000, "TIECLX", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        row = sheet_to_dicts(wb[ReviewSheets.ERRORES])[0]
        assert row[ErroresCols.CODIGO_TECNICO] == "extract_tie_max_fecha_limite"
        assert "misma fecha" in str(row[ErroresCols.DESCRIPCION]).lower()
        assert "extract_tie_max_fecha_limite" not in str(row.get(ErroresCols.DESCRIPCION, ""))

    asyncio.run(run_test())


def test_phase12_v2_fecha_limite_column_comes_from_pdf_not_amortization_table():
    """En v2, Distribución «Fecha límite» es la del PDF seleccionado aunque la tabla tenga otra."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("PDFLIM", is_folder=True)]
        client.folder_children["clientes/PDFLIM"] = [make_item("888", is_folder=True)]
        client.folder_children["clientes/PDFLIM/888"] = [
            make_item("Extracto 2026-08-10 CREDITO # 888.pdf"),
            make_item("Tabla amortizacion PDFLIM 888.xlsx"),
        ]
        client.downloaded_files["clientes/PDFLIM/888/Tabla amortizacion PDFLIM 888.xlsx"] = create_amortization_excel(
            [[date(2025, 1, 1), None, None, 99999, None]],
        )
        client.downloaded_files["clientes/PDFLIM/888/Extracto 2026-08-10 CREDITO # 888.pdf"] = create_pdf_bytes(
            "50.000", date(2026, 8, 10)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 8, 10), 100000, "PDFLIM", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        fl = dist[0][DistribucionCols.FECHA_LIMITE]
        assert fl == "2026-08-10" or (hasattr(fl, "isoformat") and fl.isoformat()[:10] == "2026-08-10")

    asyncio.run(run_test())


def test_phase12_valid_extract_ambiguous_excel_tables_still_distrib_and_warns():
    """Dos Excel «tabla amortización» empatados → observación ambigua, fila en Distribución."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("AMBIV", is_folder=True)]
        client.folder_children["clientes/AMBIV"] = [make_item("777", is_folder=True)]
        client.folder_children["clientes/AMBIV/777"] = [
            make_item("Extracto 2026-02-01 CREDITO # 777.pdf"),
            make_item("Tabla amortizacion AMBIV 777.xlsx"),
            make_item("Tabla amortizacion AMBIV 777 copia.xlsx"),
        ]
        tab_path_a = "clientes/AMBIV/777/Tabla amortizacion AMBIV 777.xlsx"
        tab_path_b = "clientes/AMBIV/777/Tabla amortizacion AMBIV 777 copia.xlsx"
        row_tabla = [[date(2026, 2, 1), None, None, 5000, None]]
        client.downloaded_files[tab_path_a] = create_amortization_excel(row_tabla)
        client.downloaded_files[tab_path_b] = create_amortization_excel(row_tabla)
        client.downloaded_files["clientes/AMBIV/777/Extracto 2026-02-01 CREDITO # 777.pdf"] = create_pdf_bytes(
            "10.000", date(2026, 2, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 2, 1), 100000, "AMBIV", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert len(dist) == 1
        assert "tabla ambigua" in str(dist[0].get(DistribucionCols.OBSERVACION, "")).lower()
        assert not any(
            "amortization_table_ambiguous" in str(r.get(ErroresCols.CODIGO_TECNICO, "")) for r in err
        )

    asyncio.run(run_test())


def test_phase12_valid_extract_without_excel_table_still_distrib_and_warns():
    """Sin tabla Excel operativa pero con extracto válido → fila + advertencia tabla no encontrada."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("NOTAB", is_folder=True)]
        client.folder_children["clientes/NOTAB"] = [make_item("666", is_folder=True)]
        client.folder_children["clientes/NOTAB/666"] = [
            make_item("Extracto 2026-03-15 CREDITO # 666.pdf"),
        ]
        client.downloaded_files["clientes/NOTAB/666/Extracto 2026-03-15 CREDITO # 666.pdf"] = create_pdf_bytes(
            "3.000", date(2026, 3, 15)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 3, 15), 50000, "NOTAB", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert "tabla no encontrada" in str(dist[0].get(DistribucionCols.OBSERVACION, "")).lower()

    asyncio.run(run_test())


def test_phase12_acimor_pdf_informativo_no_interfiere_con_tabla_xlsx():
    """PDF «tabla amortización» informativo + Excel operativo → solo Excel cuent."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("ACIMOR2", is_folder=True)]
        client.folder_children["clientes/ACIMOR2"] = [
            make_item("TABLA DE AMORTIZACION ACIMOR2.pdf"),
            make_item("Tabla de amortizacion ACIMOR2 ACIMOR2.xlsx"),
            make_item("Extracto 2026-04-01 CREDITO # 901.pdf"),
        ]
        client.downloaded_files["clientes/ACIMOR2/TABLA DE AMORTIZACION ACIMOR2.pdf"] = b"%PDF fake"
        client.downloaded_files["clientes/ACIMOR2/Tabla de amortizacion ACIMOR2 ACIMOR2.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 5000, None]])
        )
        client.downloaded_files["clientes/ACIMOR2/Extracto 2026-04-01 CREDITO # 901.pdf"] = create_pdf_bytes(
            "4.000", date(2026, 4, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "ACIMOR2", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.CREDITO] == "901"

    asyncio.run(run_test())


def test_phase12_flat_client_ambiguous_tables_valid_extract_distrib():
    """Cliente plano: dos Excel empatados + extracto válido → Distribución con advertencia."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("FLAMB", is_folder=True)]
        client.folder_children["clientes/FLAMB"] = [
            make_item("Extracto 2026-05-10 CREDITO # 55.pdf"),
            make_item("Tabla amortizacion FLAMB FLAMB.xlsx"),
            make_item("Tabla amortizacion FLAMB FLAMB duplicado.xlsx"),
        ]
        row_t = [[date(2026, 5, 10), None, None, 800, None]]
        client.downloaded_files["clientes/FLAMB/Tabla amortizacion FLAMB FLAMB.xlsx"] = create_amortization_excel(row_t)
        client.downloaded_files["clientes/FLAMB/Tabla amortizacion FLAMB FLAMB duplicado.xlsx"] = (
            create_amortization_excel(row_t)
        )
        client.downloaded_files["clientes/FLAMB/Extracto 2026-05-10 CREDITO # 55.pdf"] = create_pdf_bytes(
            "800", date(2026, 5, 10)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 5, 10), 50000, "FLAMB", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert "tabla ambigua" in str(dist[0].get(DistribucionCols.OBSERVACION, "")).lower()

    asyncio.run(run_test())


def test_phase12_flat_infer_credito_from_pdf_obligacion_gb():
    """Cliente plano: número de obligación en texto del PDF → columna Crédito."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        pdf_path = "clientes/ROOT82/Extracto abril.pdf"
        client.folder_children["clientes"] = [make_item("ROOT82", is_folder=True)]
        client.folder_children["clientes/ROOT82"] = [
            make_item("Extracto abril.pdf"),
            make_item("Tabla amortizacion ROOT82 ROOT82.xlsx"),
        ]
        client.downloaded_files[pdf_path] = (
            "PDF_MOCK:FECHA_LIMITE=2026-04-01;No, Obligación GB-1022-082\nTOTAL A PAGAR 100".encode()
        )
        client.downloaded_files["clientes/ROOT82/Tabla amortizacion ROOT82 ROOT82.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 100, None]],
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 5000, "ROOT82", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.CREDITO] == "82"

    asyncio.run(run_test())


def test_phase12_auxiliary_compare_fecha_limite_vs_last_fecha_pago_observations():
    """Comparación tabla auxiliar: observaciones según última Fecha pago; Fecha límite sigue siendo la del PDF."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        base = "clientes/CMPRV/501"
        client.folder_children["clientes"] = [make_item("CMPRV", is_folder=True)]
        client.folder_children["clientes/CMPRV"] = [make_item("501", is_folder=True)]
        client.folder_children[f"{base}"] = [
            make_item("Extracto cmp.pdf"),
            make_item("Tabla amortizacion CMPRV 501.xlsx"),
        ]
        client.downloaded_files[f"{base}/Tabla amortizacion CMPRV 501.xlsx"] = create_amortization_excel(
            [
                [date(2025, 10, 1), date(2025, 9, 10), None, 1000, None],
                [date(2025, 12, 23), date(2025, 12, 20), None, 1000, None],
            ],
        )

        pdf_fecha = date(2025, 12, 20)
        client.downloaded_files[f"{base}/Extracto cmp.pdf"] = create_pdf_bytes("5.000", pdf_fecha)

        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[pdf_fecha, 50000, "CMPRV", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "coincide con la última fecha de pago" in obs.lower()
        fl = dist[0][DistribucionCols.FECHA_LIMITE]
        assert fl == "2025-12-20" or (hasattr(fl, "isoformat") and fl.isoformat()[:10] == "2025-12-20")

    asyncio.run(run_test())


def test_phase12_auxiliary_compare_extract_fecha_before_last_pay_warns():
    """Fecha límite extracto anterior a última Fecha pago → advertencia sin bloquear."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        base = "clientes/CMPBF/503"
        client.folder_children["clientes"] = [make_item("CMPBF", is_folder=True)]
        client.folder_children["clientes/CMPBF"] = [make_item("503", is_folder=True)]
        client.folder_children[base] = [
            make_item("Extracto cmpbf.pdf"),
            make_item("Tabla amortizacion CMPBF 503.xlsx"),
        ]
        client.downloaded_files[f"{base}/Tabla amortizacion CMPBF 503.xlsx"] = create_amortization_excel(
            [
                [date(2025, 10, 1), date(2025, 12, 25), None, 1000, None],
            ],
        )
        pdf_fecha = date(2025, 11, 1)
        client.downloaded_files[f"{base}/Extracto cmpbf.pdf"] = create_pdf_bytes("2.000", pdf_fecha)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[pdf_fecha, 50000, "CMPBF", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "anterior a la última fecha de pago" in obs.lower()
        fl = dist[0][DistribucionCols.FECHA_LIMITE]
        assert fl == "2025-11-01" or (hasattr(fl, "isoformat") and fl.isoformat()[:10] == "2025-11-01")

    asyncio.run(run_test())


def test_phase12_auxiliary_compare_when_extract_after_last_pay_no_compare_warning():
    """Si fecha límite extracto > última fecha pago en tabla, no se añade advertencia de comparación."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        base = "clientes/CMPAF/502"
        client.folder_children["clientes"] = [make_item("CMPAF", is_folder=True)]
        client.folder_children["clientes/CMPAF"] = [make_item("502", is_folder=True)]
        client.folder_children[base] = [
            make_item("Extracto cmp2.pdf"),
            make_item("Tabla amortizacion CMPAF 502.xlsx"),
        ]
        client.downloaded_files[f"{base}/Tabla amortizacion CMPAF 502.xlsx"] = create_amortization_excel(
            [
                [date(2025, 10, 1), date(2025, 9, 10), None, 1000, None],
            ],
        )
        pdf_fecha = date(2027, 1, 1)
        client.downloaded_files[f"{base}/Extracto cmp2.pdf"] = create_pdf_bytes("1.000", pdf_fecha)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[pdf_fecha, 50000, "CMPAF", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "coincide con la última fecha de pago" not in obs.lower()
        assert "anterior a la última fecha de pago" not in obs.lower()

    asyncio.run(run_test())


def test_phase1_feature_flag_v2_false_still_generates():
    """Flag GENERATE_EXTRACT_SELECTION_V2=false conserva ruta legacy sin fallar en setup mínimo."""

    async def run_test():
        set_env_vars()
        old = os.environ.get("GENERATE_EXTRACT_SELECTION_V2")
        os.environ["GENERATE_EXTRACT_SELECTION_V2"] = "false"
        try:
            client = MockGraphClient()
            client.children = []
            _make_single_client_setup(
                client,
                "800",
                tabla_cuota=10000.0,
                pdf_total_str="8.000",
                due_date=date(2025, 12, 23),
                banco_date=date(2025, 12, 23),
            )
            with _run_with_pdf_mock(client):
                await generate_payment_validation(client, date(2026, 1, 1))
            wb = load_generated_workbook(client)
            assert len(sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])) == 1
        finally:
            if old is None:
                os.environ.pop("GENERATE_EXTRACT_SELECTION_V2", None)
            else:
                os.environ["GENERATE_EXTRACT_SELECTION_V2"] = old

    asyncio.run(run_test())


def test_generate_does_not_use_intereses_mora_column():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "300", tabla_cuota=10000.0, pdf_total_str="15.000",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws_d = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws_d, DistribucionCols.ID_PAGO)
        rows = list(ws_d.iter_rows(min_row=dr, values_only=True))
        assert len(rows) == 1
        val_idx = _get_dist_col_idx(wb, DistribucionCols.VALOR_EXTRACTO)
        assert rows[0][val_idx] == 15000.0, "Debe usar PDF, no intereses de mora"
    asyncio.run(run_test())


def test_generate_casos_pago_total_uses_pdf_amounts_only():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", "trx-geo"]],
        date(2025, 12, 23),
    )
    row = sheet_to_dicts(workbook[ReviewSheets.CASOS_PAGO])[0]

    assert list(row.keys()) == CasosPagoCols.HEADERS
    assert row[CasosPagoCols.MONTO_BANCO] == 25443565
    assert "validar" in str(row.get(CasosPagoCols.OBSERVACION, "")).lower()


def test_generate_adds_visual_formulas_dropdowns_and_hidden_lists_sheet():
    _, _, workbook = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", "trx-geo"]],
        date(2025, 12, 23),
    )

    assert ReviewSheets.LISTAS in workbook.sheetnames
    assert workbook[ReviewSheets.LISTAS].sheet_state == "hidden"

    ws_control = workbook[ReviewSheets.CONTROL]
    control_values = {}
    for row in ws_control.iter_rows(max_col=2, values_only=True):
        if row and row[0] and row[1] is not None:
            control_values[str(row[0]).strip()] = row[1]
    assert control_values[ControlCols.ROW_ESTADO] == "EN_REVISION"
    assert control_values[ControlCols.ROW_PROCESAR] == "NO"

    ws_lists = workbook[ReviewSheets.LISTAS]
    control_statuses = [ws_lists[f"A{idx}"].value for idx in range(1, 6)]
    process_values = [ws_lists[f"B{idx}"].value for idx in range(1, 3)]
    dist_statuses = [ws_lists[f"C{idx}"].value for idx in range(1, 6)]
    validar_vals = [ws_lists[f"D{idx}"].value for idx in range(1, 3)]
    assert control_statuses == ["EN_REVISION", "PROCESANDO", "PROCESADO", "ERROR", "CANCELADO"]
    assert process_values == ["NO", "SI"]
    assert dist_statuses == list(EstadoPago.OPTIONS_ORDERED)
    assert validar_vals == [ValidarPago.SI, ValidarPago.NO]

    ws_dist = workbook[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    assert ws_dist[f"L{dr}"].value.startswith("=")
    assert "SUMIF" in ws_dist[f"M{dr}"].value.upper()

    validation_formulas = [dv.formula1 for dv in ws_control.data_validations.dataValidation]
    assert "=_Listas!$A$1:$A$5" in validation_formulas
    assert "=_Listas!$B$1:$B$2" in validation_formulas

    dist_validation_formulas = [dv.formula1 for dv in ws_dist.data_validations.dataValidation]
    assert "=_Listas!$C$1:$C$5" in dist_validation_formulas
    assert "=_Listas!$D$1:$D$2" in dist_validation_formulas
    assert len(ws_dist.conditional_formatting) > 0


def test_generate_pendiente_mora_keeps_total_blank_and_saldo_formula():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(
            client,
            "500",
            tabla_cuota=50000.0,
            pdf_total_str="45.000",
            due_date=date(2025, 12, 23),
            banco_date=date(2025, 12, 30),
        )

        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))

        workbook = load_generated_workbook(client)
        ws = workbook[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws, DistribucionCols.ID_PAGO)

        assert ws[f"N{dr}"].value == EstadoPago.ATRASADO
        assert isinstance(ws[f"L{dr}"].value, str) and ws[f"L{dr}"].value.startswith("=")
        m_formula = ws[f"M{dr}"].value
        assert isinstance(m_formula, str) and m_formula.startswith("=")
        assert m_formula.upper().count("SUMIF") == 4
        assert "$L$" not in m_formula
        assert "$I$" in m_formula and "$J$" in m_formula and "$K$" in m_formula and "$D$" in m_formula

    asyncio.run(run_test())


def test_generate_pendiente_mora_still_blank_mora_total():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        _make_single_client_setup(client, "500", tabla_cuota=50000.0, pdf_total_str="45.000",
                                   due_date=date(2025, 12, 23), banco_date=date(2025, 12, 30))
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
        rows = list(ws.iter_rows(min_row=dr, values_only=True))
        assert len(rows) == 1
        hr = _header_row_index(ws, DistribucionCols.ID_PAGO)
        headers = [c.value for c in ws[hr]]
        estado_idx = headers.index(DistribucionCols.ESTADO_LINEA)
        mora_apl_idx = headers.index(DistribucionCols.INTERESES_MORA)
        total_apl_idx = headers.index(DistribucionCols.TOTAL_APLICADO)
        val_idx = headers.index(DistribucionCols.VALOR_EXTRACTO)
        assert rows[0][estado_idx] == EstadoPago.ATRASADO
        assert rows[0][val_idx] == 45000.0, "Valor extracto debe venir del PDF"
        assert rows[0][mora_apl_idx] in ["", None], "Intereses de mora debe quedar vacía"
        assert isinstance(rows[0][total_apl_idx], str) and rows[0][total_apl_idx].startswith("=")
    asyncio.run(run_test())


def test_generate_visual_distribucion_freeze_autofilter_headers_widths():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    hr = _header_row_index(ws, DistribucionCols.ID_PAGO)
    assert ws.freeze_panes == _distrib_freeze_panes_cell(dr)
    col_credito = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    assert ws.freeze_panes == f"{get_column_letter(col_credito + 1)}{dr}"
    assert ws.auto_filter.ref is None
    ws_casos = wb[ReviewSheets.CASOS_PAGO]
    assert ws_casos.auto_filter.ref is None
    ws_err = wb[ReviewSheets.ERRORES]
    assert ws_err.auto_filter.ref is None
    top_left = ws.cell(row=hr, column=1)
    assert top_left.font.bold
    rgb = getattr(top_left.fill.fgColor, "rgb", None) or getattr(top_left.fill.start_color, "rgb", None)
    assert rgb and str(rgb).upper() not in ("00000000", "FFFFFFFF", "00FFFFFF", "NONE")
    money_names = (
        DistribucionCols.MONTO_BANCO,
        DistribucionCols.VALOR_EXTRACTO,
        DistribucionCols.APLICAR_A_EXTRACTO,
        DistribucionCols.MORA_A_APLICAR,
        DistribucionCols.OTROS_VALORES,
        DistribucionCols.TOTAL_APLICADO,
        DistribucionCols.SALDO_POR_ASIGNAR,
    )
    for name in money_names:
        letter = get_column_letter(DistribucionCols.HEADERS.index(name) + 1)
        assert ws.column_dimensions[letter].width == DIST_MONEY_COL_WIDTH


def test_generate_visual_distribucion_hyperlinks_currency_date_formats():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=True,
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)

    def ci(name: str) -> int:
        return DistribucionCols.HEADERS.index(name) + 1

    assert ws.cell(dr, ci(DistribucionCols.MONTO_BANCO)).number_format == "#,##0.00"
    assert ws.cell(dr, ci(DistribucionCols.FECHA_BANCO)).number_format == "yyyy-mm-dd"
    credito = ws.cell(dr, ci(DistribucionCols.CREDITO)).value
    assert ws.cell(dr, ci(DistribucionCols.LINK_EXTRACTO)).value == f"📄 Ver extracto crédito {credito}"
    assert ws.cell(dr, ci(DistribucionCols.LINK_TABLA)).value == f"📄 Ver tabla crédito {credito}"
    assert ws.cell(dr, ci(DistribucionCols.LINK_EXTRACTO)).hyperlink is not None
    assert ws.cell(dr, ci(DistribucionCols.LINK_TABLA)).hyperlink is not None


def test_generate_visual_control_resumen_listas_adelantados():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_c = wb[ReviewSheets.CONTROL]
    assert ws_c["A1"].value == CONTROL_TITLE
    assert CONTROL_HELP in str(ws_c["A2"].value or "")

    ws_r = wb[ReviewSheets.RESUMEN]
    labels = [ws_r.cell(i, 1).value for i in range(1, 25) if ws_r.cell(i, 1).value]
    for name in (
        "Pagos banco",
        "Casos pago",
        "Líneas distribución",
        "Errores",
        "Atrasado (mora)",
        "Adelantado",
        "Normal",
        "Incompleto",
        "Revisión manual",
    ):
        assert name in labels, f"Falta métrica en Resumen: {name}"

    assert wb[ReviewSheets.LISTAS].sheet_state == "hidden"
    assert "Adelantados" not in wb.sheetnames

    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    dist_dvs = [dv.formula1 for dv in ws_dist.data_validations.dataValidation]
    assert any("=_Listas!$C$1:$C$5" in f for f in dist_dvs)
    assert any("=_Listas!$D$1:$D$2" in f for f in dist_dvs)
    ctrl = wb[ReviewSheets.CONTROL]
    ctrl_dvs = [dv.formula1 for dv in ctrl.data_validations.dataValidation]
    assert "=_Listas!$A$1:$A$5" in ctrl_dvs
    assert "=_Listas!$B$1:$B$2" in ctrl_dvs


def _find_control_value_cell(ws_ctrl, label: str):
    for r_idx, row in enumerate(ws_ctrl.iter_rows(max_col=2, values_only=False), start=1):
        if row and row[0].value == label:
            return row[1]
    raise AssertionError(f"No se encontró label en Control: {label}")


def test_generate_sheet_protection_and_editable_cells():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_ctrl = wb[ReviewSheets.CONTROL]
    assert ws_ctrl.protection.sheet is True
    c_proc = _find_control_value_cell(ws_ctrl, ControlCols.ROW_PROCESAR)
    c_est = _find_control_value_cell(ws_ctrl, ControlCols.ROW_ESTADO)
    assert c_proc.protection.locked is False
    assert c_est.protection.locked is True

    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    assert ws_dist.protection.sheet is True
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    editable = {
        DistribucionCols.VALOR_INTERESES,
        DistribucionCols.ABONO_K,
        DistribucionCols.INTERESES_MORA,
        DistribucionCols.ESTADO_PAGO,
        DistribucionCols.VALIDAR_PAGO,
        DistribucionCols.OBSERVACION,
    }
    locked = {
        DistribucionCols.ID_PAGO,
        DistribucionCols.CLIENTE,
        DistribucionCols.MONTO_BANCO,
        DistribucionCols.FECHA_BANCO,
        DistribucionCols.CREDITO,
        DistribucionCols.FECHA_LIMITE,
        DistribucionCols.DIAS_MORA,
        DistribucionCols.VALOR_EXTRACTO,
        DistribucionCols.TOTAL_APLICADO,
        DistribucionCols.SALDO_POR_ASIGNAR,
        DistribucionCols.LINK_EXTRACTO,
        DistribucionCols.LINK_TABLA,
    }
    for name in editable:
        c = ws_dist.cell(row=dr, column=_dist_col(name))
        assert c.protection.locked is False, f"{name} debe estar desbloqueada"
    for name in locked:
        c = ws_dist.cell(row=dr, column=_dist_col(name))
        assert c.protection.locked is not False, f"{name} debe estar bloqueada"

    for sheet_name in (ReviewSheets.RESUMEN, ReviewSheets.CASOS_PAGO, ReviewSheets.ERRORES, ReviewSheets.LISTAS):
        ws = wb[sheet_name]
        assert ws.protection.sheet is True
    assert wb[ReviewSheets.LISTAS].sheet_state == "hidden"


def test_protection_uses_web_safe_minimal_profile():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    p = ws.protection
    assert p.sheet is True
    # Perfil web-safe mínimo: sin password ni flags extras forzados.
    assert not p.password
    assert p.selectUnlockedCells in (None, False)
    assert p.selectLockedCells in (None, False)


def test_distribution_editable_columns_unlocked_all_data_rows():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    editable = (
        DistribucionCols.VALOR_INTERESES,
        DistribucionCols.ABONO_K,
        DistribucionCols.INTERESES_MORA,
        DistribucionCols.ESTADO_LINEA,
        DistribucionCols.OBSERVACION,
    )
    for row_idx in range(dr, ws_dist.max_row + 1):
        if not ws_dist.cell(row=row_idx, column=_dist_col(DistribucionCols.ID_PAGO)).value:
            continue
        for name in editable:
            c = ws_dist.cell(row=row_idx, column=_dist_col(name))
            assert c.protection.locked is False, f"{name} debe estar desbloqueada en fila {row_idx}"


def test_distribution_locked_columns_locked_all_data_rows():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    locked = (
        DistribucionCols.ID_PAGO,
        DistribucionCols.CLIENTE,
        DistribucionCols.MONTO_BANCO,
        DistribucionCols.FECHA_BANCO,
        DistribucionCols.CREDITO,
        DistribucionCols.FECHA_LIMITE,
        DistribucionCols.DIAS_MORA,
        DistribucionCols.VALOR_EXTRACTO,
        DistribucionCols.TOTAL_APLICADO,
        DistribucionCols.SALDO_POR_ASIGNAR,
        DistribucionCols.LINK_EXTRACTO,
        DistribucionCols.LINK_TABLA,
    )
    for row_idx in range(dr, ws_dist.max_row + 1):
        if not ws_dist.cell(row=row_idx, column=_dist_col(DistribucionCols.ID_PAGO)).value:
            continue
        for name in locked:
            c = ws_dist.cell(row=row_idx, column=_dist_col(name))
            assert c.protection.locked is not False, f"{name} debe estar bloqueada en fila {row_idx}"


def _dist_col(name: str) -> int:
    return DistribucionCols.HEADERS.index(name) + 1


def _fill_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color is None:
        return None
    return getattr(color, "rgb", None)


def test_distribucion_monto_banco_only_first_row_per_payment():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    c_monto = _dist_col(DistribucionCols.MONTO_BANCO)
    pid0 = ws.cell(dr, _dist_col(DistribucionCols.ID_PAGO)).value
    assert pid0
    r2 = dr + 1
    assert ws.cell(dr, _dist_col(DistribucionCols.ID_PAGO)).value == ws.cell(r2, _dist_col(DistribucionCols.ID_PAGO)).value
    assert ws.cell(dr, c_monto).value == 25443565
    assert ws.cell(r2, c_monto).value in (None, "")


def test_distribucion_saldo_por_asignar_only_first_row_per_payment():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    c_saldo = _dist_col(DistribucionCols.SALDO_POR_ASIGNAR)
    r2 = dr + 1
    v1 = ws.cell(dr, c_saldo).value
    v2 = ws.cell(r2, c_saldo).value
    assert isinstance(v1, str) and v1.startswith("=")
    assert v2 in (None, "")


def test_saldo_por_asignar_formula_uses_group_assignments():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    f = ws.cell(dr, _dist_col(DistribucionCols.SALDO_POR_ASIGNAR)).value
    assert isinstance(f, str) and f.startswith("=")
    uf = f.upper()
    assert uf.count("SUMIF") == 4
    col_a = get_column_letter(_dist_col(DistribucionCols.ID_PAGO))
    col_c = get_column_letter(_dist_col(DistribucionCols.MONTO_BANCO))
    col_i = get_column_letter(_dist_col(DistribucionCols.VALOR_INTERESES))
    col_j = get_column_letter(_dist_col(DistribucionCols.ABONO_K))
    col_k = get_column_letter(_dist_col(DistribucionCols.INTERESES_MORA))
    assert f"${col_a}$" in f
    assert f"${col_c}$" in f
    assert f"${col_i}$" in f
    assert f"${col_j}$" in f
    assert f"${col_k}$" in f
    assert "$L$" not in f
    col_l = get_column_letter(_dist_col(DistribucionCols.TOTAL_APLICADO))
    assert f"${col_l}$" not in f


def test_saldo_por_asignar_updates_for_multi_credit_payment():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    last_row = ws.max_row
    f = ws.cell(dr, _dist_col(DistribucionCols.SALDO_POR_ASIGNAR)).value
    assert isinstance(f, str)
    col_a = get_column_letter(_dist_col(DistribucionCols.ID_PAGO))
    assert f"${col_a}${dr}:${col_a}${last_row}" in f
    crit_cell = f"{col_a}{dr}"
    assert crit_cell in f
    r2 = dr + 1
    assert ws.cell(r2, _dist_col(DistribucionCols.SALDO_POR_ASIGNAR)).value in (None, "")


def test_total_aplicado_only_first_row_per_payment():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    c_total = _dist_col(DistribucionCols.TOTAL_APLICADO)
    assert isinstance(ws.cell(dr, c_total).value, str)
    assert ws.cell(dr, c_total).value.startswith("=")
    assert ws.cell(dr + 1, c_total).value in (None, "")


def test_total_aplicado_formula_uses_group_assignments():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    f = ws.cell(dr, _dist_col(DistribucionCols.TOTAL_APLICADO)).value
    assert isinstance(f, str) and f.startswith("=")
    uf = f.upper()
    assert uf.count("SUMIF") == 3
    col_a = get_column_letter(_dist_col(DistribucionCols.ID_PAGO))
    col_i = get_column_letter(_dist_col(DistribucionCols.VALOR_INTERESES))
    col_j = get_column_letter(_dist_col(DistribucionCols.ABONO_K))
    col_k = get_column_letter(_dist_col(DistribucionCols.INTERESES_MORA))
    assert f"${col_a}$" in f
    assert f"${col_i}$" in f
    assert f"${col_j}$" in f
    assert f"${col_k}$" in f
    col_m = get_column_letter(_dist_col(DistribucionCols.SALDO_POR_ASIGNAR))
    assert f"${col_m}$" not in f


def test_saldo_por_asignar_still_uses_group_assignments():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    f = ws.cell(dr, _dist_col(DistribucionCols.SALDO_POR_ASIGNAR)).value
    assert isinstance(f, str) and f.startswith("=")
    uf = f.upper()
    assert uf.count("SUMIF") == 4
    col_a = get_column_letter(_dist_col(DistribucionCols.ID_PAGO))
    col_c = get_column_letter(_dist_col(DistribucionCols.MONTO_BANCO))
    col_i = get_column_letter(_dist_col(DistribucionCols.VALOR_INTERESES))
    col_j = get_column_letter(_dist_col(DistribucionCols.ABONO_K))
    col_k = get_column_letter(_dist_col(DistribucionCols.INTERESES_MORA))
    assert f"${col_a}$" in f
    assert f"${col_c}$" in f
    assert f"${col_i}$" in f
    assert f"${col_j}$" in f
    assert f"${col_k}$" in f
    col_l = get_column_letter(_dist_col(DistribucionCols.TOTAL_APLICADO))
    assert f"${col_l}$" not in f


def _top_border_style(cell) -> str | None:
    top = cell.border.top
    if top is None:
        return None
    return top.style


def _bottom_border_style(cell) -> str | None:
    bottom = cell.border.bottom
    if bottom is None:
        return None
    return bottom.style


def _distrib_rows_by_cliente(ws, dr: int) -> list[tuple[int, str]]:
    cliente_col = _dist_col(DistribucionCols.CLIENTE)
    rows: list[tuple[int, str]] = []
    for row in range(dr, ws.max_row + 1):
        cliente = str(ws.cell(row=row, column=cliente_col).value or "").strip()
        if not cliente:
            continue
        rows.append((row, cliente))
    return rows


def test_distribucion_client_separator_border_on_client_change():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    cliente_col = _dist_col(DistribucionCols.CLIENTE)
    ncols = len(DistribucionCols.HEADERS)
    prev_cliente = None
    first_row = dr
    separator_row = None
    for row in range(dr, ws.max_row + 1):
        cliente = ws.cell(row=row, column=cliente_col).value
        if not cliente:
            continue
        if prev_cliente and cliente != prev_cliente:
            separator_row = row
            break
        prev_cliente = cliente
    assert separator_row is not None
    for c in range(1, ncols + 1):
        assert _top_border_style(ws.cell(row=separator_row, column=c)) == "medium"
    assert _top_border_style(ws.cell(row=first_row, column=1)) == "thin"


def test_distribucion_client_bottom_border_on_each_block_end():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    ncols = len(DistribucionCols.HEADERS)
    rows = _distrib_rows_by_cliente(ws, dr)
    assert len(rows) >= 2
    block_ends: list[int] = []
    for i, (row, cliente) in enumerate(rows):
        next_cliente = rows[i + 1][1] if i + 1 < len(rows) else None
        if next_cliente is None or next_cliente != cliente:
            block_ends.append(row)
    assert block_ends
    for end_row in block_ends:
        for c in range(1, ncols + 1):
            assert _bottom_border_style(ws.cell(row=end_row, column=c)) == "medium"


def test_distribucion_last_client_has_bottom_border():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    rows = _distrib_rows_by_cliente(ws, dr)
    last_row, _ = rows[-1]
    assert _bottom_border_style(ws.cell(row=last_row, column=1)) == "medium"


def test_distribucion_no_client_separator_on_first_data_row():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    for c in range(1, len(DistribucionCols.HEADERS) + 1):
        assert _top_border_style(ws.cell(row=dr, column=c)) == "thin"


def test_estado_linea_fill_colors_are_distinct():
    fills = list(_ESTADO_PAGO_FILLS.values())
    rgbs = [f.fgColor.rgb for f in fills]
    assert len(rgbs) == len(set(rgbs))


def test_estado_linea_colored_only_in_estado_column():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    estado_col = _dist_col(DistribucionCols.ESTADO_LINEA)
    ref_col = _dist_col(DistribucionCols.MONTO_BANCO)
    estado_rgb = _fill_rgb(ws.cell(row=dr, column=estado_col))
    ref_rgb = _fill_rgb(ws.cell(row=dr, column=ref_col))
    editable_rgb = _FILL_EDITABLE_COL.fgColor.rgb
    assert estado_rgb == editable_rgb
    assert estado_rgb != ref_rgb


def test_estado_linea_has_state_colors():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    rules = []
    for cf in ws.conditional_formatting._cf_rules.values():
        rules.extend(cf)
    assert rules, "Se esperaban reglas de formato condicional para Estado línea"
    col_estado = get_column_letter(
        DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_LINEA) + 1
    )
    for state, fill in _ESTADO_PAGO_FILLS.items():
        expected_formula = f'INDIRECT("{col_estado}"&ROW())="{state}"'
        matching = [
            r for r in rules
            if expected_formula in (r.formula or []) and getattr(getattr(r, "dxf", None), "fill", None) is not None
        ]
        assert matching, f"No se encontró regla de color para estado {state}"
        assert matching[0].dxf.fill.fgColor.rgb == fill.fgColor.rgb


def test_distribution_rows_grouped_by_cliente_have_alternating_fill():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    cliente_col = _dist_col(DistribucionCols.CLIENTE)
    band_col = _dist_col(DistribucionCols.MONTO_BANCO)
    first_seen_fill_by_cliente = {}
    ordered_group_fills = []
    for row in range(dr, ws.max_row + 1):
        cliente = ws.cell(row=row, column=cliente_col).value
        if not cliente:
            continue
        fill = _fill_rgb(ws.cell(row=row, column=band_col))
        if cliente not in first_seen_fill_by_cliente:
            first_seen_fill_by_cliente[cliente] = fill
            ordered_group_fills.append(fill)
        else:
            assert first_seen_fill_by_cliente[cliente] == fill
    assert len(first_seen_fill_by_cliente) >= 2
    assert ordered_group_fills[0] != ordered_group_fills[1]


def test_editable_columns_have_distinct_fill():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    editable_names = (
        DistribucionCols.APLICAR_A_EXTRACTO,
        DistribucionCols.MORA_A_APLICAR,
        DistribucionCols.OTROS_VALORES,
        DistribucionCols.ESTADO_LINEA,
    )
    band_ref = _fill_rgb(ws.cell(row=dr, column=_dist_col(DistribucionCols.MONTO_BANCO)))
    editable_rgb = _FILL_EDITABLE_COL.fgColor.rgb
    for name in editable_names:
        rgb = _fill_rgb(ws.cell(row=dr, column=_dist_col(name)))
        assert rgb == editable_rgb
        assert rgb != band_ref
    saldo_rgb = _fill_rgb(ws.cell(row=dr, column=_dist_col(DistribucionCols.SALDO_POR_ASIGNAR)))
    assert saldo_rgb == _FILL_SALDO_COL.fgColor.rgb
    assert saldo_rgb != editable_rgb
    obs_rgb = _fill_rgb(ws.cell(row=dr, column=_dist_col(DistribucionCols.OBSERVACION)))
    assert obs_rgb == editable_rgb


def test_validar_pago_column_matches_control_procesar_fill():
    """Mismo verde brillante que la celda de valor de Procesar en Control (p. ej. B7)."""
    if generate_payment_validation is None or _FILL_PROCESAR_SI is None:
        pytest.skip("generate module not available")
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    vp_rgb = _fill_rgb(ws.cell(row=dr, column=_dist_col(DistribucionCols.VALIDAR_PAGO)))
    assert vp_rgb == _FILL_PROCESAR_SI.fgColor.rgb


def test_editable_columns_remain_unlocked():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    editable = (
        DistribucionCols.VALOR_INTERESES,
        DistribucionCols.ABONO_K,
        DistribucionCols.INTERESES_MORA,
        DistribucionCols.ESTADO_LINEA,
        DistribucionCols.OBSERVACION,
    )
    for row_idx in range(dr, ws_dist.max_row + 1):
        if not ws_dist.cell(row=row_idx, column=_dist_col(DistribucionCols.ID_PAGO)).value:
            continue
        for name in editable:
            c = ws_dist.cell(row=row_idx, column=_dist_col(name))
            assert c.protection.locked is False


def test_non_editable_columns_remain_locked():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_dist = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws_dist, DistribucionCols.ID_PAGO)
    locked = (
        DistribucionCols.ID_PAGO,
        DistribucionCols.CLIENTE,
        DistribucionCols.MONTO_BANCO,
        DistribucionCols.FECHA_BANCO,
        DistribucionCols.CREDITO,
        DistribucionCols.FECHA_LIMITE,
        DistribucionCols.DIAS_MORA,
        DistribucionCols.VALOR_EXTRACTO,
        DistribucionCols.TOTAL_APLICADO,
        DistribucionCols.SALDO_POR_ASIGNAR,
        DistribucionCols.LINK_EXTRACTO,
        DistribucionCols.LINK_TABLA,
    )
    for row_idx in range(dr, ws_dist.max_row + 1):
        if not ws_dist.cell(row=row_idx, column=_dist_col(DistribucionCols.ID_PAGO)).value:
            continue
        for name in locked:
            c = ws_dist.cell(row=row_idx, column=_dist_col(name))
            assert c.protection.locked is not False


def test_distribucion_headers_do_not_include_legacy_names():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    hr = _header_row_index(ws, DistribucionCols.ID_PAGO)
    headers = [c.value for c in ws[hr]]

    assert headers == DistribucionCols.HEADERS
    assert "Valor intereses" not in headers
    assert "Abono a K" not in headers
    assert "Intereses de mora" not in headers
    assert DistribucionCols.APLICAR_A_EXTRACTO in headers
    assert DistribucionCols.MORA_A_APLICAR in headers
    assert DistribucionCols.OTROS_VALORES in headers
    assert DistribucionCols.ESTADO_PAGO in headers
    assert DistribucionCols.VALIDAR_PAGO in headers
    assert DistribucionCols.ESTADO_PAGO == "Estado Pago"
    assert "Estado línea" not in headers


def test_generate_result_includes_validation_file_url_when_graph_returns_weburl():
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    expected_url = "https://comwareec.sharepoint.com/sites/demo/val_2026-05-10.xlsx"

    async def run_test():
        set_env_vars()
        client = MockGraphClientPutReturnsWebUrl(expected_url)
        client.children = []
        setup_client_structure(client)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [],
        )
        extractor = make_pdf_extractor_mock(client)
        fake_fecha = _fake_fecha_limite_from_marker_bytes()
        with mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_total_a_pagar_from_pdf",
            side_effect=extractor,
        ), mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_fecha_limite_pago_from_pdf",
            side_effect=fake_fecha,
        ):
            result = await generate_payment_validation(client, date(2026, 5, 10))
        assert result["validation_file_url"] == expected_url

    asyncio.run(run_test())


@pytest.mark.parametrize("put_return", [{}, None])
def test_generate_result_sets_validation_file_url_null_when_graph_does_not_return_weburl(put_return):
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    async def run_test():
        set_env_vars()
        client = MockGraphClientPutCustomReturn(put_return)
        client.children = []
        setup_client_structure(client)
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [],
        )
        extractor = make_pdf_extractor_mock(client)
        fake_fecha = _fake_fecha_limite_from_marker_bytes()
        with mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_total_a_pagar_from_pdf",
            side_effect=extractor,
        ), mock.patch(
            "app.application.use_cases.payment_validation_generate.extract_fecha_limite_pago_from_pdf",
            side_effect=fake_fecha,
        ):
            result = await generate_payment_validation(client, date(2026, 5, 10))
        assert result.get("validation_file_url") is None

    asyncio.run(run_test())


def test_credit_unit_inversiones_empty_credit_root_extract():
    """CREDITO # sin extractos → Errores; extracto en raíz → Distribución."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("INVPRO", is_folder=True)]
        client.folder_children["clientes/INVPRO"] = [
            make_item("CREDITO # 318", is_folder=True),
            make_item("Extracto 2026-04-01 CREDITO # 801.pdf"),
            make_item("Tabla amortizacion INVPRO INVPRO.xlsx"),
        ]
        client.folder_children["clientes/INVPRO/CREDITO # 318"] = [
            make_item("Tabla amortizacion INVPRO 318.xlsx"),
        ]
        client.downloaded_files["clientes/INVPRO/Tabla amortizacion INVPRO 318.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/INVPRO/Tabla amortizacion INVPRO INVPRO.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 1000, None]]
        )
        client.downloaded_files["clientes/INVPRO/Extracto 2026-04-01 CREDITO # 801.pdf"] = create_pdf_bytes(
            "4.000", date(2026, 4, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "INVPRO", "pago"]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        err = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.CREDITO] == "801"
        assert "raíz del cliente" in str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert any(
            str(r.get(ErroresCols.CODIGO_TECNICO)) == "extract_not_found"
            and "318" in str(r.get(ErroresCols.CREDITO, ""))
            for r in err
        )

    asyncio.run(run_test())


def test_credit_unit_agrecar_root_and_credit_folders():
    """Subcarpetas crédito + extracto en raíz → al menos dos filas Distribución."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("AGRECAR", is_folder=True)]
        client.folder_children["clientes/AGRECAR"] = [
            make_item("CREDITO # 100", is_folder=True),
            make_item("Extracto root AGRECAR.pdf"),
        ]
        client.folder_children["clientes/AGRECAR/CREDITO # 100"] = [
            make_item("Extracto 2026-03-01 CREDITO # 100.pdf"),
            make_item("Tabla amortizacion AGRECAR 100.xlsx"),
        ]
        client.downloaded_files["clientes/AGRECAR/CREDITO # 100/Tabla amortizacion AGRECAR 100.xlsx"] = (
            create_amortization_excel([[date(2026, 3, 1), None, None, 2000, None]])
        )
        client.downloaded_files["clientes/AGRECAR/CREDITO # 100/Extracto 2026-03-01 CREDITO # 100.pdf"] = (
            create_pdf_bytes("2.000", date(2026, 3, 1))
        )
        client.downloaded_files["clientes/AGRECAR/Extracto root AGRECAR.pdf"] = create_pdf_bytes(
            "2.000", date(2026, 3, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 3, 1), 500000, "AGRECAR", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 2
        assert any("raíz del cliente" in str(r.get(DistribucionCols.OBSERVACION, "")) for r in dist)

    asyncio.run(run_test())


def test_credit_unit_ericcol_non_standard_folder_infers_credit():
    """Subcarpeta sin CREDITO en nombre pero con contenido operativo."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        base = "clientes/AGRECAR/ERICCOL MINERIA"
        client.folder_children["clientes"] = [make_item("AGRECAR", is_folder=True)]
        client.folder_children["clientes/AGRECAR"] = [make_item("ERICCOL MINERIA", is_folder=True)]
        client.folder_children[base] = [
            make_item("Extracto 2026-04-01 CREDITO # 555.pdf"),
            make_item("Tabla amortizacion ERICCOL 555.xlsx"),
        ]
        client.downloaded_files[f"{base}/Tabla amortizacion ERICCOL 555.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 3000, None]]
        )
        client.downloaded_files[f"{base}/Extracto 2026-04-01 CREDITO # 555.pdf"] = create_pdf_bytes(
            "3.000", date(2026, 4, 1)
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "AGRECAR", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        assert len(dist) == 1
        assert dist[0][DistribucionCols.CREDITO] == "555"
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "no sigue el estándar CREDITO #" in obs

    asyncio.run(run_test())


def test_infra_extractos_folder_not_credit_unit():
    """EXTRACTOS en raíz del cliente no se procesa como unidad de crédito."""

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("ONLYX", is_folder=True)]
        client.folder_children["clientes/ONLYX"] = [
            make_item("EXTRACTOS", is_folder=True),
            make_item("CREDITO # 1", is_folder=True),
        ]
        client.folder_children["clientes/ONLYX/CREDITO # 1"] = [
            make_item("Extracto 2026-01-01 CREDITO # 1.pdf"),
            make_item("Tabla amortizacion ONLYX 1.xlsx"),
        ]
        client.folder_children["clientes/ONLYX/EXTRACTOS"] = []
        client.downloaded_files["clientes/ONLYX/CREDITO # 1/Tabla amortizacion ONLYX 1.xlsx"] = (
            create_amortization_excel([[date(2026, 1, 1), None, None, 1000, None]])
        )
        client.downloaded_files["clientes/ONLYX/CREDITO # 1/Extracto 2026-01-01 CREDITO # 1.pdf"] = (
            create_pdf_bytes("1.000", date(2026, 1, 1))
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 1, 1), 100000, "ONLYX", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        err = sheet_to_dicts(wb[ReviewSheets.ERRORES])
        assert not any(str(r.get(ErroresCols.CREDITO)) == "EXTRACTOS" for r in err)

    asyncio.run(run_test())


def test_credit_folder_vigente_no_terminal_observation():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("VIGCLI", is_folder=True)]
        client.folder_children["clientes/VIGCLI"] = [make_item("CREDITO # 318 VIGENTE", is_folder=True)]
        client.folder_children["clientes/VIGCLI/CREDITO # 318 VIGENTE"] = [
            make_item("Extracto 2026-04-01 CREDITO # 318.pdf"),
            make_item("Tabla amortizacion VIGCLI 318.xlsx"),
        ]
        client.downloaded_files["clientes/VIGCLI/CREDITO # 318 VIGENTE/Tabla amortizacion VIGCLI 318.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 1000, None]])
        )
        client.downloaded_files["clientes/VIGCLI/CREDITO # 318 VIGENTE/Extracto 2026-04-01 CREDITO # 318.pdf"] = (
            create_pdf_bytes("1.000", date(2026, 4, 1))
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "VIGCLI", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "PAGADO/LIQUIDADO" not in obs

    asyncio.run(run_test())


def test_credit_folder_pagado_terminal_observation():
    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("PAGCLI", is_folder=True)]
        client.folder_children["clientes/PAGCLI"] = [make_item("CREDITO # 318 PAGADO", is_folder=True)]
        client.folder_children["clientes/PAGCLI/CREDITO # 318 PAGADO"] = [
            make_item("Extracto 2026-04-01 CREDITO # 318.pdf"),
            make_item("Tabla amortizacion PAGCLI 318.xlsx"),
        ]
        client.downloaded_files["clientes/PAGCLI/CREDITO # 318 PAGADO/Tabla amortizacion PAGCLI 318.xlsx"] = (
            create_amortization_excel([[date(2026, 4, 1), None, None, 1000, None]])
        )
        client.downloaded_files["clientes/PAGCLI/CREDITO # 318 PAGADO/Extracto 2026-04-01 CREDITO # 318.pdf"] = (
            create_pdf_bytes("1.000", date(2026, 4, 1))
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "PAGCLI", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        dist = sheet_to_dicts(wb[ReviewSheets.DISTRIBUCION])
        obs = str(dist[0].get(DistribucionCols.OBSERVACION, ""))
        assert "PAGADO/LIQUIDADO" in obs

    asyncio.run(run_test())


def test_dedupe_credit_candidates_by_ruta():
    from app.application.use_cases.payment_validation_generate import _dedupe_credit_candidates

    cands = [
        {
            "credito": "100",
            "ruta_extracto_pdf": "clientes/X/a.pdf",
            "fecha_limite": date(2026, 3, 1),
            "valor_extracto": 2000.0,
            "link_extracto": "http://x/a.pdf",
        },
        {
            "credito": "ROOT",
            "ruta_extracto_pdf": "clientes\\X\\a.pdf",
            "fecha_limite": date(2026, 3, 1),
            "valor_extracto": 2000.0,
            "link_extracto": "http://x/a.pdf",
        },
    ]
    assert len(_dedupe_credit_candidates(cands)) == 1


def test_casos_pago_observacion_column_is_wider_with_wrap():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.CASOS_PAGO]
    dr = _first_data_row(ws, CasosPagoCols.ID_PAGO)
    obs_c = CasosPagoCols.HEADERS.index(CasosPagoCols.OBSERVACION) + 1
    width = ws.column_dimensions[get_column_letter(obs_c)].width or 0
    assert width >= CASOS_OBS_COL_WIDTH
    max_line = 0
    for row in range(dr, ws.max_row + 1):
        txt = str(ws.cell(row=row, column=obs_c).value or "")
        for line in txt.splitlines():
            max_line = max(max_line, len(line))
    if max_line:
        assert width >= min(100.0, max_line * 1.05 + 2)
    cell = ws.cell(row=dr, column=obs_c)
    assert cell.alignment.wrap_text is True


def test_distrib_observacion_cells_use_editable_green_pastel_fill():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    obs_c = DistribucionCols.HEADERS.index(DistribucionCols.OBSERVACION) + 1
    obs_rgb = _fill_rgb(ws.cell(row=dr, column=obs_c))
    assert obs_rgb == _FILL_EDITABLE_COL.fgColor.rgb
    assert ws.cell(row=dr, column=obs_c).alignment.wrap_text is True


def test_distrib_observation_warning_fill_only_on_observation_cell():
    from app.application.use_cases.payment_validation_generate import _FILL_OBS_WARNING

    async def run_test():
        set_env_vars()
        client = MockGraphClient()
        client.children = []
        client.folder_children["clientes"] = [make_item("WARNCLI", is_folder=True)]
        client.folder_children["clientes/WARNCLI"] = [
            make_item("501", is_folder=True),
        ]
        base = "clientes/WARNCLI/501"
        client.folder_children[base] = [
            make_item("Extracto warn.pdf"),
            make_item("Tabla A amortizacion WARNCLI 501.xlsx"),
            make_item("Tabla B amortizacion WARNCLI 501.xlsx"),
        ]
        client.downloaded_files[f"{base}/Extracto warn.pdf"] = create_pdf_bytes("1.000", date(2026, 4, 1))
        client.downloaded_files[f"{base}/Tabla A amortizacion WARNCLI 501.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 100, None]]
        )
        client.downloaded_files[f"{base}/Tabla B amortizacion WARNCLI 501.xlsx"] = create_amortization_excel(
            [[date(2026, 4, 1), None, None, 100, None]]
        )
        client.downloaded_files["banco.xlsx"] = create_excel(
            ["Fecha", "Crédito", "Concepto", "Transacción"],
            [[date(2026, 4, 1), 100000, "WARNCLI", ""]],
        )
        with _run_with_pdf_mock(client):
            await generate_payment_validation(client, date(2026, 1, 1))
        wb = load_generated_workbook(client)
        ws = wb[ReviewSheets.DISTRIBUCION]
        dr = _first_data_row(ws)
        obs_c = DistribucionCols.HEADERS.index(DistribucionCols.OBSERVACION) + 1
        estado_c = DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_LINEA) + 1
        obs_fill = ws.cell(dr, obs_c).fill
        assert obs_fill.fgColor.rgb == _FILL_EDITABLE_COL.fgColor.rgb
        band_rgb = _fill_rgb(ws.cell(dr, DistribucionCols.HEADERS.index(DistribucionCols.MONTO_BANCO) + 1))
        assert obs_fill.fgColor.rgb != band_rgb

    asyncio.run(run_test())


def test_generate_existing_result_fields_remain_backward_compatible():
    if generate_payment_validation is None:
        pytest.fail("Not implemented")

    client, result, _ = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    assert result["process_id"]
    assert result["validation_file"] == "val_2025-12-23.xlsx"
    assert result["validation_file_path"] == "revision/val_2025-12-23.xlsx"
    assert "pagos_banco" in result["summary"]
    assert "errores" in result["summary"]
    assert "conditional_formatting" in result["summary"]
    assert "validation_file_url" in result
    assert result["validation_file_url"] is None
    assert client.put_calls


def test_distribucion_freeze_panes_after_credito():
    """Congela ID Pago, Cliente y Crédito; Monto banco queda desplazable."""
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    col_credito = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    assert ws.freeze_panes == f"{get_column_letter(col_credito + 1)}{dr}"


def test_errores_freeze_panes_unchanged_by_distrib_navigation_task():
    """Errores mantiene freeze A{primera_fila}; no hereda C4 de Distribución."""
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws_e = wb[ReviewSheets.ERRORES]
    dr = _first_data_row(ws_e, ErroresCols.ID_PAGO)
    assert ws_e.freeze_panes == f"A{dr}"


def test_distribucion_link_visible_text_includes_credito_preserves_hyperlink():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=True,
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)

    def ci(name: str) -> int:
        return DistribucionCols.HEADERS.index(name) + 1

    credito = ws.cell(dr, ci(DistribucionCols.CREDITO)).value
    c_ext = ws.cell(dr, ci(DistribucionCols.LINK_EXTRACTO))
    c_tab = ws.cell(dr, ci(DistribucionCols.LINK_TABLA))
    c_fold = ws.cell(dr, ci(DistribucionCols.LINK_CARPETA_CREDITO))
    assert c_ext.value == f"📄 Ver extracto crédito {credito}"
    assert c_tab.value == f"📄 Ver tabla crédito {credito}"
    assert c_fold.value == f"Ver carpeta crédito {credito}"
    assert str(c_ext.hyperlink.target).startswith("https://contoso/")
    assert str(c_tab.hyperlink.target).startswith("https://contoso/")
    assert str(c_fold.hyperlink.target).startswith("https://contoso/folder/")


def test_distribucion_sin_url_no_texto_link_falso():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=False,
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)

    def ci(name: str) -> int:
        return DistribucionCols.HEADERS.index(name) + 1

    for r in range(dr, ws.max_row + 1):
        c_ext = ws.cell(r, ci(DistribucionCols.LINK_EXTRACTO))
        c_tab = ws.cell(r, ci(DistribucionCols.LINK_TABLA))
        c_fold = ws.cell(r, ci(DistribucionCols.LINK_CARPETA_CREDITO))
        assert c_ext.value in ("", None)
        assert c_tab.value in ("", None)
        assert c_fold.value in ("", None)
        assert c_ext.hyperlink is None
        assert c_tab.hyperlink is None
        assert c_fold.hyperlink is None


def test_distribucion_link_carpeta_credito_column_position_and_ruta_hidden():
    ix_tab = DistribucionCols.HEADERS.index(DistribucionCols.LINK_TABLA)
    ix_fold = DistribucionCols.HEADERS.index(DistribucionCols.LINK_CARPETA_CREDITO)
    ix_ruta = DistribucionCols.HEADERS.index(DistribucionCols.RUTA)
    ix_ruta_uc = DistribucionCols.HEADERS.index(DistribucionCols.RUTA_UNIDAD_CREDITO)
    assert DistribucionCols.HEADERS[ix_fold] == "Link carpeta crédito"
    assert ix_fold == ix_tab + 1
    assert ix_ruta == ix_fold + 1
    assert ix_ruta_uc == ix_ruta + 1
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    for col_name in (
        DistribucionCols.RUTA_TABLA_AMORTIZACION,
        DistribucionCols.CREDITO_NORMALIZADO,
    ):
        letter = get_column_letter(DistribucionCols.HEADERS.index(col_name) + 1)
        assert ws.column_dimensions[letter].hidden is True
    col_ruta = ix_ruta + 1
    col_ruta_uc = ix_ruta_uc + 1
    ruta_pdf = str(ws.cell(dr, col_ruta).value or "")
    ruta_uc = str(ws.cell(dr, col_ruta_uc).value or "")
    assert ruta_pdf.lower().endswith(".pdf")
    assert ruta_uc and not ruta_uc.lower().endswith(".pdf")
    assert ruta_pdf.startswith(ruta_uc) or ruta_uc in ruta_pdf


def test_distribucion_link_carpeta_credito_with_url_descriptive_text_and_hyperlink():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=True,
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    col_fold = DistribucionCols.HEADERS.index(DistribucionCols.LINK_CARPETA_CREDITO) + 1
    col_cred = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    credito = ws.cell(dr, col_cred).value
    cell = ws.cell(dr, col_fold)
    assert cell.value == f"Ver carpeta crédito {credito}"
    assert cell.hyperlink is not None
    assert str(cell.hyperlink.target).startswith("https://contoso/folder/")


def test_errores_link_columns_wider_than_before():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
        include_web_urls=True,
    )
    ws = wb[ReviewSheets.ERRORES]
    col_ext = ErroresCols.HEADERS.index(ErroresCols.LINK_EXTRACTO) + 1
    col_fold = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
    w_ext = ws.column_dimensions[get_column_letter(col_ext)].width or 0
    w_fold = ws.column_dimensions[get_column_letter(col_fold)].width or 0
    assert w_ext >= 28
    assert w_fold >= 30
    assert w_ext <= 34
    assert w_fold <= 38


def test_errores_headers_unchanged_no_new_columns():
    assert ErroresCols.HEADERS == [
        "ID Pago",
        "Cliente",
        "Crédito",
        "Tipo de caso",
        "Descripción para revisión",
        "Qué debe hacer",
        "Requiere soporte",
        "Link extracto",
        "Link carpeta crédito",
        "Código técnico",
    ]


def test_distribucion_headers_order_and_link_carpeta_before_ruta():
    assert DistribucionCols.HEADERS[0] == DistribucionCols.ID_PAGO
    assert DistribucionCols.HEADERS[1] == DistribucionCols.CLIENTE
    assert DistribucionCols.HEADERS[2] == DistribucionCols.CREDITO
    assert DistribucionCols.HEADERS[3] == DistribucionCols.MONTO_BANCO
    assert DistribucionCols.RUTA in DistribucionCols.HEADERS
    assert DistribucionCols.LINK_CARPETA_CREDITO in DistribucionCols.HEADERS
    assert DistribucionCols.HEADERS[-1] == DistribucionCols.CREDITO_NORMALIZADO
    assert DistribucionCols.HEADERS[-2] == DistribucionCols.RUTA_TABLA_AMORTIZACION
    assert DistribucionCols.HEADERS[-3] == DistribucionCols.RUTA_UNIDAD_CREDITO
    assert DistribucionCols.HEADERS[-4] == DistribucionCols.RUTA
    assert DistribucionCols.HEADERS[-5] == DistribucionCols.LINK_CARPETA_CREDITO


def _medium_border_color(cell, side: str = "top") -> str | None:
    edge = getattr(cell.border, side, None)
    if edge is None or edge.style != "medium":
        return None
    rgb = edge.color.rgb if edge.color else None
    if rgb is None:
        return None
    s = str(rgb).upper()
    return s[-6:] if len(s) > 6 else s


def test_distribucion_client_separator_borders_are_navy_blue():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.DISTRIBUCION]
    dr = _first_data_row(ws, DistribucionCols.ID_PAGO)
    rows = _distrib_rows_by_cliente(ws, dr)
    sep_row = next(
        row for i, (row, cli) in enumerate(rows) if i > 0 and cli != rows[i - 1][1]
    )
    cell = ws.cell(row=sep_row, column=1)
    assert _medium_border_color(cell, "top") == "002060"


def test_errores_client_separator_borders_when_multiple_clients():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws = wb[ReviewSheets.ERRORES]
    dr = _first_data_row(ws, ErroresCols.ID_PAGO)
    if ws.max_row < dr:
        pytest.skip("Sin filas de error en este escenario")
    clientes = []
    for row in range(dr, ws.max_row + 1):
        cli = str(ws.cell(row, ErroresCols.HEADERS.index(ErroresCols.CLIENTE) + 1).value or "").strip()
        if cli:
            clientes.append((row, cli))
    if len({c for _, c in clientes}) < 2:
        pytest.skip("Se requieren errores de al menos dos clientes")
    sep_row = next(row for i, (row, cli) in enumerate(clientes) if i > 0 and cli != clientes[i - 1][1])
    assert _medium_border_color(ws.cell(row=sep_row, column=1), "top") == "002060"
    last_row, _ = clientes[-1]
    assert _medium_border_color(ws.cell(row=last_row, column=1), "bottom") == "002060"


def test_sheet_row2_instructions_visible_and_updated():
    _, _, wb = run_generate(
        [
            [datetime(2025, 12, 23), 25443565, "GEOEXCON", ""],
            [datetime(2025, 12, 24), 1234567, "EQUINORTE", ""],
        ],
        date(2025, 12, 24),
    )
    ws_c = wb[ReviewSheets.CASOS_PAGO]
    ws_d = wb[ReviewSheets.DISTRIBUCION]
    ws_e = wb[ReviewSheets.ERRORES]
    c2 = ws_c.cell(2, 1)
    d2 = ws_d.cell(2, 1)
    e2 = ws_e.cell(2, 1)
    assert CASOS_SUBTITLE.split(".")[0] in str(c2.value)
    assert "Aplicar a extracto" in str(d2.value)
    assert "Procesar a SI" in str(d2.value)
    if ws_e.max_row >= _first_data_row(ws_e, ErroresCols.ID_PAGO):
        assert ERRORES_HELP.split(".")[0] in str(e2.value)
    for cell in (c2, d2, e2):
        assert (cell.font.size or 0) >= 12
        assert (ws_c.row_dimensions[2].height or 0) >= 36


def test_control_procesar_b7_always_green_even_when_value_is_no():
    _, _, wb = run_generate(
        [[datetime(2025, 12, 23), 25443565, "GEOEXCON", ""]],
        date(2025, 12, 23),
    )
    ws = wb[ReviewSheets.CONTROL]
    proc_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == ControlCols.ROW_PROCESAR:
            proc_row = r
            break
    assert proc_row is not None
    assert (ws.row_dimensions[proc_row].height or 0) >= 34
    label = ws.cell(proc_row, 1)
    val_cell = ws.cell(proc_row, 2)
    assert label.font.bold is True
    assert (label.font.size or 0) >= 14
    assert label.alignment.vertical == "center"
    assert val_cell.font.bold is True
    assert (val_cell.font.size or 0) >= 14
    assert val_cell.alignment.horizontal == "center"
    assert val_cell.alignment.vertical == "center"
    val = str(val_cell.value or "").strip().upper()
    assert val == ControlCols.VAL_PROCESAR_NO
    fill_rgb = _fill_rgb(val_cell)
    font_rgb = str(val_cell.font.color.rgb or "").upper()
    assert fill_rgb in ("00B050", "B050", "00FF00B050", "FF00B050", "0032CD32", "32CD32", "00FF32CD32", "FF32CD32")
    assert font_rgb.endswith("FFFFFF")
    assert fill_rgb not in ("00FF6969", "FF6969", "FFFF6969")
    cf_rules = []
    for rules in ws.conditional_formatting._cf_rules.values():
        cf_rules.extend(rules)
    no_red_rules = [
        r
        for r in cf_rules
        if any(f'="{ControlCols.VAL_PROCESAR_NO}"' in (f or "") for f in (r.formula or []))
    ]
    assert not no_red_rules, "No debe existir CF que pinte B7 rojo para NO"
    proc_cell_ref = f"{get_column_letter(2)}{proc_row}"
    dvs = [dv for dv in ws.data_validations.dataValidation if proc_cell_ref in (dv.sqref or "")]
    assert dvs, "B7 debe conservar data validation SI/NO"
