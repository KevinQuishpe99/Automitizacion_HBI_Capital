"""
Registro operativo en bandejas pagos_adelantados / pagos_incompletos tras Finalize (hoja Pendientes).
"""

from __future__ import annotations

import io
import logging
import zipfile
from datetime import date, datetime
from typing import Any

import httpx
import openpyxl

from app.application.sharepoint_resolution import encode_graph_drive_path
from app.application.services.review_schema import (
    DistribucionCols,
    EstadoPago,
    is_validar_pago_si,
)
from app.application.use_cases.setup_payment_followup_workbooks import (
    ADELANTADOS_COLUMNS,
    INCOMPLETOS_COLUMNS,
    SHEET_HISTORICO,
    SHEET_PENDIENTES,
    adelantados_workbook_relative_path,
    incompletos_workbook_relative_path,
)
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

ESTADO_APLICACION_PENDIENTE_TABLA = "PENDIENTE_APLICAR_TABLA"
ESTADO_IBR_PENDIENTE = "PENDIENTE_IBR"
ESTADO_FINAL_ABIERTO = "ABIERTO"


def _content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


def _s(dist: dict[str, Any], key: str) -> str:
    v = dist.get(key)
    return "" if v is None else str(v).strip()


def _row_adelantados(dist: dict[str, Any], historical_relative_path: str, now_s: str) -> dict[str, Any]:
    return {
        "ID Pago": _s(dist, DistribucionCols.ID_PAGO),
        "Cliente": _s(dist, DistribucionCols.CLIENTE),
        "Crédito": _s(dist, DistribucionCols.CREDITO),
        "FechaPago": _s(dist, DistribucionCols.FECHA_BANCO),
        "FechaLimitePago": _s(dist, DistribucionCols.FECHA_LIMITE),
        "FechaIBRRequerida": "",
        "MontoBanco": dist.get(DistribucionCols.MONTO_BANCO),
        "ValorExtracto": dist.get(DistribucionCols.VALOR_EXTRACTO),
        "AplicarAExtracto": dist.get(DistribucionCols.APLICAR_A_EXTRACTO),
        "MoraAAplicar": dist.get(DistribucionCols.MORA_A_APLICAR),
        "OtrosValores": dist.get(DistribucionCols.OTROS_VALORES),
        "TotalAplicado": dist.get(DistribucionCols.TOTAL_APLICADO),
        "SaldoPorAsignar": dist.get(DistribucionCols.SALDO_POR_ASIGNAR),
        "TablaAmortizacionPath": _s(dist, DistribucionCols.LINK_TABLA),
        "RutaUnidadCredito": _s(dist, DistribucionCols.RUTA_UNIDAD_CREDITO),
        "RutaExtracto": _s(dist, DistribucionCols.RUTA) or _s(dist, DistribucionCols.LINK_EXTRACTO),
        "AsientoPdfPath": _s(dist, DistribucionCols.RUTA_ASIENTOS_CONTABLES),
        "HistoricalFilePath": historical_relative_path.strip().strip("/"),
        "FilaAplicacionPago": "",
        "FilaIBR": "",
        "EstadoAplicacionPago": ESTADO_APLICACION_PENDIENTE_TABLA,
        "EstadoIBR": ESTADO_IBR_PENDIENTE,
        "EstadoFinal": ESTADO_FINAL_ABIERTO,
        "IBRUsado": "",
        "FechaCierreIBR": "",
        "Observacion": _s(dist, DistribucionCols.OBSERVACION),
        "CreatedAt": now_s,
        "UpdatedAt": now_s,
    }


def _row_incompletos(dist: dict[str, Any], historical_relative_path: str, now_s: str) -> dict[str, Any]:
    return {
        "ID Pago": _s(dist, DistribucionCols.ID_PAGO),
        "Cliente": _s(dist, DistribucionCols.CLIENTE),
        "Crédito": _s(dist, DistribucionCols.CREDITO),
        "FechaPago": _s(dist, DistribucionCols.FECHA_BANCO),
        "FechaLimitePago": _s(dist, DistribucionCols.FECHA_LIMITE),
        "MontoBanco": dist.get(DistribucionCols.MONTO_BANCO),
        "ValorExtracto": dist.get(DistribucionCols.VALOR_EXTRACTO),
        "AplicarAExtracto": dist.get(DistribucionCols.APLICAR_A_EXTRACTO),
        "MoraAAplicar": dist.get(DistribucionCols.MORA_A_APLICAR),
        "OtrosValores": dist.get(DistribucionCols.OTROS_VALORES),
        "TotalAplicado": dist.get(DistribucionCols.TOTAL_APLICADO),
        "SaldoPorAsignar": dist.get(DistribucionCols.SALDO_POR_ASIGNAR),
        "TablaAmortizacionPath": _s(dist, DistribucionCols.LINK_TABLA),
        "RutaUnidadCredito": _s(dist, DistribucionCols.RUTA_UNIDAD_CREDITO),
        "RutaExtracto": _s(dist, DistribucionCols.RUTA) or _s(dist, DistribucionCols.LINK_EXTRACTO),
        "AsientoPdfPath": _s(dist, DistribucionCols.RUTA_ASIENTOS_CONTABLES),
        "HistoricalFilePath": historical_relative_path.strip().strip("/"),
        "EstadoAplicacionPago": ESTADO_APLICACION_PENDIENTE_TABLA,
        "EstadoFinal": ESTADO_FINAL_ABIERTO,
        "Observacion": _s(dist, DistribucionCols.OBSERVACION),
        "CreatedAt": now_s,
        "UpdatedAt": now_s,
    }


def _ordered_values(row_map: dict[str, Any], columns: tuple[str, ...]) -> list[Any]:
    return [row_map.get(h) for h in columns]


def _composite_key(row_map: dict[str, Any]) -> tuple[str, str]:
    return (_s(row_map, "ID Pago"), _s(row_map, "Crédito"))


async def _upsert_pendientes(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    rel_path: str,
    columns: tuple[str, ...],
    row_maps: list[dict[str, Any]],
) -> bool:
    if not row_maps:
        return True
    try:
        raw = await client.get_bytes(_content_endpoint(site_id, drive_id, rel_path))
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        if code == 404:
            logger.warning(
                "payment_followup: archivo no existe (ejecute setup/payment-followup-workbooks): %s",
                rel_path,
            )
        else:
            logger.warning("payment_followup: no se pudo leer %s: %s", rel_path, exc)
        return False
    except Exception as exc:
        logger.warning("payment_followup: no se pudo leer %s: %s", rel_path, exc)
        return False

    if not raw or len(raw) < 64:
        logger.warning("payment_followup: archivo vacío o inválido: %s", rel_path)
        return False

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw))
    except zipfile.BadZipFile:
        logger.warning("payment_followup: no es Excel válido: %s", rel_path)
        return False

    if SHEET_PENDIENTES not in wb.sheetnames:
        logger.warning("payment_followup: falta hoja %r en %s", SHEET_PENDIENTES, rel_path)
        return False
    if SHEET_HISTORICO not in wb.sheetnames:
        logger.warning("payment_followup: falta hoja %r en %s", SHEET_HISTORICO, rel_path)

    ws = wb[SHEET_PENDIENTES]
    ncols = len(columns)
    header_to_col: dict[str, int] = {}
    for c in range(1, ncols + 1):
        h = str(ws.cell(1, c).value or "").strip()
        if h:
            header_to_col[h] = c

    key_to_row: dict[tuple[str, str], int] = {}
    id_col = header_to_col.get("ID Pago")
    cred_col = header_to_col.get("Crédito")
    if id_col and cred_col:
        for r in range(2, (ws.max_row or 1) + 1):
            id_p = str(ws.cell(r, id_col).value or "").strip()
            cred = str(ws.cell(r, cred_col).value or "").strip()
            if id_p and cred:
                key_to_row[(id_p, cred)] = r

    changed = False
    for row_map in row_maps:
        key = _composite_key(row_map)
        if not key[0]:
            continue
        if key in key_to_row:
            r = key_to_row[key]
            created_col = header_to_col.get("CreatedAt")
            if created_col:
                created = ws.cell(r, created_col).value
                if created:
                    row_map["CreatedAt"] = created
            row_map["UpdatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for h, col_idx in header_to_col.items():
                if h in row_map:
                    ws.cell(r, col_idx, value=row_map[h])
            changed = True
        else:
            vals = _ordered_values(row_map, columns)
            ws.append(vals)
            key_to_row[key] = ws.max_row
            changed = True

    if changed:
        out = io.BytesIO()
        wb.save(out)
        await client.put_bytes(
            _content_endpoint(site_id, drive_id, rel_path),
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return True


async def register_payment_followups_after_finalize(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    *,
    process_date: date,
    distributions: list[dict[str, Any]],
    historical_relative_path: str,
) -> list[str]:
    """Upsert en Pendientes; no modifica Historico."""
    _ = process_date  # reservado para fases posteriores
    warnings: list[str] = []
    now_s = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    adel_maps: list[dict[str, Any]] = []
    inc_maps: list[dict[str, Any]] = []

    for dist in distributions:
        if not is_validar_pago_si(dist):
            continue
        ep = str(dist.get(DistribucionCols.ESTADO_PAGO, "")).strip().upper()
        if ep == EstadoPago.ADELANTADO:
            adel_maps.append(_row_adelantados(dist, historical_relative_path, now_s))
        elif ep == EstadoPago.INCOMPLETO:
            inc_maps.append(_row_incompletos(dist, historical_relative_path, now_s))

    path_ad = adelantados_workbook_relative_path()
    path_in = incompletos_workbook_relative_path()

    if adel_maps:
        ok = await _upsert_pendientes(
            client,
            site_id,
            drive_id,
            path_ad,
            ADELANTADOS_COLUMNS,
            adel_maps,
        )
        if not ok:
            warnings.append(f"payment_followup_failed:{path_ad}")

    if inc_maps:
        ok = await _upsert_pendientes(
            client,
            site_id,
            drive_id,
            path_in,
            INCOMPLETOS_COLUMNS,
            inc_maps,
        )
        if not ok:
            warnings.append(f"payment_followup_failed:{path_in}")

    return warnings
