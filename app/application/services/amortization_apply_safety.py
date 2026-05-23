"""
Validaciones de seguridad para apply de amortización (warnings, huella PDF, verificación post-upload).
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.application.services.accounting_pdf_parser import (
    ACCOUNT_CAPITAL,
    ACCOUNT_SALDOS_MENORES,
    ACCOUNT_VALOR_PAGADO_CLIENTE,
    WARNING_BANK_INFERRED,
    PaymentApplicationEvent,
)
from app.application.services.amortization_workbook import (
    ADOPTADO_EXISTENTE,
    APLICADO,
    AUTOMATION_LOG_SHEET,
    _amounts_close,
    _is_formula_value,
    compare_existing_application,
    detect_headers,
    find_header_row,
    load_automation_log_index,
)

_AMOUNT_TOLERANCE = 0.02


def compute_asiento_pdf_hash(pdf_bytes: bytes) -> str:
    return hashlib.sha256(pdf_bytes).hexdigest()


def item_warnings_allowed(item: dict[str, Any]) -> bool:
    """
    Único warning permitido: BANK_VALUE_INFERRED_OR_MISSING en abono saldos menores
    (544153159505 + 544113410519, sin línea banco, valor económico claro).
    """
    warnings = [str(w).strip() for w in (item.get("warnings") or []) if str(w).strip()]
    if not warnings:
        return True

    detected = {str(c).strip() for c in (item.get("detected_codes") or []) if str(c).strip()}
    required = {ACCOUNT_SALDOS_MENORES, ACCOUNT_CAPITAL}
    pa = item.get("payment_application") or {}
    vp = float(pa.get("valor_pagado_cliente") or 0)
    saldos = float(pa.get("saldos_menores") or 0)
    capital = float(pa.get("capital") or 0)

    for warning in warnings:
        if warning != WARNING_BANK_INFERRED:
            return False
        if not required.issubset(detected):
            return False
        if ACCOUNT_VALOR_PAGADO_CLIENTE in detected:
            return False
        if vp <= 0 or saldos <= 0 or capital <= 0:
            return False
        if abs(vp - saldos) > _AMOUNT_TOLERANCE and abs(vp - capital) > _AMOUNT_TOLERANCE:
            return False
        if detected - required:
            return False
    return True


def collect_disallowed_warning_items(dry_run: dict[str, Any]) -> list[dict[str, Any]]:
    blocked: list[dict[str, Any]] = []
    for item in dry_run.get("items") or []:
        if not isinstance(item, dict):
            continue
        if item.get("error_code"):
            continue
        if not item_warnings_allowed(item):
            blocked.append(item)
    return blocked


@dataclass(frozen=True)
class IdempotencyCheckResult:
    skip_idempotent: bool
    pdf_changed: bool
    reason: str = ""


def check_idempotency_against_log(
    item: dict[str, Any],
    log_index: dict[str, Any],
) -> IdempotencyCheckResult:
    """Compara IdempotencyKey + huella PDF con registro previo en _AUTOMATION_LOG."""
    idem_key = str(item.get("idempotency_key") or "").strip()
    if not idem_key or idem_key not in log_index:
        return IdempotencyCheckResult(skip_idempotent=False, pdf_changed=False)

    stored = log_index[idem_key]
    new_hash = str(item.get("asiento_pdf_hash") or "").strip()
    new_etag = str(item.get("asiento_pdf_etag") or "").strip()
    stored_hash = str(getattr(stored, "pdf_hash", "") or "").strip()
    stored_etag = str(getattr(stored, "pdf_etag", "") or "").strip()

    if stored_hash and new_hash and stored_hash != new_hash:
        return IdempotencyCheckResult(
            skip_idempotent=False,
            pdf_changed=True,
            reason="asiento_pdf_hash distinto al registrado",
        )
    if stored_etag and new_etag and stored_etag != new_etag:
        return IdempotencyCheckResult(
            skip_idempotent=False,
            pdf_changed=True,
            reason="asiento_pdf_etag distinto al registrado",
        )

    if stored_hash and new_hash and stored_hash == new_hash:
        return IdempotencyCheckResult(skip_idempotent=True, pdf_changed=False)
    if stored_etag and new_etag and stored_etag == new_etag:
        return IdempotencyCheckResult(skip_idempotent=True, pdf_changed=False)

    if not stored_hash and not stored_etag and not new_hash and not new_etag:
        return IdempotencyCheckResult(skip_idempotent=True, pdf_changed=False)

    if not stored_hash and not stored_etag:
        return IdempotencyCheckResult(skip_idempotent=True, pdf_changed=False)

    return IdempotencyCheckResult(skip_idempotent=False, pdf_changed=False)


def _event_from_item_dict(item: dict[str, Any]) -> PaymentApplicationEvent:
    from datetime import date

    pa = item.get("payment_application") or {}
    fecha_asiento = None
    raw_fa = item.get("fecha_asiento")
    if raw_fa:
        try:
            fecha_asiento = date.fromisoformat(str(raw_fa))
        except ValueError:
            fecha_asiento = None
    return PaymentApplicationEvent(
        id_pago=str(item.get("id_pago") or ""),
        cliente=str(item.get("cliente") or ""),
        credito=str(item.get("credito") or ""),
        asiento_pdf_path=str(item.get("asiento_pdf_path") or ""),
        comprobante=str(item.get("comprobante") or ""),
        fecha_asiento=fecha_asiento,
        valor_pagado_cliente=float(pa.get("valor_pagado_cliente") or 0),
        capital=float(pa.get("capital") or 0),
        intereses=float(pa.get("intereses") or 0),
        mora=float(pa.get("mora") or 0),
        retenciones=float(pa.get("retenciones") or 0),
        saldos_menores=float(pa.get("saldos_menores") or 0),
        raw_text="",
    )


def _verify_cell_plan(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    row: int,
    col: int | None,
    expected: str,
    field: str,
    idem_key: str,
) -> str | None:
    if col is None:
        return None
    raw_formula = ws_formulas.cell(row, col).value
    raw_value = ws_values.cell(row, col).value
    if expected == "formula":
        if not _is_formula_value(raw_formula):
            return (
                f"{idem_key}: {field} FORMULA_MISMATCH "
                f"(esperada fórmula, leído={raw_formula!r})"
            )
    elif expected == "empty":
        if raw_formula is not None and str(raw_formula).strip():
            if _is_formula_value(raw_formula):
                return f"{idem_key}: {field} FORMULA_MISMATCH (debe estar vacío)"
            num = _cell_float(raw_formula)
            if num is not None and abs(num) > _AMOUNT_TOLERANCE:
                return f"{idem_key}: {field} FORMULA_MISMATCH (debe estar vacío)"
        if raw_value is not None and str(raw_value).strip():
            num = _cell_float(raw_value)
            if num is not None and abs(num) > _AMOUNT_TOLERANCE:
                return f"{idem_key}: {field} debe estar vacío, leído {raw_value}"
    elif expected == "value":
        if raw_value is None or (isinstance(raw_value, str) and not str(raw_value).strip()):
            if not _is_formula_value(raw_formula):
                return f"{idem_key}: {field} valor ausente tras upload"
    return None


def verify_uploaded_table(
    wb: Workbook,
    *,
    sheet_name: str,
    tabla_path: str,
    applied_items: list[dict[str, Any]],
    wb_formulas: Workbook | None = None,
    dry_run: dict[str, Any] | None = None,
) -> list[str]:
    """
    Verifica IBR, aplicación y presencia en log tras re-descarga.
    Devuelve lista de mensajes de error (vacía = ok).
    """
    errors: list[str] = []
    if sheet_name not in wb.sheetnames:
        return [f"hoja {sheet_name!r} no encontrada tras upload"]

    ws = wb[sheet_name]
    ws_f = (
        wb_formulas[sheet_name]
        if wb_formulas is not None and sheet_name in wb_formulas.sheetnames
        else ws
    )
    header_row = find_header_row(ws)
    headers = detect_headers(ws, header_row=header_row)
    log_index = load_automation_log_index(wb)

    for item in applied_items:
        idem_key = str(item.get("idempotency_key") or "").strip()
        app_row = item.get("application_row")
        ibr_row = item.get("ibr_row")
        if app_row is None:
            errors.append(f"{idem_key}: sin application_row")
            continue

        item_payment_raw = str(item.get("payment_date_iso") or "").strip()
        if not item_payment_raw and dry_run:
            item_payment_raw = str(dry_run.get("report_date_iso") or "").strip()
        item_payment_date = None
        if item_payment_raw:
            try:
                from datetime import date as date_cls

                item_payment_date = date_cls.fromisoformat(item_payment_raw)
            except ValueError:
                pass

        event = _event_from_item_dict(item)
        compare = compare_existing_application(
            ws,
            int(app_row),
            headers,
            event,
            payment_date=item_payment_date,
            detected_codes=frozenset(str(c) for c in (item.get("detected_codes") or [])),
            warnings=frozenset(str(w) for w in (item.get("warnings") or [])),
        )
        apply_status = item.get("apply_status")
        write_plan = item.get("write_plan") or {}
        if (
            apply_status == "APPLIED"
            and not write_plan
            and compare not in (APLICADO, ADOPTADO_EXISTENTE)
        ):
            errors.append(
                f"{idem_key}: fila {app_row} no coincide tras upload (estado={compare})"
            )
        elif apply_status == "ADOPTED" and compare != ADOPTADO_EXISTENTE:
            errors.append(
                f"{idem_key}: fila {app_row} adoptada esperada, estado={compare}"
            )

        if apply_status == "APPLIED" and write_plan:
            fecha_col = headers.get("fecha_pago")
            if write_plan.get("fecha_pago") == "value" and item_payment_date and fecha_col:
                cell_date = ws.cell(int(app_row), fecha_col).value
                parsed = None
                if hasattr(cell_date, "date"):
                    parsed = cell_date.date()  # type: ignore[union-attr]
                elif isinstance(cell_date, date):
                    parsed = cell_date
                if parsed != item_payment_date:
                    errors.append(
                        f"{idem_key}: Fecha pago esperada {item_payment_date}, leída {parsed}"
                    )
            for field in (
                "saldo_a_capital",
                "valor_pagado_cliente",
                "saldos_menores",
            ):
                expected = write_plan.get(field)
                if not expected:
                    continue
                col = headers.get(field)
                mismatch = _verify_cell_plan(
                    ws,
                    ws_f,
                    int(app_row),
                    col,
                    expected,
                    field,
                    idem_key,
                )
                if mismatch:
                    errors.append(mismatch)

        ibr_block = item.get("ibr") or {}
        if ibr_block.get("status") == "WOULD_WRITE_IBR" and ibr_row is not None:
            ibr_col = headers.get("ibr_i")
            ibr_val = ibr_block.get("value")
            if ibr_col is None:
                errors.append(f"{idem_key}: columna IBR +i ausente")
            elif ibr_val is not None:
                existing = ws.cell(int(ibr_row), ibr_col).value
                if existing is None or not _amounts_close(
                    float(ibr_val), _cell_float(existing)
                ):
                    errors.append(
                        f"{idem_key}: IBR fila {ibr_row} esperado {ibr_val}, leído {existing}"
                    )

        if idem_key and idem_key not in log_index:
            errors.append(f"{idem_key}: no aparece en _AUTOMATION_LOG tras upload")

    return errors


def _cell_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def build_table_apply_summary(
    tabla_path: str,
    items: list[dict[str, Any]],
    *,
    upload_status: str,
    verification_status: str,
) -> dict[str, Any]:
    summary = {
        "tabla_amortizacion_path": tabla_path,
        "eventos_aplicados": 0,
        "eventos_adoptados": 0,
        "eventos_idempotentes": 0,
        "ibr_escritos": 0,
        "errores": 0,
        "upload_status": upload_status,
        "verification_status": verification_status,
    }
    for it in items:
        st = it.get("apply_status")
        if st == "APPLIED":
            summary["eventos_aplicados"] += 1
        elif st == "ADOPTED":
            summary["eventos_adoptados"] += 1
        elif st == "SKIPPED_IDEMPOTENT":
            summary["eventos_idempotentes"] += 1
        elif st == "ERROR":
            summary["errores"] += 1
        ibr = it.get("ibr") or {}
        if it.get("apply_ibr_written") or (
            st == "APPLIED" and ibr.get("status") == "WOULD_WRITE_IBR"
        ):
            summary["ibr_escritos"] += 1
    return summary
