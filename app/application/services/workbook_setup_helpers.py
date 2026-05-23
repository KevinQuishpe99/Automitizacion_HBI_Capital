"""
Utilidades para workbooks de setup compatibles con Excel Online (sin Table/ListObject).
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

_FILL_HEADER = PatternFill(fill_type="solid", fgColor="002060")
_FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_BODY = Font(name="Calibri", size=11)
_BORDER = Border(
    left=Side(style="thin", color="C8C8C8"),
    right=Side(style="thin", color="C8C8C8"),
    top=Side(style="thin", color="C8C8C8"),
    bottom=Side(style="thin", color="C8C8C8"),
)
_ALIGN_HEADER = Alignment(vertical="center", horizontal="center", wrap_text=True)
_ALIGN_BODY = Alignment(vertical="center", horizontal="left", wrap_text=True)
_PROT_LOCKED = Protection(locked=True)
_PROT_UNLOCKED = Protection(locked=False)


def apply_header_row(ws: Any, columns: tuple[str, ...], *, row: int = 1) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_HEADER
        cell.border = _BORDER


def _set_column_widths(ws: Any, ncols: int, *, default: float = 16.0) -> None:
    for col_idx in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = default


def _sheet_cosmetics(ws: Any) -> None:
    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass
    try:
        ws.sheet_properties.tabColor = Color(rgb="002060")
    except Exception:
        pass


def remove_all_tables(ws: Any) -> None:
    for name in list(ws.tables.keys()):
        del ws.tables[name]


def configure_internal_automation_sheet(ws: Any, columns: tuple[str, ...]) -> None:
    """
    Archivos internos (minúsculas): solo encabezados; todas las celdas bloqueadas.
    Sin ListObject — evita «LIBRO REPARADO» en Excel Online.
    """
    remove_all_tables(ws)
    ncols = len(columns)
    apply_header_row(ws, columns, row=1)
    _set_column_widths(ws, ncols)
    last_col = get_column_letter(ncols)
    max_row = max(ws.max_row or 1, 1)
    for r in range(1, max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).protection = _PROT_LOCKED
    ws.auto_filter.ref = f"A1:{last_col}{max_row}"
    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    _sheet_cosmetics(ws)


def configure_secretary_editable_sheet(ws: Any, columns: tuple[str, ...]) -> None:
    """
    Archivos editables por usuario (MAYÚSCULAS): fila 1 bloqueada; filas >= 2 editables.
    """
    remove_all_tables(ws)
    ncols = len(columns)
    apply_header_row(ws, columns, row=1)
    _set_column_widths(ws, ncols, default=14.0)
    last_col = get_column_letter(ncols)
    max_row = max(ws.max_row or 1, 2)
    if max_row < 2:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=2, column=c, value=None)
            cell.font = _FONT_BODY
            cell.alignment = _ALIGN_BODY
            cell.border = _BORDER
        max_row = 2
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c).protection = _PROT_LOCKED
    for r in range(2, max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).protection = _PROT_UNLOCKED
    ws.auto_filter.ref = f"A1:{last_col}{max_row}"
    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    _sheet_cosmetics(ws)


def sheet_headers_match(ws: Any, expected: tuple[str, ...]) -> bool:
    for i, exp in enumerate(expected, start=1):
        if str(ws.cell(row=1, column=i).value or "").strip() != exp:
            return False
    return True
