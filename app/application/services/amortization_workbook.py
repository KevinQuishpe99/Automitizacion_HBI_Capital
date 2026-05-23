"""
Utilidades sobre tablas de amortización Excel para la futura API de llenado automático.
"""

from __future__ import annotations

import calendar
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.utils import get_column_letter

from app.application.services.accounting_pdf_parser import (
    ACCOUNT_SALDOS_MENORES,
    ACCOUNT_VALOR_PAGADO_CLIENTE,
    WARNING_BANK_INFERRED,
    PaymentApplicationEvent,
)

AUTOMATION_LOG_SHEET = "_AUTOMATION_LOG"
AUTOMATION_LOG_HEADERS = (
    "Timestamp",
    "IdPago",
    "Cliente",
    "Credito",
    "Fila",
    "Accion",
    "Detalle",
    "IdempotencyKey",
    "AsientoPdfPath",
    "ApplicationRow",
    "IbrRow",
    "AsientoPdfHash",
    "AsientoPdfETag",
    "Estado",
)

_AUTOMATION_LOG_COL = {name: idx + 1 for idx, name in enumerate(AUTOMATION_LOG_HEADERS)}

APLICADO = "APLICADO"
ADOPTADO_EXISTENTE = "ADOPTADO_EXISTENTE"
REVISION_MANUAL = "REVISION_MANUAL"

_AMOUNT_TOLERANCE = 0.02

# Claves canónicas internas -> patrones de encabezado (normalizados sin tildes)
_HEADER_PATTERNS: dict[str, tuple[str, ...]] = {
    "ibr_i": ("IBR +I", "IBR+I", "IBR I", "IBR TI", "IBR"),
    "spread": ("SPREAD",),
    "tasa_i": ("TASA + I", "TASA+I", "TASA I"),
    "cuota_i": ("CUOTA + I", "CUOTA+I", "CUOTA I"),
    "cuota": ("CUOTA",),
    "dia": ("DIA", "DÍA"),
    "mes": ("MES",),
    "anio": ("AÑO", "ANO", "ANIO"),
    "fecha_pago": ("FECHA PAGO",),
    "valor_intereses": ("VALOR INTERESES",),
    "abono_k": ("ABONO A K", "ABONO K"),
    "intereses_mora": (
        "INTERESES DE MORA",
        "INTERESES MORA",
        "INTERES DE MORA",
        "INTERES MORA",
    ),
    "valor_pagado_cliente": ("VALOR PAGADO CLIENTE", "VALOR PAGADO"),
    "saldos_menores": ("SALDOS MENORES", "SALDO MENOR"),
    "retenciones": ("RETENCIONES",),
    "saldo_a_capital": ("SALDO A CAPITAL", "SALDO CAPITAL"),
}

_REQUIRED_HEADER_KEYS: frozenset[str] = frozenset(
    {
        "dia",
        "mes",
        "anio",
        "fecha_pago",
        "valor_intereses",
        "abono_k",
        "valor_pagado_cliente",
    }
)

_HEADER_SCAN_MAX_ROW = 10

_DATE_PART_KEYS = ("dia", "mes", "anio")

_PAYMENT_HEADER_KEYS = (
    "fecha_pago",
    "valor_intereses",
    "abono_k",
    "intereses_mora",
    "retenciones",
    "valor_pagado_cliente",
    "saldos_menores",
    "saldo_a_capital",
)

_APPLICATION_SEARCH_KEYS = _PAYMENT_HEADER_KEYS


def _accent_fold_upper(text: str) -> str:
    raw = unicodedata.normalize("NFC", str(text or ""))
    nfkd = unicodedata.normalize("NFD", raw)
    folded = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", folded).strip().upper()


def _header_matches(cell: str, patterns: tuple[str, ...]) -> bool:
    key = _accent_fold_upper(cell)
    if not key:
        return False
    for p in patterns:
        if key == _accent_fold_upper(p):
            return True
    return False


def _match_header_key(label: str) -> str | None:
    for key, patterns in _HEADER_PATTERNS.items():
        if _header_matches(label, patterns):
            return key
    return None


def _labels_in_row(ws: Worksheet, row: int) -> list[str]:
    labels: list[str] = []
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        label = str(ws.cell(row, col).value or "").strip()
        if label:
            labels.append(label)
    return labels


def _all_header_matches_in_row(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    """Todas las columnas por clave canónica (admite encabezados duplicados)."""
    matches: dict[str, list[int]] = defaultdict(list)
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        label = str(ws.cell(header_row, col).value or "").strip()
        if not label:
            continue
        key = _match_header_key(label)
        if key:
            matches[key].append(col)
    return dict(matches)


def _contiguous_date_triplets(matches: dict[str, list[int]]) -> list[tuple[int, int, int]]:
    """Grupos dia, mes, año en columnas consecutivas."""
    groups: list[tuple[int, int, int]] = []
    for d_col in matches.get("dia", []):
        for m_col in matches.get("mes", []):
            if m_col != d_col + 1:
                continue
            for a_col in matches.get("anio", []):
                if a_col == m_col + 1:
                    groups.append((d_col, m_col, a_col))
    return groups


def _min_col_for_keys(matches: dict[str, list[int]], keys: tuple[str, ...]) -> int | None:
    cols: list[int] = []
    for key in keys:
        cols.extend(matches.get(key, []))
    return min(cols) if cols else None


def _score_date_triplet(
    triplet: tuple[int, int, int],
    matches: dict[str, list[int]],
) -> int:
    """Mayor = mejor candidato para fecha de corte."""
    d_col, _m_col, a_col = triplet
    score = 1000 - d_col
    fecha_cols = matches.get("fecha_pago", [])
    if fecha_cols and a_col < min(fecha_cols):
        score += 100
    aplicacion_col = _min_col_for_keys(
        matches,
        tuple(k for k in matches if k not in _DATE_PART_KEYS),
    )
    if aplicacion_col is not None:
        for col in matches.get("fecha_pago", []):
            if d_col < col < aplicacion_col:
                score += 30
    for label_col in (
        matches.get("valor_intereses", [])
        + matches.get("valor_pagado_cliente", [])
    ):
        if d_col < label_col:
            score += 5
    return score


def _select_date_column_group(
    matches: dict[str, list[int]],
) -> tuple[int, int, int] | None:
    """Elige el grupo contiguo dia → mes → año de la tabla principal."""
    triplets = _contiguous_date_triplets(matches)
    if not triplets:
        return None
    return max(triplets, key=lambda t: _score_date_triplet(t, matches))


def _detect_headers_in_row(ws: Worksheet, header_row: int) -> dict[str, int]:
    """
    Mapea claves canónicas a índice de columna (1-based).
    Usa el primer grupo contiguo dia/mes/año; ignora «dia» sueltos posteriores.
    """
    matches = _all_header_matches_in_row(ws, header_row)
    date_group = _select_date_column_group(matches)
    if date_group is None:
        return {}

    d_col, m_col, a_col = date_group
    stray_dia_cols = {c for c in matches.get("dia", []) if c != d_col}
    found: dict[str, int] = {"dia": d_col, "mes": m_col, "anio": a_col}

    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        if col in (d_col, m_col, a_col) or col in stray_dia_cols:
            continue
        label = str(ws.cell(header_row, col).value or "").strip()
        if not label:
            continue
        key = _match_header_key(label)
        if key and key not in found:
            found[key] = col
    return found


def _score_header_set(found: dict[str, int]) -> int:
    """Mayor puntaje = mejor candidato; -1 si faltan obligatorios."""
    missing = _REQUIRED_HEADER_KEYS - found.keys()
    if missing:
        return -1
    score = len(_REQUIRED_HEADER_KEYS) * 10
    if "ibr_i" in found:
        score += 5
    for optional in ("intereses_mora", "saldos_menores", "spread", "tasa_i", "cuota_i", "cuota"):
        if optional in found:
            score += 1
    return score


def detect_headers(ws: Worksheet, *, header_row: int | None = None) -> dict[str, int]:
    """
    Mapea claves canónicas a índice de columna (1-based).
    Si ``header_row`` es None, busca la mejor fila de encabezados entre 1 y 10.
    """
    if header_row is not None:
        return _detect_headers_in_row(ws, header_row)

    best: dict[str, int] = {}
    best_score = -1
    last_row = min(_HEADER_SCAN_MAX_ROW, ws.max_row or _HEADER_SCAN_MAX_ROW)
    for row in range(1, last_row + 1):
        found = _detect_headers_in_row(ws, row)
        score = _score_header_set(found)
        if score > best_score:
            best_score = score
            best = found
    return best


def find_header_row(ws: Worksheet) -> int:
    """Fila (1-based) con la mejor coincidencia de encabezados de amortización."""
    best_row = 1
    best_score = -1
    last_row = min(_HEADER_SCAN_MAX_ROW, ws.max_row or _HEADER_SCAN_MAX_ROW)
    for row in range(1, last_row + 1):
        found = _detect_headers_in_row(ws, row)
        score = _score_header_set(found)
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


class AmortizationSheetNotFoundError(ValueError):
    """Ninguna hoja visible cumple encabezados mínimos de tabla de amortización."""

    def __init__(
        self,
        message: str = "amortization_sheet_not_found",
        *,
        workbook_sheets: list[str] | None = None,
        headers_detected_by_sheet: dict[str, Any] | None = None,
        required_headers_missing: list[str] | None = None,
        tabla_amortizacion_path: str = "",
    ) -> None:
        super().__init__(message)
        self.workbook_sheets = workbook_sheets or []
        self.headers_detected_by_sheet = headers_detected_by_sheet or {}
        self.required_headers_missing = required_headers_missing or []
        self.tabla_amortizacion_path = tabla_amortizacion_path


@dataclass(frozen=True)
class AmortizationSheetMatch:
    worksheet: Worksheet
    headers: dict[str, int]
    header_row: int


def _sheet_is_candidate(ws: Worksheet) -> bool:
    if ws.title == AUTOMATION_LOG_SHEET:
        return False
    state = getattr(ws, "sheet_state", "visible")
    return state in (None, "visible")


def detect_amortization_sheet(
    workbook: Workbook,
    *,
    tabla_amortizacion_path: str = "",
) -> AmortizationSheetMatch:
    """
    Selecciona la hoja visible con mejor coincidencia de encabezados (sin usar el nombre).
    """
    sheet_names = [ws.title for ws in workbook.worksheets]
    per_sheet_debug: dict[str, dict[str, Any]] = {}
    best_match: AmortizationSheetMatch | None = None
    best_score = -1

    for ws in workbook.worksheets:
        if not _sheet_is_candidate(ws):
            continue
        last_row = min(_HEADER_SCAN_MAX_ROW, ws.max_row or _HEADER_SCAN_MAX_ROW)
        row_candidates: dict[str, Any] = {}
        sheet_best_score = -1
        sheet_best_row = 1
        sheet_best_headers: dict[str, int] = {}

        for row in range(1, last_row + 1):
            found = _detect_headers_in_row(ws, row)
            score = _score_header_set(found)
            row_candidates[str(row)] = {
                "labels": _labels_in_row(ws, row),
                "matched_keys": sorted(found.keys()),
                "score": score,
            }
            if score > sheet_best_score:
                sheet_best_score = score
                sheet_best_row = row
                sheet_best_headers = found

        per_sheet_debug[ws.title] = {
            "best_row": sheet_best_row,
            "best_score": sheet_best_score,
            "rows": row_candidates,
        }

        if sheet_best_score > best_score:
            best_score = sheet_best_score
            best_match = AmortizationSheetMatch(
                worksheet=ws,
                headers=sheet_best_headers,
                header_row=sheet_best_row,
            )

    if best_match is not None and best_score >= 0:
        return best_match

    missing = sorted(_REQUIRED_HEADER_KEYS)
    raise AmortizationSheetNotFoundError(
        "amortization_sheet_not_found",
        workbook_sheets=sheet_names,
        headers_detected_by_sheet=per_sheet_debug,
        required_headers_missing=missing,
        tabla_amortizacion_path=tabla_amortizacion_path,
    )


def ensure_automation_log(workbook: Workbook) -> Worksheet:
    if AUTOMATION_LOG_SHEET in workbook.sheetnames:
        return workbook[AUTOMATION_LOG_SHEET]
    ws = workbook.create_sheet(AUTOMATION_LOG_SHEET)
    for col, name in enumerate(AUTOMATION_LOG_HEADERS, start=1):
        ws.cell(1, col, value=name)
    return ws


def _is_formula_cell_value(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().startswith("=")
    return False


def _parse_int_cell(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, str) and _is_formula_cell_value(value):
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        if abs(v - round(v)) < 1e-9:
            return int(round(v))
        return None
    text = str(value).strip().replace(",", ".")
    if _is_formula_cell_value(text):
        return None
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        v = float(text)
        if abs(v - round(v)) < 1e-9:
            return int(round(v))
    return None


def _parse_int_cell_as_month(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.month
    if isinstance(value, date):
        return value.month
    return _parse_int_cell(value)


def _parse_int_cell_as_year(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    parsed = _parse_int_cell(value)
    if parsed is None:
        return None
    if parsed < 100:
        return parsed + 2000
    return parsed


def _cell_value_as_date(value: Any) -> date | None:
    """Fecha completa si la celda es datetime/date o serial Excel (no enteros pequeños)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = float(value)
        # Solo seriales plausibles (evita confundir 22 / 4 / 2026 con día/mes/año)
        if serial > 30000:
            try:
                from openpyxl.utils.datetime import from_excel

                converted = from_excel(serial)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except (ValueError, TypeError, OverflowError):
                pass
    return None


def _row_has_formula_in_date_columns(ws: Worksheet, row: int, headers: dict[str, int]) -> bool:
    for key in _DATE_PART_KEYS:
        col = headers.get(key)
        if col is None:
            continue
        if _is_formula_cell_value(ws.cell(row, col).value):
            return True
    return False


def _row_due_date(ws: Worksheet, row: int, headers: dict[str, int]) -> date | None:
    """Fecha de corte desde valores directos (no interpreta números dentro de fórmulas)."""
    if "dia" not in headers or "mes" not in headers or "anio" not in headers:
        return None
    if _row_has_formula_in_date_columns(ws, row, headers):
        return None

    dia_val = ws.cell(row, headers["dia"]).value
    mes_val = ws.cell(row, headers["mes"]).value
    anio_val = ws.cell(row, headers["anio"]).value

    for val in (dia_val, mes_val, anio_val):
        if isinstance(val, (datetime, date)):
            parsed = _cell_value_as_date(val)
            if parsed is not None:
                return parsed

    dia = _parse_int_cell(dia_val)
    mes = _parse_int_cell_as_month(mes_val)
    anio = _parse_int_cell_as_year(anio_val)
    if dia is None or mes is None or anio is None:
        return None
    try:
        return date(anio, mes, dia)
    except ValueError:
        return None


def _add_calendar_months(anchor: date, months: int) -> date:
    """Suma meses calendario conservando el día (ajusta fin de mes)."""
    y = anchor.year
    m = anchor.month + months
    while m > 12:
        m -= 12
        y += 1
    while m < 1:
        m += 12
        y -= 1
    last_day = calendar.monthrange(y, m)[1]
    day = min(anchor.day, last_day)
    return date(y, m, day)


def _row_has_data_signal(ws: Worksheet, row: int, headers: dict[str, int]) -> bool:
    """Fila con algún dato en columnas de fecha o aplicación del pago."""
    keys = (
        *_DATE_PART_KEYS,
        "fecha_pago",
        "valor_intereses",
        "abono_k",
        "valor_pagado_cliente",
    )
    for key in keys:
        col = headers.get(key)
        if col is None:
            continue
        val = ws.cell(row, col).value
        if val is not None and str(val).strip() != "":
            return True
    return False


def _build_monthly_schedule(
    ws: Worksheet,
    headers: dict[str, int],
    header_row: int,
) -> tuple[dict[int, date], dict[str, Any]] | tuple[None, dict[str, Any]]:
    """
    Proyecta fechas mensuales desde la primera fila con fecha real cuando las
    siguientes filas usan fórmulas o celdas vacías en dia/mes/año.
    """
    meta: dict[str, Any] = {
        "monthly_schedule_fallback_used": False,
        "monthly_schedule_base_row": None,
        "monthly_schedule_base_date": None,
    }
    if "dia" not in headers or "mes" not in headers or "anio" not in headers:
        meta["reason"] = "missing_date_headers"
        return None, meta

    data_start = header_row + 1
    max_row = ws.max_row or header_row
    base_row: int | None = None
    base_date: date | None = None
    for r in range(data_start, max_row + 1):
        if not _row_has_data_signal(ws, r, headers):
            continue
        if _row_has_formula_in_date_columns(ws, r, headers):
            continue
        row_date = _row_due_date(ws, r, headers)
        if row_date is not None:
            base_row = r
            base_date = row_date
            break

    if base_row is None or base_date is None:
        meta["reason"] = "no_base_date"
        return None, meta

    needs_fallback = False
    if max_row > base_row:
        for r in range(base_row + 1, max_row + 1):
            if _row_has_formula_in_date_columns(ws, r, headers):
                needs_fallback = True
                break
            if _row_due_date(ws, r, headers) is None:
                needs_fallback = True
                break

    if not needs_fallback:
        meta["reason"] = "direct_dates_sufficient"
        return None, meta

    schedule: dict[int, date] = {}
    for r in range(base_row, max_row + 1):
        month_offset = r - base_row
        schedule[r] = _add_calendar_months(base_date, month_offset)

    if not schedule:
        meta["reason"] = "empty_schedule"
        return None, meta

    meta["monthly_schedule_fallback_used"] = True
    meta["monthly_schedule_base_row"] = base_row
    meta["monthly_schedule_base_date"] = base_date.isoformat()
    return schedule, meta


@dataclass(frozen=True)
class FindRowResult:
    row: int | None
    formulas_detected: bool = False
    monthly_schedule_fallback_used: bool = False
    monthly_schedule_base_row: int | None = None
    monthly_schedule_base_date: str | None = None
    monthly_schedule: dict[int, date] | None = None


def _resolve_row_date(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    schedule: dict[int, date] | None,
) -> date | None:
    if schedule is not None and row in schedule:
        return schedule[row]
    return _row_due_date(ws, row, headers)


def find_row_by_due_date_detailed(
    ws: Worksheet,
    headers: dict[str, int],
    fecha_limite_pago: date,
    *,
    exclude_rows: frozenset[int] | None = None,
    header_row: int = 1,
) -> FindRowResult:
    """Busca fila de corte: valores directos, luego proyección mensual si aplica."""
    if "dia" not in headers or "mes" not in headers or "anio" not in headers:
        return FindRowResult(row=None)

    excluded = exclude_rows or frozenset()
    data_start = header_row + 1
    max_row = ws.max_row or header_row
    formulas_detected = False

    for r in range(data_start, max_row + 1):
        if r in excluded:
            continue
        if _row_has_formula_in_date_columns(ws, r, headers):
            formulas_detected = True
        row_date = _row_due_date(ws, r, headers)
        if row_date == fecha_limite_pago:
            return FindRowResult(
                row=r,
                formulas_detected=formulas_detected,
            )

    schedule, sched_meta = _build_monthly_schedule(ws, headers, header_row)
    if schedule:
        for r, projected in schedule.items():
            if r in excluded:
                continue
            if projected == fecha_limite_pago:
                return FindRowResult(
                    row=r,
                    formulas_detected=formulas_detected or True,
                    monthly_schedule_fallback_used=True,
                    monthly_schedule_base_row=sched_meta.get("monthly_schedule_base_row"),
                    monthly_schedule_base_date=sched_meta.get("monthly_schedule_base_date"),
                    monthly_schedule=schedule,
                )

    return FindRowResult(
        row=None,
        formulas_detected=formulas_detected,
        monthly_schedule=schedule,
        monthly_schedule_fallback_used=bool(sched_meta.get("monthly_schedule_fallback_used")),
        monthly_schedule_base_row=sched_meta.get("monthly_schedule_base_row"),
        monthly_schedule_base_date=sched_meta.get("monthly_schedule_base_date"),
    )


def _date_header_columns(headers: dict[str, int]) -> dict[str, int]:
    return {k: headers[k] for k in _DATE_PART_KEYS if k in headers}


def build_target_row_search_debug(
    ws: Worksheet,
    headers: dict[str, int],
    fecha_limite_pago: date,
    *,
    header_row: int = 1,
    sheet_name: str = "",
    tabla_amortizacion_path: str = "",
    exclude_rows: frozenset[int] | None = None,
    max_candidates: int = 25,
    workbook_loaded_data_only: bool = False,
    find_result: FindRowResult | None = None,
) -> dict[str, Any]:
    """Diagnóstico seguro cuando no hay fila de corte para la fecha objetivo."""
    excluded = exclude_rows or frozenset()
    data_start = header_row + 1
    max_row = ws.max_row or header_row
    schedule = (find_result.monthly_schedule if find_result else None) or None
    if schedule is None:
        schedule, _ = _build_monthly_schedule(ws, headers, header_row)

    formulas_detected = find_result.formulas_detected if find_result else False
    if not formulas_detected:
        for r in range(data_start, max_row + 1):
            if _row_has_formula_in_date_columns(ws, r, headers):
                formulas_detected = True
                break

    candidates: list[dict[str, Any]] = []
    for r in range(data_start, max_row + 1):
        if r in excluded:
            continue
        row_date = _resolve_row_date(ws, r, headers, schedule)
        if row_date is None:
            continue
        source = "monthly_schedule" if schedule and r in schedule else "direct"
        candidates.append(
            {
                "row": r,
                "date": row_date.isoformat(),
                "source": source,
                "dia": ws.cell(r, headers["dia"]).value,
                "mes": ws.cell(r, headers["mes"]).value,
                "anio": ws.cell(r, headers["anio"]).value,
            }
        )
        if len(candidates) >= max_candidates:
            break

    monthly_used = (
        find_result.monthly_schedule_fallback_used
        if find_result
        else bool(schedule)
    )
    base_row = find_result.monthly_schedule_base_row if find_result else None
    base_date = find_result.monthly_schedule_base_date if find_result else None
    if schedule and base_row is None:
        _, sched_meta = _build_monthly_schedule(ws, headers, header_row)
        base_row = sched_meta.get("monthly_schedule_base_row")
        base_date = sched_meta.get("monthly_schedule_base_date")

    return {
        "target_date": fecha_limite_pago.isoformat(),
        "date_header_columns": _date_header_columns(headers),
        "candidate_date_rows": candidates,
        "header_row": header_row,
        "sheet_name": sheet_name,
        "tabla_amortizacion_path": tabla_amortizacion_path,
        "workbook_loaded_data_only": workbook_loaded_data_only,
        "formulas_detected": formulas_detected,
        "monthly_schedule_fallback_used": monthly_used,
        "monthly_schedule_base_row": base_row,
        "monthly_schedule_base_date": base_date,
    }


def build_amortization_idempotency_key(
    id_pago: str,
    credito: str,
    asiento_pdf_path: str,
    comprobante: str,
    *,
    pdf_hash: str = "",
) -> str:
    """
    Clave de idempotencia futura: no asumir unicidad por ID Pago + Crédito solos.
    """
    parts = [
        str(id_pago or "").strip(),
        str(credito or "").strip(),
        str(asiento_pdf_path or "").strip().strip("/").casefold(),
        str(comprobante or "").strip(),
        str(pdf_hash or "").strip(),
    ]
    return "|".join(parts)


def find_row_by_due_date(
    ws: Worksheet,
    headers: dict[str, int],
    fecha_limite_pago: date,
    *,
    exclude_rows: frozenset[int] | None = None,
    header_row: int = 1,
) -> int | None:
    """Fila del corte cuya fecha (dia/mes/año del grupo seleccionado) coincide."""
    return find_row_by_due_date_detailed(
        ws,
        headers,
        fecha_limite_pago,
        exclude_rows=exclude_rows,
        header_row=header_row,
    ).row


@dataclass(frozen=True)
class FindApplicationRowResult:
    row: int | None
    compare_status: str | None = None
    requires_new_row: bool = False
    suggested_row: int | None = None
    search_start_row: int | None = None


def _application_search_start_row(
    due_date_row: int | None,
    header_row: int,
) -> int:
    if due_date_row is not None:
        return due_date_row
    return header_row + 1


def _row_has_application_column_data(ws: Worksheet, row: int, headers: dict[str, int]) -> bool:
    return not is_payment_application_empty(ws, row, headers)


def find_application_row_detailed(
    ws: Worksheet,
    headers: dict[str, int],
    event: PaymentApplicationEvent,
    *,
    due_date_row: int | None,
    exclude_rows: frozenset[int] | None = None,
    header_row: int = 1,
    payment_date: date | None = None,
    detected_codes: frozenset[str] | set[str] | tuple[str, ...] | None = None,
    warnings: frozenset[str] | set[str] | tuple[str, ...] | None = None,
) -> FindApplicationRowResult:
    """
    Busca fila en el bloque «Aplicación del Pago» (columnas de pago, no dia/mes/año).
    Puede estar desplazada respecto a due_date_row.
    """
    if not any(headers.get(k) for k in _APPLICATION_SEARCH_KEYS):
        return FindApplicationRowResult(row=None, compare_status=None)

    start = _application_search_start_row(due_date_row, header_row)
    max_row = ws.max_row or start
    excluded = exclude_rows or frozenset()

    first_revision_row: int | None = None
    compare_kwargs = {
        "payment_date": payment_date,
        "detected_codes": detected_codes,
        "warnings": warnings,
    }

    for r in range(start, max_row + 1):
        if (
            compare_existing_application(ws, r, headers, event, **compare_kwargs)
            == ADOPTADO_EXISTENTE
        ):
            return FindApplicationRowResult(
                row=r,
                compare_status=ADOPTADO_EXISTENTE,
                search_start_row=start,
            )

    for r in range(start, max_row + 1):
        if r in excluded:
            continue
        status = compare_existing_application(ws, r, headers, event, **compare_kwargs)
        if status == APLICADO:
            return FindApplicationRowResult(
                row=r,
                compare_status=APLICADO,
                search_start_row=start,
            )
        if status == REVISION_MANUAL and first_revision_row is None:
            first_revision_row = r

    if first_revision_row is not None and not any(
        r not in excluded
        and compare_existing_application(ws, r, headers, event, **compare_kwargs) == APLICADO
        for r in range((first_revision_row or start) + 1, max_row + 1)
    ):
        return FindApplicationRowResult(
            row=first_revision_row,
            compare_status=REVISION_MANUAL,
            search_start_row=start,
        )

    suggested = max_row + 1
    return FindApplicationRowResult(
        row=None,
        compare_status=None,
        requires_new_row=True,
        suggested_row=suggested,
        search_start_row=start,
    )


def find_application_row(
    ws: Worksheet,
    headers: dict[str, int],
    event: PaymentApplicationEvent,
    *,
    due_date_row: int | None,
    exclude_rows: frozenset[int] | None = None,
    header_row: int = 1,
) -> int | None:
    return find_application_row_detailed(
        ws,
        headers,
        event,
        due_date_row=due_date_row,
        exclude_rows=exclude_rows,
        header_row=header_row,
    ).row


def build_application_row_search_debug(
    ws: Worksheet,
    headers: dict[str, int],
    event: PaymentApplicationEvent,
    *,
    due_date_row: int | None,
    header_row: int = 1,
    sheet_name: str = "",
    tabla_amortizacion_path: str = "",
    exclude_rows: frozenset[int] | None = None,
    find_result: FindApplicationRowResult | None = None,
    max_candidates: int = 25,
) -> dict[str, Any]:
    """Diagnóstico cuando no hay fila de aplicación disponible."""
    excluded = exclude_rows or frozenset()
    start = (
        find_result.search_start_row
        if find_result and find_result.search_start_row is not None
        else _application_search_start_row(due_date_row, header_row)
    )
    max_row = ws.max_row or start
    candidates: list[dict[str, Any]] = []
    for r in range(start, max_row + 1):
        status = compare_existing_application(ws, r, headers, event)
        candidates.append(
            {
                "row": r,
                "compare_status": status,
                "excluded": r in excluded,
                "has_application_data": _row_has_application_column_data(ws, r, headers),
            }
        )
        if len(candidates) >= max_candidates:
            break

    return {
        "due_date_row": due_date_row,
        "search_start_row": start,
        "application_header_columns": {
            k: headers[k] for k in _APPLICATION_SEARCH_KEYS if k in headers
        },
        "candidate_application_rows": candidates,
        "header_row": header_row,
        "sheet_name": sheet_name,
        "tabla_amortizacion_path": tabla_amortizacion_path,
        "requires_new_row": bool(find_result and find_result.requires_new_row),
        "suggested_row": find_result.suggested_row if find_result else None,
    }


def _cell_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _amounts_close(a: float | None, b: float | None) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= _AMOUNT_TOLERANCE


def is_payment_application_empty(ws: Worksheet, row: int, headers: dict[str, int]) -> bool:
    for key in _PAYMENT_HEADER_KEYS:
        if key == "saldo_a_capital":
            continue
        col = headers.get(key)
        if col is None:
            continue
        val = ws.cell(row, col).value
        if _is_formula_value(val):
            continue
        if val is None or str(val).strip() == "":
            continue
        num = _cell_float(val)
        if num is not None and abs(num) > _AMOUNT_TOLERANCE:
            return False
        if num is None and str(val).strip():
            return False
    return True


def should_skip_valor_pagado_cliente(
    detected_codes: frozenset[str] | set[str] | tuple[str, ...],
    warnings: frozenset[str] | set[str] | tuple[str, ...],
) -> bool:
    """Sin cuenta banco y warning de inferencia: Valor pagado cliente queda vacío."""
    codes = frozenset(str(c) for c in detected_codes)
    warns = frozenset(str(w) for w in warnings)
    return (
        ACCOUNT_VALOR_PAGADO_CLIENTE not in codes
        and WARNING_BANK_INFERRED in warns
    )


def compare_existing_application(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    event: PaymentApplicationEvent,
    *,
    payment_date: date | None = None,
    detected_codes: frozenset[str] | set[str] | tuple[str, ...] | None = None,
    warnings: frozenset[str] | set[str] | tuple[str, ...] | None = None,
) -> str:
    """
    Devuelve APLICADO (vacío), ADOPTADO_EXISTENTE (coincide) o REVISION_MANUAL (distinto).
    """
    if is_payment_application_empty(ws, row, headers):
        return APLICADO

    codes = frozenset(
        str(c) for c in (detected_codes or getattr(event, "detected_codes", ()) or ())
    )
    warns = frozenset(
        str(w) for w in (warnings or getattr(event, "parse_warnings", ()) or ())
    )
    skip_vp = should_skip_valor_pagado_cliente(codes, warns)

    mapping = {
        "valor_intereses": event.intereses,
        "abono_k": event.capital,
        "intereses_mora": event.mora,
        "retenciones": event.retenciones,
        "saldos_menores": event.saldos_menores,
    }
    if not skip_vp:
        mapping["valor_pagado_cliente"] = event.valor_pagado_cliente

    for key, expected in mapping.items():
        col = headers.get(key)
        if col is None:
            continue
        raw = ws.cell(row, col).value
        if _is_formula_value(raw):
            continue
        existing = _cell_float(raw)
        if existing is not None and abs(existing) > _AMOUNT_TOLERANCE:
            if not _amounts_close(existing, expected):
                return REVISION_MANUAL

    if skip_vp:
        vp_col = headers.get("valor_pagado_cliente")
        if vp_col is not None:
            raw = ws.cell(row, vp_col).value
            if raw is not None and str(raw).strip():
                num = _cell_float(raw)
                if num is not None and abs(num) > _AMOUNT_TOLERANCE:
                    return REVISION_MANUAL

    fecha_ref = payment_date or event.fecha_asiento
    if headers.get("fecha_pago") and fecha_ref:
        existing_date = ws.cell(row, headers["fecha_pago"]).value
        parsed: date | None = None
        if isinstance(existing_date, datetime):
            parsed = existing_date.date()
        elif isinstance(existing_date, date):
            parsed = existing_date
        if parsed and parsed != fecha_ref:
            return REVISION_MANUAL

    return ADOPTADO_EXISTENTE


@dataclass(frozen=True)
class PaymentApplicationWriteOptions:
    """Opciones de escritura alineadas con el llenado manual de secretaria."""

    payment_date: date
    detected_codes: frozenset[str] = frozenset()
    warnings: frozenset[str] = frozenset()


def _is_formula_value(value: Any) -> bool:
    return isinstance(value, str) and str(value).strip().startswith("=")


def _shift_formula_rows(formula: str, row_delta: int) -> str:
    if not formula or not _is_formula_value(formula) or row_delta == 0:
        return formula
    pattern = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)", re.IGNORECASE)

    def repl(match: re.Match[str]) -> str:
        col_abs, col, row_abs, row_s = match.group(1), match.group(2), match.group(3), match.group(4)
        if row_abs:
            return match.group(0)
        new_row = int(row_s) + row_delta
        return f"{col_abs}{col.upper()}{row_abs}{max(1, new_row)}"

    return pattern.sub(repl, formula)


def _find_formula_template(
    ws: Worksheet,
    col: int,
    target_row: int,
    header_row: int,
) -> str | None:
    for r in range(target_row - 1, header_row, -1):
        val = ws.cell(r, col).value
        if _is_formula_value(val):
            return _shift_formula_rows(str(val), target_row - r)
    return None


def _default_saldo_capital_formula(headers: dict[str, int], row: int) -> str | None:
    saldo_col = headers.get("saldo_a_capital")
    abono_col = headers.get("abono_k")
    if saldo_col is None or abono_col is None or row <= 1:
        return None
    return (
        f"=+{get_column_letter(saldo_col)}{row - 1}-"
        f"{get_column_letter(abono_col)}{row}"
    )


def _default_valor_pagado_formula(headers: dict[str, int], row: int) -> str | None:
    if headers.get("valor_pagado_cliente") is None:
        return None
    parts: list[str] = []
    for key in ("valor_intereses", "abono_k", "intereses_mora"):
        col = headers.get(key)
        if col is not None:
            parts.append(f"{get_column_letter(col)}{row}")
    if not parts:
        return None
    return "=+" + "+".join(parts)


def _default_saldos_menores_formula(headers: dict[str, int], row: int) -> str | None:
    sm_col = headers.get("saldos_menores")
    abono_col = headers.get("abono_k")
    if sm_col is None or abono_col is None:
        return None
    return f"=+{get_column_letter(abono_col)}{row}"


def _clear_application_cell(ws: Worksheet, row: int, col: int) -> None:
    ws.cell(row, col).value = None


def _write_manual_amount(
    ws: Worksheet,
    row: int,
    col: int,
    value: float,
) -> str:
    if abs(value) > _AMOUNT_TOLERANCE:
        ws.cell(row, col, value=value)
        return "value"
    _clear_application_cell(ws, row, col)
    return "empty"


def write_payment_application(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
    event: PaymentApplicationEvent,
    *,
    write_options: PaymentApplicationWriteOptions | None = None,
    header_row: int = 1,
) -> dict[str, str]:
    """
    Escribe aplicación del pago como la secretaria (montos manuales + fórmulas).
    Devuelve write_plan para verificación post-upload.
    """
    plan: dict[str, str] = {}
    opts = write_options
    payment_date = opts.payment_date if opts else event.fecha_asiento
    detected = (
        opts.detected_codes
        if opts
        else frozenset(str(c) for c in getattr(event, "detected_codes", ()) or ())
    )
    warns = (
        opts.warnings
        if opts
        else frozenset(str(w) for w in getattr(event, "parse_warnings", ()) or ())
    )
    has_bank = ACCOUNT_VALOR_PAGADO_CLIENTE in detected
    skip_vp = should_skip_valor_pagado_cliente(detected, warns)

    if headers.get("fecha_pago") and payment_date:
        ws.cell(row, headers["fecha_pago"], value=payment_date)
        plan["fecha_pago"] = "value"

    if headers.get("valor_intereses"):
        plan["valor_intereses"] = _write_manual_amount(
            ws, row, headers["valor_intereses"], event.intereses
        )
    if headers.get("abono_k"):
        plan["abono_k"] = _write_manual_amount(ws, row, headers["abono_k"], event.capital)
    if headers.get("intereses_mora"):
        plan["intereses_mora"] = _write_manual_amount(
            ws, row, headers["intereses_mora"], event.mora
        )
    if headers.get("retenciones"):
        plan["retenciones"] = _write_manual_amount(
            ws, row, headers["retenciones"], event.retenciones
        )

    saldo_col = headers.get("saldo_a_capital")
    if saldo_col is not None:
        formula = _find_formula_template(ws, saldo_col, row, header_row)
        if not formula:
            formula = _default_saldo_capital_formula(headers, row)
        if formula:
            ws.cell(row, saldo_col, value=formula)
            plan["saldo_a_capital"] = "formula"

    vp_col = headers.get("valor_pagado_cliente")
    if vp_col is not None:
        if skip_vp:
            _clear_application_cell(ws, row, vp_col)
            plan["valor_pagado_cliente"] = "empty"
        elif has_bank:
            formula = _find_formula_template(ws, vp_col, row, header_row)
            if not formula:
                formula = _default_valor_pagado_formula(headers, row)
            if formula:
                ws.cell(row, vp_col, value=formula)
                plan["valor_pagado_cliente"] = "formula"
            else:
                _clear_application_cell(ws, row, vp_col)
                plan["valor_pagado_cliente"] = "empty"
        else:
            _clear_application_cell(ws, row, vp_col)
            plan["valor_pagado_cliente"] = "empty"

    sm_col = headers.get("saldos_menores")
    if sm_col is not None:
        if ACCOUNT_SALDOS_MENORES in detected and (
            abs(event.saldos_menores) > _AMOUNT_TOLERANCE
            or abs(event.capital) > _AMOUNT_TOLERANCE
        ):
            formula = _find_formula_template(ws, sm_col, row, header_row)
            if not formula and skip_vp:
                formula = _default_saldos_menores_formula(headers, row)
            if formula:
                ws.cell(row, sm_col, value=formula)
                plan["saldos_menores"] = "formula"
            else:
                amount = (
                    event.saldos_menores
                    if abs(event.saldos_menores) > _AMOUNT_TOLERANCE
                    else event.capital
                )
                plan["saldos_menores"] = _write_manual_amount(ws, row, sm_col, amount)
        else:
            _clear_application_cell(ws, row, sm_col)
            plan["saldos_menores"] = "empty"

    return plan


def write_ibr(ws: Worksheet, row: int, headers: dict[str, int], ibr_rate: float) -> None:
    """Escribe IBR normalizado en la fila del corte (columna IBR +i si existe)."""
    col = headers.get("ibr_i")
    if col is not None:
        ws.cell(row, col, value=ibr_rate)


@dataclass(frozen=True)
class AutomationLogRecord:
    idempotency_key: str
    pdf_hash: str = ""
    pdf_etag: str = ""
    pdf_path: str = ""
    application_row: int | None = None
    ibr_row: int | None = None
    estado: str = ""


def _automation_log_header_map(ws: Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        label = str(ws.cell(1, col).value or "").strip()
        if label:
            header_map[label] = col
    return header_map


def load_automation_log_index(workbook: Workbook) -> dict[str, AutomationLogRecord]:
    """Último registro por IdempotencyKey (incluye huella PDF si existe)."""
    if AUTOMATION_LOG_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[AUTOMATION_LOG_SHEET]
    if (ws.max_row or 1) < 2:
        return {}
    header_map = _automation_log_header_map(ws)
    col_key = header_map.get("IdempotencyKey")
    if col_key is None:
        return {}
    index: dict[str, AutomationLogRecord] = {}
    for row in range(2, (ws.max_row or 1) + 1):
        key = str(ws.cell(row, col_key).value or "").strip()
        if not key:
            continue

        def _cell(name: str) -> str:
            c = header_map.get(name)
            if c is None:
                return ""
            return str(ws.cell(row, c).value or "").strip()

        app_raw = _cell("ApplicationRow") or _cell("Fila")
        ibr_raw = _cell("IbrRow")
        index[key] = AutomationLogRecord(
            idempotency_key=key,
            pdf_hash=_cell("AsientoPdfHash"),
            pdf_etag=_cell("AsientoPdfETag"),
            pdf_path=_cell("AsientoPdfPath"),
            application_row=int(app_raw) if app_raw.isdigit() else None,
            ibr_row=int(ibr_raw) if ibr_raw.isdigit() else None,
            estado=_cell("Estado") or _cell("Accion"),
        )
    return index


def load_automation_log_idempotency_keys(workbook: Workbook) -> set[str]:
    """Claves ya registradas en _AUTOMATION_LOG (columna IdempotencyKey)."""
    return set(load_automation_log_index(workbook).keys())


def append_automation_log(workbook: Workbook, log_entry: dict[str, Any]) -> None:
    ws = ensure_automation_log(workbook)
    r = (ws.max_row or 1) + 1
    ts = log_entry.get("timestamp") or datetime.now().isoformat(timespec="seconds")
    application_row = log_entry.get("application_row", log_entry.get("fila"))
    ibr_row = log_entry.get("ibr_row")
    accion = str(log_entry.get("accion", "") or "")
    values = {
        "Timestamp": ts,
        "IdPago": log_entry.get("id_pago", ""),
        "Cliente": log_entry.get("cliente", ""),
        "Credito": log_entry.get("credito", ""),
        "Fila": application_row if application_row is not None else "",
        "Accion": accion,
        "Detalle": log_entry.get("detalle", ""),
        "IdempotencyKey": log_entry.get("idempotency_key", ""),
        "AsientoPdfPath": log_entry.get("asiento_pdf_path", ""),
        "ApplicationRow": application_row if application_row is not None else "",
        "IbrRow": ibr_row if ibr_row is not None else "",
        "AsientoPdfHash": log_entry.get("asiento_pdf_hash", ""),
        "AsientoPdfETag": log_entry.get("asiento_pdf_etag", ""),
        "Estado": log_entry.get("estado", accion),
    }
    for name, col in _AUTOMATION_LOG_COL.items():
        ws.cell(r, col, value=values.get(name, ""))


@dataclass(frozen=True)
class ApplyPaymentResult:
    accion: str
    fila: int
