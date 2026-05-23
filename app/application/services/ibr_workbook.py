"""
Lectura de IBR_DIARIO.xlsx (hoja IBR) para la futura API de amortización.
"""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from typing import Any

import openpyxl

from app.application.use_cases.setup_ibr_workbook import IBR_COLUMNS, IBR_SHEET_NAME

_PERCENT_THRESHOLD = 1.0


def _accent_fold_upper(text: str) -> str:
    raw = unicodedata.normalize("NFC", str(text or ""))
    nfkd = unicodedata.normalize("NFD", raw)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn").strip().upper()


def normalize_ibr_value(value: Any) -> float | None:
    """
    Normaliza IBR a decimal (ej. 10,580% -> 0.10580; 0.10580 -> 0.10580).
    Devuelve None si el valor no es interpretable.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        if n == 0:
            return 0.0
        if abs(n) > _PERCENT_THRESHOLD:
            return n / 100.0
        return n

    text = str(value).strip()
    if not text:
        return None

    has_percent = "%" in text
    cleaned = text.replace("%", "").strip()
    cleaned = cleaned.replace(",", ".")
    cleaned = re.sub(r"[^\d.\-]", "", cleaned)
    if not cleaned or cleaned in (".", "-"):
        return None
    try:
        n = float(cleaned)
    except ValueError:
        return None

    if has_percent or abs(n) > _PERCENT_THRESHOLD:
        return n / 100.0
    return n


def _parse_excel_date_value(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def find_ibr_for_date(workbook_bytes: bytes, target_date: date) -> float | None:
    """
    Busca en hoja IBR la fila donde Inicio <= target_date <= Fin.
    Devuelve Valor normalizado o None si no hay rango aplicable.
    """
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    try:
        if IBR_SHEET_NAME not in wb.sheetnames:
            return None
        ws = wb[IBR_SHEET_NAME]
        col_inicio = col_fin = col_valor = None
        for c in range(1, (ws.max_column or 1) + 1):
            h = _accent_fold_upper(str(ws.cell(1, c).value or ""))
            if h == _accent_fold_upper(IBR_COLUMNS[0]):
                col_inicio = c
            elif h == _accent_fold_upper(IBR_COLUMNS[1]):
                col_fin = c
            elif h == _accent_fold_upper(IBR_COLUMNS[2]):
                col_valor = c
        if not col_inicio or not col_fin or not col_valor:
            return None

        for r in range(2, (ws.max_row or 1) + 1):
            inicio = _parse_excel_date_value(ws.cell(r, col_inicio).value)
            fin = _parse_excel_date_value(ws.cell(r, col_fin).value)
            if inicio is None or fin is None:
                continue
            if inicio <= target_date <= fin:
                return normalize_ibr_value(ws.cell(r, col_valor).value)
        return None
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()
