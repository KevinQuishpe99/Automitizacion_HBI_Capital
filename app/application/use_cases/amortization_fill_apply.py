"""
Apply real: escribe tablas de amortización en SharePoint tras preflight (dry-run interno).
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Any

import httpx
import openpyxl

from app.application.services.accounting_pdf_parser import PaymentApplicationEvent
from app.application.services.amortization_apply_safety import (
    build_table_apply_summary,
    check_idempotency_against_log,
    collect_disallowed_warning_items,
    verify_uploaded_table,
)
from app.application.services.amortization_workbook import (
    ADOPTADO_EXISTENTE,
    APLICADO,
    REVISION_MANUAL,
    AmortizationSheetNotFoundError,
    PaymentApplicationWriteOptions,
    append_automation_log,
    compare_existing_application,
    detect_amortization_sheet,
    load_automation_log_index,
    write_ibr,
    write_payment_application,
)
from app.application.use_cases.amortization_fill_dry_run import (
    run_amortization_fill_dry_run,
    _drive_context,
)
from app.application.use_cases.validate_payment_report import (
    _graph_download_by_path,
    _graph_upload_by_path,
)
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

APPLY_STATUS_APPLIED = "APPLIED"
APPLY_STATUS_ADOPTED = "ADOPTED"
APPLY_STATUS_SKIPPED_IDEMPOTENT = "SKIPPED_IDEMPOTENT"
APPLY_STATUS_ERROR = "ERROR"

UPLOAD_STATUS_UPLOADED = "uploaded"
UPLOAD_STATUS_EXCEL_LOCKED = "EXCEL_LOCKED"
UPLOAD_STATUS_FAILED = "failed"
UPLOAD_STATUS_SKIPPED = "skipped"

VERIFICATION_OK = "ok"
VERIFICATION_FAILED = "POST_UPLOAD_VERIFICATION_FAILED"
VERIFICATION_FORMULA_FAILED = "POST_UPLOAD_VERIFICATION_FAILED_FORMULA_MISMATCH"
VERIFICATION_SKIPPED = "skipped"


class AmortizationPreflightError(ValueError):
    """Dry-run interno no cumple reglas de seguridad para apply."""

    def __init__(self, error_code: str, message: str, dry_run: dict[str, Any]) -> None:
        super().__init__(message)
        self.error_code = error_code
        self.dry_run = dry_run


class AmortizationApplySafetyError(ValueError):
    """La tabla cambió respecto al dry-run; no se sube el archivo."""

    def __init__(self, message: str, *, tabla_path: str, item: dict[str, Any]) -> None:
        super().__init__(message)
        self.tabla_path = tabla_path
        self.item = item


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def validate_amortization_preflight(dry_run: dict[str, Any]) -> None:
    summary = dry_run.get("summary") or {}
    if int(summary.get("errors") or 0) > 0:
        raise AmortizationPreflightError(
            "preflight_errors",
            "El dry-run interno reportó errores; no se escribe ninguna tabla.",
            dry_run,
        )
    if int(summary.get("revision_manual") or 0) > 0:
        raise AmortizationPreflightError(
            "preflight_revision_manual",
            "Hay filas en REVISION_MANUAL; no se escribe ninguna tabla.",
            dry_run,
        )
    blocked = collect_disallowed_warning_items(dry_run)
    if blocked:
        raise AmortizationPreflightError(
            "preflight_warnings_not_allowed",
            (
                f"Hay {len(blocked)} evento(s) con warnings no permitidos para apply; "
                "no se escribe ninguna tabla."
            ),
            dry_run,
        )


def _payment_date_from_item(
    item: dict[str, Any], dry_run: dict[str, Any] | None = None
) -> date | None:
    raw = str(item.get("payment_date_iso") or "").strip()
    if not raw and dry_run:
        raw = str(dry_run.get("report_date_iso") or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _write_options_from_item(
    item: dict[str, Any], dry_run: dict[str, Any] | None = None
) -> PaymentApplicationWriteOptions | None:
    payment_date = _payment_date_from_item(item, dry_run)
    if payment_date is None:
        return None
    return PaymentApplicationWriteOptions(
        payment_date=payment_date,
        detected_codes=frozenset(str(c) for c in (item.get("detected_codes") or [])),
        warnings=frozenset(str(w) for w in (item.get("warnings") or [])),
    )


def _event_from_planned_item(item: dict[str, Any]) -> PaymentApplicationEvent:
    pa = item.get("payment_application") or {}
    fecha_asiento: date | None = None
    raw_fa = item.get("fecha_asiento")
    if raw_fa:
        try:
            fecha_asiento = date.fromisoformat(str(raw_fa))
        except ValueError:
            fecha_asiento = None
    return PaymentApplicationEvent(
        id_pago=str(item.get("id_pago") or ""),
        cliente=str(item.get("cliente") or ""),
        credito=str(item.get("credito") or ""),
        asiento_pdf_path=str(item.get("asiento_pdf_path") or ""),
        comprobante=str(item.get("comprobante") or ""),
        fecha_asiento=fecha_asiento,
        valor_pagado_cliente=float(pa.get("valor_pagado_cliente") or 0),
        capital=float(pa.get("capital") or 0),
        intereses=float(pa.get("intereses") or 0),
        mora=float(pa.get("mora") or 0),
        retenciones=float(pa.get("retenciones") or 0),
        saldos_menores=float(pa.get("saldos_menores") or 0),
        raw_text="",
    )


def _resolve_worksheet(
    wb: openpyxl.Workbook, item: dict[str, Any], tabla_path: str
) -> tuple[Any, dict[str, int], int, str]:
    from app.application.services.amortization_workbook import detect_headers, find_header_row

    sheet_name = str(item.get("sheet_name") or "").strip()
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        headers = detect_headers(ws, header_row=header_row)
        return ws, headers, header_row, sheet_name
    match = detect_amortization_sheet(wb, tabla_amortizacion_path=tabla_path)
    return match.worksheet, match.headers, match.header_row, match.worksheet.title


def _apply_summarize(apply_items: list[dict[str, Any]]) -> dict[str, int]:
    summary = {
        "total": len(apply_items),
        "applied": 0,
        "adopted": 0,
        "skipped_idempotent": 0,
        "errors": 0,
    }
    for it in apply_items:
        st = it.get("apply_status")
        if st == APPLY_STATUS_APPLIED:
            summary["applied"] += 1
        elif st == APPLY_STATUS_ADOPTED:
            summary["adopted"] += 1
        elif st == APPLY_STATUS_SKIPPED_IDEMPOTENT:
            summary["skipped_idempotent"] += 1
        elif st == APPLY_STATUS_ERROR:
            summary["errors"] += 1
    return summary


def _writable_planned_items(dry_run: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    by_table: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in dry_run.get("items") or []:
        if not isinstance(item, dict):
            continue
        if item.get("error_code"):
            continue
        st = item.get("application_status")
        if st not in ("WOULD_APPLY", "WOULD_ADOPT_EXISTING"):
            continue
        path = str(item.get("tabla_amortizacion_path") or "").strip()
        if path:
            by_table[path].append(item)
    for path in by_table:
        by_table[path].sort(key=lambda x: int(x.get("event_index") or 0))
    return dict(by_table)


def _mark_table_items_error(
    results: list[dict[str, Any]],
    *,
    error_code: str,
    message: str,
) -> None:
    for row in results:
        if row.get("apply_status") in (APPLY_STATUS_APPLIED, APPLY_STATUS_ADOPTED):
            row["apply_status"] = APPLY_STATUS_ERROR
            row["apply_error_code"] = error_code
            row["apply_message"] = message


async def _apply_one_table(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    tabla_path: str,
    planned_items: list[dict[str, Any]],
    *,
    dry_run: dict[str, Any] | None = None,
) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    upload_status = UPLOAD_STATUS_SKIPPED
    verification_status = VERIFICATION_SKIPPED
    tabla_bytes = await _graph_download_by_path(graph, site_id, drive_id, tabla_path)
    wb = openpyxl.load_workbook(io.BytesIO(tabla_bytes), data_only=False)
    sheet_name = ""
    try:
        try:
            ws, headers, header_row, sheet_name = _resolve_worksheet(
                wb, planned_items[0], tabla_path
            )
        except AmortizationSheetNotFoundError as exc:
            for item in planned_items:
                results.append(
                    {
                        **item,
                        "apply_status": APPLY_STATUS_ERROR,
                        "apply_error_code": "AMORTIZATION_SHEET_NOT_FOUND",
                        "apply_message": str(exc),
                    }
                )
            return {
                "items": results,
                "uploaded": False,
                "tabla_path": tabla_path,
                "upload_status": UPLOAD_STATUS_FAILED,
                "verification_status": VERIFICATION_SKIPPED,
            }

        log_index = load_automation_log_index(wb)
        ibr_written_for_cut: set[str] = set()
        applied_for_verify: list[dict[str, Any]] = []

        for item in planned_items:
            base = {k: v for k, v in item.items() if not str(k).startswith("apply_")}
            idem_key = str(item.get("idempotency_key") or "").strip()
            application_row = item.get("application_row")
            ibr_row = item.get("ibr_row")
            app_status = item.get("application_status")

            idem_check = check_idempotency_against_log(item, log_index)
            if idem_check.pdf_changed:
                results.append(
                    {
                        **base,
                        "apply_status": APPLY_STATUS_ERROR,
                        "apply_error_code": "PDF_CHANGED_SAME_PATH",
                        "apply_message": idem_check.reason or "PDF cambió con la misma ruta",
                    }
                )
                continue
            if idem_check.skip_idempotent:
                results.append(
                    {
                        **base,
                        "apply_status": APPLY_STATUS_SKIPPED_IDEMPOTENT,
                        "apply_message": "Evento ya registrado en _AUTOMATION_LOG (misma huella)",
                    }
                )
                continue

            if application_row is None or ibr_row is None:
                results.append(
                    {
                        **base,
                        "apply_status": APPLY_STATUS_ERROR,
                        "apply_error_code": item.get("error_code") or "ROW_NOT_PLANNED",
                    }
                )
                continue

            event = _event_from_planned_item(item)
            write_opts = _write_options_from_item(item, dry_run)
            compare = compare_existing_application(
                ws,
                int(application_row),
                headers,
                event,
                payment_date=write_opts.payment_date if write_opts else None,
                detected_codes=write_opts.detected_codes if write_opts else None,
                warnings=write_opts.warnings if write_opts else None,
            )
            if compare == REVISION_MANUAL:
                raise AmortizationApplySafetyError(
                    f"Fila de aplicación {application_row} con valores distintos; apply abortado.",
                    tabla_path=tabla_path,
                    item=item,
                )

            accion_log = ADOPTADO_EXISTENTE
            apply_ibr_written = False
            write_plan: dict[str, str] = {}
            if app_status == "WOULD_APPLY":
                if compare != APLICADO:
                    raise AmortizationApplySafetyError(
                        f"Se esperaba fila vacía en {application_row}; estado={compare}.",
                        tabla_path=tabla_path,
                        item=item,
                    )
                if write_opts is not None:
                    write_plan = write_payment_application(
                        ws,
                        int(application_row),
                        headers,
                        event,
                        write_options=write_opts,
                        header_row=header_row,
                    )
                else:
                    write_payment_application(
                        ws, int(application_row), headers, event, header_row=header_row
                    )
                accion_log = APLICADO
                apply_status = APPLY_STATUS_APPLIED
            else:
                apply_status = APPLY_STATUS_ADOPTED
                write_plan = {}

            ibr_block = item.get("ibr") or {}
            ibr_status = ibr_block.get("status")
            fecha_limite = str(item.get("fecha_limite_pago") or "")
            ibr_cut_key = f"{int(ibr_row)}|{fecha_limite}"
            if ibr_status == "WOULD_WRITE_IBR" and ibr_cut_key not in ibr_written_for_cut:
                ibr_value = ibr_block.get("value")
                if ibr_value is not None:
                    write_ibr(ws, int(ibr_row), headers, float(ibr_value))
                    ibr_written_for_cut.add(ibr_cut_key)
                    apply_ibr_written = True

            append_automation_log(
                wb,
                {
                    "timestamp": _utc_now_iso(),
                    "id_pago": item.get("id_pago"),
                    "cliente": item.get("cliente"),
                    "credito": item.get("credito"),
                    "fila": application_row,
                    "application_row": application_row,
                    "ibr_row": ibr_row,
                    "accion": accion_log,
                    "estado": accion_log,
                    "detalle": f"apply|{app_status}|ibr={ibr_status}",
                    "idempotency_key": idem_key,
                    "asiento_pdf_path": item.get("asiento_pdf_path"),
                    "asiento_pdf_hash": item.get("asiento_pdf_hash"),
                    "asiento_pdf_etag": item.get("asiento_pdf_etag"),
                },
            )

            result_row = {
                **base,
                "apply_status": apply_status,
                "apply_accion": accion_log,
                "apply_ibr_written": apply_ibr_written,
                "sheet_name": sheet_name or base.get("sheet_name"),
                "write_plan": write_plan if apply_status == APPLY_STATUS_APPLIED else {},
            }
            results.append(result_row)
            if apply_status in (APPLY_STATUS_APPLIED, APPLY_STATUS_ADOPTED):
                applied_for_verify.append(result_row)

        if not applied_for_verify:
            return {
                "items": results,
                "uploaded": False,
                "tabla_path": tabla_path,
                "upload_status": UPLOAD_STATUS_SKIPPED,
                "verification_status": VERIFICATION_SKIPPED,
            }

        out_buf = io.BytesIO()
        wb.save(out_buf)
        content = out_buf.getvalue()

        try:
            await _graph_upload_by_path(
                graph, site_id, drive_id, tabla_path, content
            )
            upload_status = UPLOAD_STATUS_UPLOADED
        except httpx.HTTPStatusError as exc:
            if exc.response is not None and exc.response.status_code == 423:
                _mark_table_items_error(
                    results,
                    error_code="EXCEL_LOCKED",
                    message=str(exc)[:500],
                )
                return {
                    "items": results,
                    "uploaded": False,
                    "tabla_path": tabla_path,
                    "upload_status": UPLOAD_STATUS_EXCEL_LOCKED,
                    "verification_status": VERIFICATION_SKIPPED,
                }
            raise

        verify_bytes = await _graph_download_by_path(
            graph, site_id, drive_id, tabla_path
        )
        wb_verify_values = openpyxl.load_workbook(
            io.BytesIO(verify_bytes), data_only=True
        )
        wb_verify_formulas = openpyxl.load_workbook(
            io.BytesIO(verify_bytes), data_only=False
        )
        try:
            verify_errors = verify_uploaded_table(
                wb_verify_values,
                sheet_name=sheet_name,
                tabla_path=tabla_path,
                applied_items=applied_for_verify,
                wb_formulas=wb_verify_formulas,
                dry_run=dry_run,
            )
        finally:
            for wb_v in (wb_verify_values, wb_verify_formulas):
                closer = getattr(wb_v, "close", None)
                if callable(closer):
                    closer()

        if verify_errors:
            if any("FORMULA_MISMATCH" in e for e in verify_errors):
                verification_status = VERIFICATION_FORMULA_FAILED
            else:
                verification_status = VERIFICATION_FAILED
            _mark_table_items_error(
                results,
                error_code=verification_status,
                message="; ".join(verify_errors[:5]),
            )
        else:
            verification_status = VERIFICATION_OK

        return {
            "items": results,
            "uploaded": upload_status == UPLOAD_STATUS_UPLOADED,
            "tabla_path": tabla_path,
            "upload_status": upload_status,
            "verification_status": verification_status,
        }
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


async def run_amortization_fill_apply(
    graph: GraphApiPort,
    *,
    report_date_iso: str | None = None,
    merge_manifest_path: str | None = None,
    historical_file_path: str | None = None,
) -> dict[str, Any]:
    """
    Preflight (dry-run) y escritura real en tablas de amortización.
    """
    dry_run = await run_amortization_fill_dry_run(
        graph,
        report_date_iso=report_date_iso,
        merge_manifest_path=merge_manifest_path,
        historical_file_path=historical_file_path,
    )

    try:
        validate_amortization_preflight(dry_run)
    except AmortizationPreflightError as exc:
        return {
            "status": "preflight_failed",
            "mode": "apply",
            "preflight_error_code": exc.error_code,
            "message": str(exc),
            "preflight": dry_run,
            "items": [],
            "tables_uploaded": [],
            "tables_summary": [],
            "summary": _apply_summarize([]),
        }

    site_id, drive_id = await _drive_context(graph)
    by_table = _writable_planned_items(dry_run)
    apply_items: list[dict[str, Any]] = []
    tables_uploaded: list[str] = []
    tables_summary: list[dict[str, Any]] = []
    apply_errors: list[dict[str, Any]] = []

    for tabla_path, planned in by_table.items():
        try:
            table_result = await _apply_one_table(
                graph, site_id, drive_id, tabla_path, planned, dry_run=dry_run
            )
            apply_items.extend(table_result["items"])
            upload_status = str(table_result.get("upload_status") or UPLOAD_STATUS_SKIPPED)
            verification_status = str(
                table_result.get("verification_status") or VERIFICATION_SKIPPED
            )
            if table_result.get("uploaded"):
                tables_uploaded.append(tabla_path)
            tables_summary.append(
                build_table_apply_summary(
                    tabla_path,
                    table_result["items"],
                    upload_status=upload_status,
                    verification_status=verification_status,
                )
            )
            if verification_status in (
                VERIFICATION_FAILED,
                VERIFICATION_FORMULA_FAILED,
            ):
                apply_errors.append(
                    {
                        "tabla_amortizacion_path": tabla_path,
                        "error_code": verification_status,
                        "message": "Verificación post-upload falló",
                    }
                )
            if upload_status == UPLOAD_STATUS_EXCEL_LOCKED:
                apply_errors.append(
                    {
                        "tabla_amortizacion_path": tabla_path,
                        "error_code": "EXCEL_LOCKED",
                        "message": "SharePoint devolvió 423 Locked tras reintentos",
                    }
                )
        except AmortizationApplySafetyError as exc:
            logger.error("apply abortado tabla %s: %s", tabla_path, exc)
            apply_errors.append(
                {
                    "tabla_amortizacion_path": tabla_path,
                    "error_code": "APPLY_SAFETY_ABORT",
                    "message": str(exc),
                    "item": exc.item,
                }
            )
            for item in planned:
                apply_items.append(
                    {
                        **item,
                        "apply_status": APPLY_STATUS_ERROR,
                        "apply_error_code": "APPLY_SAFETY_ABORT",
                        "apply_message": str(exc),
                    }
                )
            tables_summary.append(
                build_table_apply_summary(
                    tabla_path,
                    apply_items[-len(planned) :],
                    upload_status=UPLOAD_STATUS_FAILED,
                    verification_status=VERIFICATION_SKIPPED,
                )
            )
        except Exception as exc:
            logger.exception("apply falló tabla %s", tabla_path)
            apply_errors.append(
                {
                    "tabla_amortizacion_path": tabla_path,
                    "error_code": "TABLE_APPLY_FAILED",
                    "message": str(exc)[:500],
                }
            )
            for item in planned:
                apply_items.append(
                    {
                        **item,
                        "apply_status": APPLY_STATUS_ERROR,
                        "apply_error_code": "TABLE_APPLY_FAILED",
                        "apply_message": str(exc)[:500],
                    }
                )
            tables_summary.append(
                build_table_apply_summary(
                    tabla_path,
                    apply_items[-len(planned) :],
                    upload_status=UPLOAD_STATUS_FAILED,
                    verification_status=VERIFICATION_SKIPPED,
                )
            )

    summary = _apply_summarize(apply_items)
    status = "ok"
    if apply_errors and not tables_uploaded:
        status = "failed"
    elif apply_errors:
        status = "partial"

    return {
        "status": status,
        "mode": "apply",
        "preflight": dry_run,
        "items": apply_items,
        "tables_uploaded": tables_uploaded,
        "tables_summary": tables_summary,
        "apply_errors": apply_errors,
        "summary": summary,
        "manifest_path": dry_run.get("manifest_path"),
        "historical_file_path": dry_run.get("historical_file_path"),
    }

