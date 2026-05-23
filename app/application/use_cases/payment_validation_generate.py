import io
import logging
import os
import re
import unicodedata
import uuid
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.application.services.payment_helpers import (
    _normalize_str,
    extract_client_from_bank_row,
    extract_credit_id_from_extract_pdf,
    extract_fecha_limite_pago_from_pdf,
    extract_total_a_pagar_from_pdf,
    filter_amortization_excel_filenames,
    find_best_amortization_table,
    parse_bank_amount,
    parse_bank_date,
    parse_statement_name,
)
from app.application.services.review_schema import (
    CasosPagoCols,
    ControlCols,
    DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS,
    DistribucionCols,
    ErroresCols,
    EstadoPago,
    ReviewSheets,
    ValidarPago,
    normalize_credito_digits,
)
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)
def _build_content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


def _normalize_header(value: Any) -> str:
    return _normalize_str(str(value or "")).replace(" ", "")


def _find_header_index(headers: list[Any], aliases: list[str]) -> int:
    normalized_headers = [_normalize_header(header) for header in headers]
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for index, header in enumerate(normalized_headers):
        if header in normalized_aliases:
            return index
    return -1


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_cell_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _parse_cell_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip():
        try:
            return parse_bank_amount(value)
        except ValueError:
            return None
    return None


def _score_sheet(worksheet: Any, client_name: str = "") -> int:
    score = 0
    ws_title = worksheet.title.lower()
    
    if client_name and _normalize_str(client_name) in _normalize_str(ws_title):
        score += 5
    if "amortiza" in ws_title or "tabla" in ws_title or "credito" in ws_title:
        score += 3
        
    for row in worksheet.iter_rows(min_row=1, max_row=15, values_only=True):
        if not any(row):
            continue
        row_str = " ".join([_normalize_str(str(v)) for v in row if v])
        for kw in ["cuota", "pagado", "dia", "mes", "ano", "fecha", "saldo", "kf", "ki"]:
            if kw in row_str:
                score += 1
    return score


def _extract_pending_installment(table_bytes: bytes, client_name: str = "") -> dict[str, Any]:
    if not table_bytes:
        raise ValueError("amortization_table_not_found")

    workbook = openpyxl.load_workbook(io.BytesIO(table_bytes), data_only=True)
    
    best_sheet = None
    best_score = -1
    for sheet_name in workbook.sheetnames:
        lower_name = sheet_name.lower()
        if "tasa" in lower_name or "hoja1" in lower_name or "hoja2" in lower_name:
            continue
            
        ws = workbook[sheet_name]
        score = _score_sheet(ws, client_name)
        if score > best_score:
            best_score = score
            best_sheet = ws
            
    if best_sheet is None or best_score == 0:
        raise ValueError("amortization_sheet_not_found")
        
    worksheet = best_sheet
    
    header_row_idx = -1
    headers = []
    best_header_score = -1
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if not any(row):
            continue
        texts = [_normalize_str(str(v)).replace(" ", "") for v in row if v is not None]
        row_score = 0
        for text in texts:
            if text in ["ki", "cuota", "kf", "dia", "mes", "ano", "fechapago", "fechadepago", "valorpagado", "valorpagadocliente", "saldoacapital", "interesesdemora", "fechalimite"]:
                row_score += 1
        
        if row_score > best_header_score and row_score >= 2:
            best_header_score = row_score
            header_row_idx = row_idx
            headers = [cell for cell in row]
            
    if header_row_idx == -1:
        raise ValueError("pending_installment_not_found")
        
    due_index = _find_header_index(headers, ["Fecha límite", "Fecha limite"])
    day_index = _find_header_index(headers, ["dia", "día"])
    month_index = _find_header_index(headers, ["mes"])
    year_index = _find_header_index(headers, ["año", "ano"])
    
    pay_index = _find_header_index(headers, ["Fecha pago", "Fecha de pago"])
    total_paid_index = _find_header_index(headers, ["Total pagado", "Valor pagado cliente"])
    
    extract_value_index = _find_header_index(headers, ["Cuota"])
    if extract_value_index == -1:
        extract_value_index = _find_header_index(headers, ["Valor extracto", "Valor cuota base", "Total a pagar", "Valor cuota"])

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        if not any(row):
            continue
            
        pay_value = row[pay_index] if pay_index >= 0 and pay_index < len(row) else None
        total_paid_value = row[total_paid_index] if total_paid_index >= 0 and total_paid_index < len(row) else None
        
        is_paid = False
        if pay_index != -1:
            is_paid = not _is_blank(pay_value)
        elif total_paid_index != -1:
            is_paid = not _is_blank(total_paid_value)
            
        if not is_paid:
            due_date = None
            if due_index != -1 and due_index < len(row):
                due_date = _parse_cell_date(row[due_index])
                
            if due_date is None and day_index != -1 and month_index != -1 and year_index != -1:
                try:
                    if not _is_blank(row[day_index]) and not _is_blank(row[month_index]) and not _is_blank(row[year_index]):
                        d = int(float(row[day_index]))
                        m = int(float(row[month_index]))
                        y = int(float(row[year_index]))
                        due_date = date(y, m, d)
                except (ValueError, TypeError):
                    pass
                    
            if due_date is None:
                continue
                
            value = None
            if extract_value_index != -1 and extract_value_index < len(row):
                value = _parse_cell_amount(row[extract_value_index])
                
            if value is None:
                continue
                
            return {
                "fecha_limite": due_date, 
                "valor_extracto": value,
                "row_number": row_idx,
                "sheet_name": worksheet.title,
                "header_row": header_row_idx
            }

    raise ValueError("pending_installment_not_found")


def _extract_last_payment_date_from_amortization(table_bytes: bytes, client_name: str = "") -> date | None:
    """
    Última fecha (máximo) encontrada en columna «Fecha pago» / «Fecha de pago» con valor parseable.
    None si no hay columna o no hay fechas fiables.
    """
    if not table_bytes:
        return None
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(table_bytes), data_only=True)
    except Exception:
        return None

    best_sheet = None
    best_score = -1
    for sheet_name in workbook.sheetnames:
        lower_name = sheet_name.lower()
        if "tasa" in lower_name or "hoja1" in lower_name or "hoja2" in lower_name:
            continue
        ws = workbook[sheet_name]
        score = _score_sheet(ws, client_name)
        if score > best_score:
            best_score = score
            best_sheet = ws

    if best_sheet is None or best_score == 0:
        return None

    worksheet = best_sheet
    header_row_idx = -1
    headers: list[Any] = []
    best_header_score = -1

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if not any(row):
            continue
        texts = [_normalize_str(str(v)).replace(" ", "") for v in row if v is not None]
        row_score = 0
        for text in texts:
            if text in [
                "ki",
                "cuota",
                "kf",
                "dia",
                "mes",
                "ano",
                "fechapago",
                "fechadepago",
                "valorpagado",
                "valorpagadocliente",
                "saldoacapital",
                "interesesdemora",
                "fechalimite",
            ]:
                row_score += 1
        if row_score > best_header_score and row_score >= 2:
            best_header_score = row_score
            header_row_idx = row_idx
            headers = [cell for cell in row]

    if header_row_idx == -1:
        return None

    pay_index = _find_header_index(headers, ["Fecha pago", "Fecha de pago"])
    if pay_index < 0:
        return None

    last_pay: date | None = None
    for row in worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
        if pay_index >= len(row):
            continue
        pay_value = row[pay_index]
        if _is_blank(pay_value):
            continue
        parsed = _parse_cell_date(pay_value)
        if parsed is None:
            continue
        if last_pay is None or parsed > last_pay:
            last_pay = parsed

    return last_pay


def _find_statement_item(items: list[dict[str, Any]], credit_id: str, due_date: date | None) -> dict[str, Any] | None:
    extract_items = [item for item in items if "extracto" in str(item.get("name", "")).lower()]
    for item in extract_items:
        parsed_date, parsed_credit = parse_statement_name(item.get("name", ""))
        if parsed_credit == str(credit_id) and (due_date is None or parsed_date == due_date):
            return item
    for item in extract_items:
        parsed_date, parsed_credit = parse_statement_name(item.get("name", ""))
        if parsed_credit == str(credit_id):
            return item
        if due_date is not None and parsed_date == due_date:
            return item
    return extract_items[0] if extract_items else None


def _use_extract_selection_v2() -> bool:
    """GENERATE_EXTRACT_SELECTION_V2: default true; false|0|no desactiva la selección por fecha en PDF."""
    v = os.getenv("GENERATE_EXTRACT_SELECTION_V2", "true").strip().lower()
    return v not in ("0", "false", "no", "off")


def _is_strict_extract_pdf_file_item(item: dict[str, Any]) -> bool:
    if "folder" in item:
        return False
    name = str(item.get("name", ""))
    if "extracto" not in name.lower():
        return False
    return name.lower().endswith(".pdf")


def _find_extractos_folder_item(items: list[dict[str, Any]]) -> dict[str, Any] | None:
    for it in items:
        if "folder" not in it:
            continue
        if str(it.get("name", "")).casefold() == "extractos":
            return it
    return None


_INFRA_FOLDER_LABELS = frozenset(
    {
        "extractos",
        "asientos contables",
        "asiento contables",
        "asientos contables generados",
        "email",
        "historico",
        "control",
    }
)

_TERMINAL_FOLDER_SAFE_TOKENS = frozenset({"vigente", "repuestos", "reestructuracion"})


def _normalize_folder_label(name: str) -> str:
    v = unicodedata.normalize("NFD", str(name or "").strip().lower())
    v = "".join(ch for ch in v if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", v)


def _is_infra_folder(folder_name: str) -> bool:
    return _normalize_folder_label(folder_name) in _INFRA_FOLDER_LABELS


def _looks_like_standard_credit_folder(folder_name: str) -> bool:
    n = _normalize_folder_label(folder_name)
    if not n:
        return False
    if "credito" in n or "obligacion" in n:
        return True
    if re.search(r"#\s*\d+", n):
        return True
    if re.search(r"\b\d{3,}\b", n):
        return True
    return False


def _root_has_strict_extract_pdfs(items: list[dict[str, Any]]) -> bool:
    return any(_is_strict_extract_pdf_file_item(it) for it in items if "folder" not in it)


async def _items_have_operational_signal(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder_path: str,
    items: list[dict[str, Any]],
    cliente_folder: str,
    credit_name: str,
) -> bool:
    if _root_has_strict_extract_pdfs(items):
        return True
    file_names = [str(it.get("name", "")) for it in items if it.get("name")]
    excel_only = filter_amortization_excel_filenames(file_names)
    if excel_only:
        try:
            find_best_amortization_table(excel_only, cliente_folder, credit_name)
            return True
        except ValueError:
            pass
    extractos_item = _find_extractos_folder_item(items)
    if extractos_item is not None:
        ex_name = str(extractos_item.get("name", "EXTRACTOS"))
        extractos_path = f"{folder_path}/{ex_name}".replace("//", "/")
        children = await _get_drive_folder_children(client, site_id, drive_id, extractos_path)
        if any(_is_strict_extract_pdf_file_item(it) for it in children):
            return True
    return False


def _resolve_credit_id_for_unit(
    credit_name: str,
    cliente_folder: str,
    *,
    is_flat_unit: bool,
    is_root_unit: bool,
    statement_filename: str,
    statement_bytes: bytes,
) -> tuple[str, bool]:
    """Devuelve (credit_id, is_non_standard_folder)."""
    _, parsed_credit_fn = parse_statement_name(statement_filename)
    inferred_pdf = extract_credit_id_from_extract_pdf(statement_bytes)
    is_non_standard = (
        not is_flat_unit
        and not is_root_unit
        and not _looks_like_standard_credit_folder(credit_name)
    )
    if is_flat_unit or is_root_unit:
        credit_id = inferred_pdf or parsed_credit_fn or cliente_folder
    elif is_non_standard:
        credit_id = inferred_pdf or parsed_credit_fn or credit_name
    else:
        credit_id = credit_name
    return str(credit_id), is_non_standard


def _norm_ruta_rel(ruta: str | None) -> str:
    return str(ruta or "").strip().replace("\\", "/").lower()


def _norm_ruta_extracto(ruta: str | None) -> str:
    return _norm_ruta_rel(ruta)


def _dedupe_credit_candidates(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen_ruta: set[str] = set()
    seen_secondary: set[tuple[Any, ...]] = set()
    out: list[dict[str, Any]] = []
    for cand in candidates:
        ruta = _norm_ruta_extracto(cand.get("ruta_extracto_pdf"))
        if ruta:
            if ruta in seen_ruta:
                continue
            seen_ruta.add(ruta)
            out.append(cand)
            continue
        fecha = cand.get("fecha_limite")
        fecha_key = fecha.isoformat() if hasattr(fecha, "isoformat") else str(fecha or "")
        valor = cand.get("valor_extracto")
        try:
            valor_key = round(float(valor), 2) if valor is not None else None
        except (TypeError, ValueError):
            valor_key = None
        link_key = str(cand.get("link_extracto") or "").strip().lower()
        sec_key = (
            str(cand.get("credito") or ""),
            fecha_key,
            valor_key,
            link_key,
        )
        if sec_key in seen_secondary:
            continue
        seen_secondary.add(sec_key)
        out.append(cand)
    return out


def _possibly_finalized_observation(credit_folder_name: str) -> str | None:
    n = _normalize_folder_label(credit_folder_name)
    for safe in _TERMINAL_FOLDER_SAFE_TOKENS:
        if re.search(rf"\b{re.escape(safe)}\b", n):
            return None
    for pattern in (
        r"\bterminad[oa]\b",
        r"\bfinalizad[oa]\b",
        r"\bcancelad[oa]\b",
        r"\bpagad[oa]\b",
        r"\bliquidad[oa]\b",
    ):
        if re.search(pattern, n):
            return _OBS_POSSIBLE_FINALIZED
    return None


_OBS_TABLA_NO_VERIFICAR_NOT_FOUND = (
    "No se pudo verificar contra tabla de amortización: tabla no encontrada."
)
_OBS_TABLA_NO_VERIFICAR_AMBIGUOUS = (
    "No se pudo verificar contra tabla de amortización: tabla ambigua."
)
_OBS_TABLA_NO_VERIFICAR_NOPARSE = (
    "No se pudo verificar contra tabla de amortización: no parseable."
)
_OBS_EXTRACTO_FECHA_VS_TABLA_IGUAL = (
    "Advertencia: la fecha límite del extracto coincide con la última fecha de pago "
    "registrada en la tabla de amortización."
)
_OBS_EXTRACTO_FECHA_VS_TABLA_ANTERIOR = (
    "Advertencia: la fecha límite del extracto es anterior a la última fecha de pago "
    "registrada en la tabla de amortización."
)
_OBS_ROOT_UNIT = (
    "Unidad de crédito detectada en la raíz del cliente; revisar estructura documental."
)
_OBS_NON_STANDARD_FOLDER = (
    "Carpeta de crédito detectada por contenido, pero el nombre no sigue el estándar CREDITO #."
)
_OBS_POSSIBLE_FINALIZED = (
    "Carpeta marcada como TERMINADO/FINALIZADO/CANCELADO/PAGADO/LIQUIDADO: "
    "revisar si el crédito sigue vigente."
)


async def _get_drive_folder_children(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder_path: str,
) -> list[dict[str, Any]]:
    enc = encode_graph_drive_path(folder_path)
    resp = await client.get(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children")
    return list(resp.get("value", []))


async def _resolve_extract_pdf_pool(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    credit_path: str,
    credit_folder_items: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], str]:
    """
    Prioridad EXTRACTOS (case-insensitive): si existe y hay PDFs extracto válidos, solo esos;
    si existe pero sin candidatos válidos, fallback a la carpeta del crédito.
    """
    strict_in_credit = [it for it in credit_folder_items if _is_strict_extract_pdf_file_item(it)]
    extractos_item = _find_extractos_folder_item(credit_folder_items)
    if extractos_item is None:
        return strict_in_credit, credit_path
    ex_name = str(extractos_item.get("name", "EXTRACTOS"))
    extractos_path = f"{credit_path}/{ex_name}"
    children = await _get_drive_folder_children(client, site_id, drive_id, extractos_path)
    strict_in_extractos = [it for it in children if _is_strict_extract_pdf_file_item(it)]
    if strict_in_extractos:
        return strict_in_extractos, extractos_path
    return strict_in_credit, credit_path


async def _select_extract_by_max_fecha_limite_v2(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    extract_base_path: str,
    pool: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, bytes | None, date | None, str | None]:
    """
    Selecciona un único PDF por fecha límite de pago máxima leída del contenido.
    Retorna (item, bytes_del_pdf, fecha_limite_pdf, código_error).
    """
    if not pool:
        return None, None, None, "extract_not_found"
    scored: list[tuple[dict[str, Any], date, bytes]] = []
    for item in pool:
        name = item.get("name", "")
        fpath = f"{extract_base_path}/{name}"
        try:
            pdf_bytes = await client.get_bytes(
                _build_content_endpoint(site_id, drive_id, fpath)
            )
        except Exception:
            logger.debug("extract_candidate_skip download failed name=%s", name, exc_info=True)
            continue
        fe = extract_fecha_limite_pago_from_pdf(pdf_bytes)
        if fe is None:
            logger.debug(
                "extract_candidate_skip fecha_limite_not_readable drive_path=%s", fpath
            )
        if fe is not None:
            scored.append((item, fe, pdf_bytes))
    if not scored:
        return None, None, None, "fecha_limite_extracto_not_readable"
    max_d = max(t[1] for t in scored)
    winners = [(it, dt, b) for it, dt, b in scored if dt == max_d]
    if len(winners) > 1:
        return None, None, None, "extract_tie_max_fecha_limite"
    it, dt, b = winners[0]
    return it, b, dt, None


def _item_link(item: dict[str, Any], fallback_path: str) -> str:
    return item.get("webUrl") or fallback_path


def _http_url_only(raw: Any) -> str:
    s = str(raw or "").strip()
    return s if s.lower().startswith("http") else ""


DIST_MONEY_COL_WIDTH = 16.0
CASOS_OBS_COL_WIDTH = 48.0
CASOS_OBS_COL_CAP_WIDTH = 100.0


def _distrib_freeze_panes_cell(first_data_row: int) -> str:
    """Congela ID Pago, Cliente y Crédito; el scroll horizontal empieza en Monto banco."""
    col_credito = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    return f"{get_column_letter(col_credito + 1)}{first_data_row}"


def _link_label_suffix(credito: Any, cliente: Any) -> str:
    """Sufijo descriptivo: 'crédito 264' si es numérico; si no, identificador (p. ej. ACIMOR)."""
    cred = str(credito or "").strip()
    cli = str(cliente or "").strip()
    if cred and re.fullmatch(r"\d+", cred):
        return f"crédito {cred}"
    if cred:
        return cred
    return cli


def _format_distrib_link_visible_text(link_kind: str, credito: Any, cliente: Any) -> str:
    suffix = _link_label_suffix(credito, cliente)
    if link_kind == "extracto":
        return f"📄 Ver extracto {suffix}".strip() if suffix else "📄 Ver extracto"
    if link_kind == "tabla":
        return f"📄 Ver tabla {suffix}".strip() if suffix else "📄 Ver tabla"
    if link_kind == "carpeta":
        return f"Ver carpeta {suffix}".strip() if suffix else "Ver carpeta"
    return ""


def _format_errores_link_visible_text(link_kind: str, credito: Any, cliente: Any) -> str:
    suffix = _link_label_suffix(credito, cliente)
    if link_kind == "extracto":
        return f"Ver extracto {suffix}".strip() if suffix else "Ver extracto"
    if link_kind == "carpeta":
        return f"Ver carpeta {suffix}".strip() if suffix else "Ver carpeta"
    return ""


def _item_link_url(item: dict[str, Any], fallback_path: str) -> str:
    return _http_url_only(_item_link(item, fallback_path))


async def _fetch_folder_web_url(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder_path: str,
) -> str | None:
    enc = encode_graph_drive_path(folder_path)
    try:
        resp = await client.get(
            f"/sites/{site_id}/drives/{drive_id}/root:/{enc}",
            params={"$select": "webUrl"},
        )
        if isinstance(resp, dict):
            web = resp.get("webUrl")
            if web and str(web).strip():
                return str(web).strip()
    except Exception:
        logger.debug(
            "fetch_folder_web_url failed path=%s", folder_path, exc_info=True
        )
    return None


async def _resolve_client_folder_web_url(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    clients_path: str,
    cliente_folder: str,
) -> str | None:
    """webUrl de la carpeta del cliente: primero desde listado padre, luego GET del ítem."""
    cliente_folder_path = f"{clients_path}/{cliente_folder}"
    try:
        enc_parent = encode_graph_drive_path(clients_path)
        resp = await client.get(
            f"/sites/{site_id}/drives/{drive_id}/root:/{enc_parent}:/children"
        )
        for it in resp.get("value") or []:
            if str(it.get("name", "")).strip() == cliente_folder:
                web = it.get("webUrl")
                if web and str(web).strip():
                    return str(web).strip()
    except Exception:
        logger.debug(
            "resolve_client_folder_web_url: parent children failed cliente=%s",
            cliente_folder,
            exc_info=True,
        )
    return await _fetch_folder_web_url(client, site_id, drive_id, cliente_folder_path)


def _carpeta_link_url_for_errores(
    credit_folder_drive_item: dict[str, Any] | None,
    unit_path: str,
    *,
    client_folder_web_url: str | None,
    is_flat_unit: bool,
    is_root_unit: bool,
) -> str:
    """Solo URL http(s) para hipervínculo en Errores; vacío si Graph no devolvió webUrl."""
    if is_flat_unit or is_root_unit:
        return _http_url_only(client_folder_web_url)
    return _http_url_only((credit_folder_drive_item or {}).get("webUrl"))


def _errores_row_meta(code: str) -> tuple[str, str, str, str]:
    return _ERRORES_GUIDE_BY_CODE.get(code, _ERRORES_GUIDE_FALLBACK)


def _normalized_error_record_to_sheet_row(rec: dict[str, Any]) -> list[Any]:
    code = str(rec.get("code") or "").strip()
    tipo, descr, hacer, soporte = _errores_row_meta(code)
    values_by_col = {
        ErroresCols.ID_PAGO: rec.get("id_pago"),
        ErroresCols.CLIENTE: rec.get("cliente"),
        ErroresCols.CREDITO: rec.get("credito"),
        ErroresCols.TIPO_CASO: tipo,
        ErroresCols.DESCRIPCION: descr,
        ErroresCols.QUE_DEBE_HACER: hacer,
        ErroresCols.REQUIERE_SOPORTE: soporte,
        ErroresCols.LINK_EXTRACTO: "",
        ErroresCols.LINK_CARPETA_CREDITO: "",
        ErroresCols.CODIGO_TECNICO: code,
    }
    return [values_by_col[c] for c in ErroresCols.HEADERS]


def _round_money(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value), 2)


def _is_close(left: float | None, right: float | None) -> bool:
    if left is None or right is None:
        return False
    return abs(float(left) - float(right)) < 0.01


def _build_distribution_rows(payment: dict[str, Any], candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for candidate in candidates:
        due_date = candidate.get("fecha_limite")
        extract_value = _round_money(candidate.get("valor_extracto"))
        line = {
            DistribucionCols.ID_PAGO: payment["id_pago"],
            DistribucionCols.CLIENTE: payment["cliente"],
            DistribucionCols.MONTO_BANCO: payment["monto_banco"],
            DistribucionCols.CREDITO: candidate["credito"],
            DistribucionCols.FECHA_BANCO: payment["fecha_banco"].isoformat(),
            DistribucionCols.FECHA_LIMITE: due_date.isoformat() if due_date else "",
            DistribucionCols.DIAS_MORA: 0,
            DistribucionCols.VALOR_EXTRACTO: extract_value,
            DistribucionCols.VALOR_INTERESES: "",
            DistribucionCols.ABONO_K: "",
            DistribucionCols.INTERESES_MORA: "",
            DistribucionCols.TOTAL_APLICADO: "",
            DistribucionCols.SALDO_POR_ASIGNAR: "",
            DistribucionCols.ESTADO_PAGO: EstadoPago.REVISION_MANUAL,
            DistribucionCols.VALIDAR_PAGO: ValidarPago.NO,
            DistribucionCols.OBSERVACION: "Información incompleta para aplicar automáticamente",
            DistribucionCols.LINK_EXTRACTO: candidate.get("link_extracto", ""),
            DistribucionCols.LINK_TABLA: candidate.get("link_tabla", ""),
            DistribucionCols.LINK_CARPETA_CREDITO: candidate.get("link_carpeta_credito", ""),
            DistribucionCols.RUTA: candidate.get("ruta_extracto_pdf") or "",
            DistribucionCols.RUTA_UNIDAD_CREDITO: candidate.get("ruta_unidad_credito") or "",
            DistribucionCols.RUTA_TABLA_AMORTIZACION: candidate.get("ruta_tabla_amortizacion") or "",
            DistribucionCols.CREDITO_NORMALIZADO: candidate.get("credito_normalizado") or "",
        }

        def finalize_line(row: dict[str, Any]) -> None:
            obs_extra = candidate.get("observacion_extra")
            if obs_extra:
                cur = row.get(DistribucionCols.OBSERVACION) or ""
                row[DistribucionCols.OBSERVACION] = f"{cur} | {obs_extra}" if cur else obs_extra
            rows.append(row)

        if due_date is None:
            finalize_line(line)
            continue

        if payment["fecha_banco"] > due_date:
            line[DistribucionCols.DIAS_MORA] = (payment["fecha_banco"] - due_date).days
            line[DistribucionCols.ESTADO_PAGO] = EstadoPago.ATRASADO
            line[DistribucionCols.VALIDAR_PAGO] = ValidarPago.SI
            line[DistribucionCols.OBSERVACION] = "Confirmar mora por crédito"
            finalize_line(line)
            continue

        if payment["fecha_banco"] < due_date:
            line[DistribucionCols.ESTADO_PAGO] = EstadoPago.ADELANTADO
            line[DistribucionCols.VALIDAR_PAGO] = ValidarPago.SI
            line[DistribucionCols.OBSERVACION] = "Pago adelantado; requiere reprogramación"
            finalize_line(line)
            continue

        if extract_value is not None:
            line[DistribucionCols.ESTADO_PAGO] = EstadoPago.NORMAL
            line[DistribucionCols.VALIDAR_PAGO] = ValidarPago.SI
            line[DistribucionCols.OBSERVACION] = "Listo para validar"
            finalize_line(line)
            continue

        finalize_line(line)

    return rows


def _build_case_row(payment: dict[str, Any], distribution_rows: list[dict[str, Any]]) -> dict[str, Any]:
    observaciones = []
    for row in distribution_rows:
        observation = row.get(DistribucionCols.OBSERVACION)
        if observation and observation not in observaciones:
            observaciones.append(observation)

    return {
        CasosPagoCols.ID_PAGO: payment["id_pago"],
        CasosPagoCols.FECHA_BANCO: payment["fecha_banco"].isoformat(),
        CasosPagoCols.CLIENTE: payment["cliente"],
        CasosPagoCols.CONCEPTO_BANCO: payment["concepto"],
        CasosPagoCols.MONTO_BANCO: payment["monto_banco"],
        CasosPagoCols.OBSERVACION: " | ".join(observaciones),
    }


# Filas 1–2: bloque título; fila 3: encabezados de tabla o banda de sección (como referencia visual Claude)
_SHEET_BANNER_ROWS = 2

CONTROL_TITLE = "VALIDACIÓN DE PAGOS"
CONTROL_SUBTITLE = "Panel de control del proceso"
CONTROL_HELP = (
    "Revise el archivo, complete la distribución en la hoja «Distribucion» y cambie Procesar a SI cuando esté listo."
)
RESUMEN_TITLE = "RESUMEN DEL PROCESO"
RESUMEN_SUBTITLE = "Vista ejecutiva del lote de validación"
RESUMEN_SECTION = "MÉTRICAS GENERALES"
DISTRIB_TITLE = "DISTRIBUCIÓN DE PAGOS — HOJA PRINCIPAL"
DISTRIB_HELP = (
    "Complete únicamente las celdas editables: Aplicar a extracto, Mora a aplicar, Otros valores, "
    "Estado Pago y Validar Pago. "
    "Revise los links si necesita validar documentos. Al terminar, vaya a Control y cambie Procesar a SI."
)
CASOS_TITLE = "CASOS DE PAGO"
CASOS_SUBTITLE = (
    "Consulte aquí el resumen de pagos detectados. Esta hoja es informativa; "
    "complete la distribución en la hoja Distribución."
)
ERRORES_HELP = (
    "Revise estos casos manualmente. Use la descripción, la acción recomendada y los links "
    "para corregir documentos o carpetas antes de volver a generar."
)
ERRORES_TITLE = "REGISTRO DE ERRORES"
ERRORES_OK_MSG = "Sin errores técnicos detectados en este proceso."
ERRORES_BAD_MSG = (
    "Se registraron casos que requieren acción en las filas siguientes. "
    "Consulte «Descripción para revisión» y «Qué debe hacer»."
)

_ERRORES_GUIDE_FALLBACK: tuple[str, str, str, str] = (
    "Revisión",
    "El caso requiere revisión manual.",
    "Revise la información del cliente y, si no puede corregirlo, comuníquelo al equipo encargado.",
    "SI, si persiste",
)
_ERRORES_GUIDE_BY_CODE: dict[str, tuple[str, str, str, str]] = {
    "fecha_limite_extracto_not_readable": (
        "Extracto",
        "No se pudo leer la fecha límite de pago del extracto del crédito.",
        "Revise el PDF del extracto. Si está escaneado, borroso o no contiene texto seleccionable, "
        "cargue un extracto válido en la carpeta EXTRACTOS y vuelva a ejecutar la generación.",
        "SI, si persiste",
    ),
    "extract_not_found": (
        "Extracto",
        "No se encontró un PDF de extracto para este crédito.",
        "Cargue el extracto correspondiente en la carpeta EXTRACTOS del crédito, "
        "o en la carpeta del crédito si todavía no existe EXTRACTOS. "
        "Luego vuelva a ejecutar la generación.",
        "NO",
    ),
    "extract_tie_max_fecha_limite": (
        "Extracto",
        "Hay más de un extracto con la misma fecha límite máxima y el sistema no puede escoger uno automáticamente.",
        "Revise los extractos del crédito y deje únicamente el extracto correcto, "
        "o mueva los duplicados a una carpeta de respaldo. Luego vuelva a ejecutar la generación.",
        "NO",
    ),
    "credit_folder_not_found": (
        "Crédito",
        "No se pudo identificar una unidad de crédito válida para este cliente.",
        "Verifique que el cliente tenga una carpeta de crédito, una carpeta EXTRACTOS o extractos válidos "
        "en la raíz del cliente. Si la estructura no corresponde al estándar, comuníquelo al equipo encargado.",
        "SI, si estructura no estándar",
    ),
    "amortization_table_not_found": (
        "Tabla de amortización",
        "No se encontró una tabla de amortización Excel para este crédito.",
        "Verifique que exista un archivo Excel de tabla de amortización en la carpeta del crédito o del cliente.",
        "SI, si persiste",
    ),
    "amortization_table_ambiguous": (
        "Tabla de amortización",
        "Se encontraron varias tablas de amortización posibles y no se pudo elegir una de forma segura.",
        "Revise la carpeta y deje identificada claramente la tabla de amortización vigente.",
        "SI, si persiste",
    ),
    "extract_amount_not_found": (
        "Extracto",
        "No se pudo leer el monto «Total a pagar» en el PDF del extracto.",
        "Revise el PDF (texto seleccionable, sin cortes); cargue un extracto válido si es necesario "
        "y vuelva a ejecutar la generación.",
        "SI, si persiste",
    ),
}

_FILL_NAVY = PatternFill(fill_type="solid", fgColor="002060")
_FILL_HEADER_STRONG = _FILL_NAVY
_FILL_HEADER_SOFT = PatternFill(fill_type="solid", fgColor="5B9BD5")
_FILL_SECTION_BAND = PatternFill(fill_type="solid", fgColor="D9E2F3")
_FILL_CASOS_SUBBAR = PatternFill(fill_type="solid", fgColor="E7E8ED")
_FILL_PANEL = PatternFill(fill_type="solid", fgColor="F5F6F8")
_FILL_CONTROL_LABEL = PatternFill(fill_type="solid", fgColor="E8EDF2")
_FILL_DISTRIB_HINT = PatternFill(fill_type="solid", fgColor="E6EEF5")
_FILL_ZEBRA = PatternFill(fill_type="solid", fgColor="F9FAFB")
_FILL_SALDO_COL = PatternFill(fill_type="solid", fgColor="FFF9E6")
_FILL_DIAS_MORA_WARN = PatternFill(fill_type="solid", fgColor="FDE9D9")
_FILL_STATUS_OK = PatternFill(fill_type="solid", fgColor="D4EDDA")
_FILL_STATUS_BAD = PatternFill(fill_type="solid", fgColor="F8D7DA")
_FILL_GROUP_A = PatternFill(fill_type="solid", fgColor="FCFCFD")
_FILL_GROUP_B = PatternFill(fill_type="solid", fgColor="F6F8FA")
_FILL_EDITABLE_COL = PatternFill(fill_type="solid", fgColor="E2F4E8")
_FILL_SHEET_INSTRUCTION = PatternFill(fill_type="solid", fgColor="E8F2FA")
_TAB_COLOR_DISTRIB = "FF00B050"
_TAB_COLOR_ERRORES = "FFFF6969"
_FILL_CONTROL_PROCESAR_ROW = PatternFill(fill_type="solid", fgColor="E8EEF5")
_FILL_PROCESAR_SI = PatternFill(fill_type="solid", fgColor="FF32CD32")
_FILL_PROCESAR_NO = PatternFill(fill_type="solid", fgColor="FFFF6969")
_FILL_ESTADO_VALIDAR = PatternFill(fill_type="solid", fgColor="C8E6C9")
_FILL_ESTADO_VALIDAR_PARCIAL = PatternFill(fill_type="solid", fgColor="BBDEFB")
_FILL_ESTADO_PENDIENTE_MORA = PatternFill(fill_type="solid", fgColor="FFF9C4")
_FILL_ESTADO_REPROGRAMAR = PatternFill(fill_type="solid", fgColor="FFE0B2")
_FILL_ESTADO_NO_VALIDAR = PatternFill(fill_type="solid", fgColor="FFCDD2")
_FILL_ESTADO_REVISION_MANUAL = PatternFill(fill_type="solid", fgColor="E1BEE7")
_FILL_OBS_WARNING = PatternFill(fill_type="solid", fgColor="FFF9ED")
_FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_TITLE_NAVY = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
_FONT_SUB_NAVY = Font(name="Calibri", size=11, color="FFFFFF")
_FONT_SECTION = Font(name="Calibri", bold=True, size=11, color="002060")
_FONT_TITLE = Font(name="Calibri", bold=True, size=20, color="1A2F36")
_FONT_SUBTITLE = Font(name="Calibri", size=11, color="3D4F5C", italic=True)
_FONT_BODY = Font(name="Calibri", size=11)
_FONT_LABEL_BOLD = Font(name="Calibri", bold=True, size=11, color="000000")
_FONT_ESTADO_CTRL = Font(name="Calibri", size=11, color="2F5597")
_FONT_PROCESAR_LABEL = Font(name="Calibri", bold=True, size=14, color="002060")
_FONT_PROCESAR_VALUE = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
_FONT_SHEET_INSTRUCTION = Font(name="Calibri", size=12, color="1A2F36")
_THIN = Side(style="thin", color="D0D5DD")
_MEDIUM_CLIENT_EDGE = Side(style="medium", color="002060")
_BORDER_PROCESAR_VALUE = Border(
    left=Side(style="medium", color="002060"),
    right=Side(style="medium", color="002060"),
    top=Side(style="medium", color="002060"),
    bottom=Side(style="medium", color="002060"),
)
_BORDER_LIGHT = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALIGN_PROCESAR_LABEL = Alignment(vertical="center", horizontal="left")
_ALIGN_PROCESAR_VALUE = Alignment(vertical="center", horizontal="center")


def _distrib_row_border(*, client_top: bool = False, client_bottom: bool = False) -> Border:
    top = _MEDIUM_CLIENT_EDGE if client_top else _THIN
    bottom = _MEDIUM_CLIENT_EDGE if client_bottom else _THIN
    return Border(left=_THIN, right=_THIN, top=top, bottom=bottom)


def _apply_sheet_row2_instruction(ws: Any, ncols: int, text: str, *, height: float = 48.0) -> None:
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    cell = ws.cell(row=2, column=1, value=text)
    cell.fill = _FILL_SHEET_INSTRUCTION
    cell.font = _FONT_SHEET_INSTRUCTION
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = height


def _sheet_client_block_edges(
    ws: Any,
    first_data_row: int,
    last_row: int,
    col_cliente: int,
    col_row_key: int,
) -> dict[int, tuple[bool, bool]]:
    """Por fila de datos: (borde superior de bloque, borde inferior de bloque)."""
    indexed: list[tuple[int, str]] = []
    for r in range(first_data_row, last_row + 1):
        cliente = str(ws.cell(row=r, column=col_cliente).value or "").strip()
        row_key = ws.cell(row=r, column=col_row_key).value
        if not cliente and (row_key is None or str(row_key).strip() == ""):
            continue
        indexed.append((r, cliente))
    edges: dict[int, tuple[bool, bool]] = {}
    for i, (r, cliente) in enumerate(indexed):
        prev_cliente = indexed[i - 1][1] if i > 0 else None
        next_cliente = indexed[i + 1][1] if i + 1 < len(indexed) else None
        client_top = i > 0 and bool(cliente) and cliente != prev_cliente
        client_bottom = bool(cliente) and (next_cliente is None or cliente != next_cliente)
        edges[r] = (client_top, client_bottom)
    return edges
_ALIGN_WRAP = Alignment(vertical="center", horizontal="left", wrap_text=True)
_ALIGN_VCENTER = Alignment(vertical="center", horizontal="left")
_ALIGN_RIGHT = Alignment(vertical="center", horizontal="right")
_ALIGN_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
# Código OOXML/Excel estándar: Excel aplica miles/decimales según configuración regional
# (p. ej. en español se ve 1.234.567,89). No usar «#.##0,00» en el archivo: en muchos
# entornos se interpreta mal y aparece «,000» u otros artefactos.
_FMT_MONEY = "#,##0.00"
_FMT_DATE = "yyyy-mm-dd"
_HLINK_FONT = Font(name="Calibri", color="0563C1", size=11, underline="single")
_GROUP_BAND_FILLS = (_FILL_GROUP_A, _FILL_GROUP_B)
_ESTADO_PAGO_FILLS = {
    EstadoPago.NORMAL: _FILL_ESTADO_VALIDAR,
    EstadoPago.INCOMPLETO: _FILL_ESTADO_VALIDAR_PARCIAL,
    EstadoPago.ATRASADO: _FILL_ESTADO_PENDIENTE_MORA,
    EstadoPago.ADELANTADO: _FILL_ESTADO_REPROGRAMAR,
    EstadoPago.REVISION_MANUAL: _FILL_ESTADO_REVISION_MANUAL,
}

_TAB_COLOR_CONTROL = "FF002060"
_TAB_COLOR_RESUMEN = "FF4472C4"
_TAB_COLOR_CASOS = "FF595959"
_TAB_COLOR_LISTAS = "FFB4B4B4"


def _table_header_row() -> int:
    return _SHEET_BANNER_ROWS + 1


def _table_first_data_row() -> int:
    return _table_header_row() + 1


def _merge_navy_title_row(ws: Any, row: int, end_column: int, text: str, font: Font) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.fill = _FILL_NAVY
    cell.alignment = _ALIGN_CENTER


def _apply_tab_colors(workbook: Any) -> None:
    mapping = {
        ReviewSheets.CONTROL: _TAB_COLOR_CONTROL,
        ReviewSheets.RESUMEN: _TAB_COLOR_RESUMEN,
        ReviewSheets.CASOS_PAGO: _TAB_COLOR_CASOS,
        ReviewSheets.DISTRIBUCION: _TAB_COLOR_DISTRIB,
        ReviewSheets.ERRORES: _TAB_COLOR_ERRORES,
        ReviewSheets.LISTAS: _TAB_COLOR_LISTAS,
    }
    for name, rgb in mapping.items():
        if name in workbook.sheetnames:
            workbook[name].sheet_properties.tabColor = Color(rgb=rgb)


def _apply_casos_top_banner(ws: Any) -> None:
    n = len(CasosPagoCols.HEADERS)
    _merge_navy_title_row(ws, 1, n, CASOS_TITLE, _FONT_TITLE_NAVY)
    ws.row_dimensions[1].height = 32
    _apply_sheet_row2_instruction(ws, n, CASOS_SUBTITLE, height=46.0)


def _apply_distrib_top_banner(ws: Any) -> None:
    endc = len(DistribucionCols.HEADERS)
    _merge_navy_title_row(ws, 1, endc, DISTRIB_TITLE, _FONT_TITLE_NAVY)
    ws.row_dimensions[1].height = 34
    _apply_sheet_row2_instruction(ws, endc, DISTRIB_HELP, height=52.0)


def _apply_errores_top_banner(ws: Any, ncols: int, has_errors: bool) -> None:
    _merge_navy_title_row(ws, 1, ncols, ERRORES_TITLE, _FONT_TITLE_NAVY)
    ws.row_dimensions[1].height = 32
    if has_errors:
        _apply_sheet_row2_instruction(ws, ncols, ERRORES_HELP, height=50.0)
    else:
        _apply_sheet_row2_instruction(ws, ncols, ERRORES_OK_MSG, height=36.0)


def _write_list_values(ws_lists: Any) -> None:
    for row_idx, value in enumerate(ControlCols.ESTADO_OPTIONS, start=1):
        ws_lists.cell(row=row_idx, column=1, value=value)
    for row_idx, value in enumerate(ControlCols.PROCESAR_OPTIONS, start=1):
        ws_lists.cell(row=row_idx, column=2, value=value)
    for row_idx, value in enumerate(EstadoPago.OPTIONS_ORDERED, start=1):
        ws_lists.cell(row=row_idx, column=3, value=value)
    for row_idx, value in enumerate((ValidarPago.SI, ValidarPago.NO), start=1):
        ws_lists.cell(row=row_idx, column=4, value=value)


def _sheet_hide_gridlines(ws: Any) -> None:
    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass


def _protect_sheet(ws: Any) -> None:
    # Perfil mínimo para Excel Online/SharePoint: solo activar protección de hoja.
    ws.protection.sheet = True


def _set_cell_locked(ws: Any, row: int, col: int, locked: bool) -> None:
    ws.cell(row=row, column=col).protection = Protection(locked=locked)


def _configure_distrib_technical_path_columns(ws_distribution: Any) -> None:
    """Oculta columnas técnicas de amortización (RutaTablaAmortizacion, CreditoNormalizado)."""
    for col_name in DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS:
        try:
            cidx = DistribucionCols.HEADERS.index(col_name) + 1
            letter = get_column_letter(cidx)
            wd = ws_distribution.column_dimensions[letter]
            wd.hidden = True
            wd.width = min(float(wd.width or 9.0), 12.0)
        except Exception:
            logger.debug(
                "configure_distrib_technical_path_columns skip col=%s", col_name, exc_info=True
            )


def _protect_control_sheet(ws_control: Any, procesar_cell: str, estado_cell: str) -> None:
    # Desbloquear solo Procesar (editable). Estado queda bloqueado/informativo aunque tenga dropdown.
    ws_control[procesar_cell].protection = Protection(locked=False)
    ws_control[estado_cell].protection = Protection(locked=True)
    _protect_sheet(ws_control)


def _protect_distribution_sheet(ws_distribution: Any, first_data_row: int) -> None:
    ncols = len(DistribucionCols.HEADERS)
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        _protect_sheet(ws_distribution)
        return

    editable_names = {
        DistribucionCols.APLICAR_A_EXTRACTO,
        DistribucionCols.MORA_A_APLICAR,
        DistribucionCols.OTROS_VALORES,
        DistribucionCols.ESTADO_PAGO,
        DistribucionCols.VALIDAR_PAGO,
        DistribucionCols.OBSERVACION,
    }
    editable_cols = {DistribucionCols.HEADERS.index(name) + 1 for name in editable_names}
    for r in range(first_data_row, last_row + 1):
        for c in editable_cols:
            _set_cell_locked(ws_distribution, r, c, locked=False)
    _protect_sheet(ws_distribution)


def _apply_column_widths(ws: Any, widths: dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _visible_cell_char_len(value: Any) -> int:
    if value is None:
        return 0
    s = str(value)
    if s.startswith("="):
        return min(42, len(s))
    return len(s)


def _auto_fit_columns(
    ws: Any,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    floor_w: float = 9.0,
    cap_w: float = 56.0,
) -> None:
    """Aproxima ancho de columna según el contenido (openpyxl no tiene AutoFit nativo)."""
    for col in range(min_col, max_col + 1):
        best = floor_w
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            ln = _visible_cell_char_len(cell.value)
            est = ln * 1.12 + 2.5
            if est > best:
                best = est
        letter = get_column_letter(col)
        best = min(cap_w, max(floor_w, best))
        cur = ws.column_dimensions[letter].width
        if cur is None or cur < best:
            ws.column_dimensions[letter].width = best


def _fit_single_column_width_from_content(
    ws: Any,
    col: int,
    min_row: int,
    max_row: int,
    *,
    floor_w: float,
    cap_w: float,
) -> float:
    """Ajusta una columna al texto más largo (p. ej. Observación en Casos_Pago)."""
    best = floor_w
    for row in range(min_row, max_row + 1):
        raw = ws.cell(row=row, column=col).value
        if raw is None:
            continue
        for line in str(raw).splitlines():
            est = len(line) * 1.12 + 2.5
            if est > best:
                best = est
    width = min(cap_w, max(floor_w, best))
    ws.column_dimensions[get_column_letter(col)].width = width
    return width


def _set_range_border(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_LIGHT


def _apply_table_header_row(ws: Any, row_idx: int, headers: list[str], strong: bool = True) -> None:
    fill = _FILL_HEADER_STRONG if strong else _FILL_HEADER_SOFT
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=name)
        cell.fill = fill
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = _BORDER_LIGHT
    ws.row_dimensions[row_idx].height = 22


def _apply_body_style(
    ws: Any,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    number_formats: dict[int, str] | None = None,
    wrap_cols: set[int] | None = None,
) -> None:
    number_formats = number_formats or {}
    wrap_cols = wrap_cols or set()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_LIGHT
            if c in number_formats and cell.value not in (None, ""):
                if not (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                    cell.number_format = number_formats[c]
            cell.alignment = _ALIGN_WRAP if c in wrap_cols else _ALIGN_VCENTER


def _parse_iso_date_cell(value: Any) -> datetime | date | str | None:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value
    text = str(value).strip()
    if not text or text.startswith("="):
        return value
    try:
        return date.fromisoformat(text)
    except ValueError:
        return value


def _write_control_sheet(ws: Any, process_id: str, process_date: date) -> tuple[int, int]:
    """Panel de control estilo referencia: título navy, banda de sección y filas Campo/Valor."""
    endc = 2
    _merge_navy_title_row(ws, 1, endc, CONTROL_TITLE, _FONT_TITLE_NAVY)
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=endc)
    sub = ws.cell(row=2, column=1, value=f"{CONTROL_SUBTITLE}\n{CONTROL_HELP}")
    sub.font = _FONT_SUB_NAVY
    sub.fill = _FILL_NAVY
    sub.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 48
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=endc)
    sec = ws.cell(row=3, column=1, value="INFORMACIÓN DEL PROCESO")
    sec.fill = _FILL_SECTION_BAND
    sec.font = _FONT_SECTION
    sec.alignment = _ALIGN_CENTER
    ws.row_dimensions[3].height = 22

    dr = 4
    rows_data = [
        (ControlCols.ROW_ID_PROCESO, process_id),
        (ControlCols.ROW_ESTADO, ControlCols.ESTADO_OPTIONS[0]),
        (ControlCols.ROW_FECHA, str(process_date)),
        (ControlCols.ROW_PROCESAR, ControlCols.VAL_PROCESAR_NO),
    ]
    estado_row = 0
    procesar_row = 0
    for i, (label, val) in enumerate(rows_data):
        r = dr + i
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=val)
        a.fill = _FILL_CONTROL_LABEL
        b.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        a.font = _FONT_LABEL_BOLD
        a.border = _BORDER_LIGHT
        a.alignment = _ALIGN_VCENTER
        b.border = _BORDER_LIGHT
        b.alignment = _ALIGN_VCENTER
        if label == ControlCols.ROW_ESTADO:
            b.font = _FONT_ESTADO_CTRL
            estado_row = r
        else:
            b.font = _FONT_BODY
        if label == ControlCols.ROW_PROCESAR:
            procesar_row = r

    _apply_column_widths(ws, {1: 30, 2: 48})
    ws.freeze_panes = f"A{dr}"
    return estado_row, procesar_row


def _style_control_procesar_row(ws_control: Any, procesar_row: int) -> None:
    """Resalta fila Procesar (fila 7): label destacado y valor SI/NO con color de pestaña."""
    if procesar_row < 1:
        return
    ws_control.row_dimensions[procesar_row].height = max(
        ws_control.row_dimensions[procesar_row].height or 0, 34.0
    )
    label_cell = ws_control.cell(row=procesar_row, column=1)
    label_cell.fill = _FILL_CONTROL_PROCESAR_ROW
    label_cell.border = _BORDER_LIGHT
    label_cell.font = _FONT_PROCESAR_LABEL
    label_cell.alignment = _ALIGN_PROCESAR_LABEL

    val_cell = ws_control.cell(row=procesar_row, column=2)
    val_cell.border = _BORDER_PROCESAR_VALUE
    val_cell.font = _FONT_PROCESAR_VALUE
    val_cell.alignment = _ALIGN_PROCESAR_VALUE
    val_cell.fill = _FILL_PROCESAR_SI


def _write_resumen_sheet(ws: Any, metrics: list[tuple[str, Any]]) -> None:
    _merge_navy_title_row(ws, 1, 2, RESUMEN_TITLE, _FONT_TITLE_NAVY)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    st = ws.cell(row=2, column=1, value=RESUMEN_SUBTITLE)
    st.fill = _FILL_NAVY
    st.font = Font(name="Calibri", size=10, color="FFFFFF")
    st.alignment = _ALIGN_CENTER
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    band = ws.cell(row=3, column=1, value=RESUMEN_SECTION)
    band.fill = _FILL_SECTION_BAND
    band.font = _FONT_SECTION
    band.alignment = _ALIGN_CENTER
    ws.row_dimensions[3].height = 22

    start_data = 4
    last = start_data + len(metrics) - 1
    for i, (name, val) in enumerate(metrics):
        r = start_data + i
        c1 = ws.cell(row=r, column=1, value=name)
        c2 = ws.cell(row=r, column=2, value=val)
        c1.font = _FONT_BODY
        c2.font = _FONT_BODY
        c1.border = _BORDER_LIGHT
        c2.border = _BORDER_LIGHT
        c1.alignment = _ALIGN_VCENTER
        c2.alignment = _ALIGN_CENTER
        c1.fill = _FILL_PANEL
        c2.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    _apply_column_widths(ws, {1: 38, 2: 14})
    _set_range_border(ws, 3, last, 1, 2)
    ws.freeze_panes = f"A{start_data}"


def _build_resumen_metrics(
    payment_cases: list[dict[str, Any]],
    distribution_rows: list[dict[str, Any]],
    error_rows: list[dict[str, Any]],
) -> list[tuple[str, Any]]:
    def count_estado(estado: str) -> int:
        return sum(1 for r in distribution_rows if r.get(DistribucionCols.ESTADO_PAGO) == estado)

    pagos_banco = len(payment_cases) + len(error_rows)
    return [
        ("Pagos banco", pagos_banco),
        ("Casos pago", len(payment_cases)),
        ("Líneas distribución", len(distribution_rows)),
        ("Errores", len(error_rows)),
        ("Atrasado (mora)", count_estado(EstadoPago.ATRASADO)),
        ("Adelantado", count_estado(EstadoPago.ADELANTADO)),
        ("Normal", count_estado(EstadoPago.NORMAL)),
        ("Incompleto", count_estado(EstadoPago.INCOMPLETO)),
        ("Revisión manual", count_estado(EstadoPago.REVISION_MANUAL)),
    ]


def _apply_distrib_monto_only_leading_rows(rows: list[dict[str, Any]]) -> None:
    """Monto banco pertenece al pago (ID Pago), no a cada crédito: solo la primera fila del grupo lleva valor."""
    seen: set[str] = set()
    for row in rows:
        pid = row.get(DistribucionCols.ID_PAGO)
        if pid is None:
            continue
        k = str(pid).strip()
        if not k:
            continue
        if k in seen:
            row[DistribucionCols.MONTO_BANCO] = None
        else:
            seen.add(k)


def _add_dropdowns(
    ws_control: Any,
    ws_distribution: Any,
    estado_cell: str,
    procesar_cell: str,
    dist_first_data_row: int,
) -> None:
    dv_estado = DataValidation(
        type="list",
        formula1=f"={ReviewSheets.LISTAS}!$A$1:$A${len(ControlCols.ESTADO_OPTIONS)}",
    )
    dv_procesar = DataValidation(
        type="list",
        formula1=f"={ReviewSheets.LISTAS}!$B$1:$B${len(ControlCols.PROCESAR_OPTIONS)}",
    )
    dv_estado_pago = DataValidation(
        type="list",
        formula1=f"={ReviewSheets.LISTAS}!$C$1:$C${len(EstadoPago.OPTIONS_ORDERED)}",
    )
    dv_validar_pago = DataValidation(
        type="list",
        formula1=f"={ReviewSheets.LISTAS}!$D$1:$D$2",
    )

    ws_control.add_data_validation(dv_estado)
    ws_control.add_data_validation(dv_procesar)
    ws_distribution.add_data_validation(dv_estado_pago)
    ws_distribution.add_data_validation(dv_validar_pago)

    dv_estado.add(estado_cell)
    dv_procesar.add(procesar_cell)
    last_row = max(ws_distribution.max_row, dist_first_data_row)
    col_ep = get_column_letter(DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_PAGO) + 1)
    col_vp = get_column_letter(DistribucionCols.HEADERS.index(DistribucionCols.VALIDAR_PAGO) + 1)
    dv_estado_pago.add(f"{col_ep}{dist_first_data_row}:{col_ep}{last_row}")
    dv_validar_pago.add(f"{col_vp}{dist_first_data_row}:{col_vp}{last_row}")


def _apply_distribution_formulas(ws_distribution: Any, first_data_row: int) -> None:
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        return

    ix = DistribucionCols.HEADERS.index
    col_id = get_column_letter(ix(DistribucionCols.ID_PAGO) + 1)
    col_c = get_column_letter(ix(DistribucionCols.MONTO_BANCO) + 1)
    col_i = get_column_letter(ix(DistribucionCols.APLICAR_A_EXTRACTO) + 1)
    col_j = get_column_letter(ix(DistribucionCols.MORA_A_APLICAR) + 1)
    col_k = get_column_letter(ix(DistribucionCols.OTROS_VALORES) + 1)
    col_l = get_column_letter(ix(DistribucionCols.TOTAL_APLICADO) + 1)
    col_m = get_column_letter(ix(DistribucionCols.SALDO_POR_ASIGNAR) + 1)
    col_a_num = ix(DistribucionCols.ID_PAGO) + 1

    first_row_of_id: dict[str, int] = {}
    for row_idx in range(first_data_row, last_row + 1):
        pid = ws_distribution.cell(row=row_idx, column=col_a_num).value
        if pid is None or str(pid).strip() == "":
            continue
        k = str(pid).strip()
        if k not in first_row_of_id:
            first_row_of_id[k] = row_idx

    rng_a = f"${col_id}${first_data_row}:${col_id}${last_row}"
    rng_c = f"${col_c}${first_data_row}:${col_c}${last_row}"
    rng_i = f"${col_i}${first_data_row}:${col_i}${last_row}"
    rng_j = f"${col_j}${first_data_row}:${col_j}${last_row}"
    rng_k = f"${col_k}${first_data_row}:${col_k}${last_row}"

    for row_idx in range(first_data_row, last_row + 1):
        pid = ws_distribution.cell(row=row_idx, column=col_a_num).value
        if pid is None or str(pid).strip() == "":
            ws_distribution[f"{col_l}{row_idx}"] = None
            ws_distribution[f"{col_m}{row_idx}"] = None
            continue
        k = str(pid).strip()
        crit = f"{col_id}{row_idx}"
        if first_row_of_id.get(k) != row_idx:
            ws_distribution[f"{col_l}{row_idx}"] = None
        else:
            ws_distribution[f"{col_l}{row_idx}"] = (
                f"=SUMIF({rng_a},{crit},{rng_i})"
                f"+SUMIF({rng_a},{crit},{rng_j})"
                f"+SUMIF({rng_a},{crit},{rng_k})"
            )
        if first_row_of_id.get(k) != row_idx:
            ws_distribution[f"{col_m}{row_idx}"] = None
            continue

        ws_distribution[f"{col_m}{row_idx}"] = (
            f"=SUMIF({rng_a},{crit},{rng_c})"
            f"-SUMIF({rng_a},{crit},{rng_i})"
            f"-SUMIF({rng_a},{crit},{rng_j})"
            f"-SUMIF({rng_a},{crit},{rng_k})"
        )


def _apply_distribution_conditional_formatting(ws_distribution: Any, first_data_row: int) -> str:
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        return "dynamic"

    col_estado = get_column_letter(
        DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_PAGO) + 1
    )
    range_ref = f"{col_estado}{first_data_row}:{col_estado}{last_row}"
    fills = _ESTADO_PAGO_FILLS
    for estado, fill in fills.items():
        ws_distribution.conditional_formatting.add(
            range_ref,
            FormulaRule(
                formula=[f'INDIRECT("{col_estado}"&ROW())="{estado}"'],
                fill=fill,
            ),
        )
    return "dynamic"


def _observation_has_non_blocking_warning(obs_text: str) -> bool:
    obs = str(obs_text or "").strip()
    if not obs:
        return False
    markers = (
        "No se pudo verificar contra tabla de amortización",
        _OBS_ROOT_UNIT,
        _OBS_NON_STANDARD_FOLDER,
        _OBS_POSSIBLE_FINALIZED,
        "Advertencia:",
    )
    return any(m in obs for m in markers)


def _apply_distrib_observation_warning_fills(
    ws_distribution: Any, first_data_row: int
) -> None:
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        return
    obs_c = DistribucionCols.HEADERS.index(DistribucionCols.OBSERVACION) + 1
    for r in range(first_data_row, last_row + 1):
        cell = ws_distribution.cell(row=r, column=obs_c)
        if _observation_has_non_blocking_warning(str(cell.value or "")):
            cell.fill = _FILL_OBS_WARNING


def _apply_distrib_dias_mora_conditional(ws_distribution: Any, first_data_row: int) -> None:
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        return
    col = get_column_letter(DistribucionCols.HEADERS.index(DistribucionCols.DIAS_MORA) + 1)
    ref = f"{col}{first_data_row}:{col}{last_row}"
    # Fila relativa para que cada celda compare su propio valor de mora
    top = f"${col}{first_data_row}"
    ws_distribution.conditional_formatting.add(
        ref,
        FormulaRule(formula=[f"={top}>0"], fill=_FILL_DIAS_MORA_WARN),
    )


def _apply_distrib_hyperlinks(ws_distribution: Any, first_data_row: int) -> None:
    last_row = ws_distribution.max_row
    if last_row < first_data_row:
        return
    col_ext = DistribucionCols.HEADERS.index(DistribucionCols.LINK_EXTRACTO) + 1
    col_tab = DistribucionCols.HEADERS.index(DistribucionCols.LINK_TABLA) + 1
    col_fold = DistribucionCols.HEADERS.index(DistribucionCols.LINK_CARPETA_CREDITO) + 1
    col_credito = DistribucionCols.HEADERS.index(DistribucionCols.CREDITO) + 1
    col_cliente = DistribucionCols.HEADERS.index(DistribucionCols.CLIENTE) + 1

    def _apply_link_cell(cell: Any, url_raw: Any, link_kind: str, credito: Any, cliente: Any) -> None:
        if url_raw and isinstance(url_raw, str) and str(url_raw).strip():
            target = str(url_raw).strip()
            if target.startswith("http"):
                cell.value = _format_distrib_link_visible_text(link_kind, credito, cliente)
                cell.hyperlink = target
                cell.font = _HLINK_FONT
                return
        cell.value = ""
        cell.hyperlink = None

    for r in range(first_data_row, last_row + 1):
        credito = ws_distribution.cell(row=r, column=col_credito).value
        cliente = ws_distribution.cell(row=r, column=col_cliente).value
        _apply_link_cell(
            ws_distribution.cell(row=r, column=col_ext),
            ws_distribution.cell(row=r, column=col_ext).value,
            "extracto",
            credito,
            cliente,
        )
        _apply_link_cell(
            ws_distribution.cell(row=r, column=col_tab),
            ws_distribution.cell(row=r, column=col_tab).value,
            "tabla",
            credito,
            cliente,
        )
        _apply_link_cell(
            ws_distribution.cell(row=r, column=col_fold),
            ws_distribution.cell(row=r, column=col_fold).value,
            "carpeta",
            credito,
            cliente,
        )


def _style_distrib_sheet(ws_distribution: Any, header_row: int, first_data_row: int) -> None:
    ncols = len(DistribucionCols.HEADERS)
    saldo_c = DistribucionCols.HEADERS.index(DistribucionCols.SALDO_POR_ASIGNAR) + 1
    estado_c = DistribucionCols.HEADERS.index(DistribucionCols.ESTADO_PAGO) + 1
    money_cols = {
        DistribucionCols.HEADERS.index(DistribucionCols.MONTO_BANCO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.VALOR_EXTRACTO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.APLICAR_A_EXTRACTO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.MORA_A_APLICAR) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.OTROS_VALORES) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.TOTAL_APLICADO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.SALDO_POR_ASIGNAR) + 1,
    }
    money_editable_cols = {
        DistribucionCols.HEADERS.index(DistribucionCols.APLICAR_A_EXTRACTO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.MORA_A_APLICAR) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.OTROS_VALORES) + 1,
    }
    date_cols = {
        DistribucionCols.HEADERS.index(DistribucionCols.FECHA_BANCO) + 1,
        DistribucionCols.HEADERS.index(DistribucionCols.FECHA_LIMITE) + 1,
    }
    obs_c = DistribucionCols.HEADERS.index(DistribucionCols.OBSERVACION) + 1
    validar_c = DistribucionCols.HEADERS.index(DistribucionCols.VALIDAR_PAGO) + 1
    wrap_cols = {obs_c}
    col_cliente = DistribucionCols.HEADERS.index(DistribucionCols.CLIENTE) + 1
    col_id_pago = DistribucionCols.HEADERS.index(DistribucionCols.ID_PAGO) + 1
    last_row = ws_distribution.max_row
    if last_row >= first_data_row:
        block_edges = _sheet_client_block_edges(
            ws_distribution, first_data_row, last_row, col_cliente, col_id_pago
        )
        editable_cols = money_editable_cols | {estado_c, obs_c}
        prev_cliente: str | None = None
        group_idx = 0
        for r in range(first_data_row, last_row + 1):
            cli_value = ws_distribution.cell(row=r, column=col_cliente).value
            cliente = str(cli_value).strip() if cli_value is not None else ""
            if cliente:
                if prev_cliente and cliente != prev_cliente:
                    group_idx += 1
                prev_cliente = cliente
            group_fill = _GROUP_BAND_FILLS[group_idx % len(_GROUP_BAND_FILLS)]
            client_top, client_bottom = block_edges.get(r, (False, False))
            row_border = _distrib_row_border(
                client_top=client_top, client_bottom=client_bottom
            )
            for c in range(1, ncols + 1):
                cell = ws_distribution.cell(row=r, column=c)
                cell.font = _FONT_BODY
                if c == validar_c:
                    # Mismo verde brillante que el valor de Procesar en Control (B7)
                    cell.fill = _FILL_PROCESAR_SI
                elif c in editable_cols:
                    cell.fill = _FILL_EDITABLE_COL
                elif c == saldo_c:
                    cell.fill = _FILL_SALDO_COL
                else:
                    cell.fill = group_fill
                cell.border = row_border
                if c in money_cols:
                    cell.number_format = _FMT_MONEY
                if c in date_cols:
                    parsed = _parse_iso_date_cell(cell.value)
                    if parsed is not None and not (isinstance(parsed, str) and parsed.startswith("=")):
                        cell.value = parsed
                        cell.number_format = _FMT_DATE
                if c in wrap_cols:
                    cell.alignment = _ALIGN_WRAP
                else:
                    cell.alignment = _ALIGN_VCENTER
    if last_row >= header_row:
        _auto_fit_columns(ws_distribution, header_row, last_row, 1, ncols, floor_w=9.0, cap_w=58.0)
        _dc = DistribucionCols.HEADERS.index
        width_floor = {
            _dc(DistribucionCols.ID_PAGO) + 1: 12,
            _dc(DistribucionCols.CLIENTE) + 1: 20,
            _dc(DistribucionCols.MONTO_BANCO) + 1: 14,
            _dc(DistribucionCols.FECHA_BANCO) + 1: 12,
            _dc(DistribucionCols.CREDITO) + 1: 11,
            _dc(DistribucionCols.FECHA_LIMITE) + 1: 12,
            _dc(DistribucionCols.DIAS_MORA) + 1: 10,
            _dc(DistribucionCols.VALOR_EXTRACTO) + 1: 14,
            _dc(DistribucionCols.APLICAR_A_EXTRACTO) + 1: int(DIST_MONEY_COL_WIDTH),
            _dc(DistribucionCols.MORA_A_APLICAR) + 1: int(DIST_MONEY_COL_WIDTH),
            _dc(DistribucionCols.OTROS_VALORES) + 1: int(DIST_MONEY_COL_WIDTH),
            _dc(DistribucionCols.TOTAL_APLICADO) + 1: int(DIST_MONEY_COL_WIDTH),
            _dc(DistribucionCols.SALDO_POR_ASIGNAR) + 1: int(DIST_MONEY_COL_WIDTH),
            _dc(DistribucionCols.ESTADO_PAGO) + 1: 16,
            _dc(DistribucionCols.VALIDAR_PAGO) + 1: 14,
            _dc(DistribucionCols.OBSERVACION) + 1: 26,
            _dc(DistribucionCols.LINK_EXTRACTO) + 1: 26,
            _dc(DistribucionCols.LINK_TABLA) + 1: 28,
            _dc(DistribucionCols.LINK_CARPETA_CREDITO) + 1: 32,
            _dc(DistribucionCols.RUTA) + 1: 8,
            _dc(DistribucionCols.RUTA_UNIDAD_CREDITO) + 1: 8,
        }
        for c_idx, wmin in width_floor.items():
            letter = get_column_letter(c_idx)
            cur = ws_distribution.column_dimensions[letter].width or 0
            if cur < wmin:
                ws_distribution.column_dimensions[letter].width = wmin
        for c_idx in money_cols:
            letter = get_column_letter(c_idx)
            ws_distribution.column_dimensions[letter].width = DIST_MONEY_COL_WIDTH
        obs_c = DistribucionCols.HEADERS.index(DistribucionCols.OBSERVACION) + 1
        obs_w = ws_distribution.column_dimensions[get_column_letter(obs_c)].width or 28
        chars_per_line = max(10, int(obs_w * 1.15))
        for r in range(first_data_row, last_row + 1):
            txt = str(ws_distribution.cell(row=r, column=obs_c).value or "")
            if not txt.strip():
                lines = 1
            else:
                lines = max(1, (len(txt) + chars_per_line - 1) // chars_per_line + txt.count("\n"))
            ws_distribution.row_dimensions[r].height = min(120.0, max(15.0, 14.0 * lines))
    ws_distribution.row_dimensions[header_row].height = max(
        ws_distribution.row_dimensions[header_row].height or 0, 26.0
    )
    ws_distribution.freeze_panes = _distrib_freeze_panes_cell(first_data_row)
    ws_distribution.auto_filter.ref = None


def _style_casos_sheet(ws: Any, header_row: int, first_data_row: int) -> None:
    ncols = len(CasosPagoCols.HEADERS)
    money_col = CasosPagoCols.HEADERS.index(CasosPagoCols.MONTO_BANCO) + 1
    date_col = CasosPagoCols.HEADERS.index(CasosPagoCols.FECHA_BANCO) + 1
    wrap_cols = {CasosPagoCols.HEADERS.index(CasosPagoCols.OBSERVACION) + 1}
    last_row = ws.max_row
    if last_row >= first_data_row:
        for r in range(first_data_row, last_row + 1):
            stripe = (r - first_data_row) % 2 == 1
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _FONT_BODY
                if stripe:
                    cell.fill = _FILL_ZEBRA
                cell.border = _BORDER_LIGHT
                if c == money_col and cell.value not in (None, ""):
                    cell.number_format = _FMT_MONEY
                if c == date_col:
                    parsed = _parse_iso_date_cell(cell.value)
                    if parsed is not None and not (isinstance(parsed, str) and str(parsed).startswith("=")):
                        cell.value = parsed
                        cell.number_format = _FMT_DATE
                cell.alignment = _ALIGN_WRAP if c in wrap_cols else _ALIGN_VCENTER
    if last_row >= header_row:
        _auto_fit_columns(ws, header_row, last_row, 1, ncols, floor_w=9.0, cap_w=58.0)
        obs_c = CasosPagoCols.HEADERS.index(CasosPagoCols.OBSERVACION) + 1
        for c_idx, wmin in {1: 11, 2: 11, 3: 18, 4: 22, 5: 14}.items():
            letter = get_column_letter(c_idx)
            cur = ws.column_dimensions[letter].width or 0
            if cur < wmin:
                ws.column_dimensions[letter].width = wmin
        obs_w = _fit_single_column_width_from_content(
            ws,
            obs_c,
            header_row,
            last_row,
            floor_w=CASOS_OBS_COL_WIDTH,
            cap_w=CASOS_OBS_COL_CAP_WIDTH,
        )
        chars_per_line = max(10, int(obs_w * 1.12))
        for r in range(first_data_row, last_row + 1):
            txt = str(ws.cell(row=r, column=obs_c).value or "")
            lines = 1 if not txt.strip() else max(1, (len(txt) + chars_per_line - 1) // chars_per_line + txt.count("\n"))
            ws.row_dimensions[r].height = min(110.0, max(15.0, 14.0 * lines))
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 24.0)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = None


def _apply_errores_link_cells(
    ws_errors: Any,
    first_data_row: int,
    error_records: list[dict[str, Any]],
) -> None:
    """Asigna texto visible y cell.hyperlink desde URLs explícitas (no desde el valor de celda)."""
    if not error_records:
        return
    col_ext = ErroresCols.HEADERS.index(ErroresCols.LINK_EXTRACTO) + 1
    col_fold = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
    for i, rec in enumerate(error_records):
        r = first_data_row + i
        ext_url = _http_url_only(
            rec.get("link_extracto_url")
            or rec.get("link_extracto")
        )
        fold_url = _http_url_only(
            rec.get("link_carpeta_credito_url")
            or rec.get("link_carpeta")
        )
        credito = rec.get("credito")
        cliente = rec.get("cliente")
        c_ext = ws_errors.cell(row=r, column=col_ext)
        if ext_url:
            c_ext.value = _format_errores_link_visible_text("extracto", credito, cliente)
            c_ext.hyperlink = ext_url
            c_ext.font = _HLINK_FONT
        else:
            c_ext.value = ""
            c_ext.hyperlink = None
        c_fold = ws_errors.cell(row=r, column=col_fold)
        if fold_url:
            c_fold.value = _format_errores_link_visible_text("carpeta", credito, cliente)
            c_fold.hyperlink = fold_url
            c_fold.font = _HLINK_FONT
        else:
            c_fold.value = ""
            c_fold.hyperlink = None


def _style_errores_sheet(ws: Any, header_row: int, first_data_row: int) -> None:
    ncols = len(ErroresCols.HEADERS)
    desc_c = ErroresCols.HEADERS.index(ErroresCols.DESCRIPCION) + 1
    hacer_c = ErroresCols.HEADERS.index(ErroresCols.QUE_DEBE_HACER) + 1
    wrap_cols = {desc_c, hacer_c}
    col_ext = ErroresCols.HEADERS.index(ErroresCols.LINK_EXTRACTO) + 1
    col_fold = ErroresCols.HEADERS.index(ErroresCols.LINK_CARPETA_CREDITO) + 1
    _ec = ErroresCols.HEADERS.index
    _apply_column_widths(
        ws,
        {
            _ec(ErroresCols.ID_PAGO) + 1: 11,
            _ec(ErroresCols.CLIENTE) + 1: 20,
            _ec(ErroresCols.CREDITO) + 1: 16,
            _ec(ErroresCols.TIPO_CASO) + 1: 16,
            _ec(ErroresCols.DESCRIPCION) + 1: 40,
            _ec(ErroresCols.QUE_DEBE_HACER) + 1: 42,
            _ec(ErroresCols.REQUIERE_SOPORTE) + 1: 18,
            col_ext: 32,
            col_fold: 36,
            _ec(ErroresCols.CODIGO_TECNICO) + 1: 22,
        },
    )
    col_cliente = ErroresCols.HEADERS.index(ErroresCols.CLIENTE) + 1
    col_id_pago = ErroresCols.HEADERS.index(ErroresCols.ID_PAGO) + 1
    last_row = ws.max_row
    if last_row >= first_data_row:
        block_edges = _sheet_client_block_edges(
            ws, first_data_row, last_row, col_cliente, col_id_pago
        )
        for r in range(first_data_row, last_row + 1):
            stripe = (r - first_data_row) % 2 == 1
            client_top, client_bottom = block_edges.get(r, (False, False))
            row_border = _distrib_row_border(
                client_top=client_top, client_bottom=client_bottom
            )
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = _FONT_BODY
                if stripe:
                    cell.fill = _FILL_ZEBRA
                cell.border = row_border
                cell.alignment = _ALIGN_WRAP if c in wrap_cols else _ALIGN_VCENTER
        desc_w = ws.column_dimensions[get_column_letter(desc_c)].width or 38
        hacer_w = ws.column_dimensions[get_column_letter(hacer_c)].width or 40
        cl_desc = max(10, int(desc_w * 1.05))
        cl_hacer = max(10, int(hacer_w * 1.05))
        for r in range(first_data_row, last_row + 1):
            dtxt = str(ws.cell(row=r, column=desc_c).value or "")
            htxt = str(ws.cell(row=r, column=hacer_c).value or "")
            ld = 1 if not dtxt.strip() else max(1, (len(dtxt) + cl_desc - 1) // cl_desc + dtxt.count("\n"))
            lh = 1 if not htxt.strip() else max(1, (len(htxt) + cl_hacer - 1) // cl_hacer + htxt.count("\n"))
            ws.row_dimensions[r].height = min(140.0, max(15.0, 14.0 * max(ld, lh)))
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 26.0)
    ws.freeze_panes = f"A{first_data_row}"
    ws.auto_filter.ref = None


async def _load_credit_candidates(
    client: GraphApiPort,
    site_search: str,
    drive_name: str,
    clients_path: str,
    cliente_folder: str,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """Devuelve candidatos distribuibles por unidad de crédito + incidencias por unidad (hoja Errores)."""
    clients_info = await resolve_sharepoint_path(client, site_search, drive_name, clients_path)
    cliente_folder_path = f"{clients_path}/{cliente_folder}"
    cliente_folder_encoded = encode_graph_drive_path(cliente_folder_path)
    credit_children = await client.get(
        f"/sites/{clients_info['site_id']}/drives/{clients_info['drive_id']}/root:/{cliente_folder_encoded}:/children"
    )
    all_items: list[dict[str, Any]] = list(credit_children.get("value", []))

    candidates: list[dict[str, Any]] = []
    credit_issues: list[dict[str, Any]] = []
    site_id = clients_info["site_id"]
    drive_id = clients_info["drive_id"]
    client_folder_web_url = await _resolve_client_folder_web_url(
        client, site_id, drive_id, clients_path, cliente_folder
    )

    async def process_credit_unit(
        credit_name: str,
        credit_path: str,
        items: list[dict[str, Any]],
        credit_folder_drive_item: dict[str, Any] | None = None,
        *,
        is_flat_unit: bool = False,
        is_root_unit: bool = False,
    ) -> None:
        carpeta_link_url = _carpeta_link_url_for_errores(
            credit_folder_drive_item,
            credit_path,
            client_folder_web_url=client_folder_web_url,
            is_flat_unit=is_flat_unit,
            is_root_unit=is_root_unit,
        )
        if (is_flat_unit or is_root_unit) and not carpeta_link_url:
            logger.warning(
                "generate errores: link_carpeta_credito_url vacío cliente=%r unidad=%r "
                "credit_path=%r graph_client_web_url=%r",
                cliente_folder,
                credit_name,
                credit_path,
                client_folder_web_url,
            )
        file_names = [item.get("name", "") for item in items if item.get("name")]
        excel_only = filter_amortization_excel_filenames(file_names)
        use_v2 = _use_extract_selection_v2()

        if use_v2:
            warnings: list[str] = []
            pool, extract_base_path = await _resolve_extract_pdf_pool(
                client, site_id, drive_id, credit_path, items
            )
            statement_item, statement_bytes, fecha_limite_pdf, sel_err = (
                await _select_extract_by_max_fecha_limite_v2(
                    client, site_id, drive_id, extract_base_path, pool
                )
            )
            if (
                sel_err
                or statement_item is None
                or statement_bytes is None
                or fecha_limite_pdf is None
            ):
                issue_code = sel_err or "extract_not_found"
                link_extracto_url = ""
                if pool and issue_code == "fecha_limite_extracto_not_readable":
                    first_it = pool[0]
                    link_extracto_url = _item_link_url(
                        first_it, f"{extract_base_path}/{first_it.get('name', '')}"
                    )
                credit_issues.append(
                    {
                        "code": issue_code,
                        "unidad_credito": credit_name,
                        "link_extracto_url": link_extracto_url,
                        "link_carpeta_credito_url": carpeta_link_url,
                    }
                )
                return

            statement_path = f"{extract_base_path}/{statement_item.get('name', '')}"

            try:
                extract_value = extract_total_a_pagar_from_pdf(statement_bytes)
            except ValueError:
                credit_issues.append(
                    {
                        "code": "extract_amount_not_found",
                        "unidad_credito": credit_name,
                        "link_extracto_url": _item_link_url(statement_item, statement_path),
                        "link_carpeta_credito_url": carpeta_link_url,
                    }
                )
                return

            credit_id, is_non_standard = _resolve_credit_id_for_unit(
                credit_name,
                cliente_folder,
                is_flat_unit=is_flat_unit,
                is_root_unit=is_root_unit,
                statement_filename=str(statement_item.get("name", "")),
                statement_bytes=statement_bytes,
            )

            table_item = None
            table_path = ""
            table_name_res: str | None = None

            try:
                table_name_res = find_best_amortization_table(
                    excel_only, cliente_folder, credit_name
                )
            except ValueError as e:
                code = str(e)
                if code == "amortization_table_not_found":
                    warnings.append(_OBS_TABLA_NO_VERIFICAR_NOT_FOUND)
                elif code == "amortization_table_ambiguous":
                    warnings.append(_OBS_TABLA_NO_VERIFICAR_AMBIGUOUS)
                else:
                    warnings.append(_OBS_TABLA_NO_VERIFICAR_NOPARSE)
            else:
                table_item = next(
                    (item for item in items if item.get("name") == table_name_res), None
                )
                table_path = f"{credit_path}/{table_name_res}"
                try:
                    tb = await client.get_bytes(
                        _build_content_endpoint(site_id, drive_id, table_path)
                    )
                    last_pay = _extract_last_payment_date_from_amortization(tb, cliente_folder)
                    if last_pay is None:
                        warnings.append(_OBS_TABLA_NO_VERIFICAR_NOPARSE)
                    elif fecha_limite_pdf < last_pay:
                        warnings.append(_OBS_EXTRACTO_FECHA_VS_TABLA_ANTERIOR)
                    elif fecha_limite_pdf == last_pay:
                        warnings.append(_OBS_EXTRACTO_FECHA_VS_TABLA_IGUAL)
                except Exception:
                    logger.debug(
                        "amortization_aux_verify_failed credit=%s", credit_name, exc_info=True
                    )
                    warnings.append(_OBS_TABLA_NO_VERIFICAR_NOPARSE)

            obs_ter = _possibly_finalized_observation(credit_name)
            extra_parts = [w for w in warnings if w]
            if is_root_unit:
                extra_parts.append(_OBS_ROOT_UNIT)
            if is_non_standard:
                extra_parts.append(_OBS_NON_STANDARD_FOLDER)
            if obs_ter:
                extra_parts.append(obs_ter)
            obs_extra = " | ".join(extra_parts) if extra_parts else None

            link_tabla_val = ""
            if table_item is not None and table_path:
                link_tabla_val = _item_link(table_item, table_path)

            cred_norm = normalize_credito_digits(credit_id) or str(credit_id)
            ruta_tabla = table_path.replace("\\", "/").strip("/") if table_path else ""
            candidates.append(
                {
                    "credito": str(credit_id),
                    "credito_normalizado": cred_norm,
                    "fecha_limite": fecha_limite_pdf,
                    "valor_extracto": extract_value,
                    "link_extracto": _item_link(statement_item, statement_path),
                    "link_tabla": link_tabla_val,
                    "link_carpeta_credito": carpeta_link_url,
                    "ruta_extracto_pdf": statement_path.replace("\\", "/"),
                    "ruta_unidad_credito": credit_path.replace("\\", "/"),
                    "ruta_tabla_amortizacion": ruta_tabla,
                    **({"observacion_extra": obs_extra} if obs_extra else {}),
                }
            )
            return

        try:
            table_name = find_best_amortization_table(excel_only, cliente_folder, credit_name)
            table_item = next((item for item in items if item.get("name") == table_name), None)
            table_path = f"{credit_path}/{table_name}"
            table_bytes = await client.get_bytes(
                _build_content_endpoint(site_id, drive_id, table_path)
            )
            pending = _extract_pending_installment(table_bytes, cliente_folder)
            due_date = pending.get("fecha_limite")

            statement_item = _find_statement_item(items, credit_name, due_date)
            if statement_item is None:
                credit_issues.append(
                    {
                        "code": "extract_not_found",
                        "unidad_credito": credit_name,
                        "link_extracto_url": "",
                        "link_carpeta_credito_url": carpeta_link_url,
                    }
                )
                return
            statement_path = f"{credit_path}/{statement_item.get('name', '')}"
            statement_bytes = await client.get_bytes(
                _build_content_endpoint(site_id, drive_id, statement_path)
            )

            extract_value = extract_total_a_pagar_from_pdf(statement_bytes)

            parsed_date, parsed_credit = parse_statement_name(statement_item.get("name", ""))
            due_date = due_date or parsed_date
            credit_id = parsed_credit or credit_name
            if due_date is None:
                raise ValueError("pending_installment_not_found")

            obs_extra = _possibly_finalized_observation(credit_name)
            cred_norm = normalize_credito_digits(credit_id) or str(credit_id)
            ruta_tabla = table_path.replace("\\", "/").strip("/") if table_path else ""
            candidates.append(
                {
                    "credito": str(credit_id),
                    "credito_normalizado": cred_norm,
                    "fecha_limite": due_date,
                    "valor_extracto": extract_value,
                    "link_extracto": _item_link(statement_item, statement_path),
                    "link_tabla": _item_link(table_item or {}, table_path),
                    "link_carpeta_credito": carpeta_link_url,
                    "ruta_extracto_pdf": statement_path.replace("\\", "/"),
                    "ruta_unidad_credito": credit_path.replace("\\", "/"),
                    "ruta_tabla_amortizacion": ruta_tabla,
                    **({"observacion_extra": obs_extra} if obs_extra else {}),
                }
            )
        except ValueError as exc:
            credit_issues.append(
                {
                    "code": str(exc),
                    "unidad_credito": credit_name,
                    "link_extracto_url": "",
                    "link_carpeta_credito_url": carpeta_link_url,
                }
            )

    root_file_items = [it for it in all_items if "folder" not in it]
    subfolder_items = [it for it in all_items if "folder" in it]
    operational_units: list[
        tuple[str, str, list[dict[str, Any]], dict[str, Any] | None]
    ] = []

    for credit_folder in subfolder_items:
        credit_name = str(credit_folder.get("name", "") or "").strip()
        if not credit_name or _is_infra_folder(credit_name):
            continue
        credit_path = f"{cliente_folder_path}/{credit_name}"
        credit_encoded = encode_graph_drive_path(credit_path)
        files_resp = await client.get(
            f"/sites/{site_id}/drives/{drive_id}/root:/{credit_encoded}:/children"
        )
        children = list(files_resp.get("value", []))
        is_standard = _looks_like_standard_credit_folder(credit_name)
        if is_standard or await _items_have_operational_signal(
            client,
            site_id,
            drive_id,
            credit_path,
            children,
            cliente_folder,
            credit_name,
        ):
            operational_units.append((credit_name, credit_path, children, credit_folder))

    if operational_units:
        for credit_name, credit_path, children, folder_item in operational_units:
            await process_credit_unit(
                credit_name,
                credit_path,
                children,
                credit_folder_drive_item=folder_item,
                is_flat_unit=False,
                is_root_unit=False,
            )
        if _root_has_strict_extract_pdfs(root_file_items):
            await process_credit_unit(
                cliente_folder,
                cliente_folder_path,
                root_file_items,
                credit_folder_drive_item=None,
                is_flat_unit=False,
                is_root_unit=True,
            )
    else:
        await process_credit_unit(
            cliente_folder,
            cliente_folder_path,
            all_items,
            credit_folder_drive_item=None,
            is_flat_unit=True,
            is_root_unit=False,
        )

    candidates = _dedupe_credit_candidates(candidates)

    if not candidates and not credit_issues:
        raise ValueError("credit_folder_not_found")

    return candidates, credit_issues


async def generate_payment_validation(client: GraphApiPort, process_date: date) -> dict[str, Any]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    review_path = os.getenv("GRAPH_PAYMENT_VALIDATION_REVIEW_PATH", "").strip()
    bank_path = os.getenv("GRAPH_BANK_PAYMENTS_FILE_PATH", "").strip()
    clients_path = os.getenv("GRAPH_CLIENTS_BASE_PATH", "").strip()
    file_prefix = os.getenv("GRAPH_VALIDATION_FILE_PREFIX", "").strip()

    if not all([site_search, review_path, bank_path, clients_path, file_prefix]):
        raise ValueError("missing_sharepoint_folder")

    review_info = await resolve_sharepoint_path(client, site_search, drive_name, review_path)
    review_children = await client.get(
        f"/sites/{review_info['site_id']}/drives/{review_info['drive_id']}/root:/{review_info['path_encoded']}:/children"
    )
    valid_children = [item for item in review_children.get("value", []) if not item.get("name", "").startswith("~$")]
    if valid_children:
        raise ValueError("review_folder_not_empty")

    bank_info = await resolve_sharepoint_path(client, site_search, drive_name, bank_path)
    bank_bytes = await client.get_bytes(
        _build_content_endpoint(bank_info["site_id"], bank_info["drive_id"], bank_info["file_path"])
    )

    bank_workbook = openpyxl.load_workbook(io.BytesIO(bank_bytes), data_only=True)
    bank_sheet = bank_workbook.active

    col_map = {}
    start_row = 2
    for row_index, row in enumerate(bank_sheet.iter_rows(values_only=True), 1):
        if not row:
            continue
        texts = [_normalize_str(str(value)) if value else "" for value in row]
        if "fecha" in texts and ("credito" in texts or "monto" in texts):
            col_map = {
                "fecha": texts.index("fecha"),
                "credito": texts.index("credito") if "credito" in texts else texts.index("monto"),
                "concepto": texts.index("concepto") if "concepto" in texts else -1,
                "transaccion": texts.index("transaccion") if "transaccion" in texts else -1,
            }
            start_row = row_index + 1
            break

    if not col_map:
        raise ValueError("bank_headers_not_found")

    clients_info = await resolve_sharepoint_path(client, site_search, drive_name, clients_path)
    client_children = await client.get(
        f"/sites/{clients_info['site_id']}/drives/{clients_info['drive_id']}/root:/{clients_info['path_encoded']}:/children"
    )
    client_folders = [item for item in client_children.get("value", []) if "folder" in item]
    client_index = {_normalize_str(item.get("name", "")): item.get("name", "") for item in client_folders}

    payment_cases: list[dict[str, Any]] = []
    distribution_rows: list[dict[str, Any]] = []
    error_records: list[dict[str, Any]] = []

    for row_index in range(start_row, bank_sheet.max_row + 1):
        row = [cell.value for cell in bank_sheet[row_index]]
        if not any(row):
            continue

        payment_id = str(uuid.uuid4())[:8]
        concepto = row[col_map["concepto"]] if col_map["concepto"] >= 0 else ""
        transaccion = row[col_map["transaccion"]] if col_map["transaccion"] >= 0 else ""
        cliente_raw = "Desconocido"
        try:
            fecha_banco = parse_bank_date(row[col_map["fecha"]], process_date)
            monto_banco = parse_bank_amount(row[col_map["credito"]])
            cliente_raw = extract_client_from_bank_row(concepto, transaccion)
            cliente_norm = _normalize_str(cliente_raw)

            if cliente_norm in client_index:
                cliente_folder = client_index[cliente_norm]
            else:
                candidates = [name for norm, name in client_index.items() if cliente_norm in norm or norm in cliente_norm]
                if not candidates:
                    raise ValueError("customer_not_found")
                if len(candidates) > 1:
                    raise ValueError("customer_ambiguous")
                cliente_folder = candidates[0]

            payment = {
                "id_pago": payment_id,
                "fecha_banco": fecha_banco,
                "monto_banco": monto_banco,
                "cliente": cliente_folder,
                "concepto": concepto,
                "transaccion": transaccion,
            }

            credit_candidates, credit_issues = await _load_credit_candidates(
                client, site_search, drive_name, clients_path, cliente_folder
            )
            for ci in credit_issues:
                error_records.append(
                    {
                        "id_pago": payment_id,
                        "cliente": cliente_folder,
                        "credito": str(ci["unidad_credito"]),
                        "code": str(ci["code"]),
                        "link_extracto_url": _http_url_only(
                            ci.get("link_extracto_url") or ci.get("link_extracto")
                        ),
                        "link_carpeta_credito_url": _http_url_only(
                            ci.get("link_carpeta_credito_url") or ci.get("link_carpeta")
                        ),
                    }
                )

            if not credit_candidates:
                continue

            payment_distribution_rows = _build_distribution_rows(payment, credit_candidates)
            distribution_rows.extend(payment_distribution_rows)
            payment_cases.append(_build_case_row(payment, payment_distribution_rows))
        except Exception as exc:
            code = str(exc)
            logger.warning(
                "generate_payment_validation bank_row_failed payment_id=%s code=%s",
                payment_id,
                code,
                exc_info=True,
            )
            link_carpeta_exc_url = ""
            cf = locals().get("cliente_folder")
            if isinstance(cf, str) and cf and code == "credit_folder_not_found":
                link_carpeta_exc_url = ""
            error_records.append(
                {
                    "id_pago": payment_id,
                    "cliente": cliente_raw,
                    "credito": "",
                    "code": code,
                    "link_extracto_url": "",
                    "link_carpeta_credito_url": link_carpeta_exc_url,
                }
            )

    workbook = openpyxl.Workbook()
    process_id = str(uuid.uuid4())
    hr = _table_header_row()
    dr = _table_first_data_row()

    ws_control = workbook.active
    ws_control.title = ReviewSheets.CONTROL
    estado_row, procesar_row = _write_control_sheet(ws_control, process_id, process_date)
    estado_cell = f"B{estado_row}"
    procesar_cell = f"B{procesar_row}"

    ws_resumen = workbook.create_sheet(ReviewSheets.RESUMEN)
    _write_resumen_sheet(ws_resumen, _build_resumen_metrics(payment_cases, distribution_rows, error_records))

    ws_cases = workbook.create_sheet(ReviewSheets.CASOS_PAGO)
    _apply_casos_top_banner(ws_cases)
    _apply_table_header_row(ws_cases, hr, list(CasosPagoCols.HEADERS), strong=True)
    for row in payment_cases:
        ws_cases.append([row.get(header) for header in CasosPagoCols.HEADERS])
    _style_casos_sheet(ws_cases, hr, dr)

    ws_distribution = workbook.create_sheet(ReviewSheets.DISTRIBUCION)
    _apply_distrib_top_banner(ws_distribution)
    _apply_table_header_row(ws_distribution, hr, list(DistribucionCols.HEADERS), strong=True)
    _apply_distrib_monto_only_leading_rows(distribution_rows)
    for row in distribution_rows:
        ws_distribution.append([row.get(header) for header in DistribucionCols.HEADERS])

    ws_lists = workbook.create_sheet(ReviewSheets.LISTAS)
    _write_list_values(ws_lists)
    ws_lists.sheet_state = "hidden"

    _apply_distribution_formulas(ws_distribution, dr)
    _add_dropdowns(ws_control, ws_distribution, estado_cell, procesar_cell, dr)
    _style_control_procesar_row(ws_control, procesar_row)
    _style_distrib_sheet(ws_distribution, hr, dr)
    _apply_distrib_hyperlinks(ws_distribution, dr)
    _configure_distrib_technical_path_columns(ws_distribution)
    formatting_strategy = _apply_distribution_conditional_formatting(ws_distribution, dr)
    _apply_distrib_dias_mora_conditional(ws_distribution, dr)

    ws_errors = workbook.create_sheet(ReviewSheets.ERRORES)
    _apply_errores_top_banner(ws_errors, len(ErroresCols.HEADERS), bool(error_records))
    _apply_table_header_row(ws_errors, hr, list(ErroresCols.HEADERS), strong=True)
    for rec in error_records:
        ws_errors.append(_normalized_error_record_to_sheet_row(rec))
    _style_errores_sheet(ws_errors, hr, dr)
    _apply_errores_link_cells(ws_errors, dr, error_records)

    for _ws in (ws_control, ws_resumen, ws_cases, ws_distribution, ws_errors):
        _sheet_hide_gridlines(_ws)

    _protect_control_sheet(ws_control, procesar_cell=procesar_cell, estado_cell=estado_cell)
    _protect_distribution_sheet(ws_distribution, dr)
    _protect_sheet(ws_resumen)
    _protect_sheet(ws_cases)
    _protect_sheet(ws_errors)
    _protect_sheet(ws_lists)

    _apply_tab_colors(workbook)

    output = io.BytesIO()
    workbook.save(output)

    file_name = f"{file_prefix}_{process_date}.xlsx"
    upload_path = f"{review_path}/{file_name}"
    upload_resp = await client.put_bytes(
        _build_content_endpoint(review_info["site_id"], review_info["drive_id"], upload_path),
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    validation_file_url: str | None = (
        upload_resp.get("webUrl") if isinstance(upload_resp, dict) else None
    )

    return {
        "process_id": process_id,
        "validation_file": file_name,
        "validation_file_path": upload_path,
        "validation_file_url": validation_file_url,
        "summary": {
            "pagos_banco": len(payment_cases) + len(error_records),
            "errores": len(error_records),
            "conditional_formatting": formatting_strategy,
        },
    }
