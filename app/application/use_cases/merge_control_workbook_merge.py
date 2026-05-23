"""
Lectura y actualización de control_merge_pdfs.xlsx para el flujo Merge (fila 2, hoja Procesos).
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook

from app.application.sharepoint_resolution import encode_graph_drive_path
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_COLUMNS,
    SHEET_NAME,
    apply_merge_control_worksheet_protection,
    merge_control_workbook_relative_path,
)
from app.domain.ports.graph import GraphApiPort

_MERGE_ALLOWED_ESTADOS = frozenset({"PENDIENTE_ASIENTOS", "ERROR_MERGE", "MERGE_PARCIAL"})


class MergeControlValidationError(Exception):
    """Condiciones de negocio del archivo de control que impiden iniciar o continuar Merge."""

    def __init__(self, error_code: str) -> None:
        self.error_code = error_code
        super().__init__(error_code)


def _col(name: str) -> int:
    return MERGE_CONTROL_COLUMNS.index(name) + 1


def _norm_estado(value: Any) -> str:
    return str(value or "").strip()


def _is_active_cell(value: Any) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    s = str(value).strip().lower()
    return s in ("true", "1", "yes", "si", "sí", "x")


def _headers_match(ws: Any) -> bool:
    for i, exp in enumerate(MERGE_CONTROL_COLUMNS, start=1):
        if str(ws.cell(row=1, column=i).value or "").strip() != exp:
            return False
    return True


def _content_put_endpoint(site_id: str, drive_id: str, rel_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(rel_path)}:/content"


def _content_delete_endpoint(site_id: str, drive_id: str, rel_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(rel_path)}:"


async def merge_control_download_workbook_bytes(
    graph: GraphApiPort, site_id: str, drive_id: str
) -> tuple[str, bytes]:
    rel = merge_control_workbook_relative_path().strip().strip("/")
    enc = encode_graph_drive_path(rel)
    raw = await graph.get_bytes(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content")
    return rel, raw


@dataclass(frozen=True)
class MergeControlRowSnapshot:
    estado_proceso: str
    is_active: bool
    historical_file_path: str
    email_pdf_path: str


def merge_control_parse_row2_validate_for_merge(raw: bytes) -> MergeControlRowSnapshot:
    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise MergeControlValidationError("merge_control_invalid_structure")
        ws = wb[SHEET_NAME]
        if not _headers_match(ws):
            raise MergeControlValidationError("merge_control_invalid_structure")
        if ws.max_row < 2:
            raise MergeControlValidationError("merge_control_invalid_structure")

        estado = _norm_estado(ws.cell(row=2, column=_col("EstadoProceso")).value)
        is_active = _is_active_cell(ws.cell(row=2, column=_col("IsActive")).value)
        hist = str(ws.cell(row=2, column=_col("HistoricalFilePath")).value or "").strip()
        pdf = str(ws.cell(row=2, column=_col("EmailPdfPath")).value or "").strip()

        if not is_active or estado not in _MERGE_ALLOWED_ESTADOS:
            raise MergeControlValidationError("merge_control_no_pending_process")
        if not hist:
            raise MergeControlValidationError("missing_historical_file_path")
        if not pdf:
            raise MergeControlValidationError("missing_email_pdf_path")

        return MergeControlRowSnapshot(
            estado_proceso=estado,
            is_active=is_active,
            historical_file_path=hist.strip().strip("/"),
            email_pdf_path=pdf.strip().strip("/"),
        )
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


def _workbook_to_bytes(wb: Any) -> bytes:
    apply_merge_control_worksheet_protection(wb[SHEET_NAME])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def merge_control_upload_row2_updates(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    control_rel: str,
    *,
    updates: dict[str, Any],
) -> None:
    """Descarga el control, aplica updates en columnas por nombre (fila 2), protege y sube."""
    enc = encode_graph_drive_path(control_rel.strip().strip("/"))
    raw = await graph.get_bytes(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content")
    wb = load_workbook(filename=io.BytesIO(raw), data_only=False)
    try:
        ws = wb[SHEET_NAME]
        if not _headers_match(ws) or ws.max_row < 2:
            raise ValueError("merge_control_invalid_structure")
        for key, val in updates.items():
            if key not in MERGE_CONTROL_COLUMNS:
                continue
            ws.cell(row=2, column=_col(key), value=val)
        out = _workbook_to_bytes(wb)
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()

    await graph.put_bytes(
        _content_put_endpoint(site_id, drive_id, control_rel),
        out,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def merge_control_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def merge_control_set_consolidando(
    graph: GraphApiPort, site_id: str, drive_id: str, control_rel: str
) -> None:
    await merge_control_upload_row2_updates(
        graph,
        site_id,
        drive_id,
        control_rel,
        updates={
            "EstadoProceso": "CONSOLIDANDO",
            "LastUpdatedAtProceso": merge_control_now_iso(),
        },
    )


async def merge_control_set_consolidado_success(
    graph: GraphApiPort, site_id: str, drive_id: str, control_rel: str, outputs_count: int
) -> None:
    await merge_control_upload_row2_updates(
        graph,
        site_id,
        drive_id,
        control_rel,
        updates={
            "EstadoProceso": "CONSOLIDADO",
            "IsActive": False,
            "MergeOutputCount": int(outputs_count),
            "MergeSkippedCount": 0,
            "LastErrorUserMessage": "",
            "LastErrorNextAction": "",
            "LastUpdatedAtProceso": merge_control_now_iso(),
        },
    )


async def merge_control_set_merge_parcial(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    control_rel: str,
    *,
    outputs_count: int,
    skipped_count: int,
    user_message: str,
    next_action: str,
) -> None:
    await merge_control_upload_row2_updates(
        graph,
        site_id,
        drive_id,
        control_rel,
        updates={
            "EstadoProceso": "MERGE_PARCIAL",
            "IsActive": True,
            "MergeOutputCount": int(outputs_count),
            "MergeSkippedCount": int(skipped_count),
            "LastErrorUserMessage": user_message,
            "LastErrorNextAction": next_action,
            "LastUpdatedAtProceso": merge_control_now_iso(),
        },
    )


async def merge_control_set_error_merge(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    control_rel: str,
    *,
    user_message: str,
    next_action: str,
) -> None:
    await merge_control_upload_row2_updates(
        graph,
        site_id,
        drive_id,
        control_rel,
        updates={
            "EstadoProceso": "ERROR_MERGE",
            "IsActive": True,
            "LastErrorUserMessage": user_message[:4000],
            "LastErrorNextAction": next_action[:4000],
            "LastUpdatedAtProceso": merge_control_now_iso(),
        },
    )


async def graph_delete_drive_item_by_path(
    graph: GraphApiPort, site_id: str, drive_id: str, rel_path: str
) -> None:
    await graph.delete(_content_delete_endpoint(site_id, drive_id, rel_path.strip().strip("/")))
