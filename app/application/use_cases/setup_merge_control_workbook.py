"""
Setup one-time (idempotente) del Excel de control para el flujo Merge PDFs.

Sube ``control_merge_pdfs.xlsx`` bajo la carpeta 00 CONTROL con hoja Procesos,
tabla tblControlMergePDFs y protección de hoja sin contraseña (celdas bloqueadas).
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

MERGE_CONTROL_FOLDER_RELATIVE_PATH = (
    "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL"
)
MERGE_CONTROL_WORKBOOK_RELATIVE_PATH = f"{MERGE_CONTROL_FOLDER_RELATIVE_PATH}/control_merge_pdfs.xlsx"

SHEET_NAME = "Procesos"
TABLE_DISPLAY_NAME = "tblControlMergePDFs"

MERGE_CONTROL_COLUMNS: tuple[str, ...] = (
    "Title",
    "EstadoProceso",
    "IsActive",
    "HistoricalFilePath",
    "EmailPdfPath",
    "MergeOutputCount",
    "MergeSkippedCount",
    "LastErrorUserMessage",
    "LastErrorNextAction",
    "CreatedAtProceso",
    "LastUpdatedAtProceso",
)

SECURITY_WARNING = (
    "El archivo fue creado con protección de hoja tipo Generate para evitar edición manual, "
    "pero la restricción de apertura debe manejarse con permisos de SharePoint si se requiere."
)

_FILL_HEADER = PatternFill(fill_type="solid", fgColor="002060")
_FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_BODY = Font(name="Calibri", size=11)
_BORDER = Border(
    left=Side(style="thin", color="C8C8C8"),
    right=Side(style="thin", color="C8C8C8"),
    top=Side(style="thin", color="C8C8C8"),
    bottom=Side(style="thin", color="C8C8C8"),
)
_ALIGN_HEADER = Alignment(vertical="center", horizontal="center", wrap_text=True)
_ALIGN_BODY = Alignment(vertical="center", horizontal="left", wrap_text=True)
_PROT_LOCKED = Protection(locked=True)


class MergeControlSetupError(Exception):
    """Error de negocio / Graph al preparar el workbook de control."""

    def __init__(
        self,
        *,
        http_status: int,
        user_message: str,
        next_action: str,
        technical_message: str,
    ) -> None:
        self.http_status = http_status
        self.user_message = user_message
        self.next_action = next_action
        self.technical_message = technical_message
        super().__init__(technical_message)


def _content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


def _item_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:"


async def _list_folder_children(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    parent_rel: str,
) -> list[dict[str, Any]]:
    if parent_rel.strip():
        enc = encode_graph_drive_path(parent_rel.strip().strip("/"))
        endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children"
    else:
        endpoint = f"/sites/{site_id}/drives/{drive_id}/root/children"
    resp = await graph.get(endpoint)
    return list(resp.get("value") or [])


async def _create_folder_child(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    parent_rel: str,
    folder_name: str,
) -> None:
    if parent_rel.strip():
        enc = encode_graph_drive_path(parent_rel.strip().strip("/"))
        endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children"
    else:
        endpoint = f"/sites/{site_id}/drives/{drive_id}/root/children"
    body: dict[str, Any] = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail",
    }
    try:
        await graph.post_json(endpoint, body)
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 409:
            logger.info("merge-control-setup: carpeta %r ya existía (409)", folder_name)
            return
        raise


async def ensure_merge_control_folder_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder_rel: str,
) -> None:
    parts = [p for p in folder_rel.strip().split("/") if p.strip()]
    acc: list[str] = []
    for part in parts:
        parent = "/".join(acc) if acc else ""
        kids = await _list_folder_children(graph, site_id, drive_id, parent)
        folder_names = {
            str(it.get("name", ""))
            for it in kids
            if "folder" in it and str(it.get("name", "")).strip()
        }
        if part not in folder_names:
            try:
                await _create_folder_child(graph, site_id, drive_id, parent, part)
            except httpx.HTTPStatusError as exc:
                code = exc.response.status_code if exc.response else 0
                body_txt = (exc.response.text if exc.response else "") or ""
                raise MergeControlSetupError(
                    http_status=502,
                    user_message="No se pudo crear la carpeta de control en SharePoint.",
                    next_action=(
                        "Cree la carpeta manualmente o verifique permisos de escritura en la ruta de validación."
                    ),
                    technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
                ) from exc
        acc.append(part)


async def _file_exists(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> bool:
    try:
        await graph.get_bytes(_content_endpoint(site_id, drive_id, path))
        return True
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return False
        raise


async def _drive_item_web_url(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> str | None:
    try:
        meta = await graph.get(_item_endpoint(site_id, drive_id, path))
        wu = meta.get("webUrl")
        return str(wu).strip() if isinstance(wu, str) and wu.strip() else None
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return None
        raise


def _column_widths() -> dict[int, float]:
    return {
        1: 28.0,
        2: 22.0,
        3: 12.0,
        4: 62.0,
        5: 62.0,
        6: 18.0,
        7: 20.0,
        8: 48.0,
        9: 48.0,
        10: 22.0,
        11: 22.0,
    }


def merge_control_workbook_relative_path() -> str:
    """Ruta relativa al root del drive del Excel de control Merge (env o default)."""
    p = os.getenv("GRAPH_MERGE_CONTROL_WORKBOOK_PATH", "").strip()
    return p or MERGE_CONTROL_WORKBOOK_RELATIVE_PATH


def apply_merge_control_worksheet_protection(ws: Any) -> None:
    """Bloquea celdas A1:K2 y activa protección de hoja sin contraseña (mismo perfil que setup)."""
    ncols = len(MERGE_CONTROL_COLUMNS)
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).protection = _PROT_LOCKED
    ws.protection.sheet = True


def _build_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass
    try:
        ws.sheet_properties.tabColor = Color(rgb="002060")
    except Exception:
        pass

    ncols = len(MERGE_CONTROL_COLUMNS)
    for col_idx, name in enumerate(MERGE_CONTROL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_HEADER
        cell.border = _BORDER

    row2_values: list[Any] = [
        "",
        "VACIO",
        "false",
        "",
        "",
        0,
        0,
        "",
        "",
        "",
        "",
    ]
    for col_idx, val in enumerate(row2_values, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = _FONT_BODY
        cell.alignment = _ALIGN_BODY
        cell.border = _BORDER

    for col_idx, width in _column_widths().items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}2"
    ws.freeze_panes = "A2"

    ref = f"A1:{get_column_letter(ncols)}2"
    tab = Table(displayName=TABLE_DISPLAY_NAME, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    apply_merge_control_worksheet_protection(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_row_matches(ws: Any) -> bool:
    for i, exp in enumerate(MERGE_CONTROL_COLUMNS, start=1):
        v = ws.cell(row=1, column=i).value
        if str(v or "").strip() != exp:
            return False
    return True


def _validate_and_repair_existing_workbook(data: bytes) -> tuple[list[str], bool, bool, bytes | None]:
    """
    Devuelve (warnings, structure_ok, excel_protection_applied, new_bytes_if_modified).
    Reparación conservadora: tabla formal faltante o protección de hoja desactivada;
    celdas A1:K2 forzadas a locked=True sin cambiar valores.
    """
    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    modified = False
    try:
        if SHEET_NAME not in wb.sheetnames:
            warnings.append("needs_manual_review: falta la hoja Procesos")
            return warnings, False, False, None

        ws = wb[SHEET_NAME]

        if not _header_row_matches(ws):
            warnings.append("needs_manual_review: encabezados no coinciden con el contrato esperado")
            return warnings, False, ws.protection.sheet is True, None

        if ws.max_row < 2:
            warnings.append("needs_manual_review: falta la fila 2 de control")
            return warnings, False, ws.protection.sheet is True, None

        if TABLE_DISPLAY_NAME not in ws.tables:
            ref = f"A1:{get_column_letter(len(MERGE_CONTROL_COLUMNS))}2"
            tab = Table(displayName=TABLE_DISPLAY_NAME, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)
            warnings.append("repaired: se añadió la tabla tblControlMergePDFs faltante")
            modified = True

        if not ws.protection.sheet:
            warnings.append("repaired: se activó protección de hoja")
            modified = True

        need_relock = False
        for r in (1, 2):
            for c in range(1, len(MERGE_CONTROL_COLUMNS) + 1):
                cell = ws.cell(row=r, column=c)
                if getattr(cell, "protection", None) is None or cell.protection.locked is not True:
                    need_relock = True
                    break
            if need_relock:
                break
        if need_relock:
            modified = True

        if modified:
            apply_merge_control_worksheet_protection(ws)

        struct_ok = _header_row_matches(ws) and ws.max_row >= 2 and TABLE_DISPLAY_NAME in ws.tables
        protection_on = ws.protection.sheet is True

        new_bytes: bytes | None = None
        if modified:
            buf = io.BytesIO()
            wb.save(buf)
            new_bytes = buf.getvalue()

        return warnings, struct_ok, protection_on, new_bytes
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


async def setup_merge_control_workbook(graph: GraphApiPort) -> dict[str, Any]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    if not site_search:
        raise GraphConfigError("Missing environment variable: GRAPH_SHAREPOINT_SITE_SEARCH")

    base = await resolve_sharepoint_path(graph, site_search, drive_name, MERGE_CONTROL_FOLDER_RELATIVE_PATH)
    site_id = base["site_id"]
    drive_id = base["drive_id"]

    try:
        await ensure_merge_control_folder_path(graph, site_id, drive_id, MERGE_CONTROL_FOLDER_RELATIVE_PATH)
    except MergeControlSetupError:
        raise
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        body_txt = (exc.response.text if exc.response else "") or ""
        if code == 403:
            raise MergeControlSetupError(
                http_status=403,
                user_message="No se pudo crear el archivo de control de Merge en SharePoint.",
                next_action=(
                    "Verifique que la aplicación tenga permisos suficientes para crear archivos en la carpeta "
                    "de validación."
                ),
                technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
            ) from exc
        raise MergeControlSetupError(
            http_status=502,
            user_message="No se pudo preparar la carpeta de control en SharePoint.",
            next_action="Revise permisos y el detalle técnico devuelto por Graph.",
            technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
        ) from exc

    file_path = MERGE_CONTROL_WORKBOOK_RELATIVE_PATH
    warnings: list[str] = []

    try:
        exists = await _file_exists(graph, site_id, drive_id, file_path)
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        body_txt = (exc.response.text if exc.response else "") or ""
        if code == 403:
            raise MergeControlSetupError(
                http_status=403,
                user_message="No se pudo crear el archivo de control de Merge en SharePoint.",
                next_action=(
                    "Verifique que la aplicación tenga permisos suficientes para crear archivos en la carpeta "
                    "de validación."
                ),
                technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
            ) from exc
        raise

    file_url: str | None = None
    excel_protection = False
    created = False

    if exists:
        raw = await graph.get_bytes(_content_endpoint(site_id, drive_id, file_path))
        w, struct_ok, excel_protection, new_bytes = _validate_and_repair_existing_workbook(raw)
        warnings.extend(w)
        if new_bytes is not None:
            try:
                resp = await graph.put_bytes(
                    _content_endpoint(site_id, drive_id, file_path),
                    new_bytes,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                if isinstance(resp, dict):
                    wu = resp.get("webUrl")
                    file_url = str(wu).strip() if isinstance(wu, str) and wu.strip() else None
            except httpx.HTTPStatusError as exc:
                code = exc.response.status_code if exc.response else 0
                body_txt = (exc.response.text if exc.response else "") or ""
                warnings.append(
                    f"needs_manual_review: no se pudo guardar reparación (HTTP {code}): {body_txt[:500]}"
                )
        if file_url is None:
            file_url = await _drive_item_web_url(graph, site_id, drive_id, file_path)
        if not struct_ok:
            warnings.append("needs_manual_review: estructura del archivo no válida o incompleta")
        return {
            "status": "success",
            "created": False,
            "file_path": file_path,
            "file_url": file_url,
            "sheet_name": SHEET_NAME,
            "table_name": TABLE_DISPLAY_NAME,
            "columns": list(MERGE_CONTROL_COLUMNS),
            "row_mode": "single_active_row",
            "excel_protection_applied": excel_protection,
            "sharepoint_permission_applied": False,
            "security_warning": SECURITY_WARNING,
            "warnings": warnings,
        }

    payload = _build_workbook_bytes()
    try:
        resp = await graph.put_bytes(
            _content_endpoint(site_id, drive_id, file_path),
            payload,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        body_txt = (exc.response.text if exc.response else "") or ""
        if code == 403:
            raise MergeControlSetupError(
                http_status=403,
                user_message="No se pudo crear el archivo de control de Merge en SharePoint.",
                next_action=(
                    "Verifique que la aplicación tenga permisos suficientes para crear archivos en la carpeta "
                    "de validación."
                ),
                technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
            ) from exc
        raise MergeControlSetupError(
            http_status=502,
            user_message="No se pudo subir el archivo de control a SharePoint.",
            next_action="Verifique espacio, bloqueos (423) o permisos en la biblioteca.",
            technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
        ) from exc

    if isinstance(resp, dict):
        wu = resp.get("webUrl")
        file_url = str(wu).strip() if isinstance(wu, str) and wu.strip() else None
    if not file_url:
        file_url = await _drive_item_web_url(graph, site_id, drive_id, file_path)

    created = True
    excel_protection = True

    return {
        "status": "success",
        "created": True,
        "file_path": file_path,
        "file_url": file_url,
        "sheet_name": SHEET_NAME,
        "table_name": TABLE_DISPLAY_NAME,
        "columns": list(MERGE_CONTROL_COLUMNS),
        "row_mode": "single_active_row",
        "excel_protection_applied": excel_protection,
        "sharepoint_permission_applied": False,
        "security_warning": SECURITY_WARNING,
        "warnings": warnings,
    }
