"""
Correo Graph (sendMail): lee el Excel del reporte Banco Bogotá (GRAPH_SHAREPOINT_FILE_PATH) para la fecha mínima
y la tabla completa del correo (todas las columnas y filas de datos de esa hoja).
El histórico de validación se obtiene **solo** por `historical_file_path` (ruta relativa al root del drive,
igual que `result.historical_file_path` de Finalize); no hay búsqueda automática por carpetas ni por fecha.

Abre el Excel indicado, hoja "Distribución",
para filas con «Validar Pago»=SI (libros nuevos) o, en históricos sin esa columna, token en estado
(`GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS`, p. ej. VALIDAR);
columna "Ruta" → PDFs vía `_collect_pdf_paths_from_ruta_cell`.
Remitente y destinatarios: Excel CORREOS.xlsx (GRAPH_VALIDAR_NOTIFY_CORREOS_XLSX_PATH): columnas EMISOR (primer email) y RECEPTORES.
Asunto: GRAPH_VALIDAR_NOTIFY_EMAIL_SUBJECT (defecto ABONOS BANCO BOGOTA). Cuerpo: saludo configurable + tabla.
PDF exportado (solo contenido del correo, sin fusionar extractos): GRAPH_VALIDAR_NOTIFY_EXPORT_EMAIL_PDF_FOLDER_PATH
(defecto .../02 COMWARE - VALIDACION PAGOS/05 EMAIL).
"""

from __future__ import annotations

import base64
import html
import logging
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Literal
from urllib.parse import parse_qs, quote, urlparse
from xml.sax.saxutils import escape

import httpx
from openpyxl import load_workbook
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_from_env
from app.application.use_cases.merge_control_workbook_notify import (
    update_merge_control_workbook_after_notify,
)
from app.application.use_cases.validate_payment_report import (
    _find_header_map,
    _get_col,
    _graph_download_by_path,
    _norm_key,
    _parse_excel_date,
)
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)


def _accent_fold_upper(s: str) -> str:
    s = _norm_key(s)
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def _normalize_graph_mail_address(raw: str) -> str:
    """Quita comillas y extrae correo de formas tipo Nombre <correo@dominio.com>."""
    s = raw.strip().strip('"').strip("'").strip()
    m = re.search(r"<([^<>\s]+@[^<>\s]+)>", s)
    if m:
        return m.group(1).strip()
    return s


def _collect_emails_from_cell(raw: Any) -> list[str]:
    """Extrae uno o varios correos de una celda (coma, punto y coma, salto de línea, texto libre)."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    seen: set[str] = set()
    out: list[str] = []

    def add_one(addr: str) -> None:
        e = _normalize_graph_mail_address(addr.strip())
        if not e or "@" not in e:
            return
        lk = e.lower()
        if lk not in seen:
            seen.add(lk)
            out.append(e)

    for chunk in re.split(r"[\n\r;,]+", s):
        chunk = chunk.strip()
        if not chunk:
            continue
        add_one(chunk)
    for m in re.finditer(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", s):
        add_one(m.group(0))
    return out


def _dedupe_emails_preserve_order(emails: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for e in emails:
        lk = e.lower()
        if lk not in seen:
            seen.add(lk)
            out.append(e)
    return out


def _find_correos_header_row(ws: Any) -> tuple[int, int, int] | None:
    """Fila de encabezados con columnas EMISOR y RECEPTORES; devuelve (fila, col_emisor, col_receptores)."""
    max_col = ws.max_column or 1
    max_scan = min(ws.max_row or 1, 80)
    for row_idx in range(1, max_scan + 1):
        raw_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        cells = [_norm_key(str(v)) if v is not None and str(v).strip() else "" for v in raw_vals]
        if not any(cells):
            continue
        header_map: dict[str, int] = {}
        for col, name in enumerate(cells, start=1):
            if name:
                header_map[_accent_fold_upper(name)] = col
        col_em = header_map.get("EMISOR")
        col_rec = header_map.get("RECEPTORES") or header_map.get("RECEPTOR")
        if col_em and col_rec:
            return row_idx, col_em, col_rec
    return None


def _parse_correos_workbook(wb: Any) -> tuple[str, list[str]]:
    """Primer email en EMISOR = remitente; todos los emails de RECEPTORES (todas las filas) = destinatarios."""
    last_error: str | None = None
    for ws in wb.worksheets:
        found = _find_correos_header_row(ws)
        if not found:
            continue
        h_row, col_em, col_rec = found
        last = ws.max_row or h_row
        sender: str | None = None
        recipients: list[str] = []
        for r in range(h_row + 1, last + 1):
            if sender is None:
                em = _collect_emails_from_cell(ws.cell(row=r, column=col_em).value)
                if em:
                    sender = em[0]
            recipients.extend(_collect_emails_from_cell(ws.cell(row=r, column=col_rec).value))
        recipients = _dedupe_emails_preserve_order(recipients)
        if sender and recipients:
            return sender, recipients
        if sender and not recipients:
            last_error = f"Hoja {ws.title!r}: hay EMISOR pero RECEPTORES no tiene correos válidos."
        elif not sender and recipients:
            last_error = f"Hoja {ws.title!r}: hay RECEPTORES pero EMISOR no tiene un correo válido."
        else:
            last_error = f"Hoja {ws.title!r}: EMISOR y RECEPTORES sin datos útiles."
    msg = (
        'No se encontró en CORREOS.xlsx una hoja con columnas "EMISOR" y "RECEPTORES" '
        "y al menos un remitente y un destinatario."
    )
    if last_error:
        msg += f" Detalle: {last_error}"
    raise ValueError(msg)


_CORREOS_XLSX_DEFAULT = (
    "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/00 CONTROL/CORREOS.xlsx"
)


async def _load_sender_and_recipients_from_correos_xlsx(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
) -> tuple[str, list[str]]:
    rel = os.getenv("GRAPH_VALIDAR_NOTIFY_CORREOS_XLSX_PATH", "").strip().strip("/") or _CORREOS_XLSX_DEFAULT
    data = await _graph_download_by_path(graph, site_id, drive_id, rel)
    wb = load_workbook(filename=BytesIO(data), data_only=True)
    try:
        return _parse_correos_workbook(wb)
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


@dataclass(frozen=True)
class ValidarExtractosNotifyResult:
    report_date: str
    historico_excel_path: str
    historical_file_path: str
    historical_file_source: Literal["explicit"]
    rows_included: int
    subject: str
    attachments_count: int
    graph_sendmail_http_status: int
    mail_sender: str
    mail_to: str
    email_pdf_path: str | None = None
    email_pdf_error: str | None = None
    merge_control_updated: bool = False
    merge_control_file_path: str | None = None
    merge_control_status: str | None = None
    merge_control_warning: str | None = None
    merge_control_error_code: str | None = None


def _endpoint_and_params_from_next_link(next_link: str) -> tuple[str, dict[str, str] | None]:
    parsed = urlparse(next_link)
    path = parsed.path or ""
    if "/v1.0/" in path:
        endpoint = path.split("/v1.0/", 1)[1].lstrip("/")
    else:
        endpoint = path.lstrip("/")
    if not parsed.query:
        return endpoint, None
    flat = {k: v[0] for k, v in parse_qs(parsed.query).items() if v}
    return endpoint, flat or None


async def _list_drive_folder_children(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder_path: str,
) -> list[dict[str, Any]]:
    encoded = encode_graph_drive_path(folder_path.strip().strip("/"))
    path = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:/children"
    items: list[dict[str, Any]] = []
    endpoint = path.lstrip("/")
    params: dict[str, str] | None = {"$top": "200"}
    while True:
        resp = await graph.get(endpoint, params=params)
        items.extend(resp.get("value") or [])
        next_link = resp.get("@odata.nextLink")
        if not next_link:
            break
        endpoint, params = _endpoint_and_params_from_next_link(next_link)
    return items


def _norm_sheet_name(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s.strip().casefold())
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def _find_distribucion_sheet(wb: Any) -> Any:
    for ws in wb.worksheets:
        if _norm_sheet_name(ws.title) == "distribucion":
            return ws
    raise ValueError(
        'No se encontró una hoja llamada "Distribución" / "Distribucion" (sin importar mayúsculas o tilde).'
    )


_ESTADO_DISTRIB_HEADER_KEYS = frozenset({"ESTADO", "ESTADO LINEA", "ESTADOLINEA"})


def _is_distrib_estado_header(cell_norm: str) -> bool:
    if not cell_norm:
        return False
    key = _accent_fold_upper(cell_norm)
    if key in _ESTADO_DISTRIB_HEADER_KEYS:
        return True
    compact = key.replace(" ", "")
    if compact in ("ESTADOLINEA", "ESTADOPAGO"):
        return True
    return compact.startswith("ESTADO") and "LINEA" in compact


def _is_distrib_ruta_header(cell_norm: str) -> bool:
    if not cell_norm:
        return False
    return _accent_fold_upper(cell_norm) in ("RUTA", "RUTAS")


def _find_distribucion_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    max_col = ws.max_column or 1
    max_scan = min(ws.max_row or 1, 50)
    saw_estado_only = False
    saw_ruta_only = False
    for row_idx in range(1, max_scan + 1):
        raw_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        cells = [_norm_key(str(v)) if v is not None and str(v).strip() else "" for v in raw_vals]
        if not any(cells):
            continue
        has_estado = any(_is_distrib_estado_header(c) for c in cells if c)
        has_ruta = any(_is_distrib_ruta_header(c) for c in cells if c)
        if has_estado and not has_ruta:
            saw_estado_only = True
        if has_ruta and not has_estado:
            saw_ruta_only = True
        if has_estado and has_ruta:
            header_map: dict[str, int] = {}
            for col, name in enumerate(cells, start=1):
                if name:
                    key = _accent_fold_upper(name)
                    header_map[key] = col
            return row_idx, header_map
    if saw_estado_only:
        raise ValueError("missing_distribucion_route_column")
    if saw_ruta_only:
        raise ValueError("missing_distribucion_status_column")
    raise ValueError("missing_distribucion_headers")


def _get_col_distrib(header_map: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        key = _accent_fold_upper(c)
        if key in header_map:
            return header_map[key]
    return None


def distrib_row_included_for_validar_extractos(
    ws: Any,
    row: int,
    header_map: dict[str, int],
    legacy_estado_token: str,
) -> bool:
    """
    Filas incluidas en Notify/Merge para PDFs:
    - Si existe «Validar Pago»: debe ser SI; opcional filtro substring en «Estado Pago»
      (GRAPH_VALIDAR_ESTADO_PAGO_CONTAINS).
    - Si no hay «Validar Pago» (histórico viejo): token en columna de estado
      (GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS, p. ej. VALIDAR).
    """
    col_vp = _get_col_distrib(header_map, "Validar Pago", "Validar pago", "VALIDAR PAGO")
    col_est = _get_col_distrib(
        header_map,
        "Estado Pago",
        "Estado pago",
        "Estado",
        "Estado línea",
        "Estado linea",
        "ESTADO LINEA",
    )
    if col_vp is not None:
        raw_vp = ws.cell(row=row, column=col_vp).value
        svp = str(raw_vp or "").strip().upper()
        if svp not in ("SI", "SÍ"):
            return False
        extra = os.getenv("GRAPH_VALIDAR_ESTADO_PAGO_CONTAINS", "").strip()
        if not extra:
            return True
        if col_est is None:
            return False
        raw_est = ws.cell(row=row, column=col_est).value
        return _estado_linea_tiene_palabra_clave(raw_est, extra)
    if col_est is None:
        return False
    raw_est = ws.cell(row=row, column=col_est).value
    return _estado_linea_tiene_palabra_clave(raw_est, legacy_estado_token)


def _parse_bank_report_table_and_min_date(excel_bytes: bytes) -> tuple[date, list[str], list[list[str]]]:
    """
    Una sola lectura del reporte Banco Bogotá (GRAPH_SHAREPOINT_FILE_PATH, hoja 1):
    fecha mínima en columna Fecha; tabla solo con columnas que tengan algún dato y solo filas
    donde todas las celdas visibles estén llenas (sin vacíos).
    """
    wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    try:
        ws = wb.worksheets[0]
        header_row, header_map = _find_header_map(ws)
        col_fecha = _get_col(header_map, "Fecha")
        if col_fecha is None:
            raise ValueError("No hay columna Fecha en el reporte Banco Bogotá.")

        dates: list[date] = []
        last_r = ws.max_row or header_row
        for r in range(header_row + 1, last_r + 1):
            raw = ws.cell(row=r, column=col_fecha).value
            if raw is None or not str(raw).strip():
                continue
            try:
                dates.append(_parse_excel_date(raw).date())
            except Exception:
                continue
        if not dates:
            raise ValueError("No se pudo obtener ninguna fecha válida de la columna Fecha del reporte.")
        report_d = min(dates)

        ncols = ws.max_column or 0
        if ncols < 1:
            raise ValueError("El reporte Banco Bogotá no tiene columnas.")

        headers = [_excel_cell_display(ws.cell(row=header_row, column=c).value) for c in range(1, ncols + 1)]
        candidate_rows: list[list[str]] = []
        for r in range(header_row + 1, last_r + 1):
            row_vals = [_excel_cell_display(ws.cell(row=r, column=c).value) for c in range(1, ncols + 1)]
            if any(v.strip() for v in row_vals):
                candidate_rows.append(row_vals)

        if not candidate_rows:
            raise ValueError(
                "El reporte Banco Bogotá no tiene filas de datos bajo los encabezados."
            )

        kept_j = [
            j
            for j in range(ncols)
            if (headers[j] or "").strip()
            or any((row[j] if j < len(row) else "").strip() for row in candidate_rows)
        ]
        if not kept_j:
            raise ValueError("No quedó ninguna columna con datos en el reporte Banco Bogotá.")

        narrow_headers = [headers[j] for j in kept_j]
        bank_rows: list[list[str]] = []
        for row in candidate_rows:
            narrow = [(row[j] if j < len(row) else "") for j in kept_j]
            if len(narrow) == len(kept_j) and all((c or "").strip() for c in narrow):
                bank_rows.append(narrow)

        if not bank_rows:
            raise ValueError(
                "No hay filas en el reporte Banco Bogotá con todas las celdas llenas "
                "(cada columna con dato debe tener valor en la fila)."
            )
        return report_d, narrow_headers, bank_rows
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


def _historico_root_path() -> str:
    p = os.getenv("GRAPH_VALIDAR_NOTIFY_HISTORICO_PATH", "").strip()
    if p:
        return p.strip().strip("/")
    p2 = os.getenv("GRAPH_PAYMENT_VALIDATION_HISTORY_PATH", "").strip()
    if not p2:
        raise GraphConfigError(
            "Define GRAPH_VALIDAR_NOTIFY_HISTORICO_PATH o GRAPH_PAYMENT_VALIDATION_HISTORY_PATH "
            "(carpeta 02 HISTORICO bajo validación de pagos)."
        )
    return p2.strip().strip("/")


def _date_folder_tokens(d: date) -> list[str]:
    """Tokens de carpeta/archivo a probar (env + formatos extra por defecto, sin duplicados)."""
    env_line = os.getenv("GRAPH_VALIDAR_NOTIFY_HISTORICO_DATE_FORMATS", "").strip()
    user_fmts = [p.strip() for p in env_line.split(",") if p.strip()] if env_line else []
    builtins = (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d%m%Y",
        "%Y%m%d",
        "%m-%d-%Y",
        "%d.%m.%Y",
        "%Y.%m.%d",
    )
    fmts = user_fmts + [f for f in builtins if f not in user_fmts]
    seen: set[str] = set()
    out: list[str] = []
    for fmt in fmts:
        try:
            t = d.strftime(fmt)
        except ValueError:
            continue
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    if not out:
        out = [d.strftime("%Y-%m-%d")]
    return out


def _xlsx_files_in(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        it
        for it in items
        if "file" in it
        and "folder" not in it
        and str(it.get("name", "")).lower().endswith(".xlsx")
        and not str(it.get("name", "")).startswith("~$")
    ]


def _name_matches_date_tokens(filename: str, tokens: list[str], iso: str) -> bool:
    lower = filename.lower()
    if iso.lower() in lower:
        return True
    return any(t.lower() in lower for t in tokens if len(t) >= 4)


def _is_cartera_valida_workbook(filename: str, tokens: list[str], iso: str) -> bool:
    """Nombre tipo cartera_valida… / cartera_validada… con la fecha del reporte en el archivo."""
    lower = str(filename or "").lower()
    if "cartera_valida" not in lower:
        return False
    return _name_matches_date_tokens(filename, tokens, iso)


def _pick_xlsx_from_list(
    xlsx: list[dict[str, Any]],
    *,
    folder_path: str,
    name_template: str,
    report_date: date,
    tokens: list[str],
) -> str | None:
    if not xlsx:
        return None
    iso = report_date.strftime("%Y-%m-%d")
    if name_template:
        for tok in tokens:
            try:
                expected = name_template.format(
                    fecha=tok,
                    y=report_date.year,
                    m=report_date.month,
                    d=report_date.day,
                    iso=iso,
                )
            except Exception:
                expected = name_template
            match = next((it for it in xlsx if it.get("name") == expected), None)
            if match:
                return f"{folder_path}/{match['name']}"

    cartera_ok = [
        it
        for it in xlsx
        if _is_cartera_valida_workbook(str(it.get("name", "")), tokens, iso)
    ]
    if len(cartera_ok) == 1:
        m = cartera_ok[0]
        return f"{folder_path}/{m['name']}"
    if len(cartera_ok) > 1:
        names_dup = ", ".join(sorted(str(it.get("name", "")) for it in cartera_ok))
        raise ValueError(
            f"Hay varios Excel de histórico con nombre que incluye cartera_valida/cartera_validada "
            f"y la fecha del reporte ({iso}) en {folder_path!r}: {names_dup}. "
            "Deja solo uno o renombra."
        )

    if len(xlsx) == 1:
        only = xlsx[0]
        return f"{folder_path}/{only['name']}"

    names_all = ", ".join(sorted(str(it.get("name", "")) for it in xlsx))
    raise ValueError(
        f"Hay {len(xlsx)} archivos .xlsx en {folder_path!r} y no se puede elegir uno de forma única: {names_all}. "
        f"Usa un solo archivo cuyo nombre incluya cartera_valida o cartera_validada y la fecha del reporte ({iso}), "
        "o define GRAPH_VALIDAR_NOTIFY_HISTORICO_EXCEL_NAME con el nombre exacto."
    )


async def _resolve_historico_excel_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    report_date: date,
) -> str:
    root = _historico_root_path()
    name_template = os.getenv("GRAPH_VALIDAR_NOTIFY_HISTORICO_EXCEL_NAME", "").strip()
    tokens = _date_folder_tokens(report_date)
    iso = report_date.strftime("%Y-%m-%d")

    # 1) Subcarpeta = token exacto bajo HISTORICO
    for token in tokens:
        folder = f"{root}/{token}"
        try:
            children = await _list_drive_folder_children(graph, site_id, drive_id, folder)
        except Exception as exc:
            logger.debug("histórico: no se listó %s: %s", folder, exc)
            continue
        picked = _pick_xlsx_from_list(
            _xlsx_files_in(children),
            folder_path=folder,
            name_template=name_template,
            report_date=report_date,
            tokens=tokens,
        )
        if picked:
            return picked

    # 2) Listar raíz HISTORICO
    try:
        root_children = await _list_drive_folder_children(graph, site_id, drive_id, root)
    except Exception as exc:
        raise ValueError(
            f"No se pudo listar la carpeta de histórico {root!r}: {exc}. "
            "Revisa permisos (Files.Read) y la ruta en GRAPH_PAYMENT_VALIDATION_HISTORY_PATH."
        ) from exc

    # 2a) Excel en la raíz cuyo nombre incluye la fecha (p. ej. cartera_validada_2025-12-30.xlsx)
    root_xlsx = _xlsx_files_in(root_children)
    dated_root = [it for it in root_xlsx if _name_matches_date_tokens(str(it.get("name", "")), tokens, iso)]
    if dated_root:
        folder_path = root
        picked = _pick_xlsx_from_list(
            dated_root,
            folder_path=folder_path,
            name_template=name_template,
            report_date=report_date,
            tokens=tokens,
        )
        if picked:
            return picked

    # 2b) Carpeta cuyo nombre coincide con un token o contiene la fecha ISO
    for it in root_children:
        if "folder" not in it:
            continue
        fname = str(it.get("name", ""))
        fold_l = fname.lower()
        token_hit = any(fold_l == t.lower() for t in tokens) or iso.lower() in fold_l
        if not token_hit:
            continue
        sub = f"{root}/{fname}"
        try:
            subc = await _list_drive_folder_children(graph, site_id, drive_id, sub)
        except Exception as exc:
            logger.debug("histórico: no se listó subcarpeta %s: %s", sub, exc)
            continue
        picked = _pick_xlsx_from_list(
            _xlsx_files_in(subc),
            folder_path=sub,
            name_template=name_template,
            report_date=report_date,
            tokens=tokens,
        )
        if picked:
            return picked

    # 2c) Cualquier subcarpeta de primer nivel: buscar .xlsx cuyo nombre lleve la fecha
    for it in root_children:
        if "folder" not in it:
            continue
        fname = str(it.get("name", ""))
        sub = f"{root}/{fname}"
        try:
            subc = await _list_drive_folder_children(graph, site_id, drive_id, sub)
        except Exception:
            continue
        inner = [
            x
            for x in _xlsx_files_in(subc)
            if _name_matches_date_tokens(str(x.get("name", "")), tokens, iso)
        ]
        picked = _pick_xlsx_from_list(
            inner,
            folder_path=sub,
            name_template=name_template,
            report_date=report_date,
            tokens=tokens,
        )
        if picked:
            return picked

    # Mensaje útil: qué se probó
    tried = ", ".join(tokens[:12])
    if len(tokens) > 12:
        tried += ", …"
    sample = ", ".join(str(x.get("name", "")) for x in root_children[:15])
    raise ValueError(
        f"No hay Excel (.xlsx) en histórico para la fecha del reporte {iso}. "
        f"Carpeta base: {root!r}. "
        f"Subcarpetas/nombres probados (parcial): {tried}. "
        f"Contenido listado en la raíz (primeros ítems): {sample or '(vacío o sin acceso)'}. "
        "Crea una subcarpeta con uno de esos nombres de fecha y un .xlsx dentro, "
        "o pon el .xlsx en la raíz de HISTORICO con la fecha en el nombre y prefijo cartera_valida/cartera_validada, "
        "o define GRAPH_VALIDAR_NOTIFY_HISTORICO_EXCEL_NAME con el nombre exacto."
    )


def _estado_linea_tiene_palabra_clave(raw: Any, palabra: str) -> bool:
    """
    True solo si la palabra clave aparece como token (palabra completa), no como substring
    (evita coincidencias en 'NOVALIDAR', etc.).
    """
    key = _norm_key(palabra)
    if not key:
        return False
    text = _norm_key(str(raw or ""))
    if not text:
        return False
    if text == key:
        return True
    tokens = [t for t in re.split(r"[^\w]+", text, flags=re.UNICODE) if t]
    return key in tokens


async def _graph_drive_item_metadata(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    rel_path: str,
) -> dict[str, Any] | None:
    """Metadatos del ítem en la ruta relativa al drive (sin /content)."""
    try:
        enc = encode_graph_drive_path(rel_path.strip().strip("/"))
        return await graph.get(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:")
    except Exception:
        return None


def _ruta_fragments_from_cell(raw: Any) -> list[str]:
    """Varias rutas separadas por coma o salto de línea."""
    if raw is None:
        return []
    s = str(raw).strip().replace("\\", "/")
    if not s:
        return []
    return [p.strip().strip("/").replace("\\", "/") for p in re.split(r"[\n\r;,]+", s) if p.strip()]


async def _resolve_pdf_paths_for_ruta_fragment(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    rel_norm: str,
) -> list[str]:
    """
    A partir de un fragmento de columna Ruta (relativo al root del drive):
    - si termina en .pdf y existe como archivo → esa ruta;
    - si no, prueba la misma ruta + .pdf;
    - si es carpeta → todos los .pdf hijos directos.
    """
    rel_norm = rel_norm.strip().strip("/")
    if not rel_norm:
        return []
    out: list[str] = []

    if rel_norm.lower().endswith(".pdf"):
        meta = await _graph_drive_item_metadata(graph, site_id, drive_id, rel_norm)
        if meta and "folder" not in meta:
            out.append(rel_norm)
        return out

    alt = rel_norm + ".pdf"
    meta_alt = await _graph_drive_item_metadata(graph, site_id, drive_id, alt)
    if meta_alt and "folder" not in meta_alt:
        out.append(alt)
        return out

    try:
        kids = await _list_drive_folder_children(graph, site_id, drive_id, rel_norm)
    except Exception:
        return out

    for it in kids:
        if "folder" in it:
            continue
        if "file" not in it:
            continue
        name = str(it.get("name", ""))
        if not name.lower().endswith(".pdf") or name.startswith("~$"):
            continue
        out.append(f"{rel_norm}/{name}".replace("//", "/"))
    return out


async def _collect_pdf_paths_from_ruta_cell(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    ruta_raw: Any,
) -> list[str]:
    paths: list[str] = []
    for frag in _ruta_fragments_from_cell(ruta_raw):
        paths.extend(await _resolve_pdf_paths_for_ruta_fragment(graph, site_id, drive_id, frag))
    return paths


def _excel_cell_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _attach_max_bytes() -> int:
    try:
        return max(0, int(os.getenv("GRAPH_MAIL_ATTACH_MAX_BYTES", "20000000")))
    except ValueError:
        return 20_000_000


_PDF_MONTHS_ES = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]

_PDF_WEEKDAYS_ES = [
    "lunes",
    "martes",
    "miércoles",
    "jueves",
    "viernes",
    "sábado",
    "domingo",
]


def _format_spanish_datetime(dt: datetime) -> str:
    weekday = _PDF_WEEKDAYS_ES[dt.weekday()]
    month = _PDF_MONTHS_ES[dt.month - 1]
    day = dt.day
    year = dt.year
    hour24 = dt.hour
    ampm = "a. m." if hour24 < 12 else "p. m."
    hour12 = hour24 % 12
    if hour12 == 0:
        hour12 = 12
    minutes = f"{dt.minute:02d}"
    return f"{weekday}, {day} de {month} de {year} {hour12}:{minutes} {ampm}"


def _split_saludo(body_intro: str) -> tuple[str, str]:
    # Intenta separar "Buenos días." del resto para que se vea como en la foto.
    s = (body_intro or "").strip()
    m = re.match(r"(?i)^(buenos\s+di[íi]as\.?)\s*(.*)$", s)
    if not m:
        return s, ""
    saludo = m.group(1)
    resto = (m.group(2) or "").strip()
    return saludo, resto


def _rp_pdf(text: str) -> str:
    """Texto seguro para Paragraph de ReportLab (subset tipo HTML)."""
    return escape(str(text or "")).replace("\n", "<br/>")


def _sanitize_pdf_filename_component(name: str) -> str:
    """Quita caracteres no válidos en nombre de archivo en Windows/SharePoint."""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name.strip())
    return cleaned.strip(" .") or "export.pdf"


def _cover_pdf_bytes_reportlab(
    *,
    sender: str,
    sent_at: datetime,
    to_addr: str,
    cc_list: list[str],
    subject: str,
    attachment_names: list[str],
    body_intro: str,
    bank_headers: list[str],
    bank_rows: list[list[str]],
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    saludo, resto = _split_saludo(body_intro)
    cc_line = ", ".join(cc_list) if cc_list else ""
    attach_line = "; ".join(attachment_names) if attachment_names else "(sin adjuntos)"
    sent_at_str = _format_spanish_datetime(sent_at)

    buf = BytesIO()
    left_m = right_m = top_m = bottom_m = 54
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=right_m,
        leftMargin=left_m,
        topMargin=top_m,
        bottomMargin=bottom_m,
    )
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    story: list[Any] = []

    def _line(label: str, value: str) -> None:
        story.append(Paragraph(f"<b>{_rp_pdf(label)}</b> {_rp_pdf(value)}", normal))

    _line("De:", sender)
    _line("Enviado el:", sent_at_str)
    _line("Para:", to_addr)
    _line("CC:", cc_line)
    _line("Asunto:", subject)
    _line("Datos adjuntos:", attach_line)
    story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph(f"<b>{_rp_pdf(saludo)}</b>", normal))
    if resto:
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph(_rp_pdf(resto), normal))
    story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph("<b>Reporte de pagos (Banco Bogotá):</b>", normal))
    story.append(Spacer(1, 0.15 * cm))

    ncols = max(len(bank_headers), 1)
    content_w_pt = float(A4[0]) - float(left_m) - float(right_m)
    col_w = content_w_pt / float(ncols)
    col_widths = [col_w] * ncols

    hdr = [Paragraph(f"<b>{_rp_pdf(h)}</b>", normal) for h in bank_headers]
    data: list[list[Any]] = [hdr]
    for row in bank_rows:
        padded = list(row) + [""] * (ncols - len(row))
        padded = padded[:ncols]
        data.append([Paragraph(_rp_pdf(cell), normal) for cell in padded])
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ADD8E6")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


async def _pdf_attachments_from_drive_paths(
    graph: GraphApiPort,
    *,
    site_id: str,
    drive_id: str,
    paths_ordered: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Descarga PDFs por ruta relativa al root del drive (Graph) y arma adjuntos sendMail."""
    attachments: list[dict[str, Any]] = []
    failures: list[str] = []
    max_b = _attach_max_bytes()
    used_names: dict[str, int] = {}
    seen_path: set[str] = set()

    for raw in paths_ordered:
        key = raw.strip().strip("/")
        if not key or key in seen_path:
            continue
        seen_path.add(key)
        try:
            data = await _graph_download_by_path(graph, site_id, drive_id, key)
        except Exception as exc:
            logger.warning("validar extractos: no se pudo descargar PDF %s: %s", key[:120], exc)
            failures.append(f"{key.rsplit('/', 1)[-1]} ({exc})")
            continue
        base = key.rsplit("/", 1)[-1] or "extracto.pdf"
        if max_b and len(data) > max_b:
            failures.append(f"{base} (tamaño > {max_b})")
            continue
        n = used_names.get(base, 0) + 1
        used_names[base] = n
        name = base if n == 1 else f"{base.rsplit('.', 1)[0]}_{n}.pdf"
        b64 = base64.standard_b64encode(data).decode("ascii")
        attachments.append(
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": name,
                "contentType": "application/pdf",
                "contentBytes": b64,
            }
        )
    return attachments, failures


def _build_html(
    *,
    intro: str,
    bank_headers: list[str],
    bank_rows: list[list[str]],
) -> str:
    thead = "<tr>" + "".join(f"<th>{html.escape(h)}</th>" for h in bank_headers) + "</tr>"
    tbody = ""
    ncols = len(bank_headers)
    for row in bank_rows:
        cells = list(row) + [""] * (ncols - len(row))
        cells = cells[:ncols]
        tbody += "<tr>" + "".join(f"<td>{html.escape(c)}</td>" for c in cells) + "</tr>"
    table = (
        "<p><strong>Reporte de pagos (Banco Bogotá):</strong></p>"
        f'<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse">'
        f"{thead}<tbody>{tbody}</tbody></table>"
    )
    return f"<html><body>{intro}{table}</body></html>"


async def send_validar_extractos_notification_email(
    graph: GraphApiPort,
    *,
    historical_file_path: str | None,
    to_override: str | None = None,
    cc_override: str | None = None,
) -> ValidarExtractosNotifyResult:
    estado_filtro = os.getenv("GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS", "VALIDAR").strip()
    if not estado_filtro:
        estado_filtro = "VALIDAR"

    hist_raw = (historical_file_path or "").strip()
    historico_rel = hist_raw.strip("/")
    if not historico_rel:
        raise ValueError("missing_historical_file_path")

    banco = os.getenv("GRAPH_NOTIFY_BANK_NAME", "BANCO BOGOTA").strip() or "BANCO BOGOTA"

    ctx = await resolve_sharepoint_from_env(graph)
    site_id = ctx["site_id"]
    drive_id = ctx["drive_id"]

    sender, to_recipients = await _load_sender_and_recipients_from_correos_xlsx(graph, site_id, drive_id)
    if (to_override or "").strip():
        to_recipients = _dedupe_emails_preserve_order(
            _collect_emails_from_cell(to_override.strip())
        )
    to_recipients = [e for e in to_recipients if e.lower() != sender.lower()]
    if not to_recipients:
        raise ValueError(
            "No hay destinatarios: columna RECEPTORES vacía o sin correos válidos en CORREOS.xlsx "
            "(o revisa el campo `to` del body de la API si lo usaste como sustituto)."
        )
    to_display = "; ".join(to_recipients)

    cc_list: list[str] = []
    if cc_override is not None and str(cc_override).strip():
        cc_list = _dedupe_emails_preserve_order(_collect_emails_from_cell(cc_override))

    report_bytes = await graph.get_bytes(
        f"/sites/{site_id}/drives/{drive_id}/root:/{ctx['path_encoded']}:/content"
    )
    report_d, bank_headers, bank_rows = _parse_bank_report_table_and_min_date(report_bytes)
    fecha_str = report_d.strftime("%d/%m/%Y")

    subject = (
        os.getenv("GRAPH_VALIDAR_NOTIFY_EMAIL_SUBJECT", "").strip() or "ABONOS BANCO BOGOTA"
    )
    _body_intro_default = (
        "Buenos días. El día {fecha} ingresaron a la cuenta {banco} los siguientes valores, "
        "que corresponden a:"
    )
    body_intro_tpl = (
        os.getenv("GRAPH_VALIDAR_NOTIFY_BODY_INTRO_TEMPLATE", "").strip()
        or os.getenv("GRAPH_VALIDAR_NOTIFY_SUBJECT_TEMPLATE", "").strip()
        or _body_intro_default
    )
    body_intro = body_intro_tpl.format(fecha=fecha_str, banco=banco)

    try:
        hist_bytes = await _graph_download_by_path(graph, site_id, drive_id, historico_rel)
    except httpx.HTTPStatusError as exc:
        detail = (exc.response.text or "").strip()[:4000]
        tech = f"HTTP {exc.response.status_code} url={exc.request.url!s} detail={detail!r}"
        raise ValueError(f"historical_file_not_found|{tech}") from exc

    pdf_paths_ordered: list[str] = []

    wb = load_workbook(filename=BytesIO(hist_bytes), data_only=True)
    try:
        ws = _find_distribucion_sheet(wb)
        h_row, header_map = _find_distribucion_header_row(ws)
        col_estado = _get_col_distrib(
            header_map,
            "Estado Pago",
            "Estado pago",
            "Estado",
            "Estado línea",
            "Estado linea",
            "ESTADO LINEA",
        )
        col_ruta = _get_col_distrib(header_map, "Ruta", "RUTA", "Rutas", "RUTAS")

        col_vp = _get_col_distrib(header_map, "Validar Pago", "Validar pago", "VALIDAR PAGO")
        if col_estado is None and col_vp is None:
            raise ValueError("missing_distribucion_status_column")
        if col_ruta is None:
            raise ValueError("missing_distribucion_route_column")

        seen_pdf_paths: set[str] = set()

        last = ws.max_row or h_row
        for r in range(h_row + 1, last + 1):
            if not distrib_row_included_for_validar_extractos(ws, r, header_map, estado_filtro):
                continue

            cell_ruta = ws.cell(row=r, column=col_ruta)
            for p in await _collect_pdf_paths_from_ruta_cell(
                graph, site_id, drive_id, cell_ruta.value
            ):
                np = p.strip().strip("/")
                if np and np not in seen_pdf_paths:
                    seen_pdf_paths.add(np)
                    pdf_paths_ordered.append(np)

    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()

    attach_on = os.getenv("GRAPH_VALIDAR_NOTIFY_ATTACH_PDFS", "true").strip().lower() in (
        "1",
        "true",
        "yes",
        "si",
        "sí",
    )
    attachments: list[dict[str, Any]] = []
    failures: list[str] = []
    if attach_on and pdf_paths_ordered:
        attachments, failures = await _pdf_attachments_from_drive_paths(
            graph,
            site_id=site_id,
            drive_id=drive_id,
            paths_ordered=pdf_paths_ordered,
        )

    if failures:
        for msg in failures[:12]:
            logger.warning("validar extractos notify: %s", msg)

    intro_html = f"<p>{html.escape(body_intro)}</p>"
    html_body = _build_html(intro=intro_html, bank_headers=bank_headers, bank_rows=bank_rows)

    email_pdf_path: str | None = None
    email_pdf_error: str | None = None
    message: dict[str, Any] = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html_body},
        "toRecipients": [{"emailAddress": {"address": a}} for a in to_recipients],
    }
    if cc_list:
        message["ccRecipients"] = [{"emailAddress": {"address": cc}} for cc in cc_list]
    if attachments:
        message["attachments"] = attachments

    path_user = quote(sender, safe="")
    endpoint = f"/users/{path_user}/sendMail"
    logger.info(
        "validar extractos notify: sendMail",
        extra={
            "sender": sender,
            "to": to_display,
            "historico": historico_rel,
            "bank_report_rows": len(bank_rows),
            "attachments": len(attachments),
        },
    )
    _mail_body, http_st = await graph.post_json(
        endpoint, {"message": message, "saveToSentItems": True}
    )

    export_pdf_on = os.getenv("GRAPH_VALIDAR_NOTIFY_EXPORT_EMAIL_PDF", "true").strip().lower() in (
        "1",
        "true",
        "yes",
        "si",
        "sí",
    )
    if export_pdf_on:
        try:
            attachment_names = [str(a.get("name", "")).strip() for a in (attachments or []) if a.get("name")]
            attachment_names = [n for n in attachment_names if n]

            cover_pdf = _cover_pdf_bytes_reportlab(
                sender=sender,
                sent_at=datetime.now(),
                to_addr=to_display,
                cc_list=cc_list,
                subject=subject,
                attachment_names=attachment_names,
                body_intro=body_intro,
                bank_headers=bank_headers,
                bank_rows=bank_rows,
            )
            pdf_bytes = cover_pdf

            _pdf_folder_default = (
                "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/05 EMAIL"
            )
            pdf_folder = (
                os.getenv("GRAPH_VALIDAR_NOTIFY_EXPORT_EMAIL_PDF_FOLDER_PATH", "").strip().strip("/")
                or _pdf_folder_default
            )
            pdf_name_tpl = os.getenv(
                "GRAPH_VALIDAR_NOTIFY_EXPORT_EMAIL_PDF_NAME_TEMPLATE",
                "ABONOS BANCO BOGOTA {fecha}.pdf",
            ).strip()
            pdf_name = pdf_name_tpl.format(
                fecha=report_d.isoformat(),
                fecha_ddmmyyyy=fecha_str,
                banco=banco,
                subject=subject,
            )
            pdf_name = _sanitize_pdf_filename_component(pdf_name)
            if not pdf_name.lower().endswith(".pdf"):
                pdf_name = f"{pdf_name}.pdf"
            pdf_rel_path = f"{pdf_folder}/{pdf_name}"
            encoded_pdf = encode_graph_drive_path(pdf_rel_path)
            pdf_endpoint = (
                f"/sites/{site_id}/drives/{drive_id}/root:/{encoded_pdf}:/content"
            )
            await graph.put_bytes(
                pdf_endpoint,
                pdf_bytes,
                content_type="application/pdf",
            )
            email_pdf_path = pdf_rel_path
            logger.info("validar extractos notify: PDF exportado en drive: %s", pdf_rel_path)
        except Exception as exc:
            email_pdf_error = str(exc)[:4000]
            logger.exception("validar extractos notify: export PDF falló")

    mc_out = await update_merge_control_workbook_after_notify(
        graph,
        site_id,
        drive_id,
        historical_file_path=historico_rel,
        email_pdf_path=email_pdf_path,
        report_d=report_d,
    )

    return ValidarExtractosNotifyResult(
        report_date=fecha_str,
        historico_excel_path=historico_rel,
        historical_file_path=historico_rel,
        historical_file_source="explicit",
        rows_included=len(bank_rows),
        subject=subject,
        attachments_count=len(attachments),
        graph_sendmail_http_status=http_st,
        mail_sender=sender,
        mail_to=to_display,
        email_pdf_path=email_pdf_path,
        email_pdf_error=email_pdf_error,
        merge_control_updated=mc_out.merge_control_updated,
        merge_control_file_path=mc_out.merge_control_file_path,
        merge_control_status=mc_out.merge_control_status,
        merge_control_warning=mc_out.merge_control_warning,
        merge_control_error_code=mc_out.merge_control_error_code,
    )
