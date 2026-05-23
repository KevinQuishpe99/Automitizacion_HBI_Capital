"""
Por cada ID Pago con filas cuyo Estado línea coincide con GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS
(p. ej. VALIDAR), descarga y concatena PDFs: primero el PDF del correo (ruta exacta en el archivo de
control), luego por cada par asiento+extracto (puede haber varios asientos por crédito y extracto)
y el extracto. Varios extractos del mismo ID Pago = un solo PDF consolidado.
En cada carpeta ASIENTOS CONTABLES del crédito debe existir al menos un PDF válido; si hay varios,
todos deben corresponder al crédito y se incluyen en orden estable por nombre.

La fecha del reporte es la mínima de la columna Fecha del Excel de reporte (GRAPH_SHAREPOINT_FILE_PATH),
solo para nombrar salidas. El histórico y el PDF del correo se leen desde ``control_merge_pdfs.xlsx``
(fila 2), no por auto-resolución por fecha.
"""

from __future__ import annotations

import json
import logging
import os
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_from_env
from app.application.use_cases.merge_control_workbook_merge import (
    MergeControlValidationError,
    merge_control_download_workbook_bytes,
    merge_control_parse_row2_validate_for_merge,
    merge_control_set_consolidado_success,
    merge_control_set_consolidando,
    merge_control_set_error_merge,
    merge_control_set_merge_parcial,
)
from app.application.use_cases.send_validar_extractos_notification import (
    _collect_pdf_paths_from_ruta_cell,
    _excel_cell_display,
    _find_distribucion_header_row,
    _find_distribucion_sheet,
    _get_col_distrib,
    _list_drive_folder_children,
    _parse_bank_report_table_and_min_date,
    _sanitize_pdf_filename_component,
    distrib_row_included_for_validar_extractos,
)
from app.application.use_cases.setup_merge_control_workbook import merge_control_workbook_relative_path
from app.application.use_cases.validate_payment_report import _graph_download_by_path
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

_DEFAULT_OUTPUT_FOLDER = (
    "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/06 ASIENTO CONTABLES GENERADOS"
)


def _merge_pdf_bytes(parts: list[bytes]) -> bytes:
    writer = PdfWriter()
    for chunk in parts:
        if not chunk:
            continue
        reader = PdfReader(BytesIO(chunk))
        for page in reader.pages:
            writer.add_page(page)
    out = BytesIO()
    writer.write(out)
    return out.getvalue()


def _parent_dir(rel_file: str) -> str:
    rel_file = rel_file.strip().strip("/").replace("\\", "/")
    if "/" not in rel_file:
        raise ValueError(f"Ruta de PDF sin carpeta padre: {rel_file!r}")
    return rel_file.rsplit("/", 1)[0]


def _ruta_asientos_from_cell(cell: Any) -> str:
    if cell is None:
        return ""
    return _excel_cell_display(cell).strip().replace("\\", "/").strip("/")


_IGNORED_ASIENTO_SUBFOLDER_NAMES = frozenset({"PROCESADOS", "PROCESADO"})


def normalize_sharepoint_path(path: str) -> str:
    """Clave de deduplicación: sin espacios extremos, barras unificadas, casefold."""
    return str(path or "").strip().strip("/").replace("\\", "/").casefold()


def unique_paths_preserve_order(paths: list[str]) -> list[str]:
    """Deduplica por path normalizado; conserva la primera aparición y el casing original."""
    seen: set[str] = set()
    out: list[str] = []
    for raw in paths:
        canonical = str(raw or "").strip().strip("/").replace("\\", "/")
        if not canonical:
            continue
        key = normalize_sharepoint_path(canonical)
        if key in seen:
            continue
        seen.add(key)
        out.append(canonical)
    return out


def _pdf_names_in_children(children: list[dict[str, Any]]) -> list[str]:
    """PDF en el nivel actual de la carpeta; ignora subcarpetas (p. ej. PROCESADOS)."""
    out: list[str] = []
    for it in children:
        if "folder" in it:
            folder_name = str(it.get("name", "")).strip().casefold()
            if folder_name in _IGNORED_ASIENTO_SUBFOLDER_NAMES:
                continue
            continue
        if "file" not in it:
            continue
        name = str(it.get("name", "")).strip()
        if name.lower().endswith(".pdf") and not name.startswith("~$"):
            out.append(name)
    return sorted(out)


def _filename_contains_credit_isolated(filename: str, credit_digits: str) -> bool:
    """Evita que el crédito 264 coincida con 1264 o 2640 (límites de dígitos)."""
    if not credit_digits:
        return False
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    return (
        re.search(rf"(?<!\d){re.escape(credit_digits)}(?!\d)", stem, flags=re.IGNORECASE) is not None
    )


def _merge_skip_line(
    id_pago: str,
    reason: str,
    *,
    cliente: str = "",
    credito_label: str = "",
    credit_number_expected: str = "",
    asiento_pdf_found: str = "",
    asiento_folder_path: str = "",
    extracto_path: str = "",
    names_seen: str = "",
) -> str:
    def nz(x: str) -> str:
        return x if (x or "").strip() else "-"

    return (
        f"id_pago={id_pago} | reason={reason} | cliente={nz(cliente)} | credito={nz(credito_label)} | "
        f"credit_number_expected={nz(credit_number_expected)} | "
        f"asiento_pdf_found={nz(asiento_pdf_found)} | asiento_folder_path={nz(asiento_folder_path)} | "
        f"extracto_path={nz(extracto_path)} | names_seen={nz(names_seen)}"
    )


_FULLWIDTH_NUMBER_SIGN = "\uff03"

_CREDIT_FOLDER_STATUS_SUFFIXES = (
    "TERMINADO",
    "FINALIZADO",
    "CANCELADO",
    "PAGADO",
    "LIQUIDADO",
    "VIGENTE",
    "REPUESTOS",
)

_CREDIT_FOLDER_SUFFIX_PATTERN = "|".join(
    re.escape(s.casefold()) for s in _CREDIT_FOLDER_STATUS_SUFFIXES
)


def _normalize_folder_sharp(s: str) -> str:
    return s.replace(_FULLWIDTH_NUMBER_SIGN, "#")


def _fold_folder_segment(seg: str) -> str:
    raw = _normalize_folder_sharp(str(seg).strip())
    raw = unicodedata.normalize("NFC", raw)
    nfkd = unicodedata.normalize("NFD", raw)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn").strip().casefold()


def _looks_like_credit_folder_name(folded: str) -> bool:
    if not folded:
        return False
    if folded.startswith("credito") or folded.startswith("obligacion"):
        return True
    return bool(re.search(r"credito\s*#?\s*\d+", folded))


def _credit_number_from_folder_segment(seg: str) -> str:
    """
    Extrae dígitos de crédito desde el nombre de carpeta (CREDITO / CRÉDITO + # opcional + número).
    Acepta sufijos operativos (TERMINADO, FINALIZADO, etc.) y formas sin '#'.
  """
    if not seg or not str(seg).strip():
        return ""
    folded = _fold_folder_segment(seg)
    if not _looks_like_credit_folder_name(folded):
        return ""
    m = re.match(r"^credito\s*#?\s*(\d+)\s*$", folded)
    if m:
        return m.group(1)
    m_suffix = re.match(
        rf"^credito\s*#?\s*(\d+)\s+(?:{_CREDIT_FOLDER_SUFFIX_PATTERN})\s*$",
        folded,
    )
    if m_suffix:
        return m_suffix.group(1)
    m_plain = re.match(r"^credito\s+(\d+)\s*$", folded)
    if m_plain:
        return m_plain.group(1)
    return ""


def _credit_number_from_extract_parent(ep: str) -> str:
    """Crédito del directorio padre inmediato del PDF de extracto (carpeta CREDITO # n)."""
    try:
        parent = _parent_dir(ep)
    except ValueError:
        return ""
    parts = [p for p in parent.replace("\\", "/").split("/") if str(p).strip()]
    if not parts:
        return ""
    return _credit_number_from_folder_segment(parts[-1])


def _resolve_credit_digits_for_extract(extract_path: str, row_credito_digits: str) -> str:
    """Carpeta del extracto primero; si falla, columna Crédito de la fila del histórico."""
    d = _credit_number_from_extract_parent(extract_path)
    if d:
        return d
    row = str(row_credito_digits or "").strip()
    if row:
        return row
    return _credit_number_from_path_scan(extract_path)


async def _drive_item_exists(
    graph: GraphApiPort, site_id: str, drive_id: str, rel_path: str
) -> bool:
    enc = encode_graph_drive_path(rel_path.strip().strip("/"))
    try:
        await graph.get_bytes(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content")
        return True
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return False
        raise


def _finalize_credit_item(
    credit_digits: str,
    asiento_paths: list[str],
    extract_paths: list[str],
    *,
    warnings: list[str] | None = None,
) -> dict[str, Any]:
    asientos = unique_paths_preserve_order(asiento_paths)
    extractos = unique_paths_preserve_order(extract_paths)
    w = list(warnings or [])
    if len(extractos) > 1:
        w.append("MULTIPLE_EXTRACTS_FOR_CREDIT")
    return {
        "credito": credit_digits,
        "asiento_pdf_paths": asientos,
        "extracto_pdf_paths": extractos,
        "extracto_pdf_path": extractos[0] if extractos else "",
        "warnings": w,
    }


async def _prevalidate_id_pago_group(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    id_pago: str,
    rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Valida filas del ID Pago y devuelve credit_items (un bloque por crédito) y líneas skipped.
    """
    credit_accum: dict[str, dict[str, Any]] = {}
    skip_lines: list[str] = []

    for row in rows:
        cliente = str(row.get("cliente") or "").strip()
        credito_label = str(row.get("credito_label") or "").strip()
        row_cred = str(row.get("credito_digits") or "").strip()

        extract_paths: list[str] = []
        ruta_cell = row.get("ruta_cell")
        for p in await _collect_pdf_paths_from_ruta_cell(graph, site_id, drive_id, ruta_cell):
            extract_paths.append(p.strip().strip("/").replace("\\", "/"))
        extract_paths = unique_paths_preserve_order(extract_paths)

        if not extract_paths:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "extract_routes_missing",
                    cliente=cliente,
                    credito_label=credito_label,
                    credit_number_expected=row_cred or "-",
                    extracto_path="-",
                )
            )
            continue

        credit_digits = row_cred or _resolve_credit_digits_for_extract(extract_paths[0], "")
        if not credit_digits:
            credit_digits = _resolve_credit_digits_for_extract(extract_paths[0], row_cred)
        if not credit_digits:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "credit_number_not_resolved",
                    cliente=cliente,
                    credito_label=credito_label,
                    extracto_path=extract_paths[0],
                )
            )
            continue

        asientos_dir_ep = _ruta_asientos_from_cell(row.get("ruta_asientos_cell"))
        if not asientos_dir_ep:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "missing_ruta_asientos_contables",
                    cliente=cliente,
                    credito_label=credito_label,
                    credit_number_expected=credit_digits,
                    extracto_path=extract_paths[0],
                )
            )
            continue

        try:
            asiento_children = await _list_drive_folder_children(
                graph, site_id, drive_id, asientos_dir_ep
            )
        except Exception as exc:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "asiento_folder_list_failed",
                    cliente=cliente,
                    credito_label=credito_label,
                    credit_number_expected=credit_digits,
                    asiento_folder_path=asientos_dir_ep,
                    extracto_path=extract_paths[0],
                    names_seen=str(exc)[:800],
                )
            )
            continue

        names = _pdf_names_in_children(asiento_children)
        valid_names, rejected_names = _classify_asiento_pdf_names(names, credit_digits)
        for rej in rejected_names:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "asiento_contable_credit_mismatch",
                    cliente=cliente,
                    credito_label=credito_label,
                    credit_number_expected=credit_digits,
                    asiento_pdf_found=rej,
                    asiento_folder_path=asientos_dir_ep,
                    extracto_path=extract_paths[0],
                    names_seen=", ".join(names) if names else "-",
                )
            )
        if not valid_names:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "asiento_contable_not_found",
                    cliente=cliente,
                    credito_label=credito_label,
                    credit_number_expected=credit_digits,
                    asiento_folder_path=asientos_dir_ep,
                    extracto_path=extract_paths[0],
                    names_seen=", ".join(names) if names else "-",
                )
            )
            continue

        bucket = credit_accum.setdefault(
            credit_digits,
            {
                "asiento_pdf_paths": [],
                "extracto_pdf_paths": [],
                "warnings": [],
            },
        )
        for asiento_name in valid_names:
            asiento_rel = f"{asientos_dir_ep}/{asiento_name}".replace("//", "/")
            bucket["asiento_pdf_paths"].append(asiento_rel)
        bucket["extracto_pdf_paths"].extend(extract_paths)

    credit_items: list[dict[str, Any]] = []
    for credit_digits in sorted(credit_accum.keys(), key=lambda x: (len(x), x)):
        raw = credit_accum[credit_digits]
        item = _finalize_credit_item(
            credit_digits,
            raw["asiento_pdf_paths"],
            raw["extracto_pdf_paths"],
            warnings=raw.get("warnings"),
        )
        if not item["asiento_pdf_paths"] or not item["extracto_pdf_paths"]:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "asiento_contable_not_found"
                    if not item["asiento_pdf_paths"]
                    else "extract_routes_missing",
                    credit_number_expected=credit_digits,
                    extracto_path=item["extracto_pdf_path"] or "-",
                )
            )
            continue
        credit_items.append(item)

    if not credit_items:
        if not skip_lines:
            skip_lines.append(
                _merge_skip_line(
                    id_pago,
                    "extract_routes_missing",
                    extracto_path="-",
                )
            )
        return [], skip_lines
    return credit_items, skip_lines


async def _build_consolidated_pdf_parts(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    id_pago: str,
    email_bytes: bytes,
    email_rel: str,
    credit_items: list[dict[str, Any]],
) -> tuple[list[bytes], list[str], list[str]]:
    """
    Orden: email, luego por crédito (estable): todos los asientos, luego extracto(s) deduplicados.
    Devuelve (parts, labels, líneas skipped si falla una descarga).
    """
    parts: list[bytes] = [email_bytes]
    labels: list[str] = [f"email:{email_rel}"]
    build_skips: list[str] = []

    for item in sorted(credit_items, key=lambda x: str(x.get("credito") or "")):
        credito = str(item.get("credito") or "")
        for asiento_rel in item.get("asiento_pdf_paths") or []:
            try:
                asiento_bytes = await _graph_download_by_path(
                    graph, site_id, drive_id, str(asiento_rel)
                )
            except Exception as exc:
                build_skips.append(
                    _merge_skip_line(
                        id_pago,
                        "asiento_download_failed",
                        credito_label=credito,
                        credit_number_expected=credito,
                        asiento_folder_path=_parent_dir(str(asiento_rel)),
                        asiento_pdf_found=str(asiento_rel).rsplit("/", 1)[-1],
                        names_seen=str(exc)[:800],
                    )
                )
                return parts, labels, build_skips
            parts.append(asiento_bytes)
            labels.append(f"asiento:{asiento_rel}")

        for ep in item.get("extracto_pdf_paths") or []:
            try:
                extract_bytes = await _graph_download_by_path(graph, site_id, drive_id, str(ep))
            except Exception as exc:
                build_skips.append(
                    _merge_skip_line(
                        id_pago,
                        "extracto_download_failed",
                        credito_label=credito,
                        credit_number_expected=credito,
                        extracto_path=str(ep),
                        names_seen=str(exc)[:800],
                    )
                )
                return parts, labels, build_skips
            parts.append(extract_bytes)
            labels.append(f"extracto:{ep}")

    return parts, labels, build_skips


def _unique_creditos_from_group(g: dict[str, Any]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for raw in g.get("creditos") or []:
        tok = _normalize_credito_excel_value(raw)
        if not tok:
            tok = re.sub(r"\D", "", str(raw).strip())
        if tok and tok not in seen:
            seen.add(tok)
            out.append(tok)
    return out


def _group_is_multi_credit(g: dict[str, Any], extract_paths: list[str]) -> bool:
    if len(extract_paths) > 1:
        return True
    return len(_unique_creditos_from_group(g)) > 1


def _classify_asiento_pdf_names(
    names: list[str],
    credit_digits: str,
) -> tuple[list[str], list[str]]:
    """
    Clasifica PDFs de la carpeta de asientos del crédito.
    Devuelve (válidos ordenados, rechazados por no coincidir con el crédito).
    """
    valid: list[str] = []
    rejected: list[str] = []
    for name in sorted(names):
        if _filename_contains_credit_isolated(name, credit_digits):
            valid.append(name)
        else:
            rejected.append(name)
    return valid, rejected


def _pick_single_asiento_pdf(names: list[str], credit_digits: str) -> tuple[str | None, str | None]:
    """
    Compatibilidad tests/helpers: primer PDF válido o motivo de fallo.
    Varios PDF válidos del mismo crédito no son ambiguos; use ``_classify_asiento_pdf_names``.
    """
    valid, rejected = _classify_asiento_pdf_names(names, credit_digits)
    if valid:
        return valid[0], None
    if rejected:
        return None, "asiento_contable_credit_mismatch"
    return None, "asiento_contable_not_found"


def _unique_asiento_paths_from_pairs(pair_asientos: list[tuple[str, str]]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for asiento_rel, _ep in pair_asientos:
        key = asiento_rel.strip().strip("/")
        if key and key not in seen:
            seen.add(key)
            out.append(key)
    return out


_MESES_ES = (
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
)


def _mes_reporte_upper(report_d: date) -> str:
    return _MESES_ES[report_d.month - 1]


def _credit_number_from_path_scan(rel: str) -> str:
    """Último segmento de la ruta que coincide con carpeta de crédito (p. ej. tokens en nombre de archivo)."""
    parts = [p for p in rel.replace("\\", "/").split("/") if str(p).strip()]
    for part in reversed(parts):
        hit = _credit_number_from_folder_segment(part)
        if hit:
            return hit
    return ""


def _client_folder_before_credit(rel: str) -> str:
    parts = [p for p in rel.replace("\\", "/").split("/") if p.strip()]
    for i, part in enumerate(parts):
        if _credit_number_from_folder_segment(part) and i > 0:
            return parts[i - 1].strip()
    return ""


def _normalize_credito_excel_value(raw: Any) -> str:
    """Número para el nombre del PDF (columna Crédito); no se usa para rutas en SharePoint."""
    s = _excel_cell_display(raw).strip()
    if not s:
        return ""
    m = re.search(r"(?i)credito#?\s*(\d+)", s)
    if m:
        return m.group(1)
    m2 = re.search(r"\d{1,12}", s)
    if m2:
        return m2.group(0)
    return ""


def _merge_composite_client_token(s: str, max_len: int = 80) -> str:
    t = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", str(s or "").strip())
    t = re.sub(r"\s+", " ", t).strip()
    return t[:max_len] if len(t) > max_len else t


def _merge_composite_credit_for_filename_display(credit_part: str) -> str:
    raw = (credit_part or "").strip()
    if not raw:
        return "SIN_CREDITO"
    cleaned = re.sub(r"[^0-9,\s]", "", raw)
    cleaned = re.sub(r"\s*,\s*", ", ", cleaned).strip()
    cleaned = re.sub(r",\s*,+", ", ", cleaned)
    if not cleaned or not re.search(r"\d", cleaned):
        return "SIN_CREDITO"
    return _merge_composite_client_token(cleaned, max_len=100)


def _credit_tokens_ordered_for_rows(
    rows: list[dict[str, Any]], extract_paths: list[str]
) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        tok = str(row.get("credito_digits") or "").strip()
        if not tok:
            raw = row.get("credito_raw")
            tok = _normalize_credito_excel_value(raw)
            if not tok:
                tok = re.sub(r"\D", "", str(raw or "").strip())
        if tok and tok not in seen:
            seen.add(tok)
            out.append(tok)
    for ep in extract_paths:
        tok = _credit_number_from_extract_parent(ep) or _credit_number_from_path_scan(ep)
        if tok and tok not in seen:
            seen.add(tok)
            out.append(tok)
    return out


def _merge_composite_output_basename(report_d: date, client: str, credit_part: str) -> str:
    day = report_d.day
    mes = _mes_reporte_upper(report_d)
    cli = _merge_composite_client_token(client)
    cred = _merge_composite_credit_for_filename_display(credit_part)
    if not cli:
        cli = "CLIENTE"
    base = f"{day} {mes} PAGO {cli} {cred}.pdf"
    return _sanitize_pdf_filename_component(base) or base


def _allocate_duplicate_pdf_name(base: str, tallies: dict[str, int]) -> str:
    stem = base[:-4] if base.lower().endswith(".pdf") else base
    ext = ".pdf" if base.lower().endswith(".pdf") else ""
    key = stem.casefold()
    tallies[key] = tallies.get(key, 0) + 1
    n = tallies[key]
    if n == 1:
        return f"{stem}{ext}"
    return f"{stem}_{n}{ext}"


@dataclass(frozen=True)
class MergeCompositePdfOutput:
    id_pago: str
    output_relative_path: str
    bytes_written: int
    sources_summary: str
    cliente: str = ""
    credito: str = ""
    email_pdf_path: str = ""
    asiento_pdf_path: str = ""
    asiento_pdf_paths: tuple[str, ...] = ()
    extracto_pdf_path: str = ""
    credit_items: tuple[dict[str, Any], ...] = ()


@dataclass(frozen=True)
class MergeCompositeValidadoPdfsResult:
    report_date_iso: str
    historico_excel_path: str
    estado_linea_contains: str
    email_pdf_used: str
    outputs: tuple[MergeCompositePdfOutput, ...] = ()
    skipped: tuple[str, ...] = ()
    merge_control_file_path: str = ""
    merge_control_updated: bool = False
    merge_control_status: str | None = None
    outputs_count: int = 0
    skipped_count: int = 0
    merge_manifest_path: str = ""


def _merge_logs_folder_relative() -> str:
    p = os.getenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "").strip().rstrip("/")
    if p:
        return p
    return "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/04 LOGS"


def _legacy_paths_from_credit_items(credit_items: list[dict[str, Any]]) -> tuple[list[str], str]:
    asiento_paths: list[str] = []
    extract_paths: list[str] = []
    for item in credit_items:
        asiento_paths.extend(item.get("asiento_pdf_paths") or [])
        extract_paths.extend(item.get("extracto_pdf_paths") or [])
    asiento_paths = unique_paths_preserve_order(asiento_paths)
    extract_paths = unique_paths_preserve_order(extract_paths)
    extracto_legacy = " | ".join(extract_paths)
    return asiento_paths, extracto_legacy


def _merge_output_record(
    *,
    id_pago: str,
    output_relative_path: str,
    bytes_written: int,
    sources_summary: str,
    cliente: str,
    credito: str,
    email_pdf_path: str,
    credit_items: list[dict[str, Any]],
) -> MergeCompositePdfOutput:
    asiento_paths, extracto_path = _legacy_paths_from_credit_items(credit_items)
    legacy_asiento = asiento_paths[0] if asiento_paths else ""
    frozen_items = tuple(dict(ci) for ci in credit_items)
    return MergeCompositePdfOutput(
        id_pago=id_pago,
        output_relative_path=output_relative_path,
        bytes_written=bytes_written,
        sources_summary=sources_summary,
        cliente=cliente,
        credito=credito,
        email_pdf_path=email_pdf_path,
        asiento_pdf_path=legacy_asiento,
        asiento_pdf_paths=tuple(asiento_paths),
        extracto_pdf_path=extracto_path,
        credit_items=frozen_items,
    )


async def _upload_merge_manifest(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    report_date_iso: str,
    payload: dict[str, Any],
) -> str:
    folder = _merge_logs_folder_relative()
    rel = f"{folder}/merge_manifest_{report_date_iso}.json".replace("//", "/")
    enc = encode_graph_drive_path(rel)
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    await graph.put_bytes(
        f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content",
        body,
        content_type="application/json",
    )
    return rel


async def merge_composite_validado_pdfs(
    graph: GraphApiPort,
    *,
    force_rebuild: bool = False,
) -> MergeCompositeValidadoPdfsResult:
    estado = os.getenv("GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS", "VALIDAR").strip()
    if not estado:
        estado = "VALIDAR"

    control_rel = merge_control_workbook_relative_path().strip().strip("/")
    ctx = await resolve_sharepoint_from_env(graph)
    site_id = ctx["site_id"]
    drive_id = ctx["drive_id"]

    consolidando_started = False
    try:
        try:
            _, control_raw = await merge_control_download_workbook_bytes(graph, site_id, drive_id)
        except httpx.HTTPStatusError as exc:
            code = exc.response.status_code if exc.response else 0
            if code == 404:
                raise MergeControlValidationError("merge_control_workbook_not_found") from exc
            raise

        snapshot = merge_control_parse_row2_validate_for_merge(control_raw)
        historico_rel = snapshot.historical_file_path
        email_rel = snapshot.email_pdf_path

        await merge_control_set_consolidando(graph, site_id, drive_id, control_rel)
        consolidando_started = True

        report_bytes = await graph.get_bytes(
            f"/sites/{site_id}/drives/{drive_id}/root:/{ctx['path_encoded']}:/content"
        )
        report_d, _, _ = _parse_bank_report_table_and_min_date(report_bytes)
        iso = report_d.isoformat()

        hist_bytes = await _graph_download_by_path(graph, site_id, drive_id, historico_rel)
        email_bytes = await _graph_download_by_path(graph, site_id, drive_id, email_rel)

        wb = load_workbook(filename=BytesIO(hist_bytes), data_only=True)
        try:
            ws = _find_distribucion_sheet(wb)
            h_row, header_map = _find_distribucion_header_row(ws)
            col_estado = _get_col_distrib(
                header_map,
                "Estado Pago",
                "Estado pago",
                "Estado",
                "Estado línea",
                "Estado linea",
                "ESTADO LINEA",
            )
            col_ruta = _get_col_distrib(header_map, "Ruta", "RUTA", "Rutas", "RUTAS")
            col_ruta_asientos = _get_col_distrib(
                header_map,
                "RutaAsientosContables",
                "RUTA_ASIENTOS_CONTABLES",
                "Ruta asientos contables",
            )
            col_id = _get_col_distrib(
                header_map,
                "ID Pago",
                "ID pago",
                "ID PAGO",
                "Id pago",
                "ID_PAGO",
                "Id Pago",
            )
            col_cliente = _get_col_distrib(
                header_map,
                "Cliente",
                "CLIENTE",
                "Nombre Cliente",
                "Nombre cliente",
                "NOMBRE CLIENTE",
                "Razón social",
                "Razon social",
                "RAZON SOCIAL",
                "Empresa",
                "EMPRESA",
            )
            col_credito = _get_col_distrib(
                header_map,
                "Crédito",
                "Credito",
                "CREDITO",
                "CRÉDITO",
                "No Crédito",
                "No Credito",
                "NO CREDITO",
                "Número Crédito",
                "Numero Credito",
            )
            extra_client = os.getenv("GRAPH_MERGE_COMPOSITE_CLIENTE_COLUMN", "").strip()
            if extra_client and col_cliente is None:
                col_cliente = _get_col_distrib(header_map, extra_client, extra_client.upper())
            extra_cred = os.getenv("GRAPH_MERGE_COMPOSITE_CREDITO_COLUMN", "").strip()
            if extra_cred and col_credito is None:
                col_credito = _get_col_distrib(header_map, extra_cred, extra_cred.upper())
            col_vp = _get_col_distrib(header_map, "Validar Pago", "Validar pago", "VALIDAR PAGO")
            if col_estado is None and col_vp is None:
                raise ValueError(
                    'En Distribución se requieren columnas de estado ("Estado Pago" o histórico "Estado" / '
                    '"Estado línea"), "Ruta" y "ID Pago" para unir PDFs.'
                )
            if col_ruta is None or col_id is None or col_ruta_asientos is None:
                raise ValueError(
                    'En Distribución se requieren columnas "Ruta", "RutaAsientosContables" e "ID Pago" '
                    "para unir PDFs."
                )

            groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
            last = ws.max_row or h_row
            for r in range(h_row + 1, last + 1):
                if not distrib_row_included_for_validar_extractos(ws, r, header_map, estado):
                    continue
                id_raw = ws.cell(row=r, column=col_id).value
                id_str = _excel_cell_display(id_raw).strip()
                if not id_str:
                    id_str = f"__sin_id_fila_{r}__"
                cred_raw = ws.cell(row=r, column=col_credito).value if col_credito else None
                cred_digits = _normalize_credito_excel_value(cred_raw) if cred_raw is not None else ""
                if not cred_digits and cred_raw is not None:
                    cred_digits = re.sub(r"\D", "", str(cred_raw).strip())
                cliente = ""
                if col_cliente:
                    cliente = _excel_cell_display(ws.cell(row=r, column=col_cliente).value).strip()
                groups[id_str].append(
                    {
                        "cliente": cliente,
                        "credito_raw": cred_raw,
                        "credito_label": _excel_cell_display(cred_raw).strip()
                        if cred_raw is not None
                        else "",
                        "credito_digits": cred_digits,
                        "ruta_cell": ws.cell(row=r, column=col_ruta).value if col_ruta else None,
                        "ruta_asientos_cell": (
                            ws.cell(row=r, column=col_ruta_asientos).value
                            if col_ruta_asientos
                            else None
                        ),
                    }
                )
        finally:
            closer = getattr(wb, "close", None)
            if callable(closer):
                closer()

        out_folder = (
            os.getenv("GRAPH_MERGE_COMPOSITE_OUTPUT_FOLDER_PATH", "").strip().strip("/")
            or _DEFAULT_OUTPUT_FOLDER
        )

        outputs: list[MergeCompositePdfOutput] = []
        skipped: list[str] = []
        out_name_tallies: dict[str, int] = {}

        for id_pago, group_rows in sorted(groups.items(), key=lambda x: x[0]):
            credit_items, pre_skips = await _prevalidate_id_pago_group(
                graph, site_id, drive_id, id_pago, group_rows
            )
            if pre_skips:
                skipped.extend(pre_skips)
            if not credit_items:
                continue

            all_extract_paths = [
                ep for item in credit_items for ep in (item.get("extracto_pdf_paths") or [])
            ]
            credit_tokens = _credit_tokens_ordered_for_rows(group_rows, all_extract_paths)
            if not credit_tokens:
                credit_tokens = [str(item.get("credito") or "") for item in credit_items]
            credit_for_filename = ", ".join(credit_tokens)

            client_from_cells = next((str(r.get("cliente") or "").strip() for r in group_rows if r.get("cliente")), "")
            first_extract = all_extract_paths[0] if all_extract_paths else ""
            client_display = client_from_cells or _client_folder_before_credit(first_extract)
            if not client_display:
                client_display = "CLIENTE"

            out_base = _merge_composite_output_basename(report_d, client_display, credit_for_filename)
            base_rel = f"{out_folder}/{out_base}".replace("//", "/")
            already_exists = await _drive_item_exists(graph, site_id, drive_id, base_rel)

            if already_exists and not force_rebuild:
                labels_preview: list[str] = [f"email:{email_rel}"]
                for item in sorted(credit_items, key=lambda x: str(x.get("credito") or "")):
                    for a in item.get("asiento_pdf_paths") or []:
                        labels_preview.append(f"asiento:{a}")
                    for ep in item.get("extracto_pdf_paths") or []:
                        labels_preview.append(f"extracto:{ep}")
                outputs.append(
                    _merge_output_record(
                        id_pago=id_pago,
                        output_relative_path=base_rel,
                        bytes_written=0,
                        sources_summary="already_consolidated | " + " | ".join(labels_preview[1:]),
                        cliente=client_display,
                        credito=credit_for_filename,
                        email_pdf_path=email_rel,
                        credit_items=credit_items,
                    )
                )
                logger.info(
                    "merge_composite_validado: id_pago=%s ya consolidado en %s",
                    id_pago,
                    base_rel,
                )
                continue

            parts, labels, build_skips = await _build_consolidated_pdf_parts(
                graph, site_id, drive_id, id_pago, email_bytes, email_rel, credit_items
            )
            if build_skips:
                skipped.extend(build_skips)
                continue

            merged = _merge_pdf_bytes(parts)
            out_base = _merge_composite_output_basename(report_d, client_display, credit_for_filename)
            out_name = _allocate_duplicate_pdf_name(out_base, out_name_tallies)
            out_rel = f"{out_folder}/{out_name}".replace("//", "/")
            if already_exists and force_rebuild:
                out_rel = base_rel
            enc = encode_graph_drive_path(out_rel)
            try:
                await graph.put_bytes(
                    f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content",
                    merged,
                    content_type="application/pdf",
                )
            except Exception as exc:
                skipped.append(
                    _merge_skip_line(
                        id_pago,
                        "consolidated_upload_failed",
                        extracto_path=first_extract or "-",
                        names_seen=str(exc)[:800],
                    )
                )
                continue

            outputs.append(
                _merge_output_record(
                    id_pago=id_pago,
                    output_relative_path=out_rel,
                    bytes_written=len(merged),
                    sources_summary=" | ".join(labels),
                    cliente=client_display,
                    credito=credit_for_filename,
                    email_pdf_path=email_rel,
                    credit_items=credit_items,
                )
            )

            logger.info(
                "merge_composite_validado: subido %s (%s bytes) id_pago=%s force_rebuild=%s",
                out_rel,
                len(merged),
                id_pago,
                force_rebuild,
            )

        if not outputs and not skipped:
            raise ValueError(
                f"No hay filas cuyo Estado línea contenga el token {estado!r} "
                f"(GRAPH_VALIDAR_EXTRACTO_ESTADO_CONTAINS) en Distribución del histórico {historico_rel!r}."
            )

        oc = len(outputs)
        sc = len(skipped)
        merge_control_updated = True

        if oc > 0 and sc == 0:
            await merge_control_set_consolidado_success(
                graph, site_id, drive_id, control_rel, outputs_count=oc
            )
            final_status = "CONSOLIDADO"
        else:
            await merge_control_set_merge_parcial(
                graph,
                site_id,
                drive_id,
                control_rel,
                outputs_count=oc,
                skipped_count=sc,
                user_message=(
                    "La consolidación terminó con omisiones: revise los pagos listados en skipped "
                    "y cargue o corrija los PDF de asientos o extractos faltantes."
                ),
                next_action=(
                    "Corrija las rutas en el histórico o suba los archivos requeridos y ejecute de nuevo "
                    "el endpoint de Merge."
                ),
            )
            final_status = "MERGE_PARCIAL"

        manifest_path = ""
        try:
            manifest_payload: dict[str, Any] = {
                "report_date_iso": iso,
                "historico_excel_path": historico_rel,
                "email_pdf_used": email_rel,
                "merge_control_status": final_status,
                "outputs": [
                    {
                        "id_pago": o.id_pago,
                        "cliente": o.cliente,
                        "credito": o.credito,
                        "email_pdf_path": o.email_pdf_path,
                        "asiento_pdf_path": o.asiento_pdf_path,
                        "asiento_pdf_paths": list(o.asiento_pdf_paths),
                        "extracto_pdf_path": o.extracto_pdf_path,
                        "credit_items": [dict(ci) for ci in o.credit_items],
                        "output_relative_path": o.output_relative_path,
                        "bytes_written": o.bytes_written,
                        "sources_summary": o.sources_summary,
                    }
                    for o in outputs
                ],
                "skipped": list(skipped),
            }
            manifest_path = await _upload_merge_manifest(
                graph, site_id, drive_id, iso, manifest_payload
            )
            logger.info("merge_composite_validado: manifest %s", manifest_path)
        except Exception as man_exc:
            logger.warning("merge_composite_validado: no se pudo subir manifest: %s", man_exc)

        return MergeCompositeValidadoPdfsResult(
            report_date_iso=iso,
            historico_excel_path=historico_rel,
            estado_linea_contains=estado,
            email_pdf_used=email_rel,
            outputs=tuple(outputs),
            skipped=tuple(skipped),
            merge_control_file_path=control_rel,
            merge_control_updated=merge_control_updated,
            merge_control_status=final_status,
            outputs_count=oc,
            skipped_count=sc,
            merge_manifest_path=manifest_path,
        )

    except MergeControlValidationError as e:
        raise ValueError(e.error_code) from e
    except Exception:
        if consolidando_started:
            try:
                await merge_control_set_error_merge(
                    graph,
                    site_id,
                    drive_id,
                    control_rel,
                    user_message="Ocurrió un error durante la consolidación de PDFs.",
                    next_action=(
                        "Revise el detalle técnico del job, corrija el problema y vuelva a ejecutar Merge."
                    ),
                )
            except Exception:
                logger.exception("merge_composite_validado: no se pudo escribir ERROR_MERGE en el control")
        raise
