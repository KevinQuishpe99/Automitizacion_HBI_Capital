"""
Vacía solo las columnas de salida del Excel de reporte en SharePoint:
Validado, Estado, Observaciones, Rutas (filas de datos; se conservan los encabezados).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook

from app.application.sharepoint_resolution import resolve_sharepoint_from_env
from app.application.use_cases.validate_payment_report import (
    _find_header_map,
    _get_col,
)
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ClearReportOutputResult:
    file_path: str
    columns_cleared: list[str]
    data_rows_cleared: int


async def clear_report_output_columns(graph: GraphApiPort) -> ClearReportOutputResult:
    ctx = await resolve_sharepoint_from_env(graph)
    try:
        excel_bytes = await graph.get_bytes(
            f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content"
        )
    except Exception as exc:
        raise ValueError(
            "No se pudo descargar el Excel del reporte (no existe, ruta inválida o sin permisos)."
        ) from exc

    try:
        wb = load_workbook(filename=BytesIO(excel_bytes))
    except Exception as exc:
        raise ValueError("El archivo configurado no es un Excel válido (.xlsx).") from exc

    ws = wb.worksheets[0]
    if ws.protection.sheet:
        wb.close()
        raise ValueError(
            "La hoja está protegida; no se pueden limpiar columnas. Quita la protección y reintenta."
        )

    header_row, header_map = _find_header_map(ws)

    col_specs: list[tuple[str, tuple[str, ...]]] = [
        ("Validado", ("Validado", "VALIDADO", "Validación", "VALIDACION", "Validacion")),
        ("Estado", ("Estado", "ESTADO")),
        ("Observaciones", ("Observaciones", "OBSERVACIONES")),
        ("Rutas", ("Rutas", "RUTAS")),
    ]

    cols_to_clear: list[tuple[str, int]] = []
    for label, candidates in col_specs:
        idx = _get_col(header_map, *candidates)
        if idx is not None:
            cols_to_clear.append((label, idx))

    if not cols_to_clear:
        wb.close()
        raise ValueError(
            "No se encontró ninguna de las columnas: Validado, Estado, Observaciones, Rutas."
        )

    last_row = ws.max_row
    data_rows = max(0, last_row - header_row)

    for _label, col_idx in cols_to_clear:
        for r in range(header_row + 1, last_row + 1):
            ws.cell(row=r, column=col_idx).value = None

    out = BytesIO()
    wb.save(out)
    wb.close()
    out_bytes = out.getvalue()

    logger.info(
        "clear_report_output_columns: subiendo Excel sin datos en columnas de salida",
        extra={
            "site_id": ctx["site_id"],
            "drive_id": ctx["drive_id"],
            "path": ctx["file_path"],
            "columns": [c[0] for c in cols_to_clear],
            "data_rows": data_rows,
        },
    )

    await graph.put_bytes(
        f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content",
        out_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return ClearReportOutputResult(
        file_path=ctx["file_path"],
        columns_cleared=[c[0] for c in cols_to_clear],
        data_rows_cleared=data_rows,
    )
