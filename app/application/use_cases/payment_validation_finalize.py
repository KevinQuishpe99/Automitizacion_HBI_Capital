import logging
import os
import io
import re
import unicodedata
from calendar import monthrange
from urllib.parse import unquote
from datetime import date, datetime
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domain.ports.graph import GraphApiPort
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.application.services.payment_followup_finalize import register_payment_followups_after_finalize
from app.application.services.review_schema import (
    DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS,
    ReviewSheets,
    ControlCols,
    CasosPagoCols,
    DistribucionCols,
    EstadoPago,
    apply_legacy_estado_migration,
    is_validar_pago_si,
    normalize_distrib_row_keys,
)

logger = logging.getLogger(__name__)


def _include_in_validation_outputs(dist: dict[str, Any]) -> bool:
    """Filas que generan ruta extracto, filas en soporte secretaría y conteo «validadas»."""
    ep = str(dist.get(DistribucionCols.ESTADO_PAGO, "")).strip().upper()
    return ep in EstadoPago.SECRETARY_AND_RUTA and is_validar_pago_si(dist)

SECRETARY_SHEET = "Asientos_Pendientes"
ASIENTOS_FOLDER_LABEL = "ASIENTOS CONTABLES"
ASIENTOS_FOLDER_PER_CREDIT_PREFIX = "ASIENTOS CONTABLES CRED"
PENDIENTE_CREAR_ASIENTOS = "PENDIENTE_CREAR"
OBS_NO_ASIENTOS = "No se encontró carpeta ASIENTOS CONTABLES."

SECRETARY_HEADERS = [
    "Cliente",
    "Crédito",
    "ID Pago",
    "Fecha banco",
    "Fecha límite",
    "Total validado",
    "Link carpeta asientos contables",
    "Link extracto",
    "Link tabla amortización",
]

SECRETARY_TITLE = "Soporte de asientos contables"
SECRETARY_INSTRUCTION = (
    "Use esta hoja para cargar o revisar los soportes de asientos contables por crédito. "
    "Abra el link de la carpeta correspondiente, cargue el PDF del asiento contable y luego "
    "continúe con la consolidación cuando todos los soportes estén completos."
)
SECRETARY_HEADER_ROW = 3
SECRETARY_FIRST_DATA_ROW = 4

_SEC_THIN = Side(style="thin", color="C8C8C8")
_SEC_BORDER_LIGHT = Border(left=_SEC_THIN, right=_SEC_THIN, top=_SEC_THIN, bottom=_SEC_THIN)
_SEC_FILL_NAVY = PatternFill(fill_type="solid", fgColor="002060")
_SEC_FILL_HEADER = _SEC_FILL_NAVY
_SEC_FONT_TITLE = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
_SEC_FILL_INSTRUCTION = PatternFill(fill_type="solid", fgColor="E8F2FA")
_SEC_FONT_INSTRUCTION = Font(name="Calibri", size=12, color="1A2F36")
_SEC_MEDIUM_CLIENT_EDGE = Side(style="medium", color="002060")
_SEC_FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_SEC_FONT_BODY = Font(name="Calibri", size=11)
_SEC_FONT_HLINK = Font(name="Calibri", color="0563C1", size=11, underline="single")
_SEC_FONT_TOTAL_LABEL = Font(name="Calibri", bold=True, size=11)
_SEC_ALIGN_CENTER_WRAP = Alignment(vertical="center", horizontal="center", wrap_text=True)
_SEC_ALIGN_VCENTER = Alignment(vertical="center", horizontal="left")
_SEC_ALIGN_WRAP = Alignment(vertical="center", horizontal="left", wrap_text=True)
_SEC_FMT_MONEY = "#,##0.00"
_SEC_FMT_DATE = "yyyy-mm-dd"
_SEC_TAB_COLOR = "FF2E75B6"


def _secretary_coerce_date_value(v: Any) -> Any:
    if isinstance(v, datetime):
        return v.date()
    return v


def _apply_secretary_banner(ws: Any, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, SECRETARY_TITLE)
    t.font = _SEC_FONT_TITLE
    t.fill = _SEC_FILL_NAVY
    t.alignment = _SEC_ALIGN_CENTER_WRAP
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(2, 1, SECRETARY_INSTRUCTION)
    s.fill = _SEC_FILL_INSTRUCTION
    s.font = _SEC_FONT_INSTRUCTION
    s.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 56


def _apply_secretary_header_style(ws: Any, row_idx: int, headers: list[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row_idx, col_idx, value=name)
        cell.fill = _SEC_FILL_HEADER
        cell.font = _SEC_FONT_HEADER
        cell.alignment = _SEC_ALIGN_CENTER_WRAP
        cell.border = _SEC_BORDER_LIGHT
    ws.row_dimensions[row_idx].height = 22


def _visible_len_secretary(value: Any) -> int:
    if value is None:
        return 0
    s = str(value)
    if s.startswith("="):
        return min(42, len(s))
    return len(s)


def _auto_fit_secretary_columns(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for col in range(min_col, max_col + 1):
        best = 9.0
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            est = _visible_len_secretary(cell.value) * 1.12 + 2.5
            if est > best:
                best = est
        best = min(56.0, max(9.0, best))
        letter = get_column_letter(col)
        cur = ws.column_dimensions[letter].width
        if cur is None or cur < best:
            ws.column_dimensions[letter].width = best


def _apply_secretary_body_style(
    ws: Any,
    first_data_row: int,
    last_data_row: int,
    ncols: int,
    hmap: dict[str, int],
) -> None:
    if last_data_row < first_data_row:
        return
    wrap_cols = {
        hmap[name]
        for name in (
            "Cliente",
            "Link carpeta asientos contables",
            "Link extracto",
            "Link tabla amortización",
        )
        if name in hmap
    }
    for r in range(first_data_row, last_data_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _SEC_BORDER_LIGHT
            cell.font = _SEC_FONT_BODY
            cell.alignment = _SEC_ALIGN_WRAP if c in wrap_cols else _SEC_ALIGN_VCENTER


def _apply_secretary_number_formats(ws: Any, first_data_row: int, last_data_row: int, hmap: dict[str, int]) -> None:
    if last_data_row < first_data_row:
        return
    money_cols = {hmap["Total validado"]}
    date_cols = {hmap["Fecha banco"], hmap["Fecha límite"]}
    for r in range(first_data_row, last_data_row + 1):
        for c in money_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, "") and not (
                isinstance(cell.value, str) and str(cell.value).startswith("=")
            ):
                cell.number_format = _SEC_FMT_MONEY
        for c in date_cols:
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v not in (None, ""):
                coerced = _secretary_coerce_date_value(v)
                if coerced is not v:
                    cell.value = coerced
                if not (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                    cell.number_format = _SEC_FMT_DATE


def _secretary_http_url(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.lower().startswith("http://") or s.lower().startswith("https://"):
        return s
    return None


def _secretary_link_label_suffix(credito: Any, cliente: Any) -> str:
    cred = str(credito or "").strip()
    cli = str(cliente or "").strip()
    if cred and re.fullmatch(r"\d+", cred):
        return f"crédito {cred}"
    if cred:
        return cred
    return cli


def _secretary_link_visible_text(link_kind: str, credito: Any, cliente: Any) -> str:
    suffix = _secretary_link_label_suffix(credito, cliente)
    if link_kind == "extracto":
        return f"Ver extracto {suffix}".strip() if suffix else "Ver extracto"
    if link_kind == "tabla":
        return f"Ver tabla {suffix}".strip() if suffix else "Ver tabla"
    if link_kind == "carpeta_asientos":
        return f"Ver carpeta asientos {suffix}".strip() if suffix else "Ver carpeta asientos"
    return ""


def _hyperlink_target_from_cell(cell: Any) -> str | None:
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    tgt = getattr(hl, "target", None) or getattr(hl, "ref", None)
    return str(tgt).strip() if tgt else None


def _set_secretary_link_cell(
    dst: Any,
    src: Any,
    link_kind: str,
    credito: Any,
    cliente: Any,
) -> None:
    tgt = _secretary_http_url(_hyperlink_target_from_cell(src))
    if tgt:
        dst.hyperlink = tgt
        dst.value = _secretary_link_visible_text(link_kind, credito, cliente)
        dst.font = _SEC_FONT_HLINK
    else:
        dst.value = None


def _set_secretary_url_link_cell(
    dst: Any,
    url: Any,
    link_kind: str,
    credito: Any,
    cliente: Any,
) -> None:
    tgt = _secretary_http_url(url)
    if tgt:
        dst.hyperlink = tgt
        dst.value = _secretary_link_visible_text(link_kind, credito, cliente)
        dst.font = _SEC_FONT_HLINK
    else:
        dst.value = None


def _secretary_row_border(*, client_top: bool = False, client_bottom: bool = False) -> Border:
    top = _SEC_MEDIUM_CLIENT_EDGE if client_top else _SEC_THIN
    bottom = _SEC_MEDIUM_CLIENT_EDGE if client_bottom else _SEC_THIN
    return Border(left=_SEC_THIN, right=_SEC_THIN, top=top, bottom=bottom)


def _apply_secretary_client_borders(
    ws: Any,
    first_data_row: int,
    last_data_row: int,
    ncols: int,
    col_cliente: int,
) -> None:
    if last_data_row < first_data_row:
        return
    indexed: list[tuple[int, str]] = []
    for r in range(first_data_row, last_data_row + 1):
        cliente = str(ws.cell(row=r, column=col_cliente).value or "").strip()
        if not cliente:
            continue
        indexed.append((r, cliente))
    if not indexed:
        return
    for i, (r, cliente) in enumerate(indexed):
        prev_cliente = indexed[i - 1][1] if i > 0 else None
        next_cliente = indexed[i + 1][1] if i + 1 < len(indexed) else None
        client_top = i > 0 and bool(cliente) and cliente != prev_cliente
        client_bottom = bool(cliente) and (i == len(indexed) - 1 or cliente != next_cliente)
        border = _secretary_row_border(client_top=client_top, client_bottom=client_bottom)
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = border


def _secretary_freeze_panes_cell(hmap: dict[str, int]) -> str:
    col_credito = hmap["Crédito"]
    return f"{get_column_letter(col_credito + 1)}{SECRETARY_FIRST_DATA_ROW}"


def _apply_secretary_total_row(
    ws: Any, hmap: dict[str, int], first_data_row: int, last_data_row: int
) -> int | None:
    if last_data_row < first_data_row:
        return None
    trow = last_data_row + 1
    tc = hmap["Total validado"]
    lett = get_column_letter(tc)
    ws.cell(trow, 1, "Total validado general")
    ws.cell(trow, tc, f"=SUM({lett}{first_data_row}:{lett}{last_data_row})")
    lbl = ws.cell(trow, 1)
    lbl.font = _SEC_FONT_TOTAL_LABEL
    lbl.border = _SEC_BORDER_LIGHT
    sum_cell = ws.cell(trow, tc)
    sum_cell.font = _SEC_FONT_TOTAL_LABEL
    sum_cell.number_format = _SEC_FMT_MONEY
    sum_cell.border = _SEC_BORDER_LIGHT
    top = Side(style="medium", color="888888")
    for c in range(1, len(SECRETARY_HEADERS) + 1):
        cell = ws.cell(trow, c)
        cell.border = Border(
            left=_SEC_THIN,
            right=_SEC_THIN,
            top=top,
            bottom=_SEC_THIN,
        )
    return trow


def _find_table_header_row(ws: Any, first_header_value: str) -> int:
    marker = str(first_header_value).strip()
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and str(row[0]).strip() == marker:
            return r_idx
    raise ValueError("missing_sheet_headers")


def _normalize_process_date(process_date: date | str | None) -> date:
    if process_date is None:
        return datetime.now().date()
    if isinstance(process_date, date):
        return process_date
    return date.fromisoformat(str(process_date))


def _resolve_validation_selection(children_items: list[dict[str, Any]], prefix: str) -> str:
    valid_files = []
    for item in children_items:
        name = str(item.get("name", "")).strip()
        if not name or name.startswith("~$") or not name.lower().endswith(".xlsx"):
            continue
        if prefix and not name.startswith(prefix):
            continue
        valid_files.append(item)

    if not valid_files:
        raise ValueError("no_validation_file_found")

    def sort_key(item: dict[str, Any]) -> tuple[str, str]:
        return (str(item.get("lastModifiedDateTime", "")), str(item.get("name", "")))

    valid_files.sort(key=sort_key, reverse=True)
    return str(valid_files[0]["name"])


def _build_content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


async def _drive_item_exists(graph: GraphApiPort, site_id: str, drive_id: str, file_path: str) -> bool:
    """Pre-check idempotente de existencia (sin versionado): True si get_bytes no devuelve 404."""
    try:
        await graph.get_bytes(_build_content_endpoint(site_id, drive_id, file_path))
        return True
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return False
        raise


def _accounting_cell_filled(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    return str(value).strip() != ""


def _normalize_for_credit_match(text: str) -> str:
    if text is None:
        return ""
    value = unicodedata.normalize("NFD", str(text).strip().lower())
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = re.sub(r"\s+", " ", value)
    return value


def _extract_credit_id_for_match(credit_value: str) -> str | None:
    s = str(credit_value).strip()
    if not s:
        return None
    norm = _normalize_for_credit_match(s)
    m = re.search(r"(?:credito|obligacion)\s*#?\s*(\d+)", norm)
    if m:
        return m.group(1)
    if re.fullmatch(r"\d+", norm.strip()):
        return norm.strip()
    matches = list(re.finditer(r"(?<!\d)(\d+)(?!\d)", norm))
    if matches:
        return matches[-1].group(1)
    return None


async def _resolve_credit_folder_for_outputs(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    client_folder_path: str,
    credit_value: str,
) -> str:
    """Resuelve el nombre real de carpeta de crédito (solo lectura / rutas de salida)."""
    encoded = encode_graph_drive_path(client_folder_path)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:/children"
    try:
        resp = await client.get(endpoint)
    except Exception:
        raise ValueError("credit_folder_not_found") from None

    folders = [
        it
        for it in resp.get("value", [])
        if isinstance(it, dict) and "folder" in it and str(it.get("name", "")).strip()
    ]
    names = [str(it["name"]).strip() for it in folders]
    if not names:
        raise ValueError("credit_folder_not_found")

    norm_credit = _normalize_for_credit_match(credit_value)
    exact = [n for n in names if _normalize_for_credit_match(n) == norm_credit]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise ValueError("credit_folder_ambiguous")

    credit_id = _extract_credit_id_for_match(credit_value)
    if not credit_id:
        raise ValueError("credit_folder_not_found")

    token_re = re.compile(rf"(?<!\d){re.escape(credit_id)}(?!\d)")
    token_matches = [n for n in names if token_re.search(_normalize_for_credit_match(n))]
    if len(token_matches) == 0:
        raise ValueError("credit_folder_not_found")
    if len(token_matches) > 1:
        raise ValueError("credit_folder_ambiguous")
    return token_matches[0]


def _dist_column_map(ws: Any, header_row: int) -> dict[str, int]:
    m: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, c).value or "").strip()
        if h:
            m[h] = c
    return m


def _extract_internal_path_from_root_url(url: str) -> str | None:
    decoded = unquote(str(url))
    if "root:/" not in decoded:
        return None
    inner = decoded.split("root:/", 1)[1]
    inner = inner.split(":/", 1)[0]
    path = unquote(inner).replace("\\", "/").strip().strip("/")
    return path or None


def _looks_like_pdf_internal_path(value: str) -> bool:
    s = str(value).strip()
    if not s or s.lower().startswith("http"):
        return False
    if "/" not in s:
        return False
    return s.lower().endswith(".pdf")


def _internal_pdf_path_from_ruta_column(cell_ruta: Any | None, dist: dict[str, Any]) -> str | None:
    raw: Any = None
    if cell_ruta is not None:
        v = cell_ruta.value
        if v is not None and str(v).strip():
            raw = v
    if raw is None and dist.get(DistribucionCols.RUTA) not in (None, ""):
        raw = dist.get(DistribucionCols.RUTA)
    if raw is None:
        return None
    norm = str(raw).strip().replace("\\", "/")
    if _looks_like_pdf_internal_path(norm):
        return norm
    return None


def _extract_pdf_filename_from_sharepoint_url(url: str) -> str | None:
    m = re.search(r"(?i)[?&#]file=([^&]+)", str(url))
    if not m:
        return None
    f = unquote(m.group(1).replace("+", " ")).strip().strip("/")
    if not f.lower().endswith(".pdf"):
        return None
    return f.rsplit("/", 1)[-1]


def _extract_graph_item_reference_from_url(url: str) -> tuple[str, str] | None:
    m = re.search(r"/drives/([^/]+)/items/([^/?#]+)", str(url))
    if not m:
        return None
    return unquote(m.group(1)), unquote(m.group(2))


def _parent_reference_path_to_drive_relative(path: str) -> str | None:
    if not path:
        return None
    raw = unquote(str(path)).replace("\\", "/")
    if "root:/" not in raw:
        return None
    inner = raw.split("root:/", 1)[1]
    inner = inner.split(":/", 1)[0]
    return inner.strip("/") or None


def _drive_item_to_internal_path(item: dict[str, Any]) -> str | None:
    name = str(item.get("name", "") or "").strip()
    if not name.lower().endswith(".pdf"):
        return None
    pr = item.get("parentReference") or {}
    parent = _parent_reference_path_to_drive_relative(str(pr.get("path", "") or ""))
    if not parent:
        return None
    return f"{parent}/{name}".replace("//", "/")


def _hyperlink_target(cell: Any) -> str | None:
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    t = getattr(hl, "target", None) or getattr(hl, "ref", None)
    return str(t).strip() if t else None


def _resolve_extract_route_without_network(cell: Any) -> str | None:
    v = cell.value
    if v is not None:
        s = str(v).strip()
        if _looks_like_pdf_internal_path(s):
            return s.replace("\\", "/")
    target = _hyperlink_target(cell)
    if not target:
        return None
    gp = _extract_internal_path_from_root_url(target)
    if gp and gp.lower().endswith(".pdf") and "/" in gp:
        return gp
    if not target.lower().startswith("http") and _looks_like_pdf_internal_path(target):
        return target.replace("\\", "/")
    return None


def _parse_extract_date_from_filename(filename: str) -> date | None:
    lower = filename.strip().lower()
    m = re.search(r"(\d{4}-\d{2}-\d{2})\.pdf\s*$", lower)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            return None
    m = re.search(r"(\d{4}-\d{2})\.pdf\s*$", lower)
    if m:
        try:
            y, mo = map(int, m.group(1).split("-"))
            last = monthrange(y, mo)[1]
            return date(y, mo, last)
        except ValueError:
            return None
    found: list[date] = []
    for dm in re.finditer(r"\b(\d{4}-\d{2}-\d{2})\b", lower):
        try:
            found.append(datetime.strptime(dm.group(1), "%Y-%m-%d").date())
        except ValueError:
            continue
    if found:
        return found[-1]
    return None


def _coerce_fecha_banco_to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _pick_extract_pdf_name(
    pdf_items: list[dict[str, Any]],
    extract_keyword: str,
    fecha_banco: date | None,
    exact_basename: str | None,
) -> str | None:
    pdfs = [it for it in pdf_items if str(it.get("name", "")).lower().endswith(".pdf")]
    if not pdfs:
        return None
    if exact_basename:
        eb = exact_basename.strip().lower()
        for it in pdfs:
            if str(it.get("name", "")).strip().lower() == eb:
                return str(it["name"])
    kw = extract_keyword.lower()
    pool = [it for it in pdfs if kw in str(it.get("name", "")).lower()]
    if not pool:
        return None
    if len(pool) == 1:
        return str(pool[0]["name"])
    if fecha_banco:
        scored: list[tuple[int, str]] = []
        for it in pool:
            name = str(it.get("name", ""))
            d = _parse_extract_date_from_filename(name)
            if d is None:
                scored.append((10**9, name))
            else:
                scored.append((abs((d - fecha_banco).days), name))
        scored.sort(key=lambda x: (x[0], x[1]))
        return scored[0][1]
    return str(sorted(pool, key=lambda x: str(x.get("name", "")))[0]["name"])


async def _resolve_extract_route_from_graph_item(
    client: GraphApiPort, item_drive_id: str, item_id: str
) -> str | None:
    try:
        item = await client.get(
            f"/drives/{item_drive_id}/items/{item_id}",
            params={"$select": "name,parentReference"},
        )
        if isinstance(item, dict):
            p = _drive_item_to_internal_path(item)
            return p
    except Exception:
        return None
    return None


async def _resolve_extract_route_by_credit_folder(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    clients_path: str,
    cliente: str,
    credito: str,
    fecha_banco_raw: Any,
    extract_keyword: str,
    filename_hint: str | None,
) -> str | None:
    """
    Busca PDF de extracto; primero dentro de carpeta EXTRACTOS (si existe); si no, en raíz del crédito.
    Solo considera PDFs cuyo nombre contiene extract_keyword (vía _pick_extract_pdf_name).
    """
    if not cliente or not credito:
        return None

    credit_full: str
    try:
        real_credit = await _resolve_credit_folder_for_outputs(
            client, site_id, drive_id, f"{clients_path}/{cliente}", credito
        )
        credit_full = f"{clients_path}/{cliente}/{real_credit}"
    except Exception:
        return None

    fecha_d = _coerce_fecha_banco_to_date(fecha_banco_raw)
    hint_base = filename_hint.rsplit("/", 1)[-1] if filename_hint else None

    async def fetch_children(rel_path_full: str) -> list[dict[str, Any]]:
        enc_inner = encode_graph_drive_path(rel_path_full)
        try:
            resp = await client.get(
                f"/sites/{site_id}/drives/{drive_id}/root:/{enc_inner}:/children"
            )
            return list(resp.get("value") or [])
        except Exception:
            return []

    async def pdf_route_under(folder_full_path: str) -> str | None:
        inner = await fetch_children(folder_full_path)
        chosen = _pick_extract_pdf_name(inner, extract_keyword, fecha_d, hint_base)
        if not chosen:
            return None
        return f"{folder_full_path}/{chosen}".replace("//", "/")

    root_items = await fetch_children(credit_full)
    extractos_label = _normalize_folder_label("EXTRACTOS")
    extractos_name: str | None = None
    for it in root_items:
        if "folder" not in it:
            continue
        nm = str(it.get("name", "")).strip()
        if nm and _normalize_folder_label(nm) == extractos_label:
            extractos_name = nm
            break

    if extractos_name:
        inner_path = f"{credit_full}/{extractos_name}".replace("//", "/")
        routed = await pdf_route_under(inner_path)
        if routed:
            return routed

    return await pdf_route_under(credit_full.replace("//", "/"))


async def _resolve_extract_route_for_validar_row(
    client: GraphApiPort,
    drive_id: str,
    site_id: str,
    clients_path: str,
    extract_keyword: str,
    dist: dict[str, Any],
    cell_ext: Any,
) -> str | None:
    local = _resolve_extract_route_without_network(cell_ext)
    if local:
        return local
    target = _hyperlink_target(cell_ext)
    filename_hint: str | None = None
    if target:
        ref = _extract_graph_item_reference_from_url(target)
        if ref:
            id_drive, iid = ref
            gp = await _resolve_extract_route_from_graph_item(client, id_drive, iid)
            if gp:
                return gp
        filename_hint = _extract_pdf_filename_from_sharepoint_url(target)
    return await _resolve_extract_route_by_credit_folder(
        client,
        site_id,
        drive_id,
        clients_path,
        str(dist.get(DistribucionCols.CLIENTE, "")).strip(),
        str(dist.get(DistribucionCols.CREDITO, "")).strip(),
        dist.get(DistribucionCols.FECHA_BANCO),
        extract_keyword,
        filename_hint,
    )


def _dirname_internal_path(path: str) -> str | None:
    path = path.strip().replace("\\", "/")
    if "/" not in path:
        return None
    return path.rsplit("/", 1)[0]


def _credit_parent_dir_from_tabla_cell(cell: Any) -> str | None:
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    target = getattr(hl, "target", None)
    if not target:
        return None
    t = str(target).strip()
    p = _extract_internal_path_from_root_url(t) or (t if "/" in t and not t.lower().startswith("http") else None)
    if not p:
        return None
    return _dirname_internal_path(p)


def _normalize_folder_label(name: str) -> str:
    v = unicodedata.normalize("NFD", name.strip().lower())
    v = "".join(ch for ch in v if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", v)


def _asientos_folder_name_for_credit(credito: str, ruta_unidad_credito: str) -> str:
    """Nombre bajo la unidad de crédito: ASIENTOS CONTABLES CRED {n}."""
    credit_id = _extract_credit_id_for_match(str(credito or "").strip())
    if not credit_id:
        seg = ruta_unidad_credito.replace("\\", "/").strip("/").split("/")[-1]
        credit_id = _extract_credit_id_for_match(seg)
    if not credit_id:
        label = str(credito or "").strip()
        if not label:
            raise ValueError("credit_number_not_resolved")
        credit_id = label
    return f"{ASIENTOS_FOLDER_PER_CREDIT_PREFIX} {credit_id}"


def _folder_item_by_exact_name(children: list[dict[str, Any]], folder_name: str) -> dict[str, Any] | None:
    target = folder_name.casefold()
    for it in children:
        if "folder" not in it:
            continue
        nm = str(it.get("name", "")).strip()
        if nm.casefold() == target:
            return it
    return None


async def _list_folder_children_finalize(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    parent_rel: str,
) -> list[dict[str, Any]]:
    enc = encode_graph_drive_path(parent_rel.strip().strip("/"))
    resp = await client.get(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children")
    return list(resp.get("value") or [])


async def _ensure_asientos_folder_under_credit_unit(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    ruta_unidad_credito: str,
    folder_name: str,
) -> tuple[str, str | None]:
    """
    Crea la carpeta de asientos bajo la unidad de crédito si no existe.
    Retorna (ruta relativa al drive, webUrl opcional).
    """
    parent = ruta_unidad_credito.strip().replace("\\", "/").strip("/")
    if not parent:
        raise ValueError("missing_ruta_unidad_credito")
    rel_path = f"{parent}/{folder_name}".replace("//", "/")
    children = await _list_folder_children_finalize(client, site_id, drive_id, parent)
    existing = _folder_item_by_exact_name(children, folder_name)
    if existing is not None:
        web = existing.get("webUrl")
        return rel_path, str(web) if web else None

    enc = encode_graph_drive_path(parent)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children"
    body: dict[str, Any] = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail",
    }
    try:
        resp, code = await client.post_json(endpoint, body)
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 409:
            children = await _list_folder_children_finalize(client, site_id, drive_id, parent)
            existing = _folder_item_by_exact_name(children, folder_name)
            if existing is not None:
                web = existing.get("webUrl")
                return rel_path, str(web) if web else None
        raise
    else:
        if code not in (200, 201):
            raise RuntimeError(f"unexpected_graph_status_{code}")
        web = resp.get("webUrl") if isinstance(resp, dict) else None
        return rel_path, str(web) if web else None


async def _provision_asientos_folders_for_distribution_rows(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    ws_dist: Any,
    dist_header_row: int,
    distributions: list[dict[str, Any]],
) -> dict[int, tuple[str, str]]:
    """Crea carpetas ASIENTOS CONTABLES CRED {n} y rellena RutaAsientosContables por fila."""
    colmap = _dist_column_map(ws_dist, dist_header_row)
    col_ruta_uc = colmap.get(DistribucionCols.RUTA_UNIDAD_CREDITO)
    asientos_by_row: dict[int, tuple[str, str]] = {}

    for dist in distributions:
        if not _include_in_validation_outputs(dist):
            continue
        r = int(dist["_excel_row"])
        ruta_uc = str(dist.get(DistribucionCols.RUTA_UNIDAD_CREDITO) or "").strip().replace("\\", "/")
        if not ruta_uc and col_ruta_uc:
            cell_val = ws_dist.cell(r, col_ruta_uc).value
            ruta_uc = str(cell_val or "").strip().replace("\\", "/")
        if not ruta_uc:
            logger.error(
                "finalize missing_ruta_unidad_credito: fila=%s id_pago=%r cliente=%r credito=%r",
                r,
                dist.get(DistribucionCols.ID_PAGO),
                dist.get(DistribucionCols.CLIENTE),
                dist.get(DistribucionCols.CREDITO),
            )
            raise ValueError("missing_ruta_unidad_credito")

        credito = str(dist.get(DistribucionCols.CREDITO, "")).strip()
        try:
            folder_name = _asientos_folder_name_for_credit(credito, ruta_uc)
        except ValueError:
            logger.error(
                "finalize credit_number_not_resolved: fila=%s id_pago=%r credito=%r ruta_unidad=%r",
                r,
                dist.get(DistribucionCols.ID_PAGO),
                credito,
                ruta_uc,
            )
            raise ValueError("credit_number_not_resolved") from None

        try:
            rel_path, web_url = await _ensure_asientos_folder_under_credit_unit(
                client, site_id, drive_id, ruta_uc, folder_name
            )
        except Exception as exc:
            logger.error(
                "finalize asientos_folder_create_failed: fila=%s id_pago=%r cliente=%r credito=%r "
                "ruta_unidad=%r folder_name=%r detail=%s",
                r,
                dist.get(DistribucionCols.ID_PAGO),
                dist.get(DistribucionCols.CLIENTE),
                credito,
                ruta_uc,
                folder_name,
                exc,
                exc_info=True,
            )
            raise ValueError("asientos_folder_create_failed") from exc

        dist[DistribucionCols.RUTA_ASIENTOS_CONTABLES] = rel_path
        link_txt = web_url if web_url else rel_path
        asientos_by_row[r] = (link_txt, "")

    return asientos_by_row


def _configure_hist_distrib_technical_path_columns(
    ws_distribution: Any,
    dist_header_row: int,
) -> None:
    colmap = _dist_column_map(ws_distribution, dist_header_row)
    for col_name in (
        *DISTRIBUCION_TECHNICAL_HIDDEN_COLUMNS,
        DistribucionCols.RUTA_ASIENTOS_CONTABLES,
    ):
        cidx = colmap.get(col_name)
        if not cidx:
            continue
        letter = get_column_letter(cidx)
        wd = ws_distribution.column_dimensions[letter]
        wd.hidden = True
        wd.width = min(float(wd.width or 9.0), 12.0)


async def _apply_ruta_asientos_column_on_hist_sheet(
    ws_dist: Any,
    dist_header_row: int,
    distributions: list[dict[str, Any]],
) -> None:
    colmap = _dist_column_map(ws_dist, dist_header_row)
    col_asientos = colmap.get(DistribucionCols.RUTA_ASIENTOS_CONTABLES)
    if col_asientos is None:
        col_asientos = ws_dist.max_column + 1
        ws_dist.cell(dist_header_row, col_asientos, DistribucionCols.RUTA_ASIENTOS_CONTABLES)
        colmap[DistribucionCols.RUTA_ASIENTOS_CONTABLES] = col_asientos

    for dist in distributions:
        if not _include_in_validation_outputs(dist):
            continue
        r = int(dist["_excel_row"])
        rel = str(dist.get(DistribucionCols.RUTA_ASIENTOS_CONTABLES) or "").strip()
        if not rel:
            raise ValueError("missing_ruta_asientos_contables")
        ws_dist.cell(r, col_asientos, rel)


async def _apply_ruta_column_on_hist_sheet(
    client: GraphApiPort,
    site_id: str,
    drive_id: str,
    clients_path: str,
    extract_keyword: str,
    ws_dist: Any,
    dist_header_row: int,
    distributions: list[dict[str, Any]],
) -> None:
    colmap = _dist_column_map(ws_dist, dist_header_row)
    col_ruta = colmap.get(DistribucionCols.RUTA)
    if col_ruta is None:
        col_ruta = ws_dist.max_column + 1
        ws_dist.cell(dist_header_row, col_ruta, DistribucionCols.RUTA)
    col_link_ext = colmap.get(DistribucionCols.LINK_EXTRACTO)
    if col_link_ext is None:
        logger.error(
            "finalize missing_extract_route: columna %r no encontrada (fila encabezados=%s)",
            DistribucionCols.LINK_EXTRACTO,
            dist_header_row,
        )
        raise ValueError("missing_extract_route")

    for dist in distributions:
        r = int(dist["_excel_row"])
        if not _include_in_validation_outputs(dist):
            continue
        cell_ruta = ws_dist.cell(r, col_ruta)
        cell_ext = ws_dist.cell(r, col_link_ext)

        resolved: str | None = _internal_pdf_path_from_ruta_column(cell_ruta, dist)
        if not resolved:
            resolved = await _resolve_extract_route_for_validar_row(
                client,
                drive_id,
                site_id,
                clients_path,
                extract_keyword,
                dist,
                cell_ext,
            )

        if not resolved:
            logger.error(
                "finalize missing_extract_route: fila_excel=%s id_pago=%r cliente=%r credito=%r "
                "estado_linea=%r link_val=%r link_hyperlink=%r ruta_celda=%r dist_ruta=%r "
                "(se intentó fallback con subcarpeta EXTRACTOS antes que raíz de crédito)",
                r,
                dist.get(DistribucionCols.ID_PAGO),
                dist.get(DistribucionCols.CLIENTE),
                dist.get(DistribucionCols.CREDITO),
                dist.get(DistribucionCols.ESTADO_PAGO),
                cell_ext.value,
                _hyperlink_target(cell_ext),
                cell_ruta.value,
                dist.get(DistribucionCols.RUTA),
            )
            raise ValueError("missing_extract_route")
        cell_ruta.value = resolved


def _build_secretary_workbook(
    ws_src_dist: Any,
    dist_header_row: int,
    distributions: list[dict[str, Any]],
    asientos_by_row: dict[int, tuple[str, str]],
) -> bytes:
    colmap = _dist_column_map(ws_src_dist, dist_header_row)
    ncols = len(SECRETARY_HEADERS)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SECRETARY_SHEET
    _apply_secretary_banner(ws, ncols)
    _apply_secretary_header_style(ws, SECRETARY_HEADER_ROW, SECRETARY_HEADERS)
    hmap = {h: i + 1 for i, h in enumerate(SECRETARY_HEADERS)}

    row_idx = SECRETARY_FIRST_DATA_ROW
    for dist in distributions:
        if not _include_in_validation_outputs(dist):
            continue
        r = int(dist["_excel_row"])
        total_v = float(dist.get("total_f", 0))
        cliente = dist.get(DistribucionCols.CLIENTE)
        credito = dist.get(DistribucionCols.CREDITO)
        obs_as, _obs_note = asientos_by_row.get(r, (PENDIENTE_CREAR_ASIENTOS, OBS_NO_ASIENTOS))

        ws.cell(row_idx, hmap["Cliente"], cliente)
        ws.cell(row_idx, hmap["Crédito"], credito)
        ws.cell(row_idx, hmap["ID Pago"], dist.get(DistribucionCols.ID_PAGO))
        ws.cell(row_idx, hmap["Fecha banco"], dist.get(DistribucionCols.FECHA_BANCO))
        ws.cell(row_idx, hmap["Fecha límite"], dist.get(DistribucionCols.FECHA_LIMITE))
        ws.cell(row_idx, hmap["Total validado"], total_v)

        _set_secretary_url_link_cell(
            ws.cell(row_idx, hmap["Link carpeta asientos contables"]),
            obs_as,
            "carpeta_asientos",
            credito,
            cliente,
        )

        c_le = colmap.get(DistribucionCols.LINK_EXTRACTO)
        c_lt = colmap.get(DistribucionCols.LINK_TABLA)
        if c_le:
            _set_secretary_link_cell(
                ws.cell(row_idx, hmap["Link extracto"]),
                ws_src_dist.cell(r, c_le),
                "extracto",
                credito,
                cliente,
            )
        if c_lt:
            _set_secretary_link_cell(
                ws.cell(row_idx, hmap["Link tabla amortización"]),
                ws_src_dist.cell(r, c_lt),
                "tabla",
                credito,
                cliente,
            )

        row_idx += 1

    last_data_row = row_idx - 1
    first_data_row = SECRETARY_FIRST_DATA_ROW

    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass
    try:
        ws.sheet_properties.tabColor = Color(rgb=_SEC_TAB_COLOR)
    except Exception:
        pass

    if last_data_row >= first_data_row:
        _apply_secretary_body_style(ws, first_data_row, last_data_row, ncols, hmap)
        _apply_secretary_number_formats(ws, first_data_row, last_data_row, hmap)
        _apply_secretary_client_borders(
            ws, first_data_row, last_data_row, ncols, hmap["Cliente"]
        )
        trow = _apply_secretary_total_row(ws, hmap, first_data_row, last_data_row)
        max_r = trow if trow else last_data_row
        _auto_fit_secretary_columns(ws, 1, max_r, 1, ncols)
    else:
        _auto_fit_secretary_columns(ws, 1, SECRETARY_HEADER_ROW, 1, ncols)

    ws.freeze_panes = _secretary_freeze_panes_cell(hmap)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def finalize_payment_validation(
    client: GraphApiPort,
    validation_file: str = None,
    validation_file_path: str | None = None,
    process_date: date | str | None = None,
) -> dict[str, Any]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    review_path = os.getenv("GRAPH_PAYMENT_VALIDATION_REVIEW_PATH", "").strip()
    history_path = os.getenv("GRAPH_PAYMENT_VALIDATION_HISTORY_PATH", "").strip()
    validation_prefix = os.getenv("GRAPH_VALIDATION_FILE_PREFIX", "").strip()
    clients_path = os.getenv("GRAPH_CLIENTS_BASE_PATH", "").strip()
    extract_keyword = os.getenv("GRAPH_EXTRACT_KEYWORD", "Extracto").strip() or "Extracto"
    effective_process_date = _normalize_process_date(process_date)

    rev_info = await resolve_sharepoint_path(client, site_search, drive_name, review_path)

    if validation_file_path:
        file_path = validation_file_path
        validation_file = validation_file or file_path.rsplit("/", 1)[-1]
    else:
        if not validation_file:
            children = await client.get(
                f"/sites/{rev_info['site_id']}/drives/{rev_info['drive_id']}/root:/{rev_info['path_encoded']}:/children"
            )
            children_items = children.get("value", [])
            validation_file = _resolve_validation_selection(children_items, validation_prefix)

        file_path = f"{review_path}/{validation_file}"

    rev_bytes = await client.get_bytes(
        _build_content_endpoint(rev_info["site_id"], rev_info["drive_id"], file_path)
    )

    wb_rev = openpyxl.load_workbook(io.BytesIO(rev_bytes), data_only=False)

    if ReviewSheets.CONTROL not in wb_rev.sheetnames:
        raise ValueError("missing_control_sheet")
    if ReviewSheets.DISTRIBUCION not in wb_rev.sheetnames:
        raise ValueError("missing_distribucion_sheet")

    ws_ctrl = wb_rev[ReviewSheets.CONTROL]
    procesar = None
    estado_ctrl = None
    for row in ws_ctrl.iter_rows(values_only=True):
        if row and row[0] == ControlCols.ROW_PROCESAR:
            procesar = str(row[1]).strip().upper() if row[1] else ""
        if row and row[0] == ControlCols.ROW_ESTADO:
            estado_ctrl = str(row[1]).strip().upper() if row[1] else ""

    if procesar != ControlCols.VAL_PROCESAR_SI:
        raise ValueError("process_not_approved")

    if not estado_ctrl:
        raise ValueError("missing_control_state")
    if estado_ctrl != "EN_REVISION":
        raise ValueError("invalid_control_state")

    monto_casos: dict[str, float] = {}
    if ReviewSheets.CASOS_PAGO in wb_rev.sheetnames:
        ws_casos = wb_rev[ReviewSheets.CASOS_PAGO]
        casos_header_row = _find_table_header_row(ws_casos, CasosPagoCols.ID_PAGO)
        c_headers: list[str] = []
        for r_idx, row in enumerate(ws_casos.iter_rows(values_only=True), start=1):
            if r_idx < casos_header_row:
                continue
            if r_idx == casos_header_row:
                c_headers = [str(v).strip() if v else "" for v in row]
                continue
            if not any(row):
                continue
            rd = dict(zip(c_headers, row))
            monto_casos[str(rd.get(CasosPagoCols.ID_PAGO))] = float(rd.get(CasosPagoCols.MONTO_BANCO, 0) or 0)

    ws_dist = wb_rev[ReviewSheets.DISTRIBUCION]
    distributions: list[dict[str, Any]] = []

    dist_header_row = _find_table_header_row(ws_dist, DistribucionCols.ID_PAGO)
    headers: list[str] = []
    for r_idx, row in enumerate(ws_dist.iter_rows(values_only=True), start=1):
        if r_idx < dist_header_row:
            continue
        if r_idx == dist_header_row:
            headers = [str(v).strip() if v else "" for v in row]
            continue

        if not any(row):
            continue

        row_dict = normalize_distrib_row_keys(dict(zip(headers, row)))
        apply_legacy_estado_migration(row_dict)
        row_dict["_excel_row"] = r_idx
        distributions.append(row_dict)

    sum_aplicado: dict[str, float] = {}

    def safe_float(v: Any) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        text = str(v).strip()
        if not text or text.startswith("="):
            return 0.0
        try:
            return float(text)
        except ValueError:
            try:
                return float(text.replace(".", "").replace(",", "."))
            except ValueError:
                return 0.0

    for dist in distributions:
        estado = str(dist.get(DistribucionCols.ESTADO_PAGO, "")).strip().upper()
        if not estado or estado == "NONE":
            raise ValueError("empty_estado_pago")
        if estado not in EstadoPago.ALLOWED:
            raise ValueError("invalid_estado_pago")
        if estado in EstadoPago.FINALIZE_FORBIDDEN:
            raise ValueError("estado_pago_no_finalizable")

        id_pago = str(dist.get(DistribucionCols.ID_PAGO))
        raw_im = dist.get(DistribucionCols.OTROS_VALORES)
        raw_vi = dist.get(DistribucionCols.APLICAR_A_EXTRACTO)
        raw_ak = dist.get(DistribucionCols.MORA_A_APLICAR)
        obs = dist.get(DistribucionCols.OBSERVACION)
        vp_si = is_validar_pago_si(dist)

        if estado in EstadoPago.COUNTERS_POSITIVE_TOTAL and vp_si:
            if not _accounting_cell_filled(raw_vi):
                raise ValueError("missing_valor_intereses")
            if not _accounting_cell_filled(raw_ak):
                raise ValueError("missing_abono_k")
            if not _accounting_cell_filled(raw_im):
                raise ValueError("missing_mora")

        mora_f = safe_float(raw_im)
        int_f = safe_float(raw_vi)
        abono_f = safe_float(raw_ak)

        total_f = int_f + abono_f + mora_f
        dist[DistribucionCols.TOTAL_APLICADO] = total_f

        if id_pago not in monto_casos:
            monto_casos[id_pago] = safe_float(dist.get(DistribucionCols.MONTO_BANCO))

        if estado in (EstadoPago.NORMAL, EstadoPago.INCOMPLETO) and not vp_si:
            if not obs or str(obs).strip() == "":
                raise ValueError("no_validar_requires_observation")
        elif estado in EstadoPago.COUNTERS_POSITIVE_TOTAL and vp_si:
            if total_f <= 0:
                raise ValueError("validar_requires_positive_total")
            sum_aplicado[id_pago] = sum_aplicado.get(id_pago, 0) + total_f

        dist["mora_f"] = mora_f
        dist["int_f"] = int_f
        dist["abono_f"] = abono_f
        dist["total_f"] = total_f

    for idp, sum_ap in sum_aplicado.items():
        if idp in monto_casos:
            if abs(sum_ap - monto_casos[idp]) > 0.01:
                raise ValueError("amount_mismatch")

    validated_rows = sum(1 for d in distributions if _include_in_validation_outputs(d))

    clients_info = await resolve_sharepoint_path(client, site_search, drive_name, clients_path)
    clients_drive_id = clients_info["drive_id"]
    clients_site_id = clients_info["site_id"]
    asientos_by_row = await _provision_asientos_folders_for_distribution_rows(
        client,
        clients_site_id,
        clients_drive_id,
        ws_dist,
        dist_header_row,
        distributions,
    )

    wb_hist = openpyxl.load_workbook(io.BytesIO(rev_bytes), data_only=False)
    ws_hist_dist = wb_hist[ReviewSheets.DISTRIBUCION]
    hist_dist_header = _find_table_header_row(ws_hist_dist, DistribucionCols.ID_PAGO)
    await _apply_ruta_column_on_hist_sheet(
        client,
        clients_site_id,
        clients_drive_id,
        clients_path,
        extract_keyword,
        ws_hist_dist,
        hist_dist_header,
        distributions,
    )
    await _apply_ruta_asientos_column_on_hist_sheet(
        ws_hist_dist,
        hist_dist_header,
        distributions,
    )
    _configure_hist_distrib_technical_path_columns(ws_hist_dist, hist_dist_header)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_hist_ctrl = wb_hist[ReviewSheets.CONTROL]
    ws_hist_ctrl.append([ControlCols.ROW_ESTADO_PROCESO, ControlCols.VAL_PROCESADO])
    ws_hist_ctrl.append([ControlCols.ROW_FECHA_PROCESAMIENTO, now_str])
    ws_hist_ctrl.append([ControlCols.ROW_RESULTADO, ControlCols.VAL_FINALIZADO])

    out_hist = io.BytesIO()
    wb_hist.save(out_hist)
    hist_bytes = out_hist.getvalue()

    sec_bytes = _build_secretary_workbook(
        ws_hist_dist, hist_dist_header, distributions, asientos_by_row
    )

    hist_name = f"cartera_validada_{effective_process_date.isoformat()}.xlsx"
    sec_name = f"soporte_asientos_contables_{effective_process_date.isoformat()}.xlsx"

    hist_info = await resolve_sharepoint_path(client, site_search, drive_name, history_path)
    hist_full_path = f"{history_path}/{hist_name}"
    sec_full_path = f"{history_path}/{sec_name}"

    historical_file_already_exists = await _drive_item_exists(
        client, hist_info["site_id"], hist_info["drive_id"], hist_full_path
    )
    secretary_file_already_exists = await _drive_item_exists(
        client, hist_info["site_id"], hist_info["drive_id"], sec_full_path
    )

    historical_file_action = (
        "replaced" if historical_file_already_exists else "created"
    )
    secretary_file_action = (
        "replaced" if secretary_file_already_exists else "created"
    )

    already_exists_warning: dict[str, Any] = {}
    if historical_file_already_exists or secretary_file_already_exists:
        already_exists_warning = {
            "severity": "warning",
            "warning_code": "FINALIZE_OUTPUT_ALREADY_EXISTS",
            "user_message": (
                "Finalize terminó, pero el histórico/soporte del día ya existía en SharePoint "
                "y fue reemplazado."
            ),
            "next_action": (
                "Revise en SharePoint los archivos reemplazados (02 HISTORICO / soporte_asientos_contables) "
                "si esperaba un resultado diferente."
            ),
        }

    payment_followup_warnings = await register_payment_followups_after_finalize(
        client,
        hist_info["site_id"],
        hist_info["drive_id"],
        process_date=effective_process_date,
        distributions=distributions,
        historical_relative_path=hist_full_path,
    )

    historical_file_url: str | None = None
    secretary_file_url: str | None = None
    try:
        hist_resp = await client.put_bytes(
            _build_content_endpoint(hist_info["site_id"], hist_info["drive_id"], hist_full_path),
            hist_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        historical_file_url = hist_resp.get("webUrl") if isinstance(hist_resp, dict) else None
        sec_resp = await client.put_bytes(
            _build_content_endpoint(hist_info["site_id"], hist_info["drive_id"], sec_full_path),
            sec_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        secretary_file_url = sec_resp.get("webUrl") if isinstance(sec_resp, dict) else None
    except Exception as e:
        raise Exception(f"upload_failed|{hist_full_path}|{sec_full_path}|{str(e)}") from e

    return {
        "status": "success",
        "historical_file_path": hist_full_path,
        "historical_file_url": historical_file_url,
        "historical_file_action": historical_file_action,
        "historical_file_already_exists": historical_file_already_exists,
        "secretary_file_path": sec_full_path,
        "secretary_file_url": secretary_file_url,
        "secretary_file_action": secretary_file_action,
        "secretary_file_already_exists": secretary_file_already_exists,
        "validated_rows": validated_rows,
        "amortization_updated": False,
        "bank_cleaned": False,
        "history_file": hist_name,
        "validation_file": validation_file,
        "validation_file_path": file_path,
        "process_date": effective_process_date.isoformat(),
        "payment_followup_warnings": payment_followup_warnings,
        **already_exists_warning,
    }
