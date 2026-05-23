"""
Setup idempotente de bandejas operativas pagos_adelantados / pagos_incompletos (Pendientes + Historico).

Estructura simple sin Table/ListObject para compatibilidad con Excel Online.
Hojas totalmente protegidas (solo la API escribe vía backend).
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any

import httpx
import openpyxl

from app.application.services.workbook_setup_helpers import (
    configure_internal_automation_sheet,
    sheet_headers_match,
)
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_FOLDER_RELATIVE_PATH,
    MergeControlSetupError,
    ensure_merge_control_folder_path,
)
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

SHEET_PENDIENTES = "Pendientes"
SHEET_HISTORICO = "Historico"

FILENAME_ADELANTADOS = "pagos_adelantados.xlsx"
FILENAME_INCOMPLETOS = "pagos_incompletos.xlsx"

ADELANTADOS_COLUMNS: tuple[str, ...] = (
    "ID Pago",
    "Cliente",
    "Crédito",
    "FechaPago",
    "FechaLimitePago",
    "FechaIBRRequerida",
    "MontoBanco",
    "ValorExtracto",
    "AplicarAExtracto",
    "MoraAAplicar",
    "OtrosValores",
    "TotalAplicado",
    "SaldoPorAsignar",
    "TablaAmortizacionPath",
    "RutaUnidadCredito",
    "RutaExtracto",
    "AsientoPdfPath",
    "HistoricalFilePath",
    "FilaAplicacionPago",
    "FilaIBR",
    "EstadoAplicacionPago",
    "EstadoIBR",
    "EstadoFinal",
    "IBRUsado",
    "FechaCierreIBR",
    "Observacion",
    "CreatedAt",
    "UpdatedAt",
)

INCOMPLETOS_COLUMNS: tuple[str, ...] = (
    "ID Pago",
    "Cliente",
    "Crédito",
    "FechaPago",
    "FechaLimitePago",
    "MontoBanco",
    "ValorExtracto",
    "AplicarAExtracto",
    "MoraAAplicar",
    "OtrosValores",
    "TotalAplicado",
    "SaldoPorAsignar",
    "TablaAmortizacionPath",
    "RutaUnidadCredito",
    "RutaExtracto",
    "AsientoPdfPath",
    "HistoricalFilePath",
    "EstadoAplicacionPago",
    "EstadoFinal",
    "Observacion",
    "CreatedAt",
    "UpdatedAt",
)


class PaymentFollowupSetupError(Exception):
    def __init__(
        self,
        *,
        http_status: int,
        user_message: str,
        next_action: str,
        technical_message: str,
    ) -> None:
        self.http_status = http_status
        self.user_message = user_message
        self.next_action = next_action
        self.technical_message = technical_message
        super().__init__(technical_message)


def _content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


def _item_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:"


def _env_workbook_path(new_var: str, legacy_var: str, default_filename: str) -> str:
    p = os.getenv(new_var, "").strip()
    if not p:
        p = os.getenv(legacy_var, "").strip()
    if p:
        return p
    folder = followup_workbooks_folder_relative_path().strip().rstrip("/")
    return f"{folder}/{default_filename}"


def followup_workbooks_folder_relative_path() -> str:
    p = os.getenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", "").strip()
    return p or MERGE_CONTROL_FOLDER_RELATIVE_PATH


def adelantados_workbook_relative_path() -> str:
    return _env_workbook_path(
        "GRAPH_FOLLOWUP_PAGOS_ADELANTADOS_PATH",
        "GRAPH_AUDIT_PAGOS_ADELANTADOS_PATH",
        FILENAME_ADELANTADOS,
    )


def incompletos_workbook_relative_path() -> str:
    return _env_workbook_path(
        "GRAPH_FOLLOWUP_PAGOS_INCOMPLETOS_PATH",
        "GRAPH_AUDIT_PAGOS_INCOMPLETOS_PATH",
        FILENAME_INCOMPLETOS,
    )


async def _file_exists(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> bool:
    try:
        await graph.get_bytes(_content_endpoint(site_id, drive_id, path))
        return True
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return False
        raise


async def _drive_item_web_url(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> str | None:
    try:
        meta = await graph.get(_item_endpoint(site_id, drive_id, path))
        wu = meta.get("webUrl")
        return str(wu).strip() if isinstance(wu, str) and wu.strip() else None
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return None
        raise


def _build_followup_workbook_bytes(*, columns: tuple[str, ...]) -> bytes:
    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = SHEET_PENDIENTES
    configure_internal_automation_sheet(ws_p, columns)
    ws_h = wb.create_sheet(SHEET_HISTORICO)
    configure_internal_automation_sheet(ws_h, columns)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _validate_followup_workbook(data: bytes, columns: tuple[str, ...]) -> tuple[bool, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    try:
        for sheet in (SHEET_PENDIENTES, SHEET_HISTORICO):
            if sheet not in wb.sheetnames:
                warnings.append(f"needs_manual_review: falta la hoja {sheet!r}")
                return False, warnings
            ws = wb[sheet]
            if not sheet_headers_match(ws, columns):
                warnings.append(f"needs_manual_review: encabezados incorrectos en {sheet!r}")
                return False, warnings
            if ws.tables:
                warnings.append(f"needs_manual_review: hoja {sheet!r} contiene Table/ListObject; use force_recreate")
                return False, warnings
        if "Registros" in wb.sheetnames:
            warnings.append("needs_manual_review: hoja legacy Registros presente; use force_recreate")
            return False, warnings
        return True, warnings
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


async def _upload_workbook(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    file_path: str,
    payload: bytes,
) -> str | None:
    resp = await graph.put_bytes(
        _content_endpoint(site_id, drive_id, file_path),
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if isinstance(resp, dict):
        wu = resp.get("webUrl")
        if isinstance(wu, str) and wu.strip():
            return wu.strip()
    return await _drive_item_web_url(graph, site_id, drive_id, file_path)


async def _ensure_one_followup_workbook(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    file_path: str,
    *,
    columns: tuple[str, ...],
    force_recreate: bool,
) -> dict[str, Any]:
    warnings: list[str] = []
    exists = await _file_exists(graph, site_id, drive_id, file_path)

    if exists and not force_recreate:
        raw = await graph.get_bytes(_content_endpoint(site_id, drive_id, file_path))
        ok, w = _validate_followup_workbook(raw, columns)
        warnings.extend(w)
        return {
            "file_path": file_path,
            "created": False,
            "recreated": False,
            "file_url": await _drive_item_web_url(graph, site_id, drive_id, file_path),
            "structure_ok": ok,
            "sheet_protection": "full_locked",
            "warnings": warnings,
        }

    if exists and force_recreate:
        warnings.append("force_recreate: archivo reemplazado con estructura operativa nueva")

    payload = _build_followup_workbook_bytes(columns=columns)
    file_url = await _upload_workbook(graph, site_id, drive_id, file_path, payload)
    return {
        "file_path": file_path,
        "created": not exists,
        "recreated": exists and force_recreate,
        "file_url": file_url,
        "structure_ok": True,
        "sheet_protection": "full_locked",
        "warnings": warnings,
    }


async def setup_payment_followup_workbooks(
    graph: GraphApiPort,
    *,
    force_recreate: bool = False,
) -> dict[str, Any]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    if not site_search:
        raise GraphConfigError("Missing environment variable: GRAPH_SHAREPOINT_SITE_SEARCH")

    folder_rel = followup_workbooks_folder_relative_path()
    base = await resolve_sharepoint_path(graph, site_search, drive_name, folder_rel)
    site_id = base["site_id"]
    drive_id = base["drive_id"]

    try:
        await ensure_merge_control_folder_path(graph, site_id, drive_id, folder_rel)
    except MergeControlSetupError as exc:
        raise PaymentFollowupSetupError(
            http_status=exc.http_status,
            user_message=exc.user_message,
            next_action=exc.next_action,
            technical_message=exc.technical_message,
        ) from exc

    path_ad = adelantados_workbook_relative_path()
    path_in = incompletos_workbook_relative_path()
    global_warnings: list[str] = []

    try:
        out_ad = await _ensure_one_followup_workbook(
            graph, site_id, drive_id, path_ad, columns=ADELANTADOS_COLUMNS, force_recreate=force_recreate
        )
        out_in = await _ensure_one_followup_workbook(
            graph, site_id, drive_id, path_in, columns=INCOMPLETOS_COLUMNS, force_recreate=force_recreate
        )
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        body_txt = (exc.response.text if exc.response else "") or ""
        if code == 403:
            raise PaymentFollowupSetupError(
                http_status=403,
                user_message="No se pudieron crear las bandejas operativas en SharePoint.",
                next_action="Verifique permisos de escritura en 00 CONTROL.",
                technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
            ) from exc
        raise PaymentFollowupSetupError(
            http_status=502,
            user_message="Error al subir bandejas operativas a SharePoint.",
            next_action="Verifique bloqueos (423) o permisos en la biblioteca.",
            technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
        ) from exc

    global_warnings.extend(out_ad.get("warnings") or [])
    global_warnings.extend(out_in.get("warnings") or [])

    return {
        "status": "success",
        "force_recreate": force_recreate,
        "folder_relative_path": folder_rel,
        "pagos_adelantados": {
            "file_path": out_ad["file_path"],
            "created": out_ad["created"],
            "recreated": out_ad["recreated"],
            "file_url": out_ad["file_url"],
            "structure_ok": out_ad["structure_ok"],
            "sheets": [SHEET_PENDIENTES, SHEET_HISTORICO],
        },
        "pagos_incompletos": {
            "file_path": out_in["file_path"],
            "created": out_in["created"],
            "recreated": out_in["recreated"],
            "file_url": out_in["file_url"],
            "structure_ok": out_in["structure_ok"],
            "sheets": [SHEET_PENDIENTES, SHEET_HISTORICO],
        },
        "warnings": global_warnings,
    }
