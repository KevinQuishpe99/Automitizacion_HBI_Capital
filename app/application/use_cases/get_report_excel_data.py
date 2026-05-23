"""
Lee el Excel de reporte en SharePoint (GRAPH_SHAREPOINT_FILE_PATH) y expone filas como datos JSON.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.application.sharepoint_resolution import resolve_sharepoint_from_env
from app.application.use_cases.validate_payment_report import _find_header_map
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ReportExcelData:
    file_path: str
    sheet_name: str
    header_row: int
    columns: list[str]
    rows: list[dict[str, Any]]
    total_data_rows: int
    returned_offset: int
    returned_limit: int | None
    truncated: bool


def _unique_header_labels(ws: Any, header_row: int, max_col: int) -> list[str]:
    labels: list[str] = []
    counts: dict[str, int] = {}
    for c in range(1, max_col + 1):
        raw = ws.cell(row=header_row, column=c).value
        base = str(raw).strip() if raw is not None and str(raw).strip() else f"_col_{c}"
        n = counts.get(base, 0) + 1
        counts[base] = n
        labels.append(base if n == 1 else f"{base}_{n}")
    return labels


def _cell_to_json(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def _row_is_empty(values: list[Any]) -> bool:
    for v in values:
        if v is not None and str(v).strip():
            return False
    return True


def _iter_data_row_values(ws: Any, *, first_data: int, last_row: int, max_col: int):
    for r in range(first_data, last_row + 1):
        yield r, [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]


async def get_report_excel_data(
    graph: GraphApiPort,
    *,
    sheet_index: int = 0,
    limit: int | None = None,
    offset: int = 0,
    skip_empty_rows: bool = True,
) -> ReportExcelData:
    """
    Descarga el .xlsx del reporte, detecta la fila de encabezados (misma heurística que la validación)
    y devuelve las filas de datos como lista de dicts.

    :param sheet_index: índice 0-based de la hoja (por defecto la primera).
    :param limit: máximo de filas de datos a devolver (None = todas tras aplicar offset).
    :param offset: filas de datos a saltar desde la primera fila bajo el encabezado.
    :param skip_empty_rows: si True, no incluye filas totalmente vacías en el conteo ni en la respuesta.
    """
    if offset < 0:
        raise ValueError("offset debe ser >= 0")
    if limit is not None and limit < 0:
        raise ValueError("limit debe ser >= 0 o None")

    ctx = await resolve_sharepoint_from_env(graph)
    try:
        excel_bytes = await graph.get_bytes(
            f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content"
        )
    except Exception as exc:
        raise ValueError(
            "No se pudo descargar el Excel del reporte (no existe, ruta inválida o sin permisos)."
        ) from exc

    try:
        wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("El archivo configurado no es un Excel válido (.xlsx).") from exc

    try:
        if sheet_index < 0 or sheet_index >= len(wb.worksheets):
            raise ValueError(
                f"Hoja inválida: índice {sheet_index} (el libro tiene {len(wb.worksheets)} hoja(s))."
            )

        ws = wb.worksheets[sheet_index]
        sheet_name = ws.title
        header_row, _header_map = _find_header_map(ws)
        max_col = ws.max_column or 0
        column_labels = _unique_header_labels(ws, header_row, max_col)

        first_data = header_row + 1
        last_row = ws.max_row or header_row

        total_data_rows = 0
        for _r, values in _iter_data_row_values(ws, first_data=first_data, last_row=last_row, max_col=max_col):
            if skip_empty_rows and _row_is_empty(values):
                continue
            total_data_rows += 1

        out_rows: list[dict[str, Any]] = []

        if skip_empty_rows:
            idx = -1
            for _r, values in _iter_data_row_values(ws, first_data=first_data, last_row=last_row, max_col=max_col):
                if _row_is_empty(values):
                    continue
                idx += 1
                if idx < offset:
                    continue
                if limit is not None and len(out_rows) >= limit:
                    break
                row_dict = {
                    column_labels[i]: _cell_to_json(values[i]) for i in range(len(column_labels))
                }
                out_rows.append(row_dict)
        else:
            for r, values in _iter_data_row_values(ws, first_data=first_data, last_row=last_row, max_col=max_col):
                idx = r - first_data
                if idx < offset:
                    continue
                if limit is not None and len(out_rows) >= limit:
                    break
                row_dict = {
                    column_labels[i]: _cell_to_json(values[i]) for i in range(len(column_labels))
                }
                out_rows.append(row_dict)

        truncated = limit is not None and offset + len(out_rows) < total_data_rows

        logger.info(
            "get_report_excel_data: leído reporte",
            extra={
                "path": ctx["file_path"],
                "sheet": sheet_name,
                "total_data_rows": total_data_rows,
                "returned": len(out_rows),
            },
        )

        return ReportExcelData(
            file_path=ctx["file_path"],
            sheet_name=sheet_name,
            header_row=header_row,
            columns=list(column_labels),
            rows=out_rows,
            total_data_rows=total_data_rows,
            returned_offset=offset,
            returned_limit=limit,
            truncated=truncated,
        )
    finally:
        wb.close()
