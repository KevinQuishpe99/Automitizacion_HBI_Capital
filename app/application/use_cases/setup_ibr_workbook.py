"""
Setup idempotente de IBR_DIARIO.xlsx (hoja IBR: Inicio, Fin, Valor).
Sin Table/ListObject; encabezados bloqueados y filas de datos editables.
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any

import httpx
import openpyxl

from app.application.services.workbook_setup_helpers import (
    configure_secretary_editable_sheet,
    sheet_headers_match,
)
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_FOLDER_RELATIVE_PATH,
    MergeControlSetupError,
    ensure_merge_control_folder_path,
)
from app.application.use_cases.setup_payment_followup_workbooks import followup_workbooks_folder_relative_path
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

IBR_SHEET_NAME = "IBR"
IBR_COLUMNS: tuple[str, ...] = ("Inicio", "Fin", "Valor")
FILENAME_IBR = "IBR_DIARIO.xlsx"


class IbrWorkbookSetupError(Exception):
    def __init__(
        self,
        *,
        http_status: int,
        user_message: str,
        next_action: str,
        technical_message: str,
    ) -> None:
        self.http_status = http_status
        self.user_message = user_message
        self.next_action = next_action
        self.technical_message = technical_message
        super().__init__(technical_message)


def ibr_workbook_relative_path() -> str:
    p = os.getenv("GRAPH_IBR_DIARIO_PATH", "").strip()
    if p:
        return p
    folder = followup_workbooks_folder_relative_path().strip().rstrip("/")
    return f"{folder}/{FILENAME_IBR}"


def _content_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:/content"


def _item_endpoint(site_id: str, drive_id: str, file_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(file_path)}:"


async def _file_exists(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> bool:
    try:
        await graph.get_bytes(_content_endpoint(site_id, drive_id, path))
        return True
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return False
        raise


async def _drive_item_web_url(graph: GraphApiPort, site_id: str, drive_id: str, path: str) -> str | None:
    try:
        meta = await graph.get(_item_endpoint(site_id, drive_id, path))
        wu = meta.get("webUrl")
        return str(wu).strip() if isinstance(wu, str) and wu.strip() else None
    except httpx.HTTPStatusError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            return None
        raise


def _build_new_ibr_workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = IBR_SHEET_NAME
    configure_secretary_editable_sheet(ws, IBR_COLUMNS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _validate_ibr_workbook(data: bytes) -> tuple[bool, list[str]]:
    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    try:
        if IBR_SHEET_NAME not in wb.sheetnames:
            warnings.append(f"needs_manual_review: falta la hoja {IBR_SHEET_NAME!r}")
            return False, warnings
        ws = wb[IBR_SHEET_NAME]
        if not sheet_headers_match(ws, IBR_COLUMNS):
            warnings.append("needs_manual_review: encabezados IBR no coinciden")
            return False, warnings
        if ws.tables:
            warnings.append("needs_manual_review: hoja IBR contiene Table/ListObject; use force_recreate")
            return False, warnings
        return True, warnings
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


def _repair_existing_workbook(data: bytes) -> tuple[bytes, list[str], bool]:
    """Añade hoja/columnas IBR sin borrar datos; quita tablas si existían."""
    warnings: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    modified = False
    try:
        if IBR_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(IBR_SHEET_NAME)
            warnings.append(f"sheet_added:{IBR_SHEET_NAME}")
            modified = True
        else:
            ws = wb[IBR_SHEET_NAME]

        if ws.tables:
            for name in list(ws.tables.keys()):
                del ws.tables[name]
            modified = True
            warnings.append("removed_tables:IBR")

        before_headers = sheet_headers_match(ws, IBR_COLUMNS)
        configure_secretary_editable_sheet(ws, IBR_COLUMNS)
        if not before_headers or ws.tables:
            modified = True

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), warnings, modified
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


async def _upload_workbook(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    file_path: str,
    payload: bytes,
) -> str | None:
    resp = await graph.put_bytes(
        _content_endpoint(site_id, drive_id, file_path),
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if isinstance(resp, dict):
        wu = resp.get("webUrl")
        if isinstance(wu, str) and wu.strip():
            return wu.strip()
    return await _drive_item_web_url(graph, site_id, drive_id, file_path)


async def setup_ibr_workbook(graph: GraphApiPort, *, force_recreate: bool = False) -> dict[str, Any]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    if not site_search:
        raise GraphConfigError("Missing environment variable: GRAPH_SHAREPOINT_SITE_SEARCH")

    folder_rel = followup_workbooks_folder_relative_path()
    base = await resolve_sharepoint_path(graph, site_search, drive_name, folder_rel)
    site_id = base["site_id"]
    drive_id = base["drive_id"]
    file_path = ibr_workbook_relative_path()

    try:
        await ensure_merge_control_folder_path(graph, site_id, drive_id, folder_rel)
    except MergeControlSetupError as exc:
        raise IbrWorkbookSetupError(
            http_status=exc.http_status,
            user_message=exc.user_message,
            next_action=exc.next_action,
            technical_message=exc.technical_message,
        ) from exc

    warnings: list[str] = []
    exists = await _file_exists(graph, site_id, drive_id, file_path)

    try:
        if not exists or force_recreate:
            if exists and force_recreate:
                warnings.append("force_recreate: IBR_DIARIO.xlsx reemplazado")
            payload = _build_new_ibr_workbook_bytes()
            file_url = await _upload_workbook(graph, site_id, drive_id, file_path, payload)
            return {
                "status": "success",
                "force_recreate": force_recreate,
                "file_path": file_path,
                "created": not exists,
                "recreated": exists and force_recreate,
                "file_url": file_url,
                "sheet_name": IBR_SHEET_NAME,
                "columns": list(IBR_COLUMNS),
                "sheet_protection": "header_locked_data_editable",
                "warnings": warnings,
            }

        raw = await graph.get_bytes(_content_endpoint(site_id, drive_id, file_path))
        ok, val_warnings = _validate_ibr_workbook(raw)
        warnings.extend(val_warnings)
        if ok:
            return {
                "status": "success",
                "force_recreate": False,
                "file_path": file_path,
                "created": False,
                "recreated": False,
                "file_url": await _drive_item_web_url(graph, site_id, drive_id, file_path),
                "sheet_name": IBR_SHEET_NAME,
                "columns": list(IBR_COLUMNS),
                "sheet_protection": "header_locked_data_editable",
                "warnings": warnings,
            }

        payload, repair_warnings, modified = _repair_existing_workbook(raw)
        warnings.extend(repair_warnings)
        file_url = await _drive_item_web_url(graph, site_id, drive_id, file_path)
        if modified:
            file_url = await _upload_workbook(graph, site_id, drive_id, file_path, payload) or file_url

        return {
            "status": "success",
            "force_recreate": False,
            "file_path": file_path,
            "created": False,
            "recreated": modified,
            "file_url": file_url,
            "sheet_name": IBR_SHEET_NAME,
            "columns": list(IBR_COLUMNS),
            "sheet_protection": "header_locked_data_editable",
            "warnings": warnings,
        }
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        body_txt = (exc.response.text if exc.response else "") or ""
        raise IbrWorkbookSetupError(
            http_status=502 if code != 403 else 403,
            user_message="Error al preparar IBR_DIARIO.xlsx en SharePoint.",
            next_action="Verifique permisos o bloqueos en 00 CONTROL.",
            technical_message=f"Graph HTTP {code}: {body_txt[:2000]}",
        ) from exc
