"""
Actualiza el Excel de control Merge tras Notify exitoso (fila 2, hoja Procesos).
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Any

import httpx
from openpyxl import load_workbook

from app.application.sharepoint_resolution import encode_graph_drive_path
from app.application.use_cases.setup_merge_control_workbook import (
    MERGE_CONTROL_COLUMNS,
    SHEET_NAME,
    TABLE_DISPLAY_NAME,
    apply_merge_control_worksheet_protection,
    merge_control_workbook_relative_path,
)
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

_MERGE_CONTROL_ACTIVE_ESTADOS = frozenset(
    {"PENDIENTE_ASIENTOS", "CONSOLIDANDO", "MERGE_PARCIAL", "ERROR_MERGE"}
)


def _col(name: str) -> int:
    return MERGE_CONTROL_COLUMNS.index(name) + 1


def _norm_estado(value: Any) -> str:
    return str(value or "").strip()


def _is_active_cell(value: Any) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    s = str(value).strip().lower()
    return s in ("true", "1", "yes", "si", "sí", "x")


def _should_block_overwrite(estado: str, is_active: bool) -> bool:
    return is_active and _norm_estado(estado) in _MERGE_CONTROL_ACTIVE_ESTADOS


def _content_put_endpoint(site_id: str, drive_id: str, rel_path: str) -> str:
    return f"/sites/{site_id}/drives/{drive_id}/root:/{encode_graph_drive_path(rel_path)}:/content"


async def _download_workbook_bytes(graph: GraphApiPort, site_id: str, drive_id: str, rel: str) -> bytes:
    enc = encode_graph_drive_path(rel.strip().strip("/"))
    return await graph.get_bytes(f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content")


def _headers_match(ws: Any) -> bool:
    for i, exp in enumerate(MERGE_CONTROL_COLUMNS, start=1):
        if str(ws.cell(row=1, column=i).value or "").strip() != exp:
            return False
    return True


@dataclass(frozen=True)
class MergeControlNotifyWriteOutcome:
    merge_control_updated: bool
    merge_control_file_path: str
    merge_control_status: str | None
    merge_control_warning: str | None
    merge_control_error_code: str | None


async def update_merge_control_workbook_after_notify(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    *,
    historical_file_path: str,
    email_pdf_path: str | None,
    report_d: date,
) -> MergeControlNotifyWriteOutcome:
    control_rel = merge_control_workbook_relative_path().strip().strip("/")

    if not (email_pdf_path or "").strip():
        return MergeControlNotifyWriteOutcome(
            merge_control_updated=False,
            merge_control_file_path=control_rel,
            merge_control_status=None,
            merge_control_warning=(
                "No se registró el paso para unir PDFs: no se generó el PDF copia del correo en la carpeta 05 EMAIL. "
                "El correo sí pudo haberse enviado; revise esa carpeta en SharePoint."
            ),
            merge_control_error_code="missing_email_pdf_path_for_merge_control",
        )

    hist = historical_file_path.strip().strip("/")
    pdf = email_pdf_path.strip().strip("/")

    try:
        raw = await _download_workbook_bytes(graph, site_id, drive_id, control_rel)
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        if code == 404:
            return MergeControlNotifyWriteOutcome(
                merge_control_updated=False,
                merge_control_file_path=control_rel,
                merge_control_status=None,
                merge_control_warning=(
                    "No existe control_merge_pdfs.xlsx en 00 CONTROL. "
                    "Soporte debe crear ese archivo una vez antes de poder unir PDFs después del correo."
                ),
                merge_control_error_code="merge_control_workbook_not_found",
            )
        logger.warning("merge control: error al descargar workbook HTTP %s", code)
        return MergeControlNotifyWriteOutcome(
            merge_control_updated=False,
            merge_control_file_path=control_rel,
            merge_control_status=None,
            merge_control_warning=(
                f"No se pudo abrir control_merge_pdfs.xlsx en SharePoint (error {code}). "
                "Cierre el archivo si está abierto y verifique permisos."
            ),
            merge_control_error_code="merge_control_read_failed",
        )
    except Exception as exc:
        logger.exception("merge control: fallo inesperado al descargar")
        return MergeControlNotifyWriteOutcome(
            merge_control_updated=False,
            merge_control_file_path=control_rel,
            merge_control_status=None,
            merge_control_warning=str(exc)[:2000],
            merge_control_error_code="merge_control_read_failed",
        )

    wb = load_workbook(filename=io.BytesIO(raw), data_only=False)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return MergeControlNotifyWriteOutcome(
                merge_control_updated=False,
                merge_control_file_path=control_rel,
                merge_control_status=None,
                merge_control_warning=(
                    'El archivo control_merge_pdfs.xlsx no tiene la hoja "Procesos". '
                    "Pida a soporte restaurar la plantilla del archivo de control."
                ),
                merge_control_error_code="merge_control_invalid_structure",
            )
        ws = wb[SHEET_NAME]
        if not _headers_match(ws):
            return MergeControlNotifyWriteOutcome(
                merge_control_updated=False,
                merge_control_file_path=control_rel,
                merge_control_status=None,
                merge_control_warning=(
                    "Los encabezados de control_merge_pdfs.xlsx no coinciden con la plantilla esperada. "
                    "No modifique ese archivo a mano; pida a soporte restaurarlo."
                ),
                merge_control_error_code="merge_control_invalid_structure",
            )
        if ws.max_row < 2:
            return MergeControlNotifyWriteOutcome(
                merge_control_updated=False,
                merge_control_file_path=control_rel,
                merge_control_status=None,
                merge_control_warning=(
                    "Falta la fila 2 de datos en control_merge_pdfs.xlsx. "
                    "Pida a soporte restaurar la plantilla del archivo de control."
                ),
                merge_control_error_code="merge_control_invalid_structure",
            )

        estado = _norm_estado(ws.cell(row=2, column=_col("EstadoProceso")).value)
        is_active = _is_active_cell(ws.cell(row=2, column=_col("IsActive")).value)

        if _should_block_overwrite(estado, is_active):
            return MergeControlNotifyWriteOutcome(
                merge_control_updated=False,
                merge_control_file_path=control_rel,
                merge_control_status=estado or None,
                merge_control_warning=(
                    "En control_merge_pdfs.xlsx ya hay un proceso activo de unión de PDFs. "
                    "Termine o cancele ese proceso antes de registrar el correo de hoy."
                ),
                merge_control_error_code="merge_control_active_process_exists",
            )

        now_iso = datetime.now(timezone.utc).isoformat()
        title = f"Merge {report_d.isoformat()}"

        ws.cell(row=2, column=_col("Title"), value=title)
        ws.cell(row=2, column=_col("EstadoProceso"), value="PENDIENTE_ASIENTOS")
        ws.cell(row=2, column=_col("IsActive"), value=True)
        ws.cell(row=2, column=_col("HistoricalFilePath"), value=hist)
        ws.cell(row=2, column=_col("EmailPdfPath"), value=pdf)
        ws.cell(row=2, column=_col("MergeOutputCount"), value=0)
        ws.cell(row=2, column=_col("MergeSkippedCount"), value=0)
        ws.cell(row=2, column=_col("LastErrorUserMessage"), value="")
        ws.cell(row=2, column=_col("LastErrorNextAction"), value="")
        ws.cell(row=2, column=_col("CreatedAtProceso"), value=now_iso)
        ws.cell(row=2, column=_col("LastUpdatedAtProceso"), value=now_iso)

        if TABLE_DISPLAY_NAME not in ws.tables:
            logger.warning("merge control: tabla %s ausente; se continúa sin crear tabla", TABLE_DISPLAY_NAME)

        apply_merge_control_worksheet_protection(ws)

        buf = io.BytesIO()
        wb.save(buf)
        out_bytes = buf.getvalue()
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()

    try:
        await graph.put_bytes(
            _content_put_endpoint(site_id, drive_id, control_rel),
            out_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        logger.exception("merge control: fallo al subir workbook actualizado")
        return MergeControlNotifyWriteOutcome(
            merge_control_updated=False,
            merge_control_file_path=control_rel,
            merge_control_status=None,
            merge_control_warning=str(exc)[:2000],
            merge_control_error_code="merge_control_write_failed",
        )

    return MergeControlNotifyWriteOutcome(
        merge_control_updated=True,
        merge_control_file_path=control_rel,
        merge_control_status="PENDIENTE_ASIENTOS",
        merge_control_warning=None,
        merge_control_error_code=None,
    )
