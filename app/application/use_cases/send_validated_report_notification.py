"""
Lee el Excel del reporte en SharePoint, filtra filas Validado=SI y Estado=OK,
descarga los PDF indicados en Rutas (mismo site/drive que el reporte), los adjunta al correo,
y envía sendMail por Microsoft Graph.
Requiere permisos de aplicación Mail.Send y acceso de lectura a los PDF en el drive.
"""

from __future__ import annotations

import base64
import html
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_from_env
from app.application.use_cases.validate_payment_report import (
    _find_header_map,
    _get_col,
    _norm_key,
    _parse_excel_date,
)
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class NotifyResult:
    rows_included: int
    subject: str
    attachments_count: int


def _excel_cell_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _row_dates_for_intro(fecha_raw: Any) -> list[datetime.date]:
    try:
        dt = _parse_excel_date(fecha_raw)
        return [dt.date()]
    except Exception:
        return []


def _format_fechas_intro(dates: list[datetime.date]) -> str:
    uniq = sorted(set(dates))
    return ", ".join(d.strftime("%d/%m/%Y") for d in uniq)


def _split_rutas(value: Any) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = re.split(r",\s*", s)
    return [p.strip().replace("\\", "/") for p in parts if p.strip()]


def _collect_unique_paths(paths: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _build_html_email(
    *,
    intro_html: str,
    fechas_str: str,
    banco: str,
    rutas: list[str],
    table_rows: list[dict[str, str]],
    attachment_note: str,
) -> str:
    intro = intro_html.replace("{fechas}", html.escape(fechas_str)).replace(
        "{banco}", html.escape(banco)
    )
    ul = ""
    if rutas:
        items = "".join(f"<li>{html.escape(r)}</li>" for r in rutas)
        ul = (
            f"<p><strong>Rutas de extractos (SharePoint):</strong></p>"
            f"<p>Los PDF de esas rutas se descargan desde el mismo drive y se adjuntan a este correo "
            f"(cuando sea posible).</p><ul>{items}</ul>"
        )
    note_block = f"<p>{html.escape(attachment_note)}</p>" if attachment_note else ""
    thead = (
        "<tr>"
        "<th>Fecha</th><th>Crédito</th><th>Concepto</th><th>Transacción</th><th>Rutas</th>"
        "</tr>"
    )
    tbody = ""
    for row in table_rows:
        tbody += (
            "<tr>"
            f"<td>{html.escape(row['fecha'])}</td>"
            f"<td>{html.escape(row['credito'])}</td>"
            f"<td>{html.escape(row['concepto'])}</td>"
            f"<td>{html.escape(row['transaccion'])}</td>"
            f"<td>{html.escape(row['rutas'])}</td>"
            "</tr>"
        )
    table = (
        "<p><strong>Detalle validado (Validado=SI, Estado=OK):</strong></p>"
        f'<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse">'
        f"{thead}<tbody>{tbody}</tbody></table>"
    )
    return f"<html><body>{intro}{note_block}{ul}{table}</body></html>"


def _attach_max_bytes() -> int:
    try:
        return max(0, int(os.getenv("GRAPH_MAIL_ATTACH_MAX_BYTES", "20000000")))
    except ValueError:
        return 20_000_000


async def _pdf_attachments_from_rutas(
    graph: GraphApiPort,
    *,
    site_id: str,
    drive_id: str,
    rutas_unique: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Descarga PDF por ruta bajo el drive y arma adjuntos fileAttachment para sendMail."""
    attachments: list[dict[str, Any]] = []
    failures: list[str] = []
    max_b = _attach_max_bytes()

    for rel in rutas_unique:
        rel_n = rel.strip().replace("\\", "/")
        if not rel_n.lower().endswith(".pdf"):
            failures.append(f"{rel_n} (no es PDF, no se adjunta)")
            continue
        try:
            enc = encode_graph_drive_path(rel_n)
            endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/content"
            data = await graph.get_bytes(endpoint)
        except Exception as exc:
            logger.warning("notify email: error descargando extracto %s: %s", rel_n, exc)
            failures.append(f"{rel_n} ({exc})")
            continue

        if max_b and len(data) > max_b:
            failures.append(f"{rel_n} (tamaño > {max_b} bytes, no adjuntado)")
            continue

        name = rel_n.rsplit("/", 1)[-1] or "extracto.pdf"
        b64 = base64.standard_b64encode(data).decode("ascii")
        attachments.append(
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": name,
                "contentType": "application/pdf",
                "contentBytes": b64,
            }
        )

    return attachments, failures


async def send_validated_report_notification_email(
    graph: GraphApiPort,
    *,
    to_override: str | None = None,
    cc_override: str | None = None,
    subject_override: str | None = None,
    intro_html_override: str | None = None,
) -> NotifyResult:
    sender = os.getenv("GRAPH_MAIL_SENDER_EMAIL", "").strip()
    if not sender:
        raise GraphConfigError("Missing environment variable: GRAPH_MAIL_SENDER_EMAIL")

    to_addr = (to_override or os.getenv("GRAPH_MAIL_TO", "").strip()).strip()
    if not to_addr:
        raise GraphConfigError("Missing recipient: set GRAPH_MAIL_TO or pass to_override")

    cc_raw = (cc_override if cc_override is not None else os.getenv("GRAPH_MAIL_CC", "")).strip()
    cc_list = [e.strip() for e in cc_raw.split(",") if e.strip()] if cc_raw else []

    subject = (subject_override or os.getenv("GRAPH_MAIL_SUBJECT", "").strip()).strip() or (
        "Notificación — reporte validado (SI / OK)"
    )
    banco = os.getenv("GRAPH_NOTIFY_BANK_NAME", "BANCO BOGOTA").strip() or "BANCO BOGOTA"
    default_intro = (
        "<p>Buenos días. El día <strong>{fechas}</strong> se ingresó a la cuenta "
        "<strong>{banco}</strong> los siguientes valores, que corresponden a los extractos "
        "indicados en las rutas.</p>"
    )
    intro_template = (intro_html_override or os.getenv("GRAPH_MAIL_INTRO_HTML", "").strip()).strip()
    if not intro_template:
        intro_template = default_intro

    logger.info("notify email: resolviendo SharePoint y descargando Excel de reporte")
    ctx = await resolve_sharepoint_from_env(graph)
    excel_bytes = await graph.get_bytes(
        f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content"
    )

    wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    ws = wb.worksheets[0]
    try:
        header_row, header_map = _find_header_map(ws)
        col_fecha = _get_col(header_map, "Fecha")
        col_credito = _get_col(header_map, "Crédito", "Credito")
        col_concepto = _get_col(header_map, "Concepto")
        col_trans = _get_col(header_map, "Transacción", "Transaccion")
        col_validado = _get_col(header_map, "Validado", "VALIDADO")
        col_estado = _get_col(header_map, "Estado", "ESTADO")
        col_rutas = _get_col(header_map, "Rutas", "RUTAS")

        if not (col_fecha and col_credito and col_concepto):
            raise ValueError("Faltan columnas Fecha / Crédito / Concepto.")
        if not (col_validado and col_estado and col_rutas):
            raise ValueError("Faltan columnas Validado / Estado / Rutas.")

        all_paths: list[str] = []
        table_rows: list[dict[str, str]] = []
        intro_dates: list[datetime.date] = []

        for r in range(header_row + 1, ws.max_row + 1):
            v_ok = _norm_key(str(ws.cell(row=r, column=col_validado).value or ""))
            e_ok = _norm_key(str(ws.cell(row=r, column=col_estado).value or ""))
            if v_ok != "SI" or e_ok != "OK":
                continue

            fecha_raw = ws.cell(row=r, column=col_fecha).value
            intro_dates.extend(_row_dates_for_intro(fecha_raw))

            rutas_cell = ws.cell(row=r, column=col_rutas).value
            row_paths = _split_rutas(rutas_cell)
            all_paths.extend(row_paths)

            fecha_s = _excel_cell_display(fecha_raw)
            credito_s = _excel_cell_display(ws.cell(row=r, column=col_credito).value)
            concepto_s = _excel_cell_display(ws.cell(row=r, column=col_concepto).value)
            trans_s = _excel_cell_display(
                ws.cell(row=r, column=col_trans).value if col_trans else ""
            )
            rutas_s = ", ".join(row_paths) if row_paths else ""

            table_rows.append(
                {
                    "fecha": fecha_s,
                    "credito": credito_s,
                    "concepto": concepto_s,
                    "transaccion": trans_s,
                    "rutas": rutas_s,
                }
            )
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()

    if not table_rows:
        raise ValueError(
            "No hay filas con Validado=SI y Estado=OK para enviar en el correo."
        )

    fechas_str = _format_fechas_intro(intro_dates) if intro_dates else "(ver columna Fecha en tabla)"
    rutas_unique = _collect_unique_paths(all_paths)

    attach_on = os.getenv("GRAPH_MAIL_ATTACH_PDFS", "true").strip().lower() in (
        "1",
        "true",
        "yes",
        "si",
        "sí",
    )
    failures: list[str] = []
    attachments: list[dict[str, Any]] = []
    if attach_on and rutas_unique:
        attachments, failures = await _pdf_attachments_from_rutas(
            graph,
            site_id=ctx["site_id"],
            drive_id=ctx["drive_id"],
            rutas_unique=rutas_unique,
        )

    attachment_note = ""
    if attachments:
        attachment_note = f"Adjuntos: {len(attachments)} PDF(s) descargados desde SharePoint."
    if failures:
        extra = " Errores al adjuntar: " + "; ".join(failures[:10])
        if len(failures) > 10:
            extra += f" … y {len(failures) - 10} más."
        attachment_note = (attachment_note + extra).strip()

    html_body = _build_html_email(
        intro_html=intro_template,
        fechas_str=fechas_str,
        banco=banco,
        rutas=rutas_unique,
        table_rows=table_rows,
        attachment_note=attachment_note,
    )

    message: dict[str, Any] = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html_body},
        "toRecipients": [{"emailAddress": {"address": to_addr}}],
    }
    if cc_list:
        message["ccRecipients"] = [{"emailAddress": {"address": cc}} for cc in cc_list]
    if attachments:
        message["attachments"] = attachments

    payload = {"message": message, "saveToSentItems": True}

    path_user = quote(sender, safe="")
    endpoint = f"/users/{path_user}/sendMail"
    logger.info(
        "notify email: enviando sendMail",
        extra={
            "sender": sender,
            "to": to_addr,
            "rows": len(table_rows),
            "attachments": len(attachments),
        },
    )
    _, _ = await graph.post_json(endpoint, payload)

    return NotifyResult(
        rows_included=len(table_rows),
        subject=subject,
        attachments_count=len(attachments),
    )
