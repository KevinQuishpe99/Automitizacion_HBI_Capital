from __future__ import annotations

import logging
import os
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader

from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_from_env
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort


@dataclass(frozen=True)
class ValidationSummary:
    processed_rows: int
    ok_count: int
    no_count: int
    error_count: int
    errors: list[dict[str, Any]]


@dataclass(frozen=True)
class CachedClientResult:
    """Resultado por cliente/fecha de pago."""

    total: Decimal
    selected_date: datetime | None = None
    selected_file_name: str = ""
    selected_path: str = ""
    analyzed_paths: tuple[str, ...] = ()
    observation: str = ""
    error: str | None = None
    is_adelantado: bool = False  # fecha pago < fecha extracto (ancla)

    @staticmethod
    def fail(message: str) -> CachedClientResult:
        return CachedClientResult(total=Decimal("0"), error=message, observation=message)


logger = logging.getLogger(__name__)


def _norm_key(value: str) -> str:
    return " ".join(value.strip().upper().split())


_ES_MONTHS = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "set": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}


def _parse_excel_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if value is None:
        raise ValueError("Fecha vacía")
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    # 30-dic-2025 / 30/dic/2025
    m = re.match(r"^(\d{1,2})[-/]([a-záéíóú]{3,9})[-/](\d{2,4})$", s, re.IGNORECASE)
    if m:
        day = int(m.group(1))
        mon_token = m.group(2).lower()[:3]
        mon_token = mon_token.replace("á", "a").replace("é", "e").replace("í", "i")
        month = _ES_MONTHS.get(mon_token)
        if month:
            yraw = m.group(3)
            year = int(yraw) if len(yraw) == 4 else 2000 + int(yraw)
            return datetime(year, month, day)
    raise ValueError(f"Fecha inválida: {s!r}")


def _parse_extract_date_from_filename(filename: str) -> datetime | None:
    """
    Fecha en nombre de archivo: ... YYYY-MM-DD.pdf o ... YYYY-MM.pdf (último tramo antes de .pdf).
    Para YYYY-MM se usa el último día del mes como referencia (extracto mensual).
    """
    lower = filename.strip().lower()
    m = re.search(r"(\d{4}-\d{2}-\d{2})\.pdf\s*$", lower)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d")
        except ValueError:
            return None
    m = re.search(r"(\d{4}-\d{2})\.pdf\s*$", lower)
    if m:
        try:
            y, mo = map(int, m.group(1).split("-"))
            last = monthrange(y, mo)[1]
            return datetime(y, mo, last)
        except ValueError:
            return None
    return None


def _collect_all_extract_candidates(
    items: list[dict[str, Any]],
    extract_keyword: str,
    credit_path: str,
) -> list[tuple[datetime, str, str, dict[str, Any]]]:
    """Todos los PDF de extracto con fecha parseable en el nombre (sin filtrar por cercanía)."""
    candidates: list[tuple[datetime, str, str, dict[str, Any]]] = []
    for it in items:
        if "file" not in it:
            continue
        name = str(it.get("name", "")).strip()
        if not name.lower().endswith(".pdf"):
            continue
        if extract_keyword.lower() not in name.lower():
            continue
        d = _parse_extract_date_from_filename(name)
        if d is None:
            continue
        full_path = f"{credit_path}/{name}"
        candidates.append((d, name, full_path, it))
    return candidates


def _dedupe_extract_candidates(
    candidates: list[tuple[datetime, str, str, dict[str, Any]]],
) -> list[tuple[datetime, str, str, dict[str, Any]]]:
    seen: set[str] = set()
    out: list[tuple[datetime, str, str, dict[str, Any]]] = []
    for c in candidates:
        iid = str(c[3].get("id", "") or "").strip()
        key = iid if iid else c[2]
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return out


def _resolve_anchor_date(
    all_dates: set[datetime.date],
    payment_dt: datetime,
) -> tuple[datetime.date | None, str]:
    """
    Entre todas las fechas de extractos del cliente, compara la más cercana hacia atrás
    (fecha de extracto < día de pago) y la más cercana hacia adelante (>= día de pago).
    Elige la que esté a menos días del pago; si empatan, se usa la fecha hacia adelante.
    """
    if not all_dates:
        return None, ""
    payment_d = payment_dt.date()
    day_set = all_dates
    past = [d for d in day_set if d < payment_d]
    future = [d for d in day_set if d >= payment_d]
    nearest_past = max(past) if past else None
    nearest_future = min(future) if future else None
    if nearest_past is None and nearest_future is None:
        return None, ""
    if nearest_past is None:
        return nearest_future, (
            f"Ancla: {nearest_future} (solo extractos en/fecha posterior al pago "
            f"{_format_date_short(payment_dt)})."
        )
    if nearest_future is None:
        return nearest_past, (
            f"Ancla: {nearest_past} (solo extractos con fecha anterior al pago "
            f"{_format_date_short(payment_dt)})."
        )
    dist_p = (payment_d - nearest_past).days
    dist_f = (nearest_future - payment_d).days
    if dist_f < dist_p:
        return nearest_future, (
            f"Ancla: {nearest_future} (más cercana al pago {payment_d}: adelante {dist_f} d "
            f"vs atrás {dist_p} d desde {nearest_past})."
        )
    if dist_p < dist_f:
        return nearest_past, (
            f"Ancla: {nearest_past} (más cercana al pago {payment_d}: atrás {dist_p} d "
            f"vs adelante {dist_f} d desde {nearest_future})."
        )
    return nearest_future, (
        f"Ancla: {nearest_future} (empate {dist_p} d con {nearest_past}; se prioriza fecha hacia adelante)."
    )


def _find_best_subset_match(
    folder_rows: list[tuple[str, Decimal, list[str], list[str]]],
    target: Decimal,
    tol: Decimal,
) -> tuple[Decimal, list[int], bool]:
    """
    Busca combinación de carpetas cuya suma de montos coincida con target (tolerancia).
    Si hay varias coincidencias exactas, elige la que use menos carpetas.
    Si no hay exacta, devuelve la combinación con menor |suma - target|.
    Retorna (suma, índices de folder_rows, es_coincidencia_exacta).
    """
    n = len(folder_rows)
    if n == 0:
        return Decimal("0"), [], False
    best_exact: tuple[Decimal, list[int], int] | None = None  # sum, idxs, len
    best_any: tuple[Decimal, list[int], Decimal] | None = None  # sum, idxs, diff

    for mask in range(1, 1 << n):
        idxs = [i for i in range(n) if mask & (1 << i)]
        s = sum(folder_rows[i][1] for i in idxs)
        diff = abs(s - target)
        if best_any is None or diff < best_any[2] or (diff == best_any[2] and len(idxs) < len(best_any[1])):
            best_any = (s, idxs, diff)
        if diff <= tol:
            ln = len(idxs)
            if best_exact is None or ln < best_exact[2]:
                best_exact = (s, idxs, ln)

    if best_exact is not None:
        return best_exact[0], best_exact[1], True
    assert best_any is not None
    return best_any[0], best_any[1], False


_MONEY_RE = re.compile(r"[-+]?\d[\d\.\s]*([,]\d{1,2})?")


def _parse_money(value: Any) -> Decimal:
    if value is None:
        raise ValueError("Crédito vacío")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    s = s.replace("$", "").replace("USD", "").replace("US$", "").strip()
    m = _MONEY_RE.search(s)
    if not m:
        raise ValueError(f"No pude leer monto: {s!r}")
    num = m.group(0)
    num = num.replace(" ", "")
    num = num.replace(".", "").replace(",", ".")
    try:
        return Decimal(num)
    except InvalidOperation as exc:
        raise ValueError(f"Monto inválido: {s!r}") from exc


def _format_money(d: Decimal) -> str:
    q = d.quantize(Decimal("0.01"))
    # Formato estilo es-EC: 1.234.567,89
    s = f"{q:,.2f}"  # 1,234,567.89
    s = s.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"$ {s}"


def _format_date_short(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def _build_observation_text(
    chosen_total: Decimal,
    credito: Decimal,
    payment_dt: datetime,
    ref_dt: datetime,
    diff_days: int,
    per_extract_lines: tuple[str, ...],
    criterio: str,
) -> str:
    detalle = "; ".join(per_extract_lines)
    return (
        f"Valor calculado (extractos): {_format_money(chosen_total)}. "
        f"Por extracto: {detalle}. "
        f"Monto reporte (Crédito): {_format_money(credito)}. "
        f"Pago reporte: {_format_date_short(payment_dt)}; "
        f"referencia cercana (fecha en archivo): {_format_date_short(ref_dt)} "
        f"(diferencia {diff_days} día(s)). Criterio: {criterio}."
    )


def _find_header_map(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """
    Encuentra una fila de encabezados y mapea nombre normalizado -> índice columna (1-based).
    Busca en las primeras 20 filas.
    """
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normed = [_norm_key(str(v)) if v is not None else "" for v in row_values]
        has_fecha = "FECHA" in normed
        has_concepto = "CONCEPTO" in normed
        has_credito = ("CRÉDITO" in normed) or ("CREDITO" in normed)
        if has_fecha and has_concepto and has_credito:
            header_map: dict[str, int] = {}
            for col, name in enumerate(normed, start=1):
                if name:
                    header_map[name] = col
            return row_idx, header_map
    raise ValueError("No pude encontrar encabezados (Fecha/Crédito/Concepto).")


def _column_has_values(ws: Worksheet, header_row: int, col: int) -> bool:
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip():
            return True
    return False


def _get_col(header_map: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        key = _norm_key(c)
        if key in header_map:
            return header_map[key]
    return None


def _read_pdf_text(pdf_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(pdf_bytes))
    text_parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t:
            text_parts.append(t)
    text = "\n".join(text_parts)
    if not text.strip():
        raise ValueError("PDF sin texto (posible escaneado/OCR requerido)")
    return text


def _money_from_spanish_num_token(raw: str) -> Decimal:
    raw = raw.replace(" ", "").replace(".", "").replace(",", ".")
    return Decimal(raw)


def _extract_total_a_pagar_from_text(text: str) -> Decimal:
    patterns = [
        r"TOTAL\s*A\s*PAGAR\s*[:\-\s]*\$?\s*([\d\.\,]+)",
        r"TOTAL\s*A\s*PAGAR\s*\$?\s*([\d\.\,]+)",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return _money_from_spanish_num_token(m.group(1))
    raise ValueError("No encontré 'TOTAL A PAGAR' en el PDF")


def _extract_total_a_pagar_from_pdf(pdf_bytes: bytes) -> Decimal:
    return _extract_total_a_pagar_from_text(_read_pdf_text(pdf_bytes))


async def _graph_children_by_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    path: str,
) -> list[dict[str, Any]]:
    encoded = encode_graph_drive_path(path)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:/children"
    resp = await graph.get(endpoint)
    return list(resp.get("value") or [])


async def _graph_download_by_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    path: str,
) -> bytes:
    encoded = encode_graph_drive_path(path)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:/content"
    return await graph.get_bytes(endpoint)


async def _graph_get_item_metadata_by_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    path: str,
) -> dict[str, Any]:
    """Metadatos del ítem (eTag, size, lastModifiedDateTime) sin descargar contenido."""
    encoded = encode_graph_drive_path(path)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:"
    try:
        return await graph.get(endpoint)
    except Exception:
        return {}


async def _graph_upload_by_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    path: str,
    content: bytes,
    *,
    content_type: str = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ),
) -> dict[str, Any]:
    encoded = encode_graph_drive_path(path)
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{encoded}:/content"
    return await graph.put_bytes(endpoint, content, content_type=content_type)


async def _graph_download_by_item_id(
    graph: GraphApiPort,
    drive_id: str,
    item_id: str,
) -> bytes:
    return await graph.get_bytes(f"/drives/{drive_id}/items/{item_id}/content")


async def validate_payment_report_and_replace_excel(graph: GraphApiPort) -> ValidationSummary:
    """
    - Descarga el Excel del reporte (GRAPH_SHAREPOINT_FILE_PATH).
    - Por cliente: escanea todas las carpetas de crédito; calcula fecha ancla global (más cercana
      al día de pago del reporte comparando la mejor fecha hacia atrás vs hacia adelante).
    - Solo se usan PDFs cuya fecha (en el nombre) coincide con esa ancla; en la misma carpeta se
      suman varios PDF con la misma fecha.
    - Por PDF se usa TOTAL A PAGAR; se empareja combinación de carpetas (subset-sum) con el Crédito.
    - Validado = SI si Crédito >= suma TOTAL A PAGAR de los extractos de la combinación elegida.
    - Estado: OK (normal), REVISION (pago adelantado vs fecha extracto), ERROR (sin extracto / fallo).
    - Columnas: Validado, Estado, Observaciones, Rutas. Sube el Excel a SharePoint.
    """
    logger.info("validate_payment_report: iniciando resolución de SharePoint desde entorno")
    ctx = await resolve_sharepoint_from_env(graph)
    logger.info(
        "validate_payment_report: sitio y archivo resueltos",
        extra={
            "site_id": ctx["site_id"],
            "drive_id": ctx["drive_id"],
            "path": ctx["file_path"],
        },
    )
    try:
        excel_bytes = await graph.get_bytes(
            f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content"
        )
    except Exception as exc:
        raise ValueError(
            "No se pudo descargar el Excel del reporte (no existe, ruta inválida o sin permisos)."
        ) from exc

    clients_base = os.getenv("GRAPH_CLIENTS_BASE_PATH", "").strip()
    if not clients_base:
        raise GraphConfigError("Missing environment variable: GRAPH_CLIENTS_BASE_PATH")
    extract_keyword = os.getenv("GRAPH_EXTRACT_KEYWORD", "Extracto").strip() or "Extracto"
    try:
        match_tol = Decimal(os.getenv("GRAPH_CREDIT_MATCH_TOLERANCE", "0.02"))
    except InvalidOperation:
        match_tol = Decimal("0.02")
    logger.info(
        "validate_payment_report: ancla global (adelante/atrás) + empareje de cuotas por carpeta",
        extra={"match_tolerance": str(match_tol)},
    )

    try:
        wb = load_workbook(filename=BytesIO(excel_bytes))
    except Exception as exc:
        raise ValueError("El archivo configurado no es un Excel válido (.xlsx).") from exc
    ws = wb.worksheets[0]
    if ws.protection.sheet:
        raise ValueError(
            "La hoja del Excel está protegida y no se puede editar (Validado/Estado/Observaciones/Rutas)."
        )
    header_row, header_map = _find_header_map(ws)
    logger.info(
        "validate_payment_report: encabezados detectados",
        extra={"header_row": header_row, "headers": list(header_map.keys())},
    )

    col_fecha = _get_col(header_map, "Fecha")
    col_credito = _get_col(header_map, "Crédito", "Credito")
    col_concepto = _get_col(header_map, "Concepto")
    col_validado = _get_col(header_map, "Validado", "VALIDADO")
    col_observaciones = _get_col(header_map, "Observaciones", "OBSERVACIONES")
    col_rutas = _get_col(header_map, "Rutas", "RUTAS")
    col_estado = _get_col(header_map, "Estado", "ESTADO")

    if not (col_fecha and col_credito and col_concepto):
        raise ValueError("Faltan columnas requeridas: Fecha/Crédito/Concepto.")

    # Si no existen columnas de salida, las creamos al final.
    if not col_validado:
        col_validado = ws.max_column + 1
        ws.cell(row=header_row, column=col_validado).value = "Validado"
    if not col_observaciones:
        col_observaciones = ws.max_column + 1
        ws.cell(row=header_row, column=col_observaciones).value = "Observaciones"
    if not col_rutas:
        col_rutas = ws.max_column + 1
        ws.cell(row=header_row, column=col_rutas).value = "Rutas"
    if not col_estado:
        col_estado = ws.max_column + 1
        ws.cell(row=header_row, column=col_estado).value = "Estado"

    # Regla solicitada: si las columnas ya tienen datos, se detiene para evitar sobreescritura.
    dirty_cols: list[str] = []
    if _column_has_values(ws, header_row, col_validado):
        dirty_cols.append("Validado")
    if _column_has_values(ws, header_row, col_observaciones):
        dirty_cols.append("Observaciones")
    if _column_has_values(ws, header_row, col_rutas):
        dirty_cols.append("Rutas")
    if _column_has_values(ws, header_row, col_estado):
        dirty_cols.append("Estado")
    if dirty_cols:
        raise ValueError(
            "Las columnas de salida ya contienen datos "
            f"({', '.join(dirty_cols)}). Limpia esas columnas y vuelve a ejecutar."
        )

    # Cache de carpetas de clientes
    logger.info(
        "validate_payment_report: listando carpeta base de clientes",
        extra={"clients_base": clients_base},
    )
    client_items = await _graph_children_by_path(
        graph, ctx["site_id"], ctx["drive_id"], clients_base
    )
    client_by_norm: dict[str, str] = {}
    for item in client_items:
        if "folder" not in item:
            continue
        name = str(item.get("name", "")).strip()
        if name:
            client_by_norm[_norm_key(name)] = name

    # Cache por (cliente_norm, fecha_pago ISO)
    total_cache: dict[tuple[str, str], CachedClientResult] = {}

    processed = ok = no = err = 0
    errors: list[dict[str, Any]] = []

    logger.info(
        "validate_payment_report: comenzando recorrido de filas",
        extra={"first_row": header_row + 1, "last_row": ws.max_row},
    )

    for r in range(header_row + 1, ws.max_row + 1):
        concepto_raw = ws.cell(row=r, column=col_concepto).value
        if concepto_raw is None or not str(concepto_raw).strip():
            continue
        concepto_norm = _norm_key(str(concepto_raw))

        try:
            dt = _parse_excel_date(ws.cell(row=r, column=col_fecha).value)
            payment_key = dt.date().isoformat()
            credito = _parse_money(ws.cell(row=r, column=col_credito).value)
        except Exception as exc:
            ws.cell(row=r, column=col_validado).value = "NO"
            ws.cell(row=r, column=col_estado).value = "ERROR"
            ws.cell(row=r, column=col_observaciones).value = str(exc)
            ws.cell(row=r, column=col_rutas).value = ""
            err += 1
            processed += 1
            msg = str(exc)
            errors.append({"row": r, "concepto": str(concepto_raw), "error": msg})
            logger.warning(
                "validate_payment_report: error parseando fila",
                extra={
                    "row": r,
                    "concepto": str(concepto_raw),
                    "error": msg,
                },
            )
            continue

        cache_key = (concepto_norm, payment_key)
        if cache_key not in total_cache:
            # Resolver carpeta cliente (exacto, luego fallback contains)
            client_folder = client_by_norm.get(concepto_norm)
            if not client_folder:
                candidates = [real for nk, real in client_by_norm.items() if concepto_norm in nk]
                if len(candidates) == 1:
                    client_folder = candidates[0]
                elif len(candidates) > 1:
                    total_cache[cache_key] = CachedClientResult.fail(
                        "Carpeta cliente ambigua (varias coincidencias)"
                    )
                else:
                    total_cache[cache_key] = CachedClientResult.fail(
                        "No existe carpeta del cliente (Concepto)"
                    )

            if client_folder:
                try:
                    client_path = f"{clients_base}/{client_folder}"
                    credit_folders = await _graph_children_by_path(
                        graph, ctx["site_id"], ctx["drive_id"], client_path
                    )
                    credit_names = [
                        str(it.get("name", "")).strip()
                        for it in credit_folders
                        if "folder" in it and str(it.get("name", "")).strip()
                    ]
                    if not credit_names:
                        total_cache[cache_key] = CachedClientResult.fail(
                            "Cliente sin subcarpetas de crédito"
                        )
                    else:
                        deduped_by_folder: list[
                            tuple[str, str, list[tuple[datetime, str, str, dict[str, Any]]]]
                        ] = []
                        for credit_folder in credit_names:
                            credit_path = f"{client_path}/{credit_folder}"
                            items = await _graph_children_by_path(
                                graph, ctx["site_id"], ctx["drive_id"], credit_path
                            )
                            all_raw = _collect_all_extract_candidates(
                                items, extract_keyword, credit_path
                            )
                            deduped = _dedupe_extract_candidates(all_raw)
                            if deduped:
                                deduped_by_folder.append((credit_folder, credit_path, deduped))

                        all_dates: set[datetime.date] = set()
                        for _, _, deduped in deduped_by_folder:
                            for c in deduped:
                                all_dates.add(c[0].date())

                        anchor_d, anchor_note = _resolve_anchor_date(all_dates, dt)
                        if anchor_d is None:
                            total_cache[cache_key] = CachedClientResult.fail(
                                "No hay fechas de extracto reconocibles en las carpetas del cliente"
                            )
                        else:
                            folder_rows: list[
                                tuple[str, Decimal, list[str], list[str]]
                            ] = []
                            for credit_folder, credit_path, deduped in deduped_by_folder:
                                selected = [c for c in deduped if c[0].date() == anchor_d]
                                if not selected:
                                    continue
                                per_lines: list[str] = []
                                paths_credit: list[str] = []
                                folder_sum = Decimal("0")
                                for chosen_dt, pdf_name, pdf_path, chosen_item in selected:
                                    item_id = str(chosen_item.get("id", "") or "").strip()
                                    try:
                                        if item_id:
                                            pdf_bytes = await _graph_download_by_item_id(
                                                graph, ctx["drive_id"], item_id
                                            )
                                        else:
                                            pdf_bytes = await _graph_download_by_path(
                                                graph, ctx["site_id"], ctx["drive_id"], pdf_path
                                            )
                                        amt = _extract_total_a_pagar_from_pdf(pdf_bytes)
                                        folder_sum += amt
                                        per_lines.append(
                                            f"{pdf_name} (fecha nombre "
                                            f"{_format_date_short(chosen_dt)}): TOTAL A PAGAR "
                                            f"{_format_money(amt)} [carpeta {credit_folder}]"
                                        )
                                        paths_credit.append(pdf_path)
                                    except Exception as pdf_exc:
                                        logger.warning(
                                            "validate_payment_report: error leyendo PDF %s: %s",
                                            pdf_name,
                                            pdf_exc,
                                        )
                                if per_lines:
                                    folder_rows.append(
                                        (credit_folder, folder_sum, per_lines, paths_credit)
                                    )

                            if not folder_rows:
                                total_cache[cache_key] = CachedClientResult.fail(
                                    f"No hay extractos con fecha ancla {anchor_d} en las carpetas de crédito"
                                )
                            else:
                                matched_sum, idxs, is_exact = _find_best_subset_match(
                                    folder_rows, credito, match_tol
                                )
                                merged_lines = tuple(
                                    line for i in idxs for line in folder_rows[i][2]
                                )
                                analyzed_paths = [p for i in idxs for p in folder_rows[i][3]]
                                ref_dt = datetime.combine(anchor_d, time.min)
                                is_adelantado = dt.date() < ref_dt.date()
                                diff_days = (anchor_d - dt.date()).days
                                folders_matched = [folder_rows[i][0] for i in idxs]
                                chosen_name = ", ".join(folders_matched)
                                chosen_path = folder_rows[idxs[0]][3][0] if analyzed_paths else ""
                                criterio = (
                                    f"{anchor_note} "
                                    f"Carpetas seleccionadas: {', '.join(folders_matched)}. "
                                    + (
                                        f"Suma TOTAL A PAGAR: {_format_money(matched_sum)}. "
                                        f"Validacion: SI si Crédito >= esa suma. "
                                        f"Empareje numérico: "
                                        + (
                                            "coincide con Crédito (tolerancia)."
                                            if is_exact
                                            else (
                                                f"mejor acercamiento suma {_format_money(matched_sum)} "
                                                f"vs Crédito {_format_money(credito)}."
                                            )
                                        )
                                    )
                                )
                                obs = _build_observation_text(
                                    matched_sum,
                                    credito,
                                    dt,
                                    ref_dt,
                                    diff_days,
                                    merged_lines,
                                    criterio,
                                )
                                logger.info(
                                    "validate_payment_report: empareje cliente/fecha de pago",
                                    extra={
                                        "concepto_norm": concepto_norm,
                                        "payment_date": payment_key,
                                        "anchor": str(anchor_d),
                                        "matched_sum": str(matched_sum),
                                        "match_exact": is_exact,
                                        "folders": folders_matched,
                                    },
                                )
                                total_cache[cache_key] = CachedClientResult(
                                    total=matched_sum,
                                    selected_date=ref_dt,
                                    selected_file_name=chosen_name,
                                    selected_path=chosen_path,
                                    analyzed_paths=tuple(analyzed_paths),
                                    observation=obs,
                                    error=None,
                                    is_adelantado=is_adelantado,
                                )
                except Exception as exc:
                    msg = str(exc)
                    total_cache[cache_key] = CachedClientResult.fail(msg)
                    # El mensaje va también en el texto del log: muchas consolas no muestran `extra`.
                    logger.warning(
                        "validate_payment_report: error calculando total para cliente/fecha "
                        "(concepto=%s, fecha_pago=%s): %s",
                        concepto_norm,
                        payment_key,
                        msg,
                        exc_info=logger.isEnabledFor(logging.DEBUG),
                    )

        cached = total_cache[cache_key]
        total, total_err = cached.total, cached.error
        if total_err:
            ws.cell(row=r, column=col_validado).value = "NO"
            ws.cell(row=r, column=col_estado).value = "ERROR"
            ws.cell(row=r, column=col_observaciones).value = total_err
            ws.cell(row=r, column=col_rutas).value = ", ".join(cached.analyzed_paths)
            err += 1
            processed += 1
            errors.append(
                {
                    "row": r,
                    "concepto": str(concepto_raw),
                    "fecha_pago": payment_key,
                    "error": total_err,
                }
            )
            continue

        if credito >= total:
            ws.cell(row=r, column=col_validado).value = "SI"
            ws.cell(row=r, column=col_observaciones).value = cached.observation
            ok += 1
        else:
            ws.cell(row=r, column=col_validado).value = "NO"
            ws.cell(row=r, column=col_observaciones).value = cached.observation
            no += 1
        ws.cell(row=r, column=col_estado).value = (
            "REVISION" if cached.is_adelantado else "OK"
        )
        ws.cell(row=r, column=col_rutas).value = ", ".join(cached.analyzed_paths)
        processed += 1

    out = BytesIO()
    wb.save(out)
    out_bytes = out.getvalue()

    logger.info(
        "validate_payment_report: subiendo Excel actualizado a SharePoint",
        extra={"site_id": ctx["site_id"], "drive_id": ctx["drive_id"], "path": ctx["file_path"]},
    )
    await graph.put_bytes(
        f"/sites/{ctx['site_id']}/drives/{ctx['drive_id']}/root:/{ctx['path_encoded']}:/content",
        out_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    return ValidationSummary(
        processed_rows=processed,
        ok_count=ok,
        no_count=no,
        error_count=err,
        errors=errors,
    )

