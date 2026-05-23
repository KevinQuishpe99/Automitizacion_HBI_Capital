"""
Dry-run: plan de llenado de tablas de amortización sin escribir en SharePoint.
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
from datetime import date, datetime
from typing import Any
from urllib.parse import unquote

import httpx
import openpyxl
from openpyxl.cell.cell import Cell

from app.application.services.accounting_pdf_parser import (
    AccountingParseError,
    PdfTextNotExtractableError,
    extract_text_from_pdf,
    parse_accounting_text,
)
from app.application.services.amortization_workbook import (
    ADOPTADO_EXISTENTE,
    APLICADO,
    REVISION_MANUAL,
    AmortizationSheetNotFoundError,
    build_amortization_idempotency_key,
    build_application_row_search_debug,
    build_target_row_search_debug,
    detect_amortization_sheet,
    find_application_row_detailed,
    find_row_by_due_date_detailed,
)
from app.application.services.ibr_workbook import find_ibr_for_date
from app.application.services.review_schema import DistribucionCols, normalize_credito_digits
from app.application.sharepoint_resolution import encode_graph_drive_path, resolve_sharepoint_path
from app.application.use_cases.send_validar_extractos_notification import (
    _find_distribucion_header_row,
    _find_distribucion_sheet,
    _get_col_distrib,
)
from app.application.use_cases.setup_ibr_workbook import ibr_workbook_relative_path
from app.application.use_cases.merge_composite_validado_pdfs import (
    _credit_number_from_path_scan,
    normalize_sharepoint_path,
)
from app.application.services.amortization_apply_safety import compute_asiento_pdf_hash
from app.application.use_cases.validate_payment_report import (
    _graph_download_by_path,
    _graph_get_item_metadata_by_path,
)
from app.domain.exceptions import GraphConfigError
from app.domain.ports.graph import GraphApiPort

logger = logging.getLogger(__name__)

TABLE_PATH_NOT_FOUND = "TABLE_PATH_NOT_FOUND"
PDF_TEXT_NOT_EXTRACTABLE = "PDF_TEXT_NOT_EXTRACTABLE"

_STATUS_MAP = {
    APLICADO: "WOULD_APPLY",
    ADOPTADO_EXISTENTE: "WOULD_ADOPT_EXISTING",
    REVISION_MANUAL: "REVISION_MANUAL",
}


def _logs_folder_relative() -> str:
    p = os.getenv("GRAPH_PAYMENT_VALIDATION_LOGS_PATH", "").strip().rstrip("/")
    if p:
        return p
    return "INFORMACION CREDITOS-CLIENTES/02 COMWARE - VALIDACION PAGOS/04 LOGS"


def _hyperlink_target(cell: Cell | None) -> str | None:
    if cell is None:
        return None
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return None
    t = getattr(hl, "target", None) or getattr(hl, "ref", None)
    return str(t).strip() if t else None


def _extract_internal_path_from_url(url: str) -> str | None:
    decoded = unquote(str(url))
    if "root:/" not in decoded:
        anchor = os.getenv("GRAPH_LINK_EXTRACTO_PATH_ANCHOR", "").strip()
        if anchor and anchor in decoded:
            idx = decoded.find(anchor)
            return decoded[idx:].split("?")[0].strip().strip("/")
        return None
    inner = decoded.split("root:/", 1)[1]
    inner = inner.split(":/", 1)[0]
    path = unquote(inner).replace("\\", "/").strip().strip("/")
    return path or None


def _tabla_path_from_doc_aspx_url(url: str, ruta_unidad_credito: str) -> str | None:
    """Resuelve ruta relativa desde hipervínculo Doc.aspx?...&file=Tabla....xlsx y carpeta del crédito."""
    decoded = unquote(str(url))
    if "file=" not in decoded.lower():
        return None
    m = re.search(r"file=([^&]+)", decoded, flags=re.IGNORECASE)
    if not m:
        return None
    file_part = unquote(m.group(1).replace("+", " ")).strip().strip("/")
    if not file_part.lower().endswith((".xlsx", ".xlsm")):
        return None
    base = str(ruta_unidad_credito or "").strip().replace("\\", "/").strip("/")
    if not base:
        return None
    return f"{base}/{file_part}".replace("//", "/").strip("/")


def _resolve_tabla_path_for_row(
    ws: Any,
    row: int,
    *,
    col_ruta_tabla: int | None,
    col_link_tabla: int | None,
    col_ruta_unidad: int | None,
) -> str | None:
    """
    Prioridad: RutaTablaAmortizacion → link/hipervínculo → Doc.aspx+file+RutaUnidadCredito.
    """
    if col_ruta_tabla:
        raw = ws.cell(row, col_ruta_tabla).value
        if raw is not None and str(raw).strip():
            return str(raw).strip().replace("\\", "/").strip("/")

    link_cell = ws.cell(row, col_link_tabla) if col_link_tabla else None
    from_link = _path_from_link_cell(link_cell)
    if from_link:
        return from_link

    if link_cell and col_ruta_unidad:
        target = _hyperlink_target(link_cell)
        ruta_uc = str(ws.cell(row, col_ruta_unidad).value or "").strip().replace("\\", "/")
        if target and ruta_uc:
            doc_path = _tabla_path_from_doc_aspx_url(target, ruta_uc)
            if doc_path:
                return doc_path

    return None


def _path_from_link_cell(cell: Cell | None) -> str | None:
    if cell is None:
        return None
    raw = cell.value
    if raw is not None:
        s = str(raw).strip().replace("\\", "/")
        if s and not s.lower().startswith("http") and ("/" in s) and s.lower().endswith((".xlsx", ".xlsm")):
            return s.strip("/")
    target = _hyperlink_target(cell)
    if target:
        internal = _extract_internal_path_from_url(target)
        if internal and internal.lower().endswith((".xlsx", ".xlsm")):
            return internal
        if not target.lower().startswith("http") and target.lower().endswith((".xlsx", ".xlsm")):
            return target.replace("\\", "/").strip("/")
    return None


def _parse_date_value(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


async def _drive_context(graph: GraphApiPort) -> tuple[str, str]:
    site_search = os.getenv("GRAPH_SHAREPOINT_SITE_SEARCH", "").strip()
    drive_name = os.getenv("GRAPH_SHAREPOINT_DRIVE_NAME", "").strip()
    if not site_search:
        raise GraphConfigError("Missing environment variable: GRAPH_SHAREPOINT_SITE_SEARCH")
    anchor = (
        os.getenv("GRAPH_PAYMENT_VALIDATION_CONTROL_PATH", "").strip()
        or _logs_folder_relative()
    )
    info = await resolve_sharepoint_path(graph, site_search, drive_name, anchor)
    return info["site_id"], info["drive_id"]


async def _list_folder_json_manifests(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    folder: str,
) -> list[str]:
    enc = encode_graph_drive_path(folder.strip().strip("/"))
    endpoint = f"/sites/{site_id}/drives/{drive_id}/root:/{enc}:/children"
    try:
        resp = await graph.get(endpoint)
    except httpx.HTTPStatusError:
        return []
    names: list[str] = []
    for it in resp.get("value") or []:
        name = str(it.get("name", ""))
        if name.startswith("merge_manifest_") and name.endswith(".json"):
            names.append(f"{folder.strip().strip('/')}/{name}".replace("//", "/"))
    return sorted(names, reverse=True)


async def _resolve_manifest_rel_path(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    *,
    report_date_iso: str | None,
    merge_manifest_path: str | None,
) -> str:
    if merge_manifest_path and str(merge_manifest_path).strip():
        return str(merge_manifest_path).strip().strip("/")
    if report_date_iso and str(report_date_iso).strip():
        return f"{_logs_folder_relative()}/merge_manifest_{report_date_iso.strip()}.json"
    manifests = await _list_folder_json_manifests(graph, site_id, drive_id, _logs_folder_relative())
    if manifests:
        return manifests[0]
    raise ValueError("report_date_iso_required")


def _load_historical_index(hist_bytes: bytes) -> dict[tuple[str, str], dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(hist_bytes), data_only=True)
    try:
        ws = _find_distribucion_sheet(wb)
        h_row, header_map = _find_distribucion_header_row(ws)
        col_id = _get_col_distrib(header_map, "ID Pago", "ID pago", "ID PAGO")
        col_cred = _get_col_distrib(header_map, "Crédito", "Credito", "CREDITO")
        col_cred_norm = _get_col_distrib(
            header_map,
            DistribucionCols.CREDITO_NORMALIZADO,
            "Credito Normalizado",
            "CREDITO NORMALIZADO",
        )
        col_lim = _get_col_distrib(
            header_map,
            DistribucionCols.FECHA_LIMITE,
            "Fecha limite",
            "FECHA LIMITE",
        )
        col_tabla = _get_col_distrib(
            header_map,
            DistribucionCols.LINK_TABLA,
            "Link tabla amortización",
            "Link tabla",
            "TABLA AMORTIZACION",
        )
        col_ruta_tabla = _get_col_distrib(
            header_map,
            DistribucionCols.RUTA_TABLA_AMORTIZACION,
            "Ruta Tabla Amortizacion",
            "RUTA TABLA AMORTIZACION",
        )
        col_ruta_uc = _get_col_distrib(
            header_map,
            DistribucionCols.RUTA_UNIDAD_CREDITO,
            "RutaUnidadCredito",
            "RUTA UNIDAD CREDITO",
        )
        if not col_id or not col_cred:
            return {}

        index: dict[tuple[str, str], dict[str, Any]] = {}
        for r in range(h_row + 1, (ws.max_row or h_row) + 1):
            id_p = str(ws.cell(r, col_id).value or "").strip()
            cred_visible = str(ws.cell(r, col_cred).value or "").strip()
            if not id_p:
                continue
            cred_norm = ""
            if col_cred_norm:
                cred_norm = str(ws.cell(r, col_cred_norm).value or "").strip()
            if not cred_norm:
                cred_norm = normalize_credito_digits(cred_visible)
            lookup_cred = cred_norm or cred_visible
            fecha_lim = _parse_date_value(ws.cell(r, col_lim).value) if col_lim else None
            tabla_path = _resolve_tabla_path_for_row(
                ws,
                r,
                col_ruta_tabla=col_ruta_tabla,
                col_link_tabla=col_tabla,
                col_ruta_unidad=col_ruta_uc,
            )
            row_data = {
                "fecha_limite_pago": fecha_lim,
                "tabla_amortizacion_path": tabla_path,
                "credito_normalizado": cred_norm,
                "credito_visible": cred_visible,
                "row": r,
            }
            index[(id_p, lookup_cred)] = row_data
            if cred_visible and cred_visible != lookup_cred:
                index.setdefault((id_p, cred_visible), row_data)
        return index
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


def _plan_ibr_block(
    *,
    ibr_bytes: bytes | None,
    fecha_limite: date,
    ibr_plan_key: str,
    planned_ibr_keys: set[str],
) -> dict[str, Any]:
    """IBR en due_date_row: una sola escritura planificada por tabla + fecha límite."""
    ibr_value = None
    ibr_found = False
    if ibr_bytes:
        ibr_value = find_ibr_for_date(ibr_bytes, fecha_limite)
        if ibr_value is not None:
            ibr_found = True

    if ibr_plan_key in planned_ibr_keys:
        return {
            "required_date": fecha_limite.isoformat(),
            "found": ibr_found,
            "value": ibr_value,
            "status": "WOULD_SKIP_IBR_ALREADY_PLANNED",
        }

    if ibr_found:
        planned_ibr_keys.add(ibr_plan_key)
        return {
            "required_date": fecha_limite.isoformat(),
            "found": True,
            "value": ibr_value,
            "status": "WOULD_WRITE_IBR",
        }

    return {
        "required_date": fecha_limite.isoformat(),
        "found": False,
        "value": None,
        "status": "PENDING_IBR",
    }


def _sheet_event_meta(ws: Any | None, event: Any | None = None) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    if ws is not None:
        meta["sheet_name"] = str(getattr(ws, "title", "") or "") or None
    if event is not None and getattr(event, "fecha_asiento", None):
        meta["fecha_asiento"] = event.fecha_asiento.isoformat()
    return meta


def _payment_date_meta(payment_date_iso: str | None) -> dict[str, str]:
    raw = str(payment_date_iso or "").strip()
    return {"payment_date_iso": raw} if raw else {}


def _payment_application_dict(event: Any) -> dict[str, float]:
    return {
        "valor_pagado_cliente": event.valor_pagado_cliente,
        "capital": event.capital,
        "intereses": event.intereses,
        "mora": event.mora,
        "retenciones": event.retenciones,
        "saldos_menores": event.saldos_menores,
    }


def _infer_credit_from_asiento_path(asiento_path: str) -> str:
    """Infiere dígitos de crédito desde ruta/nombre de asiento (fallback legacy)."""
    return _credit_number_from_path_scan(asiento_path) or ""


def _credit_digits_match(a: str, b: str) -> bool:
    da = re.sub(r"\D", "", str(a or ""))
    db = re.sub(r"\D", "", str(b or ""))
    if da and db:
        return da == db
    return str(a or "").strip().casefold() == str(b or "").strip().casefold()


def _find_hist_row(
    hist_index: dict[tuple[str, str], dict[str, Any]],
    id_pago: str,
    event_credit: str,
) -> dict[str, Any] | None:
    """Busca fila histórica por ID Pago + crédito individual (CreditoNormalizado o visible)."""
    event_credit = str(event_credit or "").strip()
    digits = normalize_credito_digits(event_credit) or re.sub(r"\D", "", event_credit)
    if not digits:
        digits = event_credit

    for key in ((id_pago, event_credit), (id_pago, digits)):
        if key in hist_index:
            return hist_index[key]

    for (hid, hcred), row in hist_index.items():
        if hid != id_pago:
            continue
        row_norm = str(row.get("credito_normalizado") or "").strip()
        if row_norm and row_norm == digits:
            return row
        if _credit_digits_match(hcred, digits) or _credit_digits_match(hcred, event_credit):
            return row
        norm = normalize_credito_digits(hcred)
        if norm and norm == digits:
            return row
    return None


def _resolve_asiento_paths_from_output(output: dict[str, Any]) -> list[str]:
    raw_paths = output.get("asiento_pdf_paths")
    if isinstance(raw_paths, list):
        paths = [str(p).strip().strip("/") for p in raw_paths if str(p).strip()]
        if paths:
            return paths
    single = str(output.get("asiento_pdf_path") or "").strip()
    if not single:
        return []
    if " | " in single:
        return [p.strip().strip("/") for p in single.split(" | ") if p.strip()]
    return [single.strip("/")]


def _empty_item(
    *,
    id_pago: str,
    cliente: str,
    credito: str,
    asiento_pdf_path: str,
    application_status: str,
    error_code: str,
    warnings: list[str],
    event_index: int = 1,
    extracto_pdf_path: str | None = None,
    payment_date_iso: str | None = None,
) -> dict[str, Any]:
    item = {
        "id_pago": id_pago,
        "cliente": cliente,
        "credito": credito,
        "asiento_pdf_path": asiento_pdf_path,
        "extracto_pdf_path": extracto_pdf_path,
        "event_index": event_index,
        "idempotency_key": build_amortization_idempotency_key(
            id_pago, credito, asiento_pdf_path, ""
        ),
        "comprobante": None,
        "tabla_amortizacion_path": None,
        "fecha_limite_pago": None,
        "payment_application": None,
        "due_date_row": None,
        "ibr_row": None,
        "application_row": None,
        "target_row": None,
        "application_status": application_status,
        "ibr": {
            "required_date": None,
            "found": False,
            "value": None,
            "status": "PENDING_IBR",
        },
        "warnings": warnings,
        "error_code": error_code,
    }
    item.update(_payment_date_meta(payment_date_iso))
    return item


def _summarize(items: list[dict[str, Any]]) -> dict[str, int]:
    """Cuenta eventos (un asiento PDF = un evento), no solo outputs del manifest."""
    summary = {
        "total_events": len(items),
        "total": len(items),
        "would_apply": 0,
        "would_adopt_existing": 0,
        "pending_ibr": 0,
        "revision_manual": 0,
        "errors": 0,
    }
    for it in items:
        st = it.get("application_status")
        if st == "WOULD_APPLY":
            summary["would_apply"] += 1
        elif st == "WOULD_ADOPT_EXISTING":
            summary["would_adopt_existing"] += 1
        elif st == "REVISION_MANUAL":
            summary["revision_manual"] += 1
        elif st == "ERROR":
            summary["errors"] += 1
        ibr = it.get("ibr") or {}
        if ibr.get("status") == "PENDING_IBR" and st != "ERROR":
            summary["pending_ibr"] += 1
    return summary


async def _plan_one_asiento_event(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    *,
    id_pago: str,
    cliente: str,
    credito: str,
    asiento_path: str,
    event_index: int,
    hist_index: dict[tuple[str, str], dict[str, Any]],
    ibr_bytes: bytes | None,
    used_application_rows_by_table: dict[str, set[int]],
    planned_ibr_keys: set[str],
    extracto_pdf_path: str | None = None,
    payment_date_iso: str | None = None,
) -> dict[str, Any]:
    warnings: list[str] = []
    payment_meta = _payment_date_meta(payment_date_iso)

    if not asiento_path:
        return _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path="",
            event_index=event_index,
            application_status="ERROR",
            error_code="ASIENTO_PATH_MISSING",
            warnings=["asiento_pdf_path vacío en manifest"],
        )

    hist_row = _find_hist_row(hist_index, id_pago, credito)
    tabla_path = (hist_row or {}).get("tabla_amortizacion_path")
    fecha_limite = (hist_row or {}).get("fecha_limite_pago")
    if not tabla_path:
        return _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path=asiento_path,
            event_index=event_index,
            application_status="ERROR",
            error_code=TABLE_PATH_NOT_FOUND,
            warnings=["No se encontró Link tabla amortización en histórico para ID Pago + Crédito"],
        )

    pdf_fingerprint: dict[str, Any] = {}
    try:
        pdf_bytes = await _graph_download_by_path(graph, site_id, drive_id, asiento_path)
        pdf_meta = await _graph_get_item_metadata_by_path(
            graph, site_id, drive_id, asiento_path
        )
        pdf_fingerprint = {
            "asiento_pdf_hash": compute_asiento_pdf_hash(pdf_bytes),
            "asiento_pdf_size": len(pdf_bytes),
            "asiento_pdf_etag": str(pdf_meta.get("eTag") or pdf_meta.get("etag") or ""),
            "asiento_pdf_last_modified": pdf_meta.get("lastModifiedDateTime"),
        }
        text = extract_text_from_pdf(pdf_bytes)
    except PdfTextNotExtractableError as exc:
        item = _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path=asiento_path,
            event_index=event_index,
            application_status="ERROR",
            error_code=PDF_TEXT_NOT_EXTRACTABLE,
            warnings=[str(exc)],
        )
        item.update(pdf_fingerprint)
        item.update(payment_meta)
        return item
    except Exception as exc:
        item = _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path=asiento_path,
            event_index=event_index,
            application_status="ERROR",
            error_code="ASIENTO_DOWNLOAD_FAILED",
            warnings=[str(exc)[:500]],
        )
        item.update(pdf_fingerprint)
        item.update(payment_meta)
        return item

    try:
        event = parse_accounting_text(
            text,
            {
                "id_pago": id_pago,
                "cliente": cliente,
                "credito": credito,
                "asiento_pdf_path": asiento_path,
            },
        )
    except AccountingParseError as exc:
        err_warnings = [str(exc)]
        if exc.detected_codes:
            err_warnings.append(f"códigos detectados: {', '.join(exc.detected_codes)}")
        if exc.amounts_before_code:
            err_warnings.append("se encontraron montos antes del código contable")
        if getattr(exc, "parser_mode", ""):
            err_warnings.append(f"parser_mode={exc.parser_mode}")
        item = _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path=asiento_path,
            event_index=event_index,
            application_status="ERROR",
            error_code=getattr(exc, "error_code", None) or "ACCOUNTING_PARSE_FAILED",
            warnings=err_warnings,
        )
        if exc.detected_codes:
            item["detected_codes"] = list(exc.detected_codes)
        if getattr(exc, "parser_mode", ""):
            item["parser_mode"] = exc.parser_mode
        item.update(pdf_fingerprint)
        item.update(payment_meta)
        return item

    if event.parse_warnings:
        warnings.extend(event.parse_warnings)
    parser_meta = {
        "detected_codes": list(event.detected_codes),
        "parser_mode": event.parser_mode or None,
    }

    try:
        tabla_bytes = await _graph_download_by_path(graph, site_id, drive_id, tabla_path)
    except Exception as exc:
        return _empty_item(
            id_pago=id_pago,
            cliente=cliente,
            credito=credito,
            asiento_pdf_path=asiento_path,
            event_index=event_index,
            application_status="ERROR",
            error_code="TABLE_DOWNLOAD_FAILED",
            warnings=[str(exc)[:500]],
        )

    # data_only=True: leer valores calculados de fórmulas (solo lectura en dry-run).
    # Apply futuro debe usar data_only=True para detectar y data_only=False para escribir.
    wb = openpyxl.load_workbook(io.BytesIO(tabla_bytes), data_only=True)
    try:
        try:
            sheet_match = detect_amortization_sheet(
                wb, tabla_amortizacion_path=tabla_path
            )
        except AmortizationSheetNotFoundError as exc:
            sheet_warnings = [
                "No se pudo detectar la hoja de tabla de amortización por encabezados",
                f"tabla_amortizacion_path={tabla_path}",
                f"workbook_sheets={exc.workbook_sheets}",
            ]
            if exc.required_headers_missing:
                sheet_warnings.append(
                    f"required_headers_missing={exc.required_headers_missing}"
                )
            if exc.headers_detected_by_sheet:
                sheet_warnings.append(
                    f"headers_detected_by_sheet={exc.headers_detected_by_sheet}"
                )
            item = _empty_item(
                id_pago=id_pago,
                cliente=cliente,
                credito=credito,
                asiento_pdf_path=asiento_path,
                event_index=event_index,
                application_status="ERROR",
                error_code="AMORTIZATION_SHEET_NOT_FOUND",
                warnings=sheet_warnings,
            )
            item.update(parser_meta)
            item.update(pdf_fingerprint)
            item.update(payment_meta)
            return item

        ws = sheet_match.worksheet
        headers = sheet_match.headers
        header_row = sheet_match.header_row
        if not fecha_limite:
            return {
                "id_pago": id_pago,
                "cliente": cliente,
                "credito": credito,
                "asiento_pdf_path": asiento_path,
                "event_index": event_index,
                "idempotency_key": build_amortization_idempotency_key(
                    id_pago, credito, asiento_path, event.comprobante
                ),
                "comprobante": event.comprobante or None,
                "tabla_amortizacion_path": tabla_path,
                "fecha_limite_pago": None,
                "payment_application": _payment_application_dict(event),
                "due_date_row": None,
                "ibr_row": None,
                "application_row": None,
                "target_row": None,
                "application_status": "ERROR",
                "ibr": {
                    "required_date": None,
                    "found": False,
                    "value": None,
                    "status": "PENDING_IBR",
                },
                "warnings": ["fecha_limite_pago no disponible en histórico"],
                "error_code": "FECHA_LIMITE_NOT_FOUND",
                **parser_meta,
                **pdf_fingerprint,
                **payment_meta,
            }

        tabla_key = normalize_sharepoint_path(tabla_path)
        reserved_application_rows = used_application_rows_by_table.setdefault(tabla_key, set())
        ibr_plan_key = f"{tabla_key}|{fecha_limite.isoformat()}"

        find_due_result = find_row_by_due_date_detailed(
            ws,
            headers,
            fecha_limite,
            header_row=header_row,
        )
        due_date_row = find_due_result.row
        ibr_row = due_date_row
        if due_date_row is None:
            row_warnings = ["No hay fila de cronograma (dia/mes/año) para la fecha límite"]
            debug = build_target_row_search_debug(
                ws,
                headers,
                fecha_limite,
                header_row=header_row,
                sheet_name=ws.title,
                tabla_amortizacion_path=tabla_path,
                workbook_loaded_data_only=True,
                find_result=find_due_result,
            )
            row_warnings.append(f"due_date_row_debug={debug}")
            return {
                "id_pago": id_pago,
                "cliente": cliente,
                "credito": credito,
                "asiento_pdf_path": asiento_path,
                "event_index": event_index,
                "idempotency_key": build_amortization_idempotency_key(
                    id_pago, credito, asiento_path, event.comprobante
                ),
                "comprobante": event.comprobante or None,
                "tabla_amortizacion_path": tabla_path,
                "fecha_limite_pago": fecha_limite.isoformat(),
                "payment_application": _payment_application_dict(event),
                "due_date_row": None,
                "ibr_row": None,
                "application_row": None,
                "target_row": None,
                "application_status": "ERROR",
                "ibr": {
                    "required_date": fecha_limite.isoformat(),
                    "found": False,
                    "value": None,
                    "status": "PENDING_IBR",
                },
                "warnings": row_warnings,
                "error_code": "DUE_DATE_ROW_NOT_FOUND",
                **_sheet_event_meta(ws, event),
                **parser_meta,
                **pdf_fingerprint,
                **payment_meta,
            }

        payment_date: date | None = None
        if payment_date_iso:
            try:
                payment_date = date.fromisoformat(str(payment_date_iso).strip())
            except ValueError:
                payment_date = None

        app_result = find_application_row_detailed(
            ws,
            headers,
            event,
            due_date_row=due_date_row,
            exclude_rows=frozenset(reserved_application_rows),
            header_row=header_row,
            payment_date=payment_date,
            detected_codes=frozenset(parser_meta.get("detected_codes") or []),
            warnings=frozenset(warnings),
        )
        application_row = app_result.row
        if application_row is None:
            app_warnings = ["No hay fila libre en el bloque Aplicación del Pago"]
            if reserved_application_rows:
                app_warnings.append(
                    "Filas de aplicación ya reservadas por asientos previos del mismo crédito: "
                    f"{sorted(reserved_application_rows)}"
                )
            if app_result.requires_new_row and app_result.suggested_row:
                app_warnings.append(
                    f"Se requiere ampliar el bloque de aplicación (fila sugerida: {app_result.suggested_row})"
                )
            app_debug = build_application_row_search_debug(
                ws,
                headers,
                event,
                due_date_row=due_date_row,
                header_row=header_row,
                sheet_name=ws.title,
                tabla_amortizacion_path=tabla_path,
                exclude_rows=frozenset(reserved_application_rows),
                find_result=app_result,
            )
            app_warnings.append(f"application_row_debug={app_debug}")
            error_code = (
                "REQUIRES_APPLICATION_ROW"
                if app_result.requires_new_row
                else "APPLICATION_ROW_NOT_FOUND"
            )
            return {
                "id_pago": id_pago,
                "cliente": cliente,
                "credito": credito,
                "asiento_pdf_path": asiento_path,
                "event_index": event_index,
                "idempotency_key": build_amortization_idempotency_key(
                    id_pago, credito, asiento_path, event.comprobante
                ),
                "comprobante": event.comprobante or None,
                "tabla_amortizacion_path": tabla_path,
                "fecha_limite_pago": fecha_limite.isoformat(),
                "payment_application": _payment_application_dict(event),
                "due_date_row": due_date_row,
                "ibr_row": ibr_row,
                "application_row": None,
                "target_row": None,
                "application_status": "ERROR",
                "ibr": _plan_ibr_block(
                    ibr_bytes=ibr_bytes,
                    fecha_limite=fecha_limite,
                    ibr_plan_key=ibr_plan_key,
                    planned_ibr_keys=planned_ibr_keys,
                ),
                "warnings": warnings + app_warnings,
                "error_code": error_code,
                **_sheet_event_meta(ws, event),
                **parser_meta,
                **pdf_fingerprint,
                **payment_meta,
            }

        compare_status = app_result.compare_status or APLICADO
        application_status = _STATUS_MAP.get(compare_status, "ERROR")
        if compare_status == APLICADO:
            reserved_application_rows.add(application_row)

        ibr_block = _plan_ibr_block(
            ibr_bytes=ibr_bytes,
            fecha_limite=fecha_limite,
            ibr_plan_key=ibr_plan_key,
            planned_ibr_keys=planned_ibr_keys,
        )

        return {
            "id_pago": id_pago,
            "cliente": cliente,
            "credito": credito,
            "asiento_pdf_path": asiento_path,
            "extracto_pdf_path": extracto_pdf_path,
            "event_index": event_index,
            "idempotency_key": build_amortization_idempotency_key(
                id_pago, credito, asiento_path, event.comprobante
            ),
            "comprobante": event.comprobante or None,
            "tabla_amortizacion_path": tabla_path,
            "fecha_limite_pago": fecha_limite.isoformat(),
            "payment_application": _payment_application_dict(event),
            "due_date_row": due_date_row,
            "ibr_row": ibr_row,
            "application_row": application_row,
            "target_row": application_row,
            "application_status": application_status,
            "ibr": ibr_block,
            "warnings": warnings,
            "error_code": None,
            **_sheet_event_meta(ws, event),
            **parser_meta,
            **pdf_fingerprint,
            **payment_meta,
        }
    finally:
        closer = getattr(wb, "close", None)
        if callable(closer):
            closer()


async def _plan_events_for_manifest_output(
    graph: GraphApiPort,
    site_id: str,
    drive_id: str,
    output: dict[str, Any],
    hist_index: dict[tuple[str, str], dict[str, Any]],
    ibr_bytes: bytes | None,
    used_application_rows_by_table: dict[str, set[int]],
    planned_ibr_keys: set[str],
    payment_date_iso: str | None = None,
) -> list[dict[str, Any]]:
    id_pago = str(output.get("id_pago") or "").strip()
    cliente = str(output.get("cliente") or "").strip()
    legacy_credito = str(output.get("credito") or "").strip()

    credit_items = output.get("credit_items")
    event_specs: list[tuple[str, str, str | None]] = []

    if isinstance(credit_items, list) and credit_items:
        for ci in credit_items:
            if not isinstance(ci, dict):
                continue
            event_credit = str(ci.get("credito") or "").strip()
            extracto = str(ci.get("extracto_pdf_path") or "").strip() or None
            if not extracto:
                eps = ci.get("extracto_pdf_paths") or []
                if isinstance(eps, list) and eps:
                    extracto = str(eps[0]).strip() or None
            for asiento_path in ci.get("asiento_pdf_paths") or []:
                ap = str(asiento_path).strip().strip("/")
                if ap:
                    event_specs.append((event_credit, ap, extracto))
    else:
        for asiento_path in _resolve_asiento_paths_from_output(output):
            inferred = _infer_credit_from_asiento_path(asiento_path)
            event_credit = inferred or legacy_credito
            event_specs.append((event_credit, asiento_path, None))

    if not event_specs:
        return [
            _empty_item(
                id_pago=id_pago,
                cliente=cliente,
                credito=legacy_credito,
                asiento_pdf_path="",
                event_index=1,
                application_status="ERROR",
                error_code="ASIENTO_PATH_MISSING",
                warnings=["asiento_pdf_path / asiento_pdf_paths / credit_items vacío en manifest"],
                payment_date_iso=payment_date_iso,
            )
        ]

    events: list[dict[str, Any]] = []
    for event_index, (event_credit, asiento_path, extracto_path) in enumerate(event_specs, start=1):
        item = await _plan_one_asiento_event(
            graph,
            site_id,
            drive_id,
            id_pago=id_pago,
            cliente=cliente,
            credito=event_credit,
            asiento_path=asiento_path,
            event_index=event_index,
            hist_index=hist_index,
            ibr_bytes=ibr_bytes,
            used_application_rows_by_table=used_application_rows_by_table,
            planned_ibr_keys=planned_ibr_keys,
            extracto_pdf_path=extracto_path,
            payment_date_iso=payment_date_iso,
        )
        events.append(item)
    return events


async def run_amortization_fill_dry_run(
    graph: GraphApiPort,
    *,
    report_date_iso: str | None = None,
    merge_manifest_path: str | None = None,
    historical_file_path: str | None = None,
) -> dict[str, Any]:
    """
    Simula el llenado de tablas de amortización. No escribe ni mueve archivos en SharePoint.
    """
    site_id, drive_id = await _drive_context(graph)
    manifest_rel = await _resolve_manifest_rel_path(
        graph,
        site_id,
        drive_id,
        report_date_iso=report_date_iso,
        merge_manifest_path=merge_manifest_path,
    )

    try:
        manifest_raw = await _graph_download_by_path(graph, site_id, drive_id, manifest_rel)
        manifest = json.loads(manifest_raw.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"invalid_merge_manifest_json|{manifest_rel}") from exc
    except httpx.HTTPStatusError as exc:
        code = exc.response.status_code if exc.response else 0
        if code == 404:
            raise ValueError(f"merge_manifest_not_found|{manifest_rel}") from exc
        raise

    effective_report_date = (
        str(report_date_iso or "").strip()
        or str(manifest.get("report_date_iso") or "").strip()
        or None
    )

    outputs = manifest.get("outputs") or []
    if not isinstance(outputs, list):
        outputs = []

    hist_index: dict[tuple[str, str], dict[str, Any]] = {}
    hist_path = (historical_file_path or manifest.get("historico_excel_path") or "").strip().strip("/")
    if hist_path:
        try:
            hist_bytes = await _graph_download_by_path(graph, site_id, drive_id, hist_path)
            hist_index = _load_historical_index(hist_bytes)
        except Exception as exc:
            logger.warning("amortization dry_run: histórico no legible %s: %s", hist_path, exc)

    ibr_bytes: bytes | None = None
    ibr_path = ibr_workbook_relative_path()
    try:
        ibr_bytes = await _graph_download_by_path(graph, site_id, drive_id, ibr_path)
    except Exception as exc:
        logger.warning("amortization dry_run: IBR no disponible %s: %s", ibr_path, exc)

    used_application_rows_by_table: dict[str, set[int]] = {}
    planned_ibr_keys: set[str] = set()
    items: list[dict[str, Any]] = []
    for out in outputs:
        if not isinstance(out, dict):
            continue
        items.extend(
            await _plan_events_for_manifest_output(
                graph,
                site_id,
                drive_id,
                out,
                hist_index,
                ibr_bytes,
                used_application_rows_by_table,
                planned_ibr_keys,
                payment_date_iso=effective_report_date,
            )
        )

    return {
        "status": "ok",
        "mode": "dry_run",
        "report_date_iso": effective_report_date,
        "manifest_path": manifest_rel,
        "historical_file_path": hist_path or None,
        "ibr_workbook_path": ibr_path,
        "manifest_outputs_count": len(outputs),
        "items": items,
        "summary": _summarize(items),
    }
